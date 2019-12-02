import sys
sys.path.append('/fenics/shared/local_lib')
#from okada_wrapper.okada_wrapper import dc3d0wrapper, dc3dwrapper
import petsc4py as p4p
from petsc4py.PETSc import Mat
from dolfin import *
from hippylib import *
import scipy.sparse as sp
from scipy.spatial.distance import pdist, squareform
import logging
import numpy as np
import matplotlib.pyplot as plt
import openpyxl as pyxl
import time
from block import *
from block.dolfin_util import *
from block.iterative import *
from block.algebraic.petsc import *
import time


parameters.reorder_dofs_serial = False

dir_path = '/fenics/shared/'

def u_boundary(x, on_boundary):
    return on_boundary

def GPSlatlon2XY(data_sheet, origin, theta):
    """
    latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
    into numerical models.

    INPUTS:
    [data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
    [origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
    [theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
                counter-clockwise from latlon to XY

    Assumes first column is longitude and second column is latitude

    OUTPUTS:
    Returns transformed coordinates as numpy vectors X, Y in kilometers
    """

    lon = np.array([[data_sheet.cell(row=i, column=1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    lat = np.array([[data_sheet.cell(row=i, column=2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )

    lon_u = np.array([[data_sheet.cell(row=i, column=5).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    lat_u = np.array([[data_sheet.cell(row=i, column=6).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    Uz = np.array([[data_sheet.cell(row=i, column=4).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )

    error_E = np.array([[data_sheet.cell(row=i, column=8).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    error_N = np.array([[data_sheet.cell(row=i, column=9).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    error_Z = np.array([[data_sheet.cell(row=i, column=7).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )

    lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
    lat_in_km = (lat - origin[1]) * 111

    rho_u = np.sqrt(np.power(lon_u, 2) + np.power(lat_u, 2))
    theta_new_u = np.arctan2(lat_u, lon_u) - theta

    rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
    theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

    X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)
    Ux, Uy = rho_u * np.cos(theta_new_u), rho_u * np.sin(theta_new_u)
    error_x = error_E * np.cos(theta) + error_N * np.sin(theta)
    error_y = error_E * np.sin(theta) + error_N * np.cos(theta)

    return 1e3 * X, 1e3 * Y, 1e-3 * Ux, 1e-3 * Uy, 1e-3 * Uz, 1e-3 * error_x, 1e-3 * error_y, 1e-3 * error_Z

def GPSlatlon2XY_time(lat_u, lon_u, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	rho_u = np.sqrt(np.power(lon_u, 2) + np.power(lat_u, 2))
	theta_new_u = np.arctan2(lat_u, lon_u) - theta

	Ux, Uy = rho_u * np.cos(theta_new_u), rho_u * np.sin(theta_new_u)

	return Ux, Uy

def latlon2XY_points(lat, lon, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""
 
	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)

	return 1e3 * X, 1e3 * Y

def checkerboard(coords_b, size):
	X, Y = coords_b[:,0], coords_b[:,1]
	check_slip = np.zeros(X.shape[0],)
	xblocks = (np.round(X.max()/ size)/2 + 1).astype(int)
	yblocks = (np.round(Y.max() / size)/2 + 1).astype(int)
	for i in range (0, xblocks):
		for j in range(0, yblocks):
			checkdof = np.where((X > 2*size*i) & (X < 2*size*i + size)
			                    & (Y > 2*size*j + size) & (Y < 2*size*j + 2*size))[0]

			check_slip[checkdof] = -1.0
	for i in range (0, xblocks):
		for j in range(0, yblocks):
			checkdof = np.where((X > 2*size*i +size) & (X < 2*size*i + 2*size)
			                    & (Y > 2*size*j) & (Y < 2*size*j + size))[0]

			check_slip[checkdof] = -1.0

	return check_slip

def importGPSdata(startday, endday, origin, theta, PE = True):

	####### load GPS data from text files
    path = "/fenics/shared/"
    with open(path + 'data/GD/station_list.txt', 'r') as f:
        stations = f.readlines()

    names = [x.split()[0] for x in stations]
    gps_lon = np.asarray([x.split()[1] for x in stations], dtype=np.float32)
    gps_lat = np.asarray([x.split()[2] for x in stations], dtype=np.float32)
    X_gps_post, Y_gps_post = latlon2XY_points(gps_lat, gps_lon, origin, theta)

    data = {}
    for name in names:
        with open(path + 'data/GD/' + name + '.lat', 'r') as latt:
            latread = latt.readlines()

        lat = np.asarray([x.split()[1] for x in latread], dtype=np.float32)
        time = np.asarray([x.split()[0] for x in latread], dtype=np.float32)
        t2_ind = np.where((time > 2012.5))[0][0] - 1
        slope = (np.average(lat[t2_ind - 5:t2_ind]) - np.average(lat[0:5])) / (time[t2_ind] - time[0])

        plate_motion = (time - time[0]) * slope
        latnorm = lat - plate_motion
        latnorm -= np.average(latnorm[0:t2_ind])

        data[name + '_time'] = time
        data[name + '_lat'] = latnorm

        with open(path + 'data/GD/' + name + '.lon', 'r') as lont:
            lonread = lont.readlines()

        lon = np.asarray([x.split()[1] for x in lonread], dtype=np.float32)
        t2_ind = np.where((time > 2012.5))[0][0] - 1
        slope = (np.average(lon[t2_ind - 5:t2_ind]) - np.average(lon[0:5])) / (time[t2_ind] - time[0])

        plate_motion = (time - time[0]) * slope
        lonnorm = lon - plate_motion
        lonnorm -= np.average(lonnorm[0:t2_ind])

        data[name + '_lon'] = lonnorm

        with open(path + 'data/GD/' + name + '.rad', 'r') as radt:
            radread = radt.readlines()

        rad = np.asarray([x.split()[1] for x in radread], dtype=np.float32)
        t2_ind = np.where((time > 2012.5))[0][0] - 1
        slope = (np.average(rad[t2_ind - 5:t2_ind]) - np.average(rad[0:5])) / (time[t2_ind] - time[0])

        plate_motion = (time - time[0]) * slope
        radnorm = rad - plate_motion
        radnorm -= np.average(radnorm[0:t2_ind])

        data[name + '_rad'] = radnorm

    """load in 3D results """
    results_path = '/fenics/shared/firedrake/results/numpy_results/'
    var_path = '/fenics/shared/firedrake/results/numpy_variables/'
    x_all = np.load(var_path + "x_all_3D.npy")
    y_all = np.load(var_path + "y_all_3D.npy")
    z_all = np.load(var_path + "z_all_3D.npy")
    surface_dofs = np.load(var_path + "surface_dofs_3D.npy")
    gps_dofs = np.load(var_path + "GPS_dofs_3D.npy")
    Sol_gps_mapped = np.load(results_path + "Sol_gps_3D_EQ_mapped.npy")
    Sol_surf_p3D = 1e3 / (9.81) * np.load(results_path + "Sol_surfp_3D_EQ_mapped.npy")
    x_surf_3D, y_surf_3D = x_all[surface_dofs], y_all[surface_dofs]
    X_gps_3Dmodel, Y_gps_3Dmodel = x_all[gps_dofs], y_all[gps_dofs]

    Ux_gps_model_coseis = Sol_gps_mapped[:, 0, 0]
    Uy_gps_model_coseis = Sol_gps_mapped[:, 1, 0]
    Uz_gps_model_coseis = Sol_gps_mapped[:, 2, 0]

    Ux_gps_model_all = Sol_gps_mapped[:, 0, :]
    Uy_gps_model_all = Sol_gps_mapped[:, 1, :]
    Uz_gps_model_all = Sol_gps_mapped[:, 2, :]

    dtt3D = [15, 240, 720, 288e1, 2 * 432e2]  # , 5256e2]  # in minutes
    ntime3D = [48, 60, 60, 30, 60]  # , 5
    dtt3D = [i * 60 for i in dtt3D]  # in seconds
    dt3D = np.repeat(dtt3D, ntime3D)  # timesteps
    dtt_all3D = np.cumsum(dt3D)  # - 5256e2*50*60 #timeline
    dtt_days_3D = dtt_all3D / (3600 * 24 * 3)

    end_model_time = np.where(dtt_days_3D > endday)[0][0]

    ind_starttime = []
    ind_endtime = []
    Ux_cosies = []
    Uy_cosies = []
    Uz_cosies = []
    Ux_endtime = []
    Uy_endtime = []
    Uz_endtime = []
    X_gps_time = []
    Y_gps_time = []
    model_ind = []

    """load and rotate timeseries"""
    for i in range(0, X_gps_post.shape[0]):
        gps_time = data[names[i] + '_time']
        lon_time = 1e-2 * data[names[i] + '_lon']
        lat_time = 1e-2 * data[names[i] + '_lat']
        Uz_time = 1e-2 * data[names[i] + '_rad']

        Ux_time, Uy_time = GPSlatlon2XY_time(lat_time, lon_time, theta)
        ind_time_EQ = np.where(gps_time[:] > 2012.68)[0][0]
        ind_time_end = np.where(gps_time[:] > 2012.68 + (endday/365.))[0][0]
        ind_3Dmodel = (np.abs(np.sqrt(pow((X_gps_3Dmodel[:] - X_gps_post[i]), 2) +
                                      pow((Y_gps_3Dmodel[:] - Y_gps_post[i]), 2)))).argmin()

        if gps_time[ind_time_EQ] < 2012.9 and ind_time_end - ind_time_EQ > 0.:
	        
            if np.sqrt(pow((X_gps_3Dmodel[ind_3Dmodel] - X_gps_post[i]), 2) +
                pow((Y_gps_3Dmodel[ind_3Dmodel] - Y_gps_post[i]), 2)) < 1e3:

                ind_starttime.append(ind_time_EQ)
                ind_endtime.append(ind_time_end)

                Ux_cosies.append(Ux_time[ind_time_EQ])
                Uy_cosies.append(Uy_time[ind_time_EQ])
                Uz_cosies.append(Uz_time[ind_time_EQ])

                Ux_endtime.append(Ux_time[ind_time_end])
                Uy_endtime.append(Uy_time[ind_time_end])
                Uz_endtime.append(Uz_time[ind_time_end])

                X_gps_time.append(X_gps_post[i])
                Y_gps_time.append(Y_gps_post[i])

                model_ind.append(ind_3Dmodel)

    # make sure the non-PE type is the same as the PE type
    if PE:
        Ux_post = (np.asarray(Ux_endtime) - np.asarray(Ux_cosies)) \
                  - ((np.asarray(Ux_gps_model_all[model_ind, end_model_time])) - (np.asarray(Ux_gps_model_coseis[model_ind])))

        Uy_post = (np.asarray(Uy_endtime) - np.asarray(Uy_cosies)) \
                  - ((np.asarray(Uy_gps_model_all[model_ind, end_model_time])) - (np.asarray(Uy_gps_model_coseis[model_ind])))

        Uz_post = (np.asarray(Uz_endtime) - np.asarray(Uz_cosies)) \
                  - ((np.asarray(Uz_gps_model_all[model_ind, end_model_time])) - (np.asarray(Uz_gps_model_coseis[model_ind])))
    else:
        Ux_post = (np.asarray(Ux_endtime) - np.asarray(Ux_cosies)).astype('float64')
        Uy_post = (np.asarray(Uy_endtime) - np.asarray(Uy_cosies)).astype('float64')
        Uz_post = (np.asarray(Uz_endtime) - np.asarray(Uz_cosies)).astype('float64')

    X = np.asarray(X_gps_time)
    Y = np.asarray(Y_gps_time)

    return X, Y, Ux_post, Uy_post, Uz_post


class Elasticity:
    def __init__(self, mesh, boundaries, Vh, Vh1, u0true, targets, u_gps, invGnoise, point_var, synthetic):
        """
        Construct a model by providing
        - the mesh
        - the finite element spaces for the STATE/ADJOINT variable and the PARAMETER variable
        - the Prior information
        """
        self.results_path = 'results_deterministic/numpy_results/'
        self.var_path = 'results_deterministic/numpy_variables/'
        self.paraviewpath = 'results_deterministic/paraview/'

        if not os.path.exists(self.results_path):
            os.makedirs(self.results_path)
        if not os.path.exists(self.var_path):
            os.makedirs(self.var_path)
        if not os.path.exists(self.paraviewpath):
            os.makedirs(self.paraviewpath)


        self.mesh = mesh
        self.boundaries = boundaries
        self.Vh = Vh
        self.Vh1 = Vh1
        self.targets = targets
        self.Vdim = self.Vh1.dim()
        self.u, self.u0, self.p = TrialFunction(self.Vh[STATE]), \
                                  TrialFunction(self.Vh[PARAMETER]), TrialFunction(self.Vh[ADJOINT])
        self.u_test, self.u0_test, self.p_test = TestFunction(self.Vh[STATE]), \
                                                 TestFunction(self.Vh[PARAMETER]), TestFunction(self.Vh[ADJOINT])
        self.ds = Measure('ds', domain=self.mesh, subdomain_data=self.boundaries)
        
        ################ PARAMETERS ###############
        self.synthetic = synthetic
        self.point_variance = point_var
        self.invGnoise = invGnoise
        self.correlation_length = 2e4 # length over which points share 10% correlation
        self.corr = -np.log(.1)/self.correlation_length #regularization parameter
        self.area = assemble(Constant(1.)*self.ds(2))
        self.h = CellSize(self.mesh)
        
        check_size = 40e3
        gps_obs = True

        # #good for data inversion
        self.delta_b = 2e-6 #* self.h  # Dirichlet control regularization on bottom boundary
        self.gamma = 5e2 # Laplace regularization constant
        self.beta = 1e-9 # mass matrix multiplier in reg.
        
        #good for checkerboard
        #self.delta_b = 5e-6  # Dirichlet control regularization on bottom boundary 5e-6
        #self.gamma = 5e2 # Laplace regularization constant 5e2
        #correlation_length = 1e5
        #self.beta = self.gamma / (correlation_length ** 2)


        print "Gamma {0:5g}; Beta {1:5g}; Delta {2:5g}".format(self.gamma, self.beta, self.delta_b)
        self.delta_b *= self.h
        
        if not self.point_variance:
            self.noise_variance = 1e-6

        self.extract_coords()
        #self.E = Expression('347 * pow(-x[2], 0.516) + 5.324e4', degree = 1)
        self.E = Expression('500 * pow(-x[2], 0.516) + 2e4', degree = 1)

        self.nu = 0.4  # Poisson's ratio
        self.mu   = self.E / (2.0*(1.0 + self.nu)) # shear modulus
        self.lmbda = self.E*self.nu / ((1.0 + self.nu)*(1.0 - 2.0*self.nu))# Lame's parameter
        
        # Initialize Expressions
        self.dim = 3
        self.f = Constant(('0.0','0.0','0.0'))
        self.I = Identity(self.dim)
        self.n = FacetNormal(mesh)
        self.T = self.I - outer(self.n,self.n)

        # Set boundary conditions
        self.b0 = Constant(0.0)
        self.ud1 = Constant((0.0, 0.0, 0.0))
        self.bc1 = DirichletBC(self.Vh[STATE].sub(0), self.b0, self.boundaries, 6) # back
        self.bc2 = DirichletBC(self.Vh[STATE].sub(2), self.b0, self.boundaries, 7) # mantle
        self.bcall = DirichletBC(self.Vh[STATE], self.ud1, u_boundary) # adj var = 0 on all boundaries
        self.bcstate = [self.bc1, self.bc2] # u = 0 on sides and bottom for state eq. and everywhere for adjoint
        self.bcadj = [self.bc1, self.bc2]

        if not gps_obs:
            self.all_surf_dof = np.concatenate((self.ocean_dofs[::50], self.surface_dofs[::10]), axis=0)
            targets = self.coordinates[self.surface_dofs[::5], :]


        self.B = assemblePointwiseObservation(Vh[STATE],targets)
        self.invG = self.build_invG(self.invGnoise)

        bcprior = DirichletBC(Vh[PARAMETER], (1., 0., 0.), boundaries, 3)  # fault surface
        priortest = Function(Vh[PARAMETER])
        bcprior.apply(priortest.vector())
        self.dof_slab = np.where(priortest.vector() == 1.)[0]#.tolist()
        self.coords_all = Vh[STATE].tabulate_dof_coordinates().reshape((Vh[STATE].dim(), -1))
        self.coords_b = self.coords_all[self.dof_slab, :]

        check_slip = checkerboard(self.coords_b, check_size)
        check_slip_func = Function(FunctionSpace(mesh, 'Lagrange', 1))
        check_slip_func.vector()[:] = 0.0
        check_slip_func.vector()[self.dof_slab] = check_slip
        self.check_slip = Expression(('check', '0.0', '-1.0*check'), check= check_slip_func, degree = 1)

        self.u0true = u0true
        #self.u0true = self.check_slip
 
        self.assemble_system()
        print (time.time() - start) / 60.0, "minutes"

        self.AA = block_mat([[self.E_state, self.C_state,       0      ],   # u
                             [    0       , self.RR_grad,   self.C_grad],   # v
                             [self.Wuu    ,      0      ,    self.E_adj]])  # u0

        if self.synthetic:
            self.u_obs = self.computeObservation()
            #quit()
        if not self.synthetic:
            self.u_obs = Vector()
            self.B.init_vector(self.u_obs, 0)
            self.u_obs.set_local(u_gps) #BUG

        self.assemble_rhs()
        self.bb = block_vec([ 0, 0, self.rhs_adjoint])
        print "obs max, min: ", self.u_obs.max(), self.u_obs.min()

        Eps = MumpsSolver(self.E_state)
        Epa = MumpsSolver(self.E_adj)
        Rp = Jacobi(self.RR_grad)
        
        """Create preconditioners:"""
        AApre = block_mat([[Eps, 0, 0 ],
                           [0, Rp, 0 ],
                           [0, 0, Epa]])

        """ Create the block inverse, using the preconditioned Minimum Residual method"""
        self.AAinv = BiCGStab(self.AA, precond = AApre)

    def extract_coords(self):
        if not os.path.exists(self.var_path):
            os.makedirs(self.var_path)

        self.coordinates = self.Vh1.tabulate_dof_coordinates().reshape((self.Vdim, -1))
        self.x_all, self.y_all, self.z_all = self.coordinates[:, 0], self.coordinates[:, 1], self.coordinates[:, 2]

        bctest0 = DirichletBC(self.Vh1, (1, 1, 1), self.boundaries, 1)  # ocean surface
        bctest1 = DirichletBC(self.Vh1, (2, 2, 2), self.boundaries, 2)  # top surface
        bctest2 = DirichletBC(self.Vh1, (3, 3, 3), self.boundaries, 3)  # slab surface

        ptest = Function(self.Vh1)
        bctest0.apply(ptest.vector())
        bctest1.apply(ptest.vector())
        bctest2.apply(ptest.vector())

        self.ocean_dofs = np.where(ptest.vector() == 1)[0]
        self.surface_dofs = np.where(ptest.vector() == 2)[0]
        self.slab_dofs = np.where(ptest.vector() == 3)[0]
        
        
        self.size_gps = self.targets.shape[0]
        size_w_gps = 3 * self.size_gps
        indices_gps = np.empty((self.size_gps, 3))
        for j in range(0, self.size_gps):
            indice_gps = np.where(np.sqrt(
                pow(self.x_all[:] - self.targets[j,0], 2) + pow(self.y_all[:] - self.targets[j,1], 2) + pow(
                    self.z_all[:] - 0.0, 2)) < 1e3)[0]
            indices_gps[j, :] = indice_gps
            self.GPS_dofs = indices_gps.reshape((size_w_gps), order = 'F')
        XYZ_slab = np.vstack((self.x_all[self.slab_dofs], self.y_all[self.slab_dofs], self.z_all[self.slab_dofs])).T

        np.save(self.var_path + "XYZ_slab", XYZ_slab)
        np.save(self.var_path + "x_all_3D", self.x_all)
        np.save(self.var_path + "y_all_3D", self.y_all)
        np.save(self.var_path + "z_all_3D", self.z_all)
        np.save(self.var_path + "GPS_dofs", self.GPS_dofs)
        np.save(self.var_path + "surface_dofs_u_3D", self.surface_dofs)
        np.save(self.var_path + "ocean_dofs_u_3D", self.ocean_dofs)
        np.save(self.var_path + "slab_dofs_u_3D", self.slab_dofs)

    def build_invG(self, invGnoise):
    
        invGp4p = p4p.PETSc.Mat()
        invGp4p.create(p4p.PETSc.COMM_WORLD)
        invGp4p.setSizes([invGnoise.shape[0], invGnoise.shape[1]])
        invGp4p.setType('aij')  # sparse
        invGp4p.setUp()
    
        for j in range(0, invGnoise.shape[0]):
            invGp4p[j, j] = invGnoise[j, j]
    
        invGp4p.assemblyBegin()
        invGp4p.assemblyEnd()
        invG = PETScMatrix(invGp4p)
    
        return invG

    def build_Cov_R(self):
        
        bcprior = DirichletBC(self.Vh[PARAMETER], (1., 1., 1.), self.boundaries, 3)  # land surface
        priortest = Function(self.Vh[PARAMETER])
        bcprior.apply(priortest.vector())
        self.dof_slab_all = np.where(priortest.vector() == 1.)[0]#.tolist()

        rho = squareform(pdist(self.coords_b[:, :3], 'euclidean'))
        R = np.exp(-self.corr * rho)
        Rinv = np.linalg.inv(R)
        RR = np.empty((self.dof_slab_all.shape[0], self.dof_slab_all.shape[0]))
        RI = np.identity(3)
   
        for i in range(0,R.shape[0]):
            RRi = np.kron(RI, Rinv[i,:])
            RR[3*(i+1)-3:3*(i+1),:] = RRi
            
        RRall = np.empty((self.coords_all.shape[0], self.coords_all.shape[0]))
        RRall[self.dof_slab_all,:][:, self.dof_slab_all] = RR
        RRall += assemble(self.beta * inner(self.u0, self.u0_test)*dx).array() #add mass matrix
        
        RRp4p = p4p.PETSc.Mat()
        RRp4p.create(p4p.PETSc.COMM_WORLD)
        RRp4p.setSizes([RRall.shape[0], RRall.shape[1]])
        RRp4p.setType('dense')
        RRp4p.setUp()
        RRp4p[:,:] = RRall[:,:]
        RRp4p.assemblyBegin()
        RRp4p.assemblyEnd()
        RR_block = Matrix(PETScMatrix(RRp4p))

        return RR_block
 
    def build_Wuu(self):
        Bm = as_backend_type(self.B).mat()
        BT = Bm.transpose(Mat())
   
        if self.point_variance:
            BtG_mat = BT.matMult(self.invG.mat())
            BtGB_mat = BtG_mat.matMult(Bm)
            BtGB = Matrix(PETScMatrix(BtGB_mat))
            Wuu = BtGB #* du
            self.BtG = Matrix(PETScMatrix(BtG_mat))
        else:
            BtB_mat = BT.matMult(Bm)
            BtB = Matrix(PETScMatrix((1.0 / self.noise_variance) * BtB_mat))
            Wuu = BtB
            self.BtG = Matrix(PETScMatrix((1.0 / self.noise_variance) * BT))
            
        return Wuu

    def strain(self, w):  # strain = 1/2 (grad u + grad u^T)
        return sym(nabla_grad(w))
    
    def elasticity(self, u, v): # elasticity weak form with slip boundary conditions
        varf =  inner(2.0*self.mu*self.strain(u), self.strain(v))*dx + inner(self.lmbda*nabla_div(u), nabla_div(v))*dx\
               + (1.0/self.delta_b)*inner(self.T*u ,v)*self.ds(3) \
               + (1.0/self.delta_b)*dot(u,self.n)*dot(v,self.n)*self.ds(3)
        
        return varf

    def constraintB(self, u, v): # slip boundary condition
        return (1.0/self.delta_b)*inner(self.T*u, v)*self.ds(3)

    def regularization(self, u, v): # regularization
        varf = self.gamma * inner(nabla_grad(u), nabla_grad(v)) * self.ds(3) \
               + self.beta * inner(u,v)*dx #tiny mass matrix to make block invertable
        
        return varf

    def assemble_system(self):

        self.Wuu = self.build_Wuu()

        self.E_state = assemble(self.elasticity(self.u, self.p_test))
        [bc.apply(self.E_state) for bc in self.bcstate]

        self.E_adj = assemble(self.elasticity(self.p, self.u_test))
        [bc.apply(self.E_adj) for bc in self.bcadj]

        self.C_state = assemble(-self.constraintB(self.u0, self.p_test))
        [bc.zero(self.C_state) for bc in self.bcstate]

        # Take transpose of C_state for C_grad
        self.C_grad = Transpose(self.C_state)
        
        u0 = Function(self.Vh[PARAMETER])
        u0_test = Function(self.Vh[PARAMETER])
        RR = assemble(self.regularization(self.u0, self.u0_test))
        self.RR_grad = RR

    def assemble_rhs(self):
        self.rhs_adjoint = self.BtG * self.u_obs

    def generate_vector(self, component="ALL"):
        """
        Return the list x=[u,a,p] where:
        - u is any object that describes the state variable
        - a is a Vector object that describes the parameter.
          (Need to support linear algebra operations)
        - p is any object that describes the adjoint variable
        
        If component is STATE, PARAMETER, or ADJOINT return x[component]
        """
        if component == "ALL":
            x = [Vector(), Vector(), Vector()]
            self.B.init_vector(x[STATE],1)
            self.B.init_vector(x[PARAMETER],0)
            self.B.init_vector(x[ADJOINT], 1)
        elif component == STATE:
            x = Vector()
            self.B.init_vector(x,1)
        elif component == PARAMETER:
            x = Vector()
            self.R.init_vector(x,0)
        elif component == ADJOINT:
            x = Vector()
            self.B.init_vector(x,1)
            
        return x

    def computeObservation(self):
        """
        Compute the syntetic observation
        """
        
        u_o = Vector()
        A = self.E_state
        b = assemble(self.constraintB(self.u0true, self.p_test))
        [bc.apply(A) for bc in self.bcstate]
        [bc.apply(b) for bc in self.bcstate]

        A.init_vector(u_o, 1)
        solver = LinearSolver('mumps')
        solver.solve(A, u_o, b)
        self.u0_fwdsolve = u_o

        # Create noisy data, ud
        MAX = u_o.norm("linf")
        noise = .01 * MAX * np.random.normal(0, 1, len(u_o.array()))
        u_o.set_local(u_o.array())# + noise)

        u_obs = self.B * u_o  # check that this is no ZERO

        self.paraviewpath = 'results_deterministic/'
        self.u0file = File(self.paraviewpath + 'u_true.pvd')
        self.u0file << vector2Function(self.u0_fwdsolve, Vh[STATE], name = "true displacement")
        
        return u_obs

    def cost(self, x):
        """
        Given the list x = [u,a,p] which describes the state, parameter, and
        adjoint variable compute the cost functional as the sum of 
        the misfit functional and the regularization functional.
        
        Return the list [cost functional, regularization functional, misfit functional]
        
        Note: p is not needed to compute the cost functional
        """        
        assert x[STATE] != None
                       
        diff = self.B*x[STATE]
        diff -= self.u_obs
        if self.point_variance:
            Gdiff = self.invG*diff
            misfit = 0.5 * diff.inner(Gdiff)
        else:
            misfit = (.5 / self.noise_variance) * diff.inner(diff)
        rx = x[PARAMETER]
        reg = 0.5 * rx.inner(self.RR_grad * rx)
        c = misfit + reg
        
        return c, reg, misfit


################### SET UP MESH, MODEL PARAMETERS AND INPUT DATA ##########################################

if __name__ == "__main__":
    set_log_active(False)

    start = time.time()
    sep = "\n"+"#"*80+"\n"
    print sep, "Set up the mesh and finite element spaces: ",(time.time() - start) / 60.0, "minutes", sep
    
    ############ IMPORT MESH (made with gmsh) #######################

    path = "/fenics/shared/meshes"
    mesh_size = 'med_inv'  # fine, med, coarse
    mesh = Mesh(path + '/CR3D_' + mesh_size + '.xml')
    boundaries = MeshFunction("size_t", mesh, path + '/CR3D_' + mesh_size + '_facet_region.xml')

    Vh1 = VectorFunctionSpace(mesh, 'Lagrange', 1)
    Vh = [Vh1, Vh1, Vh1]
    print "Number of dofs: STATE={0}, PARAMETER={1}, ADJOINT={2}".\
        format(Vh[STATE].dim(), Vh[PARAMETER].dim(), Vh[ADJOINT].dim())

    dim = 3
    udim = Vh[STATE].dim()

    ################ USER INPUTS ###############
    point_var = True
    synthetic = False
    event = 'EQ'   # 'EQ' or 'afterslip'
    
    sigmab = 2e4
    xcenter = 80e3
    ycenter = 130e3
    u0d, u0s = -4.0, -0.0

    ################ DATA ###############
    print sep, "Set up the location of observation, Prior Information, and model: ",(time.time() - start) / 60.0, "minutes", sep

    origin = np.array([-85.1688, 8.6925])  # origin of top surface of mesh - used as point of rotation
    theta = 0.816  # angle of rotation between latitude line and fault trace (42 deg)
    
    if event == 'EQ':
        nobs = 37
        data_file = dir_path + 'data/Data.xlsx'
        wb = pyxl.load_workbook(data_file)
        gps_data = wb['GPS_data']
        X_gps, Y_gps, Ux_gps, Uy_gps, Uz_gps, err_x, err_y, err_z = GPSlatlon2XY(gps_data, origin, theta)
        stddev = np.concatenate((err_x.reshape(nobs, 1), err_y.reshape(nobs, 1),
                                 err_z.reshape(nobs, 1)), axis=1).reshape(3 * nobs, )
    else:
        startday = 0
        endday = 60
        print "Running afterslip model for ", str(endday), " days"
        X_gps, Y_gps, Ux_gps, Uy_gps, Uz_gps \
            = importGPSdata(startday, endday, origin, theta, PE = False)
        nobs = len(X_gps)
        error = 1e-3
        errorz = 100* error
        stddev =  np.concatenate((error * np.ones(nobs).reshape(nobs, 1), error * np.ones(nobs).reshape(nobs, 1),
                                  errorz * np.ones(nobs).reshape(nobs, 1)), axis=1).reshape(3 * nobs, )
    Z_gps = np.zeros(len(X_gps))
    targets = np.concatenate((X_gps.reshape(nobs,1), Y_gps.reshape(nobs,1), Z_gps.reshape(nobs,1)),axis=1)
    u_gps = np.concatenate((Ux_gps.reshape(nobs, 1), Uy_gps.reshape(nobs, 1),
                            Uz_gps.reshape(nobs, 1)), axis = 1).reshape(3 * nobs, )  # order="F")

    print "Number of observation points: {0}".format(targets.shape[0])
    """NOISE MATRIX """
    Gnoise = np.diag(stddev ** 2)
    invGnoise = np.linalg.inv(Gnoise)

    if synthetic:
        u0true = Expression(('u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigma,2)) '
                             '+ (pow((x[1] - ycenter),2)/(2*pow(sigma,2))))) ',
                             'u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigma,2)) '
                             '+ (pow((x[1] - ycenter),2)/(2*pow(sigma,2)))))', '0.0'),
                            sigma=sigmab,
                            xcenter=xcenter,
                            ycenter=ycenter,
                            u0d=u0d,
                            u0s=u0s,
                            degree=1)
    else:
        u0true = Expression(('0.0 ', '0.0', '0.0'), degree=1)

    """SOLVE """
    model = Elasticity(mesh, boundaries, Vh, Vh1, u0true, targets, u_gps, invGnoise, point_var, synthetic)
    sol_u, sol_u0, sol_v = model.AAinv * model.bb

    np.save(model.results_path + 'u0_slab_array', sol_u0[model.slab_dofs])
    np.save(model.results_path + 'u0_array', sol_u0)
    np.save(model.results_path + 'u_array', sol_u)

    total_cost, reg_cost, misfit_cost = model.cost([sol_u, sol_u0, sol_v])
    print "Total cost {0:5g}; Reg Cost {1:5g}; Misfit {2:5g}".format(total_cost, reg_cost, misfit_cost)
    print "state max/min: ", sol_u.array().max(), sol_u.array().min()
    print "u0 max/min: ", sol_u0.array().max(), sol_u0.array().min()

    u0 = vector2Function(sol_u0, Vh[PARAMETER])
    u_adj = vector2Function(sol_v, Vh[ADJOINT])
    ufwd = vector2Function(sol_u, Vh[STATE])

    File(model.paraviewpath + "sol_forward.pvd") << ufwd
    File(model.paraviewpath+ "sol_adjoint.pvd") << u_adj
    File(model.paraviewpath+ "u0_inversion.pvd") << u0

    output_file = HDF5File(mesh.mpi_comm(), model.results_path + 'u0slab.h5', 'w')
    output_file.write(u0, 'slip')
    output_file.write(ufwd, 'ufwd')
    output_file.write(mesh, 'mesh')
    output_file.close()
    print sep, "Solution saved as function ",(time.time() - start) / 60.0, "minutes", sep
    print sep, "ALL DONE!: ",(time.time() - start) / 60.0, "minutes", sep
