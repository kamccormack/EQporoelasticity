"""
Kimmy McCormack
Functions used by 'model_run.py' to solve coupled set of
poroelastic equations in both 2D
and 3D subduction zone domains

Units are in meters and MPa
"""

import matplotlib.pyplot as plt
import numpy as np
import openpyxl as pyxl
import os
import scipy as sp
import scipy.interpolate
import shutil
from dolfin import *
from scipy.interpolate import griddata
from tempfile import TemporaryFile
from pykrige.ok import OrdinaryKriging
import pykrige.kriging_tools as kt
import matplotlib.tri as tri
import time
import gc
import sys
sys.path.append('/EQporoelasticity/local_lib')
gc.collect()

"""Optimization options"""
dolfin.parameters.reorder_dofs_serial = False
parameters['form_compiler']['optimize'] = True
parameters["form_compiler"]["cpp_optimize"] = True
ffc_options = {"optimize": True, "eliminate_zeros": True, "precompute_basis_const": True, "precompute_ip_const": True}
parameters['allow_extrapolation'] = True


"""Data interpolation functions"""
def slabkriging(XYZslab, u0load, xvec, yvec, **kwargs):
	"""
	Interpolates inverted fault slip onto the mesh by by kriging.

		INPUTS:
			XYZslab = coordinates of fault slip solution in the model cartesian space
			u0load = slip vector at each solution node
			xvec - vector of dof x locations
			yvec - vector of dof y locations

		OUPUTS:
			u03 - vector containing inverted slip at all dofs
	"""
	
	x_all, y_all = xvec[:], yvec[:]
	
	nonzero = np.where(np.abs(u0load[:, 0]) > 0.10)[0]
	skip = (slice(None, None, 3))
	
	uxslip_load, uyslip_load, uzslip_load = u0load[:, 0][nonzero][skip], \
	                                        u0load[:, 1][nonzero][skip], u0load[:, 2][nonzero][skip]
	xslip, yslip, zslip = XYZslab[:, 0][nonzero][skip], \
	                      XYZslab[:, 1][nonzero][skip], XYZslab[:, 2][nonzero][skip]
	
	# Interpolate data using pykrige
	OK_uxslip = OrdinaryKriging(xslip, yslip, uxslip_load,
	                            variogram_model='linear',
	                            variogram_parameters=[2e-4, 1e-2])
	uxslip, ss_ux = OK_uxslip.execute('points', x_all, y_all)
	OK_uyslip = OrdinaryKriging(xslip, yslip, uyslip_load,
	                            variogram_model='linear',
	                            variogram_parameters=[2e-4, 1e-2])
	uyslip, ss_uy = OK_uyslip.execute('points', x_all, y_all)
	OK_uzslip = OrdinaryKriging(xslip, yslip, uzslip_load,
	                            variogram_model='linear',
	                            variogram_parameters=[2e-4, 1e-2])
	uzslip, ss_uz = OK_uzslip.execute('points', x_all, y_all)
	u03 = np.concatenate((uxslip.reshape((uxslip.shape[0], 1)),
	                      uyslip.reshape((uxslip.shape[0], 1)),
	                      uzslip.reshape((uxslip.shape[0], 1))), axis=1)
	
	return u03

def kappakriging( data_file, origin, theta, coords_K,  **kwargs):
	"""Interpolates surface permeability data by kriging. Take surface value to interpolate kappa with depth.

		INPUTS:
			[data_sheet] = excel sheet containing pointwise permeability data in (lon, lat, permaebility[m^2]) format.
					Default is to ignore the first row (containing headers).
			[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
			[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
					counter-clockwise from latlon to XY
			[coords_K] = coordinates of degrees of freedom onto which the permeability is mapped
			[dof_surf] = the degrees of freedom of the surface of the mesh

			OrdinaryKriging function takes variogram parameters as:
				linear - [slope, nugget]
				power - [scale, exponent, nugget]
			`   gaussian - [sill, range, nugget]
				spherical - [sill, range, nugget]
				exponential - [sill, range, nugget]

		OUPUTS:
			kappa [m^2] - the interpolated 3D permeabiltiy field in the form of a FEniCS Expression()
	"""

	x_all, y_all, z_all = coords_K[:, 0], coords_K[:, 1], coords_K[:, 2]
	wb = pyxl.load_workbook(data_file)
	well_data = wb['well_data']
	x_wells, y_wells = latlon2XY(well_data, origin, theta)
	permeability = np.array(
		[[well_data.cell(row=i, column=5).value] for i in range(2, well_data.max_row + 1)]).reshape(
		well_data.max_row - 1, )
	logk = np.log10(permeability)

	"""PyKrige interpolation"""
	# OK = OrdinaryKriging(x_wells, y_wells, logk, variogram_model = 'power', variogram_parameters = [1, 1, 2e4])
	OK = OrdinaryKriging(x_wells, y_wells, logk, variogram_model='exponential',
	                     variogram_parameters=[1.086, 1.28e4, 0.771])
	k_surf, ss = OK.execute('points', x_all, y_all)
	np.save("results/numpy_variables/k_surf", k_surf.compressed())

	return k_surf.compressed()

def kappakriging2D(data_file, coords_K, dof_surf, **kwargs):
	"""Interpolates 2D permeability data by kriging from excel file

		INPUTS:
			[data_sheet] = excel sheet containing pointwise permeability data in (X, Y, permaebility[m^2]) format.
					Default is to ignore the first row (containing headers).
			[coords_K] = coordinates of degrees of freedom onto which the permeability is mapped
			[dof_surf] = the degrees of freedom of the surface of the mesh

			OrdinaryKriging function takes variogram parameters as:
				linear - [slope, nugget]
				power - [scale, exponent, nugget]
			`   gaussian - [sill, range, nugget]
				spherical - [sill, range, nugget]
				exponential - [sill, range, nugget]
				hole-effect - [sill, range, nugget]

		OUPUTS:
			kappa [m^2] - the interpolated 3D permeabiltiy field in the form of a FEniCS Expression()
	"""

	x_surf, y_surf = coords_K[dof_surf, 0], coords_K[dof_surf, 1]
	x_all, y_all = coords_K[:, 0], coords_K[:, 1]

	wb = pyxl.load_workbook(data_file)
	data = wb['kappa_data2D']

	x_kappa = 1e3 * np.array([[data.cell(row=i, column=1).value] for i in range(2, data.max_row + 1)]).reshape(
		data.max_row - 1, )
	y_kappa = 1e3 * np.array([[data.cell(row=i, column=2).value] for i in range(2, data.max_row + 1)]).reshape(
		data.max_row - 1, )
	permeability = np.array([[data.cell(row=i, column=3).value] for i in range(2, data.max_row + 1)]).reshape(
		data.max_row - 1, )
	logk = np.log10(permeability)

	# PyKrige interpolation

	# OK = OrdinaryKriging(x_wells, y_wells, logk, variogram_model = 'power', variogram_parameters = [1, 1, 2e4])
	OK = OrdinaryKriging(x_kappa, y_kappa, logk,
	                     variogram_model='gaussian', variogram_parameters=[4, 1e5, 0.5])

	# k_surf, ss = OK.execute('points', x_surf, y_surf)
	logk_all, ss = OK.execute('points', x_all, y_all)
	k_all = np.power(10, logk_all)

	# np.save("results/numpy_variables/k_all2D", k_all.compressed())
	# np.save("saved_variables_poroelastic/ss_krige", self.ss.compressed())
	# self.k_surf = np.load("saved_variables_poroelastic/k_all2D.npy")
	# self.ss = np.load("saved_variables_poroelastic/ss_krige_med.npy")

	# return k_all
	return k_all.compressed()

def topokriging(data, origin, theta, coords_Q, dof_surf, var_path, **kwargs):
	""" Maps surface topography to boundary pore pressure

	INPUTS:
		[data] = excel sheet containing topo data formatted as (lon, lat, elevation [meters])
		[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
		[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY
		[coords_Q] = coordinates of degrees of freedom onto which the topo pressure is mapped
		[dof_surf] = the degrees of freedom of the surface of the mesh
		[var_path] = the path of stored variables

	OUTPUTS:
		FEniCS Expression() for pore pressure along surface boundary
	"""

	x_surf, y_surf, z_surf = coords_Q[dof_surf, 0], coords_Q[dof_surf, 1], coords_Q[dof_surf, 2]
	x_topo, y_topo = latlon2XY(data, origin, theta)
	topo_p = 0.7 * (9.81 / 1e3) * np.array(
		[[data.cell(row=i, column=3).value] for i in range(2, data.max_row + 1)]).reshape(data.max_row - 1, )

	# PyKrige interpolation
	OK_topo = OrdinaryKriging(x_topo, y_topo, topo_p,
	                          variogram_model='gaussian', variogram_parameters=[5e5, 5e4, 1e2])
	topo_p, sstopo = OK_topo.execute('points', x_surf, y_surf)

	np.save(var_path + 'topo_p', topo_p.compressed())
	# self.topo_p = np.load(var_path+'topo_p.npy')

	return topo_p.compressed()

def latlon2XY(data_sheet, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = np.array([[data_sheet.cell(row=i, column=1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat = np.array([[data_sheet.cell(row=i, column=2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)

	return 1e3 * X, 1e3 * Y

KExpression2D_cpp = '''
class KExpression2D_cpp : public Expression
{
public:

  KExpression2D_cpp() :
  Expression(),
  alphak(0),
  logkr(0)
  {
  }

    void eval(Array<double>& values, const Array<double>& x) const
    {
        Ksurf->eval(values, x);
        const double Ksurf_val = values[0];

        Zsurf->eval(values, x);
        const double Zsurf_val = values[0];

        values[0] = pow(10, (logkr + ((Ksurf_val - logkr) * pow(1 - 0.25*(1e-3 * (x[1] - Zsurf_val)), -alphak))));

    }

public:
    double alphak;
    double logkr;
    std::shared_ptr<const Function> Ksurf;
    std::shared_ptr<const Function> Zsurf;
};'''

KExpression3D_cpp = '''
class KExpression3D_cpp : public Expression
{
public:

  KExpression3D_cpp() :
  Expression(),
  alphak(0),
  logkr(0)
  {
  }

    void eval(Array<double>& values, const Array<double>& x) const
    {
        Ksurf->eval(values, x);
        const double Ksurf_val = values[0];

        Zsurf->eval(values, x);
        const double Zsurf_val = values[0];

        values[0] = pow(10, (logkr + ((Ksurf_val - logkr) * pow(1 - 0.25*(1e-3 * (x[2] - Zsurf_val)), -alphak))));

    }

public:
    double alphak;
    double logkr;
    std::shared_ptr<const Function> Ksurf;
    std::shared_ptr<const Function> Zsurf;
};'''

class SlipExpression(Expression):
	""" Interpolates an arbitrary displacement field to the fault boundary of the mesh.

	INPUTS:
		[data] = an excel sheet with slip vectors recorded as
				(lon, lat, dip-slip [meters] , strike-slip [meters]) in the first four columns
		[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
		[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY
		[coords_V] = coordinates of degrees of freedom onto which the slip vectors are mapped
		[dof_slab] = the degrees of freedom of the fault boundary of the mesh
		[var_path] = the path of stored variables

	OUTPUTS:
		FEniCS Expression() containing the interpolated vector slip displacement on the fault boundary
	"""

	def __init__(self, data, origin, theta, coords_V, dof_slab, var_path, **kwargs):
		self.x_slab, self.y_slab, self.z_slab = coords_V[dof_slab, 0], coords_V[dof_slab, 1], coords_V[dof_slab, 2]

		X_slip, Y_slip = latlon2XY(data, origin, theta)

		Ux_slip = lon = np.array([[data.cell(row=i, column=3).value] for i in range(2, data.max_row + 1)]).reshape(
			data.max_row - 1, )
		Uy_slip = lon = np.array([[data.cell(row=i, column=4).value] for i in range(2, data.max_row + 1)]).reshape(
			data.max_row - 1, )

		# Interpolate data using pykrige
		OK_uxslip = OrdinaryKriging(X_slip, Y_slip, Ux_slip, variogram_model='linear', \
		                            variogram_parameters=[2e-4, 1e-2])
		self.uxslip, ss_ux = OK_uxslip.execute('points', self.x_slab, self.y_slab)

		OK_uyslip = OrdinaryKriging(X_slip, Y_slip, Uy_slip, variogram_model='linear', \
		                            variogram_parameters=[2e-4, 1e-2])
		self.uyslip, ss_ux = OK_uyslip.execute('points', self.x_slab, self.y_slab)

		np.save(var_path + 'uxslip', self.uxslip.compressed())
		np.save(var_path + 'uyslip', self.uyslip.compressed())

	# self.uxslip = np.load("saved_variables_poroelastic/uxslip_med.npy")
	# self.uyslip = np.load("saved_variables_poroelastic/uyslip_med.npy")

	def U_X_interp(self, X, Y):
		indx = (np.abs(np.sqrt(pow((self.x_slab[:] - X), 2) + pow((self.y_slab[:] - Y), 2)))).argmin()
		return self.uxslip[indx]

	def U_Y_interp(self, X, Y):
		indy = (np.abs(np.sqrt(pow((self.x_slab[:] - X), 2) + pow((self.y_slab[:] - Y), 2)))).argmin()
		return self.uyslip[indy]

	def eval(self, values, x, **kwargs):
		values[0] = self.U_X_interp(x[0], x[1])
		values[1] = self.U_Y_interp(x[0], x[1])
		values[2] = 0.0

	def value_shape(self):
		return (3,)
	
		
"""FEM solving functions"""
class Poroelasticity2D:
	def __init__(self, data_file, ndim, mesh, boundaries, plot_figs, event, permeability, days, SSE_days,
	             sub_cycle_years, sigma_b, xcenter, SS_migrate, u0_EQ, u0_SSE, u0_sub, surface_k, ocean_k, B, loop,
	             loop_variable, **kwargs):

		self.start = time.time()
		self.data_file = data_file
		self.mesh = mesh
		self.boundaries = boundaries
		self.plot_figs = plot_figs
		self.event = event
		self.permeability = permeability
		self.days = days
		self.SSE_days = SSE_days
		self.sub_cycle_years = sub_cycle_years
		self.sigma_b = sigma_b
		self.xcenter = xcenter
		self.SS_migrate = SS_migrate
		self.u0_EQ = u0_EQ
		self.u0_SSE = u0_SSE
		self.u0_sub = u0_sub
		self.surface_k = surface_k
		self.ocean_k = ocean_k
		self.loop = loop
		self.loop_variable = loop_variable

		self.solver = LinearSolver('mumps')
		self.prm = self.solver.parameters
		self.prm['reuse_factorization'] = True  # Saves Factorization of LHS

		self.V = VectorElement('Lagrange', self.mesh.ufl_cell(), 2)
		self.Q = FiniteElement('Lagrange', self.mesh.ufl_cell(), 1)
		self.ME = FunctionSpace(self.mesh, MixedElement([self.Q, self.V]))
		self.Vspace = FunctionSpace(self.mesh, self.V)
		self.Qspace = FunctionSpace(self.mesh, self.Q)
		self.Q_gspace = VectorFunctionSpace(self.mesh, "CG", 1)
		self.Wspace = TensorFunctionSpace(self.mesh, "Lagrange", 1)
		
		self.Q2 = FiniteElement('Lagrange', self.mesh.ufl_cell(), 2)
		self.Q2space = FunctionSpace(self.mesh, self.Q2)

		self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
		self.Kspace = FunctionSpace(self.mesh, self.Kelement)

		(self.p, self.u) = TrialFunctions(self.ME)
		(self.q, self.v) = TestFunctions(self.ME)
		self.ds = Measure("ds", domain=self.mesh, subdomain_data=self.boundaries)  #
		self.n = FacetNormal(self.mesh)  # normal vector
		self.dim = self.ME.dim()
		self.pdim = self.ME.sub(0).dim()
		self.udim = self.ME.sub(1).dim()
		self.geom_dim = ndim + 1

		self.results_path = 'results/numpy_results/'
		self.var_path = 'results/numpy_variables/'
		self.paraviewpath = 'results/paraview/'
		self.inverse_results = '/fenics/shared/inverse_codes/results_deterministic/numpy_results/'
		self.inverse_var = '/fenics/shared/inverse_codes/results_deterministic/numpy_variables/'
		
		if not os.path.exists(self.paraviewpath):
			os.makedirs(self.paraviewpath)

		self.extract_coords()
		print "Coordinates extracted", (time.time() - self.start) / 60.0, "minutes"
		self.k_interpolation()
		print "Permeability interpolated", (time.time() - self.start) / 60.0, "minutes"
		self.slip_interpolation()
		print "Slip interpolated", (time.time() - self.start) / 60.0, "minutes"
		self.time_stepping()

		"""POROELASTIC PARAMETERS"""
		self.delta = 1e-6  # Dirichlet control regularization on bottom boundary
		self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
		self.B = B  # Skempton's coefficient
		self.nuu = .4  # undrained poissons ratio
		self.nu = 0.25  # Poisson's ratio
		"""Young's modulus [MPa]. Fit from Deshon paper (2006)"""
		self.E = Expression('500 * pow(-x[1], 0.516) + 2e4', degree=1)
		self.mu = Expression('E / (2.0*(1.0 + nu))', E=self.E, nu=self.nu, degree=1)  # shear modulus
		self.lmbda = Expression('E*nu / ((1.0 + nu)*(1.0 - 2.0*nu))', E=self.E, nu=self.nu,
		                        degree=1)  # Lame's parameter
		self.alpha = 3 * (self.nuu - self.nu) / (self.B * (1 - 2 * self.nu) * (1 + self.nuu))  # Biot's coefficient
		self.Se = Expression('(9*(nuu-nu)*(1-(2*nuu)))/(2*mu*pow(B,2)*(1-2*nu)*pow((1+nuu),2))', nuu=self.nuu,
		                     nu=self.nu, mu=self.mu, B=self.B, degree=1)  # mass specific storage
		self.kstick = 3e-5
		self.stick = Expression('0.8*(1/(1+exp(-kstick*(xstick-x[0]))))',
		                        xstick=self.xstick, kstick=self.kstick,
		                        degree=1)
		self.d = self.u.geometric_dimension()  # number of space dimensions
		self.I = Identity(self.d)
		self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
		self.year = 60.0 * 60.0 * 24.0 * 365.0
		self.day = 60.0 * 60.0 * 24.0

		"""INITIAL AND BOUNDARY CONDITIONS"""
		####### Initial conditions: (('p0','ux0','uy0')) #########
		self.w_init = Expression(('0', '0', '0'), degree=1)
		self.w0 = Function(self.ME)
		self.w0.interpolate(self.w_init)
		self.ub0 = Expression(('0', '0'), degree=1)
		self.ub1 = Expression(('1', '1'), degree=1)
		self.b0 = Constant('0')
		self.b1 = Constant('1.0')
		self.bc1 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 2)  # top/surface
		self.bc2 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 3)  # bottom/fault boundary
		self.bc3 = DirichletBC(self.ME.sub(1).sub(0), self.b0, self.boundaries, 4)  # right/ back side (no slip condition)
		self.bc4 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 5)  # mantle (no slip condition)
		self.bc5 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 1)  # ocean surface
		self.bcp1 = DirichletBC(self.ME.sub(0), self.b0, self.boundaries, 1)  # ocean surface (free flow condition)
		self.bcp2 = DirichletBC(self.ME.sub(0), self.b0, self.boundaries, 2)  # land surface (free flow condition)

		if self.event == 'topo':
			self.topo_flow_interp()
			self.bcp2 = DirichletBC(self.ME.sub(0), self.p_topo, self.boundaries, 2)  # land surface
			self.bcs = [self.bcp1, self.bcp2]  # These are the BC's that are actually applied
		else:
			self.bcs = [self.bc3, self.bc4, self.bcp1]  # These are the BC's that are actually applied

		self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case

		"""SET UP TO SAVE SOLUTION"""
		self.Sol_all = np.empty((self.dim, self.nsteps))  # matrix to save entire solution
		self.Sol_surf = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
		self.sea_flux = np.empty((self.ocean_dof_pgrad.shape[0], self.nsteps))  # ocean bottom solution
		self.flow_velocity = np.empty((self.pdim * ndim, self.nsteps))  # ocean bottom solution
		self.Psi = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
		self.vol_strain = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
		self.sea_flux_total = np.empty(self.nsteps)  # matrix to save surface post-seismic solution
		self.esigv_slab = np.empty((self.dof_slabQ.shape[0], self.nsteps))

		if loop == 'yes':
			if self.loop_variable == 'kappa':
				self.loop_name = "k" + str(-self.surface_k)
				pfilename = self.paraviewpath + "pressure2D_loop_k%s.pvd" % str(-self.surface_k)
				p_velfilename = self.paraviewpath + "pressure_vel2D_loop_k%s.pvd" % str(-self.surface_k)
				ufilename = self.paraviewpath + "def2D_loop_k%s.pvd" % str(-self.surface_k)
				sigfilename = self.paraviewpath + "stress2D_k%s.pvd" % str(-self.surface_k)
			elif self.loop_variable == 'sigma':
				self.loop_name = 'sig' + str(self.sigma_b)[:2]
				pfilename = self.paraviewpath + "pressure2D_loop_%s.pvd" % str(self.sigma_b)[:2]
				p_velfilename = self.paraviewpath + "pressure_vel2D_loop_%s.pvd" % str(self.sigma_b)[:2]
				ufilename = self.paraviewpath + "def2D_loop_%s.pvd" % str(self.sigma_b)[:2]
				sigfilename = self.paraviewpath + "stress2D_%s.pvd" % str(self.sigma_b)[:2]
			elif self.loop_variable == 'B':
				self.loop_name = 'B' + str(-self.B)
				pfilename = self.paraviewpath + "pressure2D_loop_B%s.pvd" % str(-self.B)
				p_velfilename = self.paraviewpath + "pressure_vel2D_loop_B%s.pvd" % str(-self.B)
				ufilename = self.paraviewpath + "def2D_loop_k%s.pvd" % str(-self.B)
				sigfilename = self.paraviewpath + "stress2D_B%s.pvd" % str(-self.B)

		else:
			pfilename = self.paraviewpath + "pressure2D" + self.name_append + ".pvd"
			p_velfilename = self.paraviewpath + "pressure_vel2D" + self.name_append + ".pvd"
			ufilename = self.paraviewpath + "def2D" + self.name_append + ".pvd"
			sigfilename = self.paraviewpath + "stress2D" + self.name_append + ".pvd"

		self.pfile = File(pfilename)
		self.p_velfile = File(p_velfilename)
		self.ufile = File(ufilename)
		self.sigfile = File(sigfilename)

		"""SET UP, RUN, SOLVE and SAVE"""
		self.assemble_system()
		print "System assembled", (time.time() - self.start) / 60.0, "minutes"
		self.solve_system()
		self.save_solution()

	def time_stepping(self):
		if self.event == 'EQ':
			"""15 min for 1/2 day,
			4 hrs for 10 days,
			12 hours for 30 days,
			2 days for 2 months,
			per month for 10 years
			"""
			self.dtt = [60, 240, 720, 288e1, 2 * 432e2]   # in minutes
			self.ntime = [24, 60, 60, 30, 60]  # , 50]
		elif self.event == 'SSE':
			"""
			12 hrs for SSE days,
			12 hrs for total time-SSE_days,
			2 months for 10 years
			"""
			self.dtt = [288e1 / 2, 432e2 * 2]
			self.ntime = [self.SSE_days / 1, (self.days - self.SSE_days) / 60]  # SSE_days, 10 years
		elif self.event == 'sub_EQ':
			"""
			1/year for 50 years followed by 'EQ' scheme
			"""
			self.dtt = [60 * 24 * 365, 15, 240, 288e1, 432e2, 60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years, 96, 120, 15, 122, 40]
		elif self.event == 'sub':  # 1/year for 50 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		elif self.event == 'topo':  # 1/year for 100 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		self.dtt = [i * 60 for i in self.dtt]  # in seconds
		self.dtt_repeated = np.repeat(self.dtt, self.ntime)
		self.dtt_comsum = np.cumsum(self.dtt_repeated)
		self.dt = self.dtt_comsum[0]
		self.nsteps = self.dtt_comsum.size

	def k_interpolation(self):
		log_kr = -18.
		alpha_k = 0.6

		self.coords_K = self.Kspace.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
		bctest1 = DirichletBC(self.Kspace, (1), self.boundaries, 1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1), self.boundaries, 2)  # land surface
		ktest = Function(self.Kspace)
		bctest1.apply(ktest.vector())
		bctest2.apply(ktest.vector())
		self.dof_surfK = np.where(ktest.vector() == 1)[0]

		if self.permeability == 'constant':
			self.k_2D_func = Function(self.Kspace)
			self.k_2D_func.vector()[:] = 1e-18
			self.kappa = Expression('kappa', kappa=self.k_2D_func, degree=1)

		elif self.permeability == 'mapped':
			kappa_surf = self.surface_k
			self.k_surf_func = Function(self.Kspace)
			self.z_surf_func = Function(self.Kspace)
			self.k_surf_func.vector()[:] = kappa_surf
			self.z_surf_func.vector()[:] = 0.

			self.kappa = Expression(KExpression2D_cpp, degree=1)
			self.kappa.alphak = alpha_k
			self.kappa.logkr = log_kr
			self.kappa.Ksurf = self.k_surf_func
			self.kappa.Zsurf = self.z_surf_func

	def topo_flow_interp(self):  # debug

		wb = pyxl.load_workbook(self.data_file)
		topo_data = wb['topo_data']
		x_topo = np.array([[topo_data.cell(row=i, column=1).value] for i in range(2, topo_data.max_row + 1)]).reshape(
			topo_data.max_row - 1, )
		p_topo = 0.7 * (9.81 / 1e3) * np.array(
			[[topo_data.cell(row=i, column=2).value] for i in range(2, topo_data.max_row + 1)]).reshape(
			topo_data.max_row - 1, )
		self.coords_K = self.Kspace.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
		bctest1 = DirichletBC(self.Kspace, (1), self.boundaries, 1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1), self.boundaries, 2)  # land surface
		qtest = Function(self.Kspace)
		bctest1.apply(qtest.vector())
		bctest2.apply(qtest.vector())
		self.dof_surfK = np.where(qtest.vector() == 1)[0]

		topo_surf_p = griddata(x_topo, p_topo, self.coords_K[self.dof_surfK, 0], method='linear')

		p_surf_func = Function(self.Kspace)
		p_surf_func.vector()[self.dof_surfK] = topo_p_surf
		self.p_topo = Expression(('ptopo'), ptopo=p_surf_func, degree=1)

	def strain(self, w):  # strain = 1/2 (grad u + grad u^T)
		return sym(nabla_grad(w))

	def sigma(self, w):  # stress = 2 mu strain + lambda tr(strain) I
		return 2.0 * self.mu * self.strain(w) + self.lmbda * div(w) * self.I

	def assemble_system(self):
		"""Assemble linear system"""
		self.a_Mass = assemble(self.Se * self.p * self.q * dx)  # Mass matrix
		self.a_K = assemble(
			((self.kappa) / self.muf) * inner(nabla_grad(self.p), nabla_grad(self.q)) * dx)  # Stiffness matrix
		self.a_Div = assemble(nabla_div(self.u) * self.q * dx)  # Divergence matrix
		self.a_Grad = assemble(-self.p * (nabla_div(self.v)) * dx)  # Gradient
		self.a_E = assemble(inner(self.sigma(self.u), nabla_grad(self.v)) * dx)  # Elasticity
		self.a_Boundary = assemble((1 / self.delta) * dot(self.u, self.n) * dot(self.v, self.n) * self.ds(self.slab)
		                           + (1 / self.delta) * inner(self.T * self.u, self.T * self.v) * self.ds(self.slab))

		self.A = self.a_Mass + self.dt * self.a_K + self.alpha * self.a_Div + self.alpha * self.a_Grad \
		         + self.a_E + self.a_Boundary  # Left hand side (LHS)
		self.L = self.a_Mass + self.alpha * self.a_Div  # RHS

	def extract_coords(self):
		if not os.path.exists(self.var_path):
			os.makedirs(self.var_path)

		self.coordinates = self.ME.tabulate_dof_coordinates().reshape((self.dim, -1))
		self.x_all, self.y_all = self.coordinates[:, 0], self.coordinates[:, 1]

		self.coords_Q = self.Qspace.tabulate_dof_coordinates().reshape((self.pdim, -1))
		self.coords_V = self.Vspace.tabulate_dof_coordinates().reshape((self.udim, -1))

		bctest0 = DirichletBC(self.ME, (1, 1, 1), self.boundaries, 1)  # ocean surface
		bctest1 = DirichletBC(self.ME, (2, 2, 2), self.boundaries, 2)  # top surface
		bctest2 = DirichletBC(self.ME, (3, 3, 3), self.boundaries, 3)  # slab surface

		ptest = Function(self.ME)
		bctest0.apply(ptest.vector())
		bctest1.apply(ptest.vector())
		bctest2.apply(ptest.vector())
		self.ocean_dofs = np.where(ptest.vector() == 1)[0]
		self.surface_dofs = np.where(ptest.vector() == 2)[0]
		self.slab_dofs = np.where(ptest.vector() == 3)[0]

		self.size_p = self.pdim
		self.size_u = self.dim - self.size_p

		self.ocean_dof_p = self.ocean_dofs[np.where(self.ocean_dofs < self.size_p)]
		self.ocean_dof_u = self.ocean_dofs[np.where(self.ocean_dofs >= self.size_p)] - self.size_p
		self.ocean_dof_pgrad = np.hstack((self.ocean_dof_p, (self.ocean_dof_p + self.ocean_dof_p.shape[0])))
		self.surface_dof_p = self.surface_dofs[np.where(self.surface_dofs < self.size_p)]
		self.surface_dof_u = self.surface_dofs[np.where(self.surface_dofs >= self.size_p)] - self.size_p
		self.surface_dof_pgrad = np.hstack((self.surface_dof_p, (self.surface_dof_p + self.surface_dof_p.shape[0])))

		bctestq = DirichletBC(self.Qspace, (1), self.boundaries, 3)  # slab surface
		qtest = Function(self.Qspace)
		bctestq.apply(qtest.vector())
		self.dof_slabQ = np.where(qtest.vector() == 1)[0]

		np.save(self.var_path + "dof_slab", self.slab_dofs)
		np.save(self.var_path + "x_all_2D", self.x_all)
		np.save(self.var_path + "y_all_2D", self.y_all)
		np.save(self.var_path + "surface_dofs_2D", self.surface_dofs)
		np.save(self.var_path + "ocean_dofs_2D", self.ocean_dofs)
		np.save(self.var_path + "size_p_CR_2D", self.size_p)

	def init_boundary_slip(self, dt):
		if self.event == 'SSE':
			self.u_addSS = self.u0_SSE
			self.u_addsub = self.u0_sub
			self.u0_SSE = self.u_addSS * (dt / self.day)
			self.u0_sub = self.u_addsub * (dt / self.year)

		elif self.event == 'sub':  # Amount of slip per day (at center of event)
			self.u_add = self.u0_sub * (dt / self.year)
		elif self.event == 'sub_EQ':  # Amount of slip per day (at center of event)
			self.u_add = self.u0_sub * (dt / self.year)

	def streamfunction(self):
		"""Stream function for a given general 2D velocity field.
		The boundary conditions are weakly imposed through the term

			inner(q, grad(psi)*n)*ds,

		where grad(psi) = [-v, u] is set on all boundaries.
		This should work for any collection of boundaries:
		walls, inlets, outlets etc.
		"""
		vel = self.q_velocity
		qtest = TestFunction(self.Qspace)
		psi = TrialFunction(self.Qspace)
		a = inner(grad(qtest), grad(psi)) * dx
		# L   = dot(qtest, u[1].dx(0) - u[0].dx(1))*dx
		L = dot(qtest, curl(vel)) * dx
		bcu = []
		L = L + qtest * (self.n[1] * vel[0] - self.n[0] * vel[1]) * self.ds

		# Compute solution
		psi = Function(self.Qspace)
		A = assemble(a)
		b = assemble(L)
		normalize(b)  # Because we only have Neumann conditions
		[bc.apply(A, b) for bc in bcu]
		solve(A, psi.vector(), b)
		normalize(psi.vector())

		return psi

	def solve_system(self):
		count = 0
		if self.event == 'EQ':
			"""For an EQ, this can be outside the time loop"""
			a_Boundary0 = assemble((1 / self.delta) * inner(self.u0slab, self.T * self.v) * self.ds(
				self.slab))

		"""TIME LOOP"""
		for i in range(self.nsteps):

			if i == 0:
				dt = self.dtt_comsum[i]
			else:
				dt = self.dtt_comsum[i] - self.dtt_comsum[i - 1]
			if i > 2:
				dt_old = self.dtt_comsum[i - 1] - self.dtt_comsum[i - 2]
			if i > 2 and dt != dt_old:
				self.A = self.a_Mass + dt * self.a_K + self.alpha * self.a_Div + self.alpha * self.a_Grad \
				         + self.a_E + self.a_Boundary  # Left hand side (LHS)

			self.b = Vector()
			self.a_Mass.init_vector(self.b, 0)
			self.L.mult(self.w0.vector(), self.b)
			if self.event != 'topo':
				self.b += a_Boundary0

			[bc.apply(self.b) for bc in self.bcs]
			[bc.apply(self.A) for bc in self.bcs]
			print "BC's applied", (time.time() - self.start) / 60.0, "minutes"

			self.w = Function(self.ME)
			self.solver.set_operator(self.A)
			self.solver.solve(self.w.vector(), self.b)
			print "System solved", (time.time() - self.start) / 60.0, "minutes"
			p, u = self.w.split()
			
			self.domain_flux = (-self.kappa / self.muf) * grad(p)  # element-wise flux vector
			self.bound_flux = assemble(inner(self.domain_flux, self.n) * self.ds(1))
			print "outward flux through seafloor: ", self.bound_flux
			self.bound_flux2 = assemble(inner(self.domain_flux, self.n) * self.ds(2))
			print "outward flux through land surface: ", self.bound_flux2

			self.q_velocity = Function(self.Q_gspace)
			self.q_velocity.assign(project(self.domain_flux, self.Q_gspace, solver_type='mumps'))
			print "Computed velocity field", (time.time() - self.start) / 60.0, "minutes"

			self.v_strain = Function(self.Qspace)
			self.v_strain.assign(project(nabla_div(u), self.Qspace, solver_type='mumps'))
			self.psi = self.streamfunction()  # compute numerical streamfunction

			stress = Function(self.Wspace)
			stress.assign(project(self.sigma(u), self.Wspace, solver_type='mumps'))

			"""SAVE SOLUTION"""
			p.rename('p', 'p')
			self.pfile << p
			u.rename('u', 'u')
			self.ufile << u
			self.p_velfile << self.q_velocity
			stress.rename('stress', 'stress')
			self.sigfile << stress

			self.sea_flux_total[count] = self.bound_flux
			self.sea_flux[:, count] = self.q_velocity.vector()[self.ocean_dof_pgrad]
			self.flow_velocity[:, count] = self.q_velocity.vector()
			self.Psi[:, count] = self.psi.vector()
			self.vol_strain[:, count] = self.v_strain.vector()
			self.Sol_surf[:, count] = self.w.vector()[self.surface_dofs]
			self.Sol_all[:, count] = self.w.vector()

			self.w0 = self.w

			count += 1
			print "Timestep:", count, "dt:", dt

	def save_solution(self):
		if not os.path.exists(self.results_path):
			os.makedirs(self.results_path)
		if self.event == 'SSE':
			np.save(self.results_path + "Sol_surf_2D_SSE.npy", self.Sol_surf)
			np.save(self.results_path + "Sol_all_2D_SSE.npy", self.Sol_all)
			np.save(self.results_path + "sea_flux_2D_SSE.npy", self.sea_flux)
			np.save(self.var_path + "dtt_comsum_SSE.npy", self.dtt_comsum)
		elif self.event == 'sub':
			np.save(self.results_path + "Sol_surf_2D_sub.npy", self.Sol_surf)
			np.save(self.results_path + "Sol_all_2D_sub.npy", self.Sol_all)
			np.save(self.results_path + "sea_flux_2D_sub.npy", self.sea_flux)
			np.save(self.results_path + "flow_velocity_2D_sub.npy", self.flow_velocity)
			np.save(self.results_path + "vol_strain_2D_sub.npy", self.vol_strain)
			np.save(self.results_path + "Streamfun_2D_sub.npy", self.Psi)
			np.save(self.results_path + "sea_flux_total_2D_sub.npy", self.sea_flux_total)
			np.save(self.var_path + "dtt_comsum_sub.npy", self.dtt_comsum)
			np.save(self.var_path + "sub_cycle_sub.npy", self.sub_cycle_years)
		elif self.event == 'sub_EQ':
			np.save(self.results_path + "Sol_surf_2D_sub_EQ.npy", self.Sol_surf)
			np.save(self.results_path + "Sol_all_2D_sub_EQ.npy", self.Sol_all)
			np.save(self.results_path + "sea_flux_2D_sub_EQ.npy", self.sea_flux)
			np.save(self.var_path + "dtt_comsum_sub_EQ.npy", self.dtt_comsum)
			np.save(self.var_path + "sub_cycle_sub_EQ.npy", self.sub_cycle_years)
		elif self.event == 'EQ':
			if self.loop == 'yes':
				np.save(self.results_path + "Sol_all_2D_EQ_loop_%s.npy" % self.loop_name, self.Sol_all)
				np.save(self.results_path + "Sol_surf_2D_EQ_loop_%s.npy" % self.loop_name, self.Sol_surf)
				np.save(self.results_path + "sea_flux_total_2D_EQ_loop_%s.npy" % self.loop_name, self.sea_flux_total)
				np.save(self.results_path + "sea_flux_2D_EQ_loop_%s.npy" % self.loop_name, self.sea_flux)
				np.save(self.results_path + "vol_strain_2D_EQ_loop_%s.npy" % self.loop_name, self.vol_strain)
				np.save(self.results_path + "flow_velocity_2D_EQ_loop_%s.npy" % self.loop_name, self.flow_velocity)
				np.save(self.results_path + "Streamfun_2D_EQ_loop_%s.npy" % self.loop_name, self.Psi)
				np.save(self.var_path + "dtt_comsum_EQ.npy", self.dtt_comsum)
			if self.loop == 'no':
				np.save(self.results_path + "Sol_surf_2D_EQ.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_all_2D_EQ.npy", self.Sol_all)
				np.save(self.results_path + "sea_flux_2D_EQ.npy", self.sea_flux)
				np.save(self.var_path + "dtt_comsum_EQ.npy", self.dtt_comsum)
				np.save(self.results_path + "flow_velocity_2D_EQ.npy", self.flow_velocity)
				np.save(self.results_path + "vol_strain_2D_EQ.npy", self.vol_strain)
				np.save(self.results_path + "Streamfun_2D_EQ.npy", self.Psi)
				np.save(self.results_path + "sea_flux_total_2D_EQ.npy", self.sea_flux_total)
		elif self.event == 'topo':
			if self.loop == 'yes':
				np.save(self.results_path + "Sol_surf_2D_topo_loop_%s.npy" % self.loop_name, self.Sol_surf)
				np.save(self.results_path + "Streamfun_2D_topo_loop_%s.npy" % self.loop_name, self.Psi)
				np.save(self.results_path + "sea_flux_total_2D_topo_loop_%s.npy" % self.loop_name,
				        self.sea_flux_total)
				np.save(self.results_path + "sea_flux_2D_topo_loop_%s.npy" % self.loop_name, self.sea_flux)
				np.save(self.results_path + "flow_velocity_2D_topo_loop_%s.npy" % self.loop_name,
				        self.flow_velocity)
				np.save(self.results_path + "vol_strain_2D_topo_loop_%s.npy" % self.loop_name, self.vol_strain)
				np.save(self.results_path + "Sol_all_2D_topo_loop_%s.npy" % self.loop_name, self.Sol_all)
				np.save(self.var_path + "dtt_comsum_topo.npy", self.dtt_comsum)
			if self.loop == 'no':
				np.save(self.results_path + "Sol_surf_2D_topo.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_all_2D_topo.npy", self.Sol_all)
				np.save(self.results_path + "sea_flux_2D_topo.npy", self.sea_flux)
				np.save(self.var_path + "dtt_comsum_topo.npy", self.dtt_comsum)
				np.save(self.results_path + "flow_velocity_2D_topo.npy", self.flow_velocity)
				np.save(self.results_path + "vol_strain_2D_topo.npy", self.vol_strain)
				np.save(self.results_path + "Streamfun_2D_topo.npy", self.Psi)
				np.save(self.results_path + "sea_flux_total_2D_topo.npy", self.sea_flux_total)

class Poroelasticity3DSynthetic:
	def __init__(self, data_file, origin, theta, ndim, mesh, boundaries, plot_figs, event, permeability, SSE_days,
	             sub_cycle_years, sigma_bx, xcenter, sigma_by, ycenter, SS_migratedip, SS_migratestrike, u0_EQdip,
	             u0_EQstrike, u0_SSEdip, u0_SSEstrike, u0_subdip, u0_substrike, surface_k, ocean_k):

		self.start = time.time()
		self.data_file = data_file
		self.origin = origin
		self.theta = theta
		self.mesh = mesh
		self.boundaries = boundaries
		self.plot_figs = plot_figs
		self.event = event
		self.permeability = permeability
		self.SSE_days = SSE_days
		self.sub_cycle_years = sub_cycle_years
		self.sigma_bx = sigma_bx
		self.xcenter = xcenter
		self.sigma_by = sigma_by
		self.ycenter = ycenter
		self.SS_migratex = SS_migratedip
		self.SS_migratey = SS_migratestrike
		self.u0_EQx = u0_EQdip
		self.u0_EQy = u0_EQstrike
		self.u0d_SSE = u0_SSEdip
		self.u0s_SSE = u0_SSEstrike
		self.u0_subx = u0_subdip
		self.u0_suby = u0_substrike
		self.surface_k = surface_k
		self.ocean_k = ocean_k
		self.solver = LUSolver('petsc')
		self.prm = self.solver.parameters
		self.prm['reuse_factorization'] = True  # Saves Factorization of LHS

		self.V = VectorElement('Lagrange', self.mesh.ufl_cell(), 2)
		self.Q = FiniteElement('Lagrange', self.mesh.ufl_cell(), 1)
		self.ME = FunctionSpace(self.mesh, MixedElement([self.Q, self.V]))
		self.Vspace = FunctionSpace(self.mesh, self.V)
		self.Qspace = FunctionSpace(self.mesh, self.Q)
		self.Q_gspace = VectorFunctionSpace(self.mesh, "Lagrange", 1)
		self.Wspace = TensorFunctionSpace(self.mesh, "Lagrange", 1)

		self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
		self.Kspace = FunctionSpace(self.mesh, self.Kelement)

		(self.p, self.u) = TrialFunctions(self.ME)
		(self.q, self.v) = TestFunctions(self.ME)
		self.ds = Measure("ds", domain=self.mesh, subdomain_data=self.boundaries)  #
		self.n = FacetNormal(self.mesh)  # normal vector
		self.dim = self.ME.dim()
		self.pdim = self.ME.sub(0).dim()
		self.udim = self.ME.sub(1).dim()
		self.geom_dim = ndim + 1

		self.results_path = 'results/numpy_results/'
		self.var_path = 'results/numpy_variables/'
		self.paraviewpath = 'results/paraview/'
		if not os.path.exists(self.paraviewpath):
			os.makedirs(self.paraviewpath)

		self.extract_coords()
		print "Coordinates extracted", (time.time() - self.start) / 60.0, "minutes"
		self.k_interpolation()
		print "Permeability interpolated", (time.time() - self.start) / 60.0, "minutes"
		self.timestepping()

		"""POROELASTIC PARAMETERS"""
		self.delta = 1e-6  # Dirichlet control regularization on bottom boundary
		self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
		self.B = 0.8  # Skempton's coefficient  ...Should this be constant?
		self.nuu = .4  # undrained poissons ratio
		self.nu = 0.25  # Poisson's ratio
		self.E = Expression('347 * pow(-x[2], 0.516) + 5.324e4', degree=1)
		self.mu = Expression('E / (2.0*(1.0 + nu))', E=self.E, nu=self.nu, degree=1)  # shear modulus
		self.lmbda = Expression('E*nu / ((1.0 + nu)*(1.0 - 2.0*nu))', E=self.E, nu=self.nu,
		                        degree=1)  # Lame's parameter
		self.alpha = 3 * (self.nuu - self.nu) / (self.B * (1 - 2 * self.nu) * (1 + self.nuu))  # Biot's coefficient
		self.Se = Expression('(9*(nuu-nu)*(1-(2*nuu)))/(2*mu*pow(B,2)*(1-2*nu)*pow((1+nuu),2))', nuu=self.nuu,
		                     nu=self.nu, mu=self.mu, B=self.B, degree=1)  # mass specific storage
		self.kstick = 8e-5
		self.stick = Expression('1/(1+exp(-kstick*(1.1e5-x[0])))', kstick=self.kstick, degree=1)
		self.d = self.u.geometric_dimension()  # number of space dimensions
		self.I = Identity(self.d)
		self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
		self.year = 60 * 60 * 24 * 365

		"""INITIAL AND BOUNDARY CONDITIONS"""
		p0d = 0  # initial pressure
		self.w_init = Expression(('p0d', '0', '0', '0'), p0d=p0d, degree=1)
		self.w0 = Function(self.ME)
		self.w0.interpolate(self.w_init)
		self.ub0 = Expression(('0', '0', '0'), degree=1)
		self.ub = Expression(('1', '1', '1'), degree=1)
		self.pb0 = Expression('1', degree=1)
		self.bc1 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 1)  # ocean surface
		self.bc2 = DirichletBC(self.ME.sub(1), self.ub, self.boundaries, 2)  # land surface
		self.bc3 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 3)  # subducting slab
		self.bc4 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 4)  # south side
		self.bc5 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 5)  # north side
		self.bc6 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 6)  # back side (no slip)
		self.bc7 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 7)  # mantle
		self.bcp1 = DirichletBC(self.ME.sub(0), ('0'), self.boundaries, 1)  # ocean surface (free flow condition)
		self.bcp2 = DirichletBC(self.ME.sub(0), ('1'), self.boundaries, 2)  # land surface

		if self.event == 'topo':
			self.topo_flow_interp()
			self.bcp2 = DirichletBC(self.ME.sub(0), self.p_topo, self.boundaries, 2)  # land surface
			self.bcs = [self.bc6, self.bc7, self.bcp1, self.bcp2]  # These are the BC's that are actually applied
		else:
			self.bcs = [self.bc6, self.bc7, self.bcp1]  # These are the BC's that are actually applied

		self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case

		if self.event == 'SSE':
			self.u_addx = self.u0_SSEx
			self.u_addy = self.u0_SSEy
		elif self.event == 'sub':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby
		elif self.event == 'sub_EQ':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby

		"""SET UP TO SAVE SOLUTION"""
		self.Sol_all = np.empty((self.dim, self.nsteps))  # matrix to save entire solution
		self.Sol_surf = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
		self.Sol_gps = np.empty((self.x_gps.shape[0] * self.geom_dim, self.nsteps))  # matrix to save surface solution
		self.sea_flux = np.empty((self.ocean_dof_pgrad.shape[0], self.nsteps))  # ocean bottom solution
		self.flow_velocity = np.empty((self.pdim * ndim, self.nsteps))  # ocean bottom solution
		self.vol_strain = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
		self.sea_flux_total = np.empty(self.nsteps)  # matrix to save surface post-seismic solution
		print "Solution matrices created", (time.time() - self.start) / 60.0, "minutes"

		pfilename = self.paraviewpath + "pressure3D_syn.pvd"
		p_velfilename = self.paraviewpath + "pressure_vel3D_syn.pvd"
		ufilename = self.paraviewpath + "def3D_syn.pvd"
		sigfilename = self.paraviewpath + "stress3D_syn.pvd"

		self.pfile = File(pfilename)
		self.p_velfile = File(p_velfilename)
		self.ufile = File(ufilename)
		self.sigfile = File(sigfilename)

		"""SET UP, RUN, SOLVE and SAVE"""
		self.assemble_system()
		print "System assembled", (time.time() - self.start) / 60.0, "minutes"
		self.solve_system()
		self.save_solution()

	def time_stepping(self):
		if self.event == 'EQ':
			"""15 min for 1/2 day,
			4 hrs for 10 days,
			12 hours for 30 days,
			2 days for 2 months,
			per month for 10 years
			"""
			self.dtt = [60, 240, 720, 288e1, 2 * 432e2]   # in minutes
			self.ntime = [24, 60, 60, 30, 60]  # , 50]
		elif self.event == 'SSE':
			"""
			12 hrs for SSE days,
			12 hrs for total time-SSE_days,
			2 months for 10 years
			"""
			self.dtt = [288e1 / 2, 432e2 * 2]
			self.ntime = [self.SSE_days / 1, (self.days - self.SSE_days) / 60]  # SSE_days, 10 years
		elif self.event == 'sub_EQ':
			"""
			1/year for 50 years followed by 'EQ' scheme
			"""
			self.dtt = [60 * 24 * 365, 15, 240, 288e1, 432e2, 60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years, 96, 120, 15, 122, 40]
		elif self.event == 'sub':  # 1/year for 50 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		elif self.event == 'topo':  # 1/year for 100 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		self.dtt = [i * 60 for i in self.dtt]  # in seconds
		self.dtt_repeated = np.repeat(self.dtt, self.ntime)
		self.dtt_comsum = np.cumsum(self.dtt_repeated)
		self.dt = self.dtt_comsum[0]
		self.nsteps = self.dtt_comsum.size

	def k_interpolation(self):
		log_kr = -18.0
		alpha_k = 0.6

		self.coords_K = self.Kspace.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
		bctest1 = DirichletBC(self.Kspace, (1), self.boundaries, 1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1), self.boundaries, 2)  # land surface
		ktest = Function(self.Kspace)
		bctest1.apply(ktest.vector())
		bctest2.apply(ktest.vector())
		self.dof_surfK = np.where(ktest.vector() == 1)[0]

		if self.permeability == 'mapped':
			kappa_surf = kappakriging(self.data_file, self.origin, self.theta,
			                          self.coords_K)

			k_surf_func = Function(self.Kspace)
			z_surf_func = Function(self.Kspace)
			k_surf_func.vector()[:] = kappa_surf
			z_surf_func.vector()[self.dof_surfK] = self.coords_K[self.dof_surfK, 2]

			self.kappa = Expression(KExpression3D_cpp, degree=1)
			self.kappa.alphak = alpha_k
			self.kappa.logkr = log_kr
			self.kappa.Ksurf = k_surf_func
			self.kappa.Zsurf = z_surf_func

		elif self.permeability == 'constant':
			k_surf_func = Function(self.Kspace)
			z_surf_func = Function(self.Kspace)
			k_surf_func.vector()[:] = self.surface_k
			z_surf_func.vector()[self.dof_surfK] = self.coords_K[self.dof_surfK, 2]

			self.kappa = Expression(KExpression3D_cpp, degree=1)
			self.kappa.alphak = alpha_k
			self.kappa.logkr = log_kr
			self.kappa.Ksurf = k_surf_func
			self.kappa.Zsurf = z_surf_func

	def strain(self, w):  # strain = 1/2 (grad u + grad u^T)
		return sym(nabla_grad(w))

	def sigma(self, w):  # stress = 2 mu strain + lambda tr(strain) I
		return 2.0 * self.mu * self.strain(w) + self.lmbda * div(w) * self.I

	def assemble_system(self):

		self.a_Mass = assemble(self.Se * self.p * self.q * dx)  # Mass matrix
		self.a_K = assemble((self.kappa / self.muf) * inner(nabla_grad(self.p),
		                                                    nabla_grad(self.q)) * dx)  # Stiffness matrix
		self.a_Div = assemble(nabla_div(self.u) * self.q * dx)  # Divergence matrix
		self.a_Grad = assemble(-self.p * (nabla_div(self.v)) * dx)  # Gradient
		self.a_E = assemble(inner(self.sigma(self.u), nabla_grad(self.v)) * dx)  # Elasticity
		self.a_Boundary = assemble((1 / self.delta) * dot(self.u, self.n) * dot(self.v, self.n)
		                           * self.ds(self.slab) + (1 / self.delta) *
		                           inner(self.T * self.u, self.T * self.v) * self.ds(self.slab))

		self.A = self.a_Mass + self.dt * self.a_K + self.alpha * self.a_Div \
		         + self.alpha * self.a_Grad + self.a_E + self.a_Boundary  # Left hand side (LHS)
		self.L = self.a_Mass + self.alpha * self.a_Div

	def extract_coords(self):

		if not os.path.exists(self.var_path):
			os.makedirs(self.var_path)

		self.coordinates = self.ME.tabulate_dof_coordinates().reshape((self.dim, -1))
		self.x_all, self.y_all, self.z_all = self.coordinates[:, 0], self.coordinates[:, 1], self.coordinates[:, 2]

		self.coords_Q = self.Qspace.tabulate_dof_coordinates().reshape((self.pdim, -1))
		self.coords_V = self.Vspace.tabulate_dof_coordinates().reshape((self.udim, -1))

		bctest0 = DirichletBC(self.ME, (1, 1, 1, 1), self.boundaries, 1)  # ocean surface
		bctest1 = DirichletBC(self.ME, (2, 2, 2, 2), self.boundaries, 2)  # top surface

		ptest = Function(self.ME)
		bctest0.apply(ptest.vector())
		bctest1.apply(ptest.vector())
		self.ocean_dofs = np.where(ptest.vector() == 1)[0]
		self.surface_dofs = np.where(ptest.vector() == 2)[0]

		self.size_p = self.pdim
		self.size_u = self.dim - self.size_p

		self.ocean_dof_p = self.ocean_dofs[np.where(self.ocean_dofs < self.size_p)]
		self.ocean_dof_u = self.ocean_dofs[np.where(self.ocean_dofs >= self.size_p)] - self.size_p
		self.ocean_dof_pgrad = np.hstack((self.ocean_dof_p, (self.ocean_dof_p + self.ocean_dof_p.shape[0]),
		                                  (self.ocean_dof_p + 2 * self.ocean_dof_p.shape[0])))
		self.surface_dof_p = self.surface_dofs[np.where(self.surface_dofs < self.size_p)]
		self.surface_dof_u = self.surface_dofs[np.where(self.surface_dofs >= self.size_p)] - self.size_p
		self.surface_dof_pgrad = np.hstack((self.surface_dof_p, (self.surface_dof_p + self.surface_dof_p.shape[0]),
		                                    (self.surface_dof_p + 2 * self.surface_dof_p.shape[0])))

		self.size_gps = self.x_gps.shape[0]
		size_w_gps = 4 * self.size_gps
		indices_gps = np.empty((self.size_gps, 4))
		for j in range(0, self.size_gps):
			indice_gps = np.where((self.x_all[:] == self.x_gps[j]) & (self.y_all[:] == self.y_gps[j]))[0]
			indices_gps[j, :] = indice_gps
			self.GPS_dofs = indices_gps.reshape((size_w_gps), order='F')

		np.save(self.var_path + "x_all_3D", self.x_all)
		np.save(self.var_path + "y_all_3D", self.y_all)
		np.save(self.var_path + "z_all_3D", self.z_all)
		np.save(self.var_path + "GPS_dofs_3D", self.GPS_dofs)
		np.save(self.var_path + "surface_dofs_3D", self.surface_dofs)
		np.save(self.var_path + "ocean_dofs_3D", self.ocean_dofs)
		np.save(self.var_path + "size_p_3D", self.size_p)
		np.save(self.var_path + "size_u_3D", self.size_u)

	def solve_system(self):
		count = 0
		if self.event == 'EQ':
			"""For an EQ, this can be outside the time loop
			since we only use it for the first time step"""
			u0slab = Expression((
			                    'u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
			                    'u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
			                    '0'), u0d=self.u0_EQx, u0s=self.u0_EQy, xcenter=self.xcenter, ycenter=self.ycenter,
			                    sigmax=self.sigma_bx, sigmay=self.sigma_by, degree=1)
			a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
				self.slab))  # Prescribed slip boundary term
		for i in range(self.nsteps):

			if i == 0:
				dt = self.dtt_comsum[i]
			else:
				dt = self.dtt_comsum[i] - self.dtt_comsum[i - 1]
			if i > 2:
				dt_old = self.dtt_comsum[i - 1] - self.dtt_comsum[i - 2]
			if i > 2 and dt != dt_old:
				self.A = self.a_Mass + dt * self.a_K + self.alpha * self.a_Div + self.alpha * self.a_Grad + self.a_E + self.a_Boundary  # Left hand side (LHS)

			if i > 0:
				"""Add slip to the boundary for a SSE
				and move center of slip along boundary"""
				if self.event == 'SSE' and count <= self.SSE_days:
					self.u0d_SSE += self.u_addx
					self.u0s_SSE += self.u_addy
					self.xcenter += self.SS_migratex
					self.ycenter += self.SS_migratey
				elif self.event == 'sub':
					self.u0_subx += self.u_addx
					self.u0_suby += self.u_addy
				elif self.event == 'sub_EQ':
					self.u0_subx += self.u_addx * (dt / self.year)
					self.u0_suby += self.u_addy * (dt / self.year)

			"""Slip boundary Expression"""
			if self.event == 'SSE':
				u0slab = Expression((
					'u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
					'u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
					'0'), u0d=self.u0d_SSE, u0s=self.u0s_SSE, xcenter=self.xcenter,
					ycenter=self.ycenter, sigmax=self.sigma_bx, sigmay=self.sigma_by, degree=1)
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
					self.slab))
			elif self.event == 'sub':
				u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                    u0_suby=self.u0_suby, stick=self.stick, degree=1)
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
					self.slab))
			elif self.event == 'sub_EQ' and i < self.sub_cycle_years:
				u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                    u0_suby=self.u0_suby, stick=self.stick, degree=1)
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
					self.slab))
			elif self.event == 'sub_EQ' and i >= self.sub_cycle_years:
				u0slab = Expression(('u0_subx*stick + u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + '
				                     '(pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
				                     'u0_suby*stick + u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) '
				                     '+ (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
				                     '0'), u0d=self.u0_subx, u0s=self.u0_suby, xcenter=self.xcenter,
				                    ycenter=self.ycenter, sigmax=self.sigma_bx, sigmay=self.Tsigma_by, degree=1)
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) * self.ds(
					self.slab))

			print "Boundary condition set", (time.time() - self.start) / 60.0, "minutes"

			b = Vector()
			self.a_Mass.init_vector(b, 0)
			self.L.mult(self.w0.vector(), b)
			if self.event != 'topo':
				b += a_Boundary0
			[bc.apply(b) for bc in self.bcs]
			[bc.apply(self.A) for bc in self.bcs]
			print "BC's applied", (time.time() - self.start) / 60.0, "minutes"
			self.w = Function(self.ME)
			self.solver.set_operator(self.A)
			self.solver.solve(self.w.vector(), b)
			print "System solved", (time.time() - self.start) / 60.0, "minutes"
			p, u = self.w.split()

			self.domain_flux = (-self.kappa / self.muf) * grad(p)  # element-wise flux vector
			self.bound_flux = assemble(inner(self.domain_flux, self.n) * self.ds(1))
			self.bound_flux2 = assemble(inner(self.domain_flux, self.n) * self.ds(2))
			print "outward flux through seafloor: ", self.bound_flux
			print "outward flux through land surface: ", self.bound_flux2

			self.q_velocity = Function(self.Q_gspace)
			self.q_velocity.assign(project(self.domain_flux, self.Q_gspace))
			print "Computed velocity field", (time.time() - self.start) / 60.0, "minutes"

			self.v_strain = Function(self.Qspace)
			self.v_strain.assign(project(nabla_div(u), self.Qspace))
			self.psi = self.streamfunction3D()  # compute numerical streamfunction

			self.stress = Function(self.Wspace)
			self.stress.assign(project(self.sigma(u), self.Wspace))

			"""SAVE SOLUTION"""
			p.rename('p', 'p')
			self.pfile << p
			u.rename('u', 'u')
			self.p_velfile << self.q_velocity
			self.ufile << u
			self.sigfile << self.stress

			self.sea_flux_total[count] = self.bound_flux
			self.sea_flux[:, count] = self.q_velocity.vector()[self.ocean_dof_pgrad]
			self.flow_velocity[:, count] = self.q_velocity.vector()
			self.Psi[:, count] = self.psi.vector()
			self.vol_strain[:, count] = self.v_strain.vector()
			self.Sol_surf[:, count] = self.w.vector()[self.surface_dofs]
			self.Sol_all[:, count] = self.w.vector()

			self.w0 = self.w
			count += 1
			print "Timestep:", count

	def save_solution(self):

		if not os.path.exists(self.results_path):
			os.makedirs(self.results_path)
		if self.permeability == 'mapped':
			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_mapped.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_mapped.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_EQ_mapped.npy", self.Sol_all)
			elif self.event == 'EQ':
				np.save(self.results_path + "Sol_surf_3D_EQ_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_EQ_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_EQ_mapped.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_EQ_mapped.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_EQ_mapped.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_EQ_mapped.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_EQ_mapped.npy", self.sea_flux_total)
		elif self.permeability == 'constant':
			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_constant.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_constant.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_EQ_constant.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_sub_EQ_constant.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_sub_EQ_constant.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_sub_EQ_constant.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_sub_EQ_constant.npy", self.sea_flux_total)
			elif self.event == 'EQ':
				np.save(self.results_path + "Sol_surf_3D_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_EQ_constant.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_EQ_constant.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_EQ_constant.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_EQ_constant.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_EQ_constant.npy", self.sea_flux_total)

class Poroelasticity3D:
	def __init__(self, data_file, origin, theta, ndim, mesh, boundaries, plot_figs, event, permeability,
	             sub_cycle_years, u0_subdip, u0_substrike, surface_k, ocean_k, **kwargs):

		self.start = time.time()
		self.data_file = data_file
		self.origin = origin
		self.theta = theta
		self.mesh = mesh
		self.boundaries = boundaries
		self.plot_figs = plot_figs
		self.event = event
		self.permeability = permeability
		self.sub_cycle_years = sub_cycle_years
		self.u0_subx = u0_subdip
		self.u0_suby = u0_substrike
		self.surface_k = surface_k
		self.ocean_k = ocean_k
		# solver = PETScKrylovSolver("gmres","petsc_amg")
		# solver = PETScKrylovSolver("cg","ilu")
		self.solver = LinearSolver('mumps')
		self.prm = self.solver.parameters
		# self.prm["form_compiler"]["representation"] = "uflacs"
		# self.prm['reuse_factorization'] = True  # Saves Factorization of LHS
		self.V = VectorElement('Lagrange', self.mesh.ufl_cell(), 2)
		self.Q = FiniteElement('Lagrange', self.mesh.ufl_cell(), 1)
		self.ME = FunctionSpace(self.mesh, MixedElement([self.Q, self.V]))

		self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
		self.Kspace = FunctionSpace(self.mesh, self.Kelement)

		self.Vspace = FunctionSpace(self.mesh, self.V)
		self.Qspace = FunctionSpace(self.mesh, self.Q)
		self.Q_gspace = VectorFunctionSpace(self.mesh, "DG", 0)
		self.Q_gspace1 = VectorFunctionSpace(self.mesh, "CG", 1)
		self.Wspace = TensorFunctionSpace(self.mesh, "Lagrange", 1)
		(self.p, self.u) = TrialFunctions(self.ME)
		(self.q, self.v) = TestFunctions(self.ME)
		self.ds = Measure("ds", domain=self.mesh, subdomain_data=self.boundaries)  #
		self.n = FacetNormal(self.mesh)  # normal vector
		self.dim = self.ME.dim()
		self.pdim = self.ME.sub(0).dim()
		self.udim = self.ME.sub(1).dim()
		self.geom_dim = ndim + 1
		
		"""GPS STATION LOCATIONS"""
		wb = pyxl.load_workbook(self.data_file)
		gps_data = wb['GPS_data']
		
		self.x_gps, self.y_gps = latlon2XY(gps_data, self.origin, self.theta)
		self.z_gps = np.zeros(len(self.x_gps))

		self.results_path = 'results/numpy_results/'
		self.var_path = 'results/numpy_variables/'
		self.paraviewpath = 'results/paraview/'
		self.inverse_results = '/fenics/shared/inverse_codes/results_deterministic/numpy_results/'
		self.inverse_var = '/fenics/shared/inverse_codes/results_deterministic/numpy_variables/'
		
		if not os.path.exists(self.paraviewpath):
			os.makedirs(self.paraviewpath)

		self.extract_coords()
		print "Coordinates extracted", (time.time() - self.start) / 60.0, "minutes"
		self.k_interpolation()
		print "Permeability interpolated", (time.time() - self.start) / 60.0, "minutes"
		self.timestepping()

		"""POROELASTIC PARAMETERS"""
		self.delta = 1e-6  # Dirichlet control regularization on bottom boundary
		self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
		self.B = .6  # Skempton's coefficient
		self.nuu = .38  # undrained poissons ratio
		self.nu = 0.27  # Poisson's ratio
		self.E = Expression('347 * pow(-x[2], 0.516) + 5.324e4', degree=1)
		self.mu = Expression('E / (2.0*(1.0 + nu))', E=self.E, nu=self.nu, degree=1)  # shear modulus
		self.lmbda = Expression('E*nu / ((1.0 + nu)*(1.0 - 2.0*nu))', E=self.E, nu=self.nu,
		                        degree=1)  # Lame's parameter
		self.alpha = 3 * (self.nuu - self.nu) / (self.B * (1 - 2 * self.nu) * (1 + self.nuu))  # Biot's coefficient
		self.Se = Expression('(9*(nuu-nu)*(1-(2*nuu)))/(2*mu*pow(B,2)*(1-2*nu)*pow((1+nuu),2))', \
		                     nuu=self.nuu, nu=self.nu, mu=self.mu, B=self.B, degree=1)  # mass specific storage
		self.kstick = 8e-5
		self.stick = Expression('1/(1+exp(-kstick*(1.1e5-x[0])))', kstick=self.kstick, degree=1)
		self.d = self.u.geometric_dimension()  # number of space dimensions
		self.I = Identity(self.d)
		self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
		self.year = 60 * 60 * 24 * 365


		""" INITIAL AND BOUNDARY CONDITIONS """
		####### Initial conditions: (('p0','ux0','uy0')) ##########
		p0d = 0  # initial pressure
		self.w_init = Expression(('p0d', '0', '0', '0'), p0d=p0d, degree=1)
		self.w0 = Function(self.ME)
		self.w0.interpolate(self.w_init)
		self.ub0 = Expression(('0', '0', '0'), degree=1)
		self.ub = Expression(('1', '1', '1'), degree=1)
		self.pb0 = Expression('1', degree=1)
		self.bc1 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 1)  # ocean surface
		self.bc2 = DirichletBC(self.ME.sub(1), self.ub, self.boundaries, 2)  # land surface
		self.bc3 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 3)  # subducting slab
		self.bc4 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 4)  # south side
		self.bc5 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 5)  # north side
		self.bc6 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 6)  # back side (no slip)
		self.bc7 = DirichletBC(self.ME.sub(1), self.ub0, self.boundaries, 7)  # mantle
		self.bcp1 = DirichletBC(self.ME.sub(0), ('0'), self.boundaries, 1)  # ocean surface (free flow condition)

		if self.event == 'topo':
			self.topo_flow_interp()
			self.bcp2 = DirichletBC(self.ME.sub(0), self.p_topo, self.boundaries, 2)  # land surface
			self.bcs = [self.bc7, self.bcp1, self.bcp2]  # These are the BC's that are actually applied
		else:
			self.bcs = [self.bc6, self.bc7, self.bcp1]  # These are the BC's that are actually applied

		self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case

		if self.event == 'SSE':
			self.u_addx = self.u0_SSEx
			self.u_addy = self.u0_SSEy
		elif self.event == 'sub':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby
		elif self.event == 'sub_EQ':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby

		"""SET UP TO SAVE SOLUTION"""
		self.Sol_all = np.empty((self.dim, self.nsteps))  # matrix to save entire solution
		self.Sol_surf = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
		self.Sol_gps = np.empty((self.x_gps.shape[0] * self.geom_dim, self.nsteps))  # matrix to save surface solution
		self.sea_flux = np.empty((self.ocean_dof_pgrad.shape[0], self.nsteps))  # ocean bottom solution
		self.flow_velocity = np.empty((self.pdim * ndim, self.nsteps))  # ocean bottom solution
		self.vol_strain = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
		self.sea_flux_total = np.empty(self.nsteps)  # matrix to save surface post-seismic solution

		pfilename = self.paraviewpath + "pressure3D.pvd"
		p_velfilename = self.paraviewpath + "flux0_comp3D.pvd"
		ufilename = self.paraviewpath + "def3D.pvd"
		sigfilename = self.paraviewpath + "stress3D.pvd"

		self.pfile = File(pfilename)
		self.qfile = File(p_velfilename)
		self.ufile = File(ufilename)
		self.sigfile = File(sigfilename)
		print "Solution matrices created", (time.time() - self.start) / 60.0, "minutes"

		"""SET UP, RUN, SOLVE and SAVE"""
		self.slip_interpolation()
		print "Slip data interpolated", (time.time() - self.start) / 60.0, "minutes"
		self.assemble_system()
		print "System assembled", (time.time() - self.start) / 60.0, "minutes"
		self.solve_system()
		self.save_solution()

	def extract_coords(self):
		if not os.path.exists(self.var_path):
			os.makedirs(self.var_path)

		self.coordinates = self.ME.tabulate_dof_coordinates().reshape((self.dim, -1))
		self.x_all, self.y_all, self.z_all = self.coordinates[:, 0], self.coordinates[:, 1], self.coordinates[:, 2]

		self.coords_Q = self.Qspace.tabulate_dof_coordinates().reshape((self.pdim, -1))
		self.coords_V = self.Vspace.tabulate_dof_coordinates().reshape((self.udim, -1))

		bctest0 = DirichletBC(self.ME, (1, 1, 1, 1), self.boundaries, 1)  # ocean surface
		bctest1 = DirichletBC(self.ME, (2, 2, 2, 2), self.boundaries, 2)  # top surface

		ptest = Function(self.ME)
		bctest0.apply(ptest.vector())
		bctest1.apply(ptest.vector())
		self.ocean_dofs = np.where(ptest.vector() == 1)[0]
		self.surface_dofs = np.where(ptest.vector() == 2)[0]

		self.size_p = self.pdim
		self.size_u = self.dim - self.size_p

		self.ocean_dof_p = self.ocean_dofs[np.where(self.ocean_dofs < self.size_p)]
		self.ocean_dof_u = self.ocean_dofs[np.where(self.ocean_dofs >= self.size_p)] - self.size_p
		self.ocean_dof_pgrad = np.hstack((self.ocean_dof_p, (self.ocean_dof_p + self.ocean_dof_p.shape[0]),
		                                  (self.ocean_dof_p + 2 * self.ocean_dof_p.shape[0])))
		self.surface_dof_p = self.surface_dofs[np.where(self.surface_dofs < self.size_p)]
		self.surface_dof_u = self.surface_dofs[np.where(self.surface_dofs >= self.size_p)] - self.size_p
		self.surface_dof_pgrad = np.hstack((self.surface_dof_p, (self.surface_dof_p + self.surface_dof_p.shape[0]),
		                                    (self.surface_dof_p + 2 * self.surface_dof_p.shape[0])))

		self.size_gps = self.x_gps.shape[0]
		size_w_gps = 4 * self.size_gps
		indices_gps = np.empty((self.size_gps, 4))
		for j in range(0, self.size_gps):
			indice_gps = np.where((self.x_all[:] == self.x_gps[j]) & (self.y_all[:] == self.y_gps[j]))[0]
			indices_gps[j, :] = indice_gps
			self.GPS_dofs = indices_gps.reshape((size_w_gps), order='F')

		np.save(self.var_path + "x_all_3D", self.x_all)
		np.save(self.var_path + "y_all_3D", self.y_all)
		np.save(self.var_path + "z_all_3D", self.z_all)
		np.save(self.var_path + "GPS_dofs_3D", self.GPS_dofs)
		np.save(self.var_path + "surface_dofs_3D", self.surface_dofs)
		np.save(self.var_path + "ocean_dofs_3D", self.ocean_dofs)
		np.save(self.var_path + "size_p_3D", self.size_p)
		np.save(self.var_path + "size_u_3D", self.size_u)

	def time_stepping(self):
		if self.event == 'EQ':
			"""15 min for 1/2 day,
			4 hrs for 10 days,
			12 hours for 30 days,
			2 days for 2 months,
			per month for 10 years
			"""
			self.dtt = [60, 240, 720, 288e1, 2 * 432e2]   # in minutes
			self.ntime = [24, 60, 60, 30, 60]  # , 50]
		elif self.event == 'SSE':
			"""
			12 hrs for SSE days,
			12 hrs for total time-SSE_days,
			2 months for 10 years
			"""
			self.dtt = [288e1 / 2, 432e2 * 2]
			self.ntime = [self.SSE_days / 1, (self.days - self.SSE_days) / 60]  # SSE_days, 10 years
		elif self.event == 'sub_EQ':
			"""
			1/year for 50 years followed by 'EQ' scheme
			"""
			self.dtt = [60 * 24 * 365, 15, 240, 288e1, 432e2, 60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years, 96, 120, 15, 122, 40]
		elif self.event == 'sub':  # 1/year for 50 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		elif self.event == 'topo':  # 1/year for 100 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		self.dtt = [i * 60 for i in self.dtt]  # in seconds
		self.dtt_repeated = np.repeat(self.dtt, self.ntime)
		self.dtt_comsum = np.cumsum(self.dtt_repeated)
		self.dt = self.dtt_comsum[0]
		self.nsteps = self.dtt_comsum.size

	def k_interpolation(self):
		log_kr = -18.0
		alpha_k = 0.6

		self.coords_K = self.Kspace.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
		bctest1 = DirichletBC(self.Kspace, (1), self.boundaries, 1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1), self.boundaries, 2)  # land surface
		ktest = Function(self.Kspace)
		bctest1.apply(ktest.vector())
		bctest2.apply(ktest.vector())
		self.dof_surfK = np.where(ktest.vector() == 1)[0]

		if self.permeability == 'mapped':
			kappa_surf = kappakriging(self.data_file, self.origin, self.theta,
			                          self.coords_K)

			k_surf_func = Function(self.Kspace)
			z_surf_func = Function(self.Kspace)
			k_surf_func.vector()[:] = kappa_surf
			z_surf_func.vector()[self.dof_surfK] = self.coords_K[self.dof_surfK, 2]

			self.kappa = Expression(KExpression3D_cpp, degree=1)
			self.kappa.alphak = alpha_k
			self.kappa.logkr = log_kr
			self.kappa.Ksurf = k_surf_func
			self.kappa.Zsurf = z_surf_func

		elif self.permeability == 'constant':
			k_surf_func = Function(self.Kspace)
			z_surf_func = Function(self.Kspace)
			k_surf_func.vector()[:] = self.surface_k
			z_surf_func.vector()[self.dof_surfK] = self.coords_K[self.dof_surfK, 2]

			self.kappa = Expression(KExpression3D_cpp, degree=1)
			self.kappa.alphak = alpha_k
			self.kappa.logkr = log_kr
			self.kappa.Ksurf = k_surf_func
			self.kappa.Zsurf = z_surf_func
			
	def topo_flow_interp(self):

		self.coords_K = self.Kspace.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
		bctest1 = DirichletBC(self.Kspace, (1), self.boundaries, 1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1), self.boundaries, 2)  # land surface
		qtest = Function(self.Kspace)
		bctest1.apply(qtest.vector())
		bctest2.apply(qtest.vector())
		self.dof_surfK = np.where(qtest.vector() == 1)[0]

		wb = pyxl.load_workbook(self.data_file)
		topo_data = wb['topo_data3D']

		topo_p_surf = topokriging(topo_data, self.origin, self.theta,
		                          self.coords_K, self.dof_surfK, self.var_path)

		low_p = np.where(topo_p_surf < 0.02)[0]
		topo_p_surf[low_p] = 0.02
		p_surf_func = Function(self.Kspace)
		p_surf_func.vector()[self.dof_surfK] = topo_p_surf
		self.p_topo = Expression(('ptopo'), ptopo=p_surf_func, degree=1)

	def slip_interpolation(self):
		
		if self.EQtype == 'inversion':
			
			u0load = np.load(self.inverse_results + 'u0_slab_array.npy')
			size_u = int(u0load.shape[0] / 3)
			u0load = u0load.reshape((size_u, 3), order='F')
			XYZslab = np.load(self.inverse_var + 'XYZ_slab.npy')[0:size_u, :]
			
			u0slab_interp = slabkriging(XYZslab, u0load, self.xvec, self.yvec)
			self.u0slab = Function(self.Q_gspace)
			uslab_func.vector()[:] = u0slab_interp
			self.p_topo = Expression(('uslab'), uslab=uslab_func, degree=1)
		
		elif self.EQtype == 'data':

			bctest1 = DirichletBC(self.Vspace, (1, 1, 1), self.boundaries, 3)  # slab surface
			stest = Function(self.Vspace)
			bctest1.apply(stest.vector())
			dof_slab = np.where(stest.vector() == 1)[0]
			wb = pyxl.load_workbook(self.data_file)
			slip_data = wb['slip_data']
	
			self.u0slab = SlipExpression(slip_data, self.origin, self.theta,
			                             self.coords_V, dof_slab, self.var_path,
			                             degree=1)

	def strain(self, w):  # strain = 1/2 (grad u + grad u^T)
		return sym(nabla_grad(w))

	def sigma(self, w):  # stress = 2 mu strain + lambda tr(strain) I
		return 2.0 * self.mu * self.strain(w) + self.lmbda * div(w) * self.I

	def assemble_system(self):

		self.a_Mass = assemble(self.Se * self.p * self.q * dx)  # Mass matrix
		self.a_K = assemble((self.kappa / self.muf) * inner(nabla_grad(self.p),
		                                                    nabla_grad(self.q)) * dx)  # Stiffness matrix
		self.a_Div = assemble(nabla_div(self.u) * self.q * dx)  # Divergence matrix
		self.a_Grad = assemble(-self.p * (nabla_div(self.v)) * dx)  # Gradient
		self.a_E = assemble(inner(self.sigma(self.u), nabla_grad(self.v)) * dx)  # Elasticity
		self.a_Boundary = assemble((1 / self.delta) * dot(self.u, self.n) * dot(self.v, self.n)
		                           * self.ds(self.slab) + (1 / self.delta) *
		                           inner(self.T * self.u, self.T * self.v) * self.ds(self.slab))

		# Left hand side (LHS)
		self.A = self.a_Mass + self.dt * self.a_K + self.alpha * self.a_Div \
		         + self.alpha * self.a_Grad + self.a_E + self.a_Boundary

		# Right Hand side
		self.L = self.a_Mass + self.alpha * self.a_Div

	def test_PSD(self):

		def is_pd(x):
			return np.all(np.linalg.eigvals(x) > 0)

		def is_psd(x, tol=1e-5):
			E, V = np.linalg.eigh(x)
			return np.all(E > -tol)

		def is_symmetric(x):
			return (x.transpose() == x).all()

		A = self.A.array()

		print('A is {}'.format('symmetric' if is_symmetric(A) else ('not symmetric')))

		np.linalg.cholesky(A)

	def solve_system(self):
		count = 0
		if self.event == 'EQ':
			"""
			For an EQ, this can be outside the time loop
			since we only use it for the first time step
			"""
			a_Boundary0 = assemble((1 / self.delta) * inner(self.T * self.u0slab, self.T * self.v) * self.ds(
				self.slab))  # Prescribed slip boundary term
		for i in range(self.nsteps):

			if i == 0:
				dt = self.dtt_comsum[i]
			else:
				dt = self.dtt_comsum[i] - self.dtt_comsum[i - 1]
			if i > 1:
				dt_old = self.dtt_comsum[i - 1] - self.dtt_comsum[i - 2]
			if i > 1 and dt != dt_old:
				self.A = self.a_Mass + dt * self.a_K + self.alpha * self.a_Div + self.alpha * self.a_Grad + self.a_E + self.a_Boundary  # Left hand side (LHS)

			if i > 0:
				if self.event == 'sub':
					self.u0_subx += self.u_addx * (dt / self.year)
					self.u0_suby += self.u_addy * (dt / self.year)
				elif self.event == 'sub_EQ':
					self.u0_subx += self.u_addx * (dt / self.year)
					self.u0_suby += self.u_addy * (dt / self.year)

			"""Slip boundary Expression"""
			if self.event == 'SSE':
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * self.u0_SSE, self.T * self.v) * self.ds(
					self.slab))

			elif self.event == 'sub':
				self.u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                         u0_suby=self.u0_suby, stick=self.stick, degree=1)
				a_Boundary0 = assemble((1 / self.delta) * inner(self.T * self.u0slab, self.T * self.v) * self.ds(
					self.slab))

			elif self.event == 'sub_EQ':
				u0slabsub = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                       u0_suby=self.u0_suby, stick=self.stick, degree=1)

				if i < self.sub_cycle_years:
					a_Boundary0 = assemble(
						(1 / self.delta) * inner(self.T * u0slabsub, self.T * self.v) * self.ds(self.slab))

				if i >= self.sub_cycle_years:
					u0slabtotal = self.u0slabsub + self.u0slab
					a_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slabtotal, self.T * self.v) * self.ds(
						self.slab))
			print "Boundary condition set", (time.time() - self.start) / 60.0, "minutes"

			b = Vector()
			self.a_Mass.init_vector(b, 0)
			self.L.mult(self.w0.vector(), b)
			if self.event != 'topo':
				b += a_Boundary0
			[bc.apply(b) for bc in self.bcs]
			[bc.apply(self.A) for bc in self.bcs]
			print "BC's applied", (time.time() - self.start) / 60.0, "minutes"
			self.w = Function(self.ME)
			self.solver.set_operator(self.A)
			self.solver.solve(self.w.vector(), b)
			print "System solved", (time.time() - self.start) / 60.0, "minutes"
			p, u = self.w.split()

			self.domain_flux = (-self.kappa / self.muf) * grad(p)  # element-wise flux vector
			self.bound_flux = assemble(inner(self.domain_flux, self.n) * self.ds(1))
			self.bound_flux2 = assemble(inner(self.domain_flux, self.n) * self.ds(2))

			q_comp = Function(self.Q_gspace)
			q_comp.assign(project(self.domain_flux, self.Q_gspace, solver_type='mumps'))
			q_comp1 = Function(self.Q_gspace1)
			q_comp1.assign(project(self.domain_flux, self.Q_gspace1, solver_type='mumps'))
			print "Computed velocity field", (time.time() - self.start) / 60.0, "minutes"

			self.v_strain = Function(self.Qspace)
			self.v_strain.assign(project(nabla_div(u), self.Qspace, solver_type='mumps'))
			self.stress = Function(self.Wspace)
			self.stress.assign(project(self.sigma(u), self.Wspace))

			""" SAVE SOLUTION """
			p.rename('p', 'p')
			self.pfile << p
			q_comp.rename('q_comp', 'q_comp')
			self.qfile << q_comp
			q_comp1.rename('q_comp1', 'q_comp1')
			self.qfile1 << q_comp1
			u.rename('u', 'u')
			self.ufile << u
			self.sigfile << self.stress

			self.sea_flux_total[count] = self.bound_flux
			self.sea_flux[:, count] = self.q_velocity.vector()[self.ocean_dof_pgrad]
			self.flow_velocity[:, count] = self.q_velocity.vector()
			self.vol_strain[:, count] = self.v_strain.vector()
			self.Sol_surf[:, count] = self.w.vector()[self.surface_dofs]
			self.Sol_all[:, count] = self.w.vector()
			self.Sol_gps[:, count] = self.w.vector()[self.GPS_dofs]
			self.w0 = self.w
			count += 1
			print "Timestep:", count

	def save_solution(self):
		if not os.path.exists(self.results_path):
			os.makedirs(self.results_path)
		if self.permeability == 'mapped':

			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_mapped.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_mapped.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_EQ_mapped.npy", self.Sol_all)
			elif self.event == 'EQ':
				np.save(self.var_path + "dtt_comsum_EQ.npy", self.dtt_comsum)
				np.save(self.results_path + "Sol_surf_3D_EQ_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_EQ_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_EQ_mapped.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_EQ_mapped.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_EQ_mapped.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_EQ_mapped.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_EQ_mapped.npy", self.sea_flux_total)
			elif self.event == 'topo':
				np.save(self.results_path + "Sol_surf_3D_topo_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_all_3D_topo_mapped.npy", self.Sol_all)
				np.save(self.results_path + "sea_flux_3D_topo_mapped.npy", self.sea_flux)
				np.save(self.var_path + "dtt_comsum_topo_mapped.npy", self.dtt_comsum)
				np.save(self.results_path + "flow_velocity_3D_topo_mapped.npy", self.flow_velocity)
				np.save(self.results_path + "vol_strain_3D_topo_mapped.npy", self.vol_strain)
				np.save(self.results_path + "sea_flux_total_3D_topo_mapped.npy", self.sea_flux_total)
		elif self.permeability == 'constant':
			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_constant.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_constant.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_EQ_constant.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_sub_EQ_constant.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_sub_EQ_constant.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_sub_EQ_constant.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_sub_EQ_constant.npy", self.sea_flux_total)
			elif self.event == 'EQ':
				np.save(self.results_path + "Sol_surf_3D_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_EQ_constant.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_EQ_constant.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_EQ_constant.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_EQ_constant.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_EQ_constant.npy", self.sea_flux_total)
