# Kimmy McCormack

# Functions used by 'model_run_file.py' to solve coupled set of poroelastic equations in both 2D
# and 3D subduction zone domains

# Units are in meters and MPa

import sys
sys.path.append('/home/fenics/shared/local_lib')
#sys.path.append('/workspace/kimmy/local_lib')
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as pyxl
import os
import scipy as sp
import scipy.interpolate
import shutil
from dolfin import *
from scipy.interpolate import griddata
from pykrige.ok import OrdinaryKriging
import pykrige.kriging_tools as kt
import time

import gc
gc.collect()

##################### Optimization options ############################
dolfin.parameters.reorder_dofs_serial = False

parameters['form_compiler']['optimize'] = True
parameters["form_compiler"]["cpp_optimize"] = True
#parameters["form_compiler"]["representation"] = "uflacs"
#ffc_options = {"optimize": True, "eliminate_zeros": True,"precompute_basis_const": True, "precompute_ip_const": True}
parameters['allow_extrapolation'] = True


#################### Data interpolation functions ######################

def kappakriging(k_ocean, data_file, origin, theta, log_kr, alpha_k, coords_K, dof_surf, **kwargs):
    """Interpolates surface permeability data by kriging. Take surface value to interpolate kappa with depth.

        INPUTS:
            [k_ocean] = log(ocean permebility) [m^2]
            [data_sheet] = excel sheet containing pointwise permeability data in (lon, lat, permaebility[m^2]) format.
                    Default is to ignore the first row (containing headers).
            [origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
            [theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
                    counter-clockwise from latlon to XY
            [log_kr] = log of residual permeabiltiy at depth (usually betwee -18 and -20) [m^2]
            [alpha_k] = decay constant of permeability as a function of depth (between 0.25 and 1.6)
            [coords_K] = coordinates of degrees of freedom onto which the permeability is mapped
            [dof_surf] = the degrees of freedom of the surface of the mesh

            OrdinaryKriging function takes variogram parameters as:
                linear - [slope, nugget]
                power - [scale, exponent, nugget]
            `   gaussian - [sill, range, nugget]
                spherical - [sill, range, nugget]
                exponential - [sill, range, nugget]
                hole-effect - [sill, range, nugget]

        OUPUTS:
            kappa [m^2] - the interpolated 3D permeabiltiy field in the form of a FEniCS Expression()
    """

    x_surf, y_surf, z_surf = coords_K[dof_surf, 0], coords_K[dof_surf, 1], coords_K[dof_surf, 2]
    x_all, y_all, z_all = coords_K[:, 0], coords_K[:, 1], coords_K[:, 2]

    wb = pyxl.load_workbook(data_file)
    well_data = wb['well_data']

    x_wells, y_wells = latlon2XY(well_data, origin, theta)
    permeability = np.array(
        [[well_data.cell(row=i, column=5).value] for i in range(2, well_data.max_row + 1)]).reshape(
        well_data.max_row - 1, )
    logk = np.log10(permeability)

    # PyKrige interpolation

    # OK = OrdinaryKriging(x_wells, y_wells, logk, variogram_model = 'power', variogram_parameters = [1, 1, 2e4])
    OK = OrdinaryKriging(x_wells, y_wells, logk, variogram_model='exponential',
                         variogram_parameters=[1.086, 1.28e4, 0.771])

    # k_surf, ss = OK.execute('points', x_surf, y_surf)
    k_surf, ss = OK.execute('points', x_all, y_all)

    np.save("results/numpy_variables/k_surf", k_surf.compressed())
    # np.save("saved_variables_poroelastic/ss_krige", self.ss.compressed())
    # self.k_surf = np.load("saved_variables_poroelastic/k_surf_med.npy")
    # self.ss = np.load("saved_variables_poroelastic/ss_krige_med.npy")

    return k_surf.compressed()

def topokriging(data, origin, theta, coords_Q, dof_surf, var_path, **kwargs):
    """ Maps surface topography to boundary pore pressure

    INPUTS:
        [data] = excel sheet containing topo data formatted as (lon, lat, elevation [meters])
        [origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
        [theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
                counter-clockwise from latlon to XY
        [coords_Q] = coordinates of degrees of freedom onto which the topo pressure is mapped
        [dof_surf] = the degrees of freedom of the surface of the mesh
        [var_path] = the path of stored variables

    OUTPUTS:
        FEniCS Expression() for pore pressure along surface boundary
    """

    x_surf, y_surf, z_surf = coords_Q[dof_surf, 0], coords_Q[dof_surf, 1], coords_Q[dof_surf, 2]
    x_topo, y_topo = latlon2XY(data, origin, theta)
    topo_p = 0.7 * (9.81 / 1e3) * np.array(
        [[data.cell(row=i, column=3).value] for i in range(2, data.max_row + 1)]).reshape(data.max_row - 1, )

    # PyKrige interpolation
    OK_topo = OrdinaryKriging(x_topo, y_topo, topo_p,
                              variogram_model='gaussian', variogram_parameters=[5e5, 5e4, 1e2])
    topo_p, sstopo = OK_topo.execute('points', x_surf, y_surf)

    np.save(var_path + 'topo_p', topo_p.compressed())
    # self.topo_p = np.load(var_path+'topo_p.npy')

    return topo_p.compressed()

def latlon2XY(data_sheet, origin, theta):
    """
    latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
    into numerical models.

    INPUTS:
    [data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
    [origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
    [theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
                counter-clockwise from latlon to XY

    Assumes first column is longitude and second column is latitude

    OUTPUTS:
    Returns transformed coordinates as numpy vectors X, Y in kilometers
    """

    lon = np.array([[data_sheet.cell(row=i, column=1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    lat = np.array([[data_sheet.cell(row=i, column=2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )

    lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
    lat_in_km = (lat - origin[1]) * 111

    rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
    theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

    X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)

    return 1e3 * X, 1e3 * Y

def latlon_disp2XY(data_sheet, origin, theta):
    """
    latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
    into numerical models.

    INPUTS:
    [data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
    [origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
    [theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
                counter-clockwise from latlon to XY

    Assumes first column is longitude and second column is latitude

    OUTPUTS:
    Returns transformed coordinates as numpy vectors X, Y in kilometers
    """

    lon_u = np.array([[data_sheet.cell(row = i, column = 3).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    lat_u = np.array([[data_sheet.cell(row = i, column = 4).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )

    rho_u = np.sqrt(np.power(lon_u, 2) + np.power(lat_u, 2))
    theta_new_u = np.arctan2(lat_u, lon_u) - theta

    Ux, Uy = rho_u * np.cos(theta_new_u), rho_u * np.sin(theta_new_u)

    return Ux, Uy

class KExpression2D(Expression):
    """Takes given surface permeabilty value and interpolates with depth.
    Normalizes the slope of the ocean surface such that the permeabilty remains
    constant as the depth increases along the ocean boundary

    Allows for the simulation of "fracking" when pore fluid pressures rise above lithostatic pressures.
    While p > lithostatic, kappa is increased by n orders of magnitude until p < lithostatic

        INPUTS:
            [k_surf] = log(land surface permeability) [m^2]
            [k_ocean] = log(ocean permebility) [m^2]
            [log_kr] = log of residual permeabiltiy at depth (usually betwee -18 and -20) [m^2]
            [alpha_k] = decay constant of permeability as a function of depth (between 0.25 and 1.6)
            [coords_k] = coordinates of degrees of freedom onto which the permeability is mapped
            [dof_surf] = the degrees of freedom of the surface of the mesh

        OUPUTS:
            kappa [m^2] - the interpolated 2D permeabiltiy field in the form of a FEniCS Expression()
    """

    def __init__(self, k_surf, k_ocean, log_kr, alpha_k, coords_K, dof_surf, **kwargs):
        self.k_surf = k_surf
        self.k_ocean = k_ocean
        self.log_kr = log_kr
        self.alpha_k = alpha_k
        self.x_surf, self.y_surf = coords_K[dof_surf, 0], coords_K[dof_surf, 1]

    def frack(self, X, Y):
        ind = (np.abs(np.sqrt(pow((self.x_surf[:] - X), 2) + pow((self.y_surf[:] - Y), 2)))).argmin()
        if p.vector()[ind] >= p_frack.vector()[ind]:
            return 1.0e2
        else:
            return 0.0

    def find_surface(self, X):
        idx = np.abs(self.x_surf[:] - X).argmin()
        return self.y_surf[idx]

    def eval(self, values, x, **kwargs):

        if self.find_surface(x[0]) < 0.0:
            values[0] = pow(10, (self.log_kr + (
                (self.k_ocean - self.log_kr) * pow((1 - (1e-3 * (x[1] - self.find_surface(x[0])))), -self.alpha_k))))

        else:
            values[0] = pow(10,
                            (self.log_kr + ((self.k_surf - self.log_kr) * pow((1 - (1e-3 * (x[1]))), -self.alpha_k))))

class KExpression3DConstant(Expression):
    """ Takes given land surface and ocean bottom permeability values and maps them to the mesh surface.
        The permeabilty decays from given surface vaule to residual vaule at depth

    INPUTS:
        [k_surf] = log(land surface permeability) [m^2]
        [k_ocean] = log(ocean permebility) [m^2]
        [log_kr] = log of residual permeabiltiy at depth (usually betwee -18 and -20) [m^2]
        [alpha_k] = decay constant of permeability as a function of depth (between 0.25 and 1.6)
        [coords_k] = coordinates of degrees of freedom onto which the permeability is mapped
        [dof_surf] = the degrees of freedom of the surface of the mesh

    OUPUTS:
        kappa [m^2] - the interpolated 3D permeabiltiy field in the form of a FEniCS Expression()

    """

    def __init__(self, k_surf, k_ocean, log_kr, alpha_k, coords_K, dof_surf, **kwargs):
        self.k_surf = k_surf
        self.k_ocean = k_ocean
        self.log_kr = log_kr
        self.alpha_k = alpha_k
        self.x_surf, self.y_surf, self.z_surf = coords_K[dof_surf, 0], coords_K[dof_surf, 1], coords_K[dof_surf, 2]

    def find_surface(self, X, Y):
        idx = (np.abs(np.sqrt(pow((self.x_surf[:] - X), 2) + pow((self.y_surf[:] - Y), 2)))).argmin()
        return self.z_surf[idx]

    def eval(self, values, x, **kwargs):

        if self.find_surface(x[0], x[1]) < 0.0:
            values[0] = pow(10, (self.log_kr + (
                (self.k_ocean - self.log_kr) * pow((1 - (1e-3 * (x[2] - self.find_surface(x[0], x[1])))), -self.alpha_k))))

        else:
            values[0] = pow(10,
                            (self.log_kr + ((self.k_surf - self.log_kr) * pow((1 - (1e-3 * (x[2]))), -self.alpha_k))))

        # values[0] = pow(10, (self.log_kr + ((self.k_ocean - self.log_kr) * pow((1 - (1e-3 * (x[2] - self.find_surface(x[0], x[1])))), -self.alpha_k))))

        # values[0] = pow(10, (((self.k_surf - self.log_kr)/60) * (1e-3 * (x[2] - self.find_surface(x[0], x[1]))) + self.k_surf))

class SlipExpression(Expression):
    """ Interpolates an arbitrary displacement field to the fault boundary of the mesh.

    INPUTS:
        [data] = an excel sheet with slip vectors recorded as
                (lon, lat, dip-slip [meters] , strike-slip [meters]) in the first four columns
        [origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
        [theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
                counter-clockwise from latlon to XY
        [coords_V] = coordinates of degrees of freedom onto which the slip vectors are mapped
        [dof_slab] = the degrees of freedom of the fault boundary of the mesh
        [var_path] = the path of stored variables

    OUTPUTS:
        FEniCS Expression() containing the interpolated vector slip displacement on the fault boundary
    """

    def __init__(self, data, origin, theta, coords_V, dof_slab, var_path, **kwargs):
        self.x_slab, self.y_slab, self.z_slab = coords_V[dof_slab, 0], coords_V[dof_slab, 1], coords_V[dof_slab, 2]

        X_slip, Y_slip = latlon2XY(data, origin, theta)

        #Ux_slip, Uy_slip = latlon_disp2XY(data, origin, theta)

        Ux_slip = lon = np.array([[data.cell(row=i, column=3).value] for i in range(2, data.max_row + 1)]).reshape(
            data.max_row - 1, )
        Uy_slip = lon = np.array([[data.cell(row=i, column=4).value] for i in range(2, data.max_row + 1)]).reshape(
            data.max_row - 1, )

        # Interpolate data using pykrige
        OK_uxslip = OrdinaryKriging(X_slip, Y_slip, Ux_slip, variogram_model='linear', \
                                    variogram_parameters=[2e-4, 1e-2])
        self.uxslip, ss_ux = OK_uxslip.execute('points', self.x_slab, self.y_slab)

        OK_uyslip = OrdinaryKriging(X_slip, Y_slip, Uy_slip, variogram_model='linear', \
                                    variogram_parameters=[2e-4, 1e-2])
        self.uyslip, ss_ux = OK_uyslip.execute('points', self.x_slab, self.y_slab)

        np.save(var_path + 'uxslip', self.uxslip.compressed())
        np.save(var_path + 'uyslip', self.uyslip.compressed())

        # self.uxslip = np.load("saved_variables_poroelastic/uxslip_med.npy")
        # self.uyslip = np.load("saved_variables_poroelastic/uyslip_med.npy")

    def U_X_interp(self, X, Y):
        indx = (np.abs(np.sqrt(pow((self.x_slab[:] - X), 2) + pow((self.y_slab[:] - Y), 2)))).argmin()
        return self.uxslip[indx]

    def U_Y_interp(self, X, Y):
        indy = (np.abs(np.sqrt(pow((self.x_slab[:] - X), 2) + pow((self.y_slab[:] - Y), 2)))).argmin()
        return self.uyslip[indy]

    def eval(self, values, x, **kwargs):
        values[0] = self.U_X_interp(x[0], x[1])
        values[1] = self.U_Y_interp(x[0], x[1])
        values[2] = 0.0

    def value_shape(self):
        return (3,)

class HtopoExpression2D(Expression):
    """ Maps surface topography to boundary pore pressure

    INPUTS:
        [x_surf] = coordinates of the surface degrees of freedom of the mesh
        [x_topo] = locations of topography measurements
        [p_topo] = pressure of water column at measurement points


    OUTPUTS:
        FEniCS Expression() for pore pressure along boundary


    """

    def __init__(self, x_surf, x_topo, p_topo, **kwargs):
        self.x_surf = x_surf
        self.x_topo = x_topo
        self.p_topo = p_topo

        self.h_surf = griddata(x_topo, p_topo, x_surf, method='linear')

    def h_topo_interp(self, X):
        ind = np.abs(np.sqrt(pow((self.x_surf[:] - X), 2))).argmin()
        return self.h_surf[ind]

    def eval(self, values, x, **kwargs):
        values[0] = self.h_topo_interp(x[0])

KExpression2D_cpp = '''
class KExpression2D_cpp : public Expression
{
public:

  KExpression2D_cpp() :
  Expression(),
  alphak(0),
  logkr(0)
  {
  }

    void eval(Array<double>& values, const Array<double>& x) const
    {
        Ksurf->eval(values, x);
        const double Ksurf_val = values[0];

        Zsurf->eval(values, x);
        const double Zsurf_val = values[0];

        values[0] = pow(10, (logkr + ((Ksurf_val - logkr) * pow(1 - 0.2*(1e-3 * (x[1] - Zsurf_val)), -alphak))));

    }

public:
    double alphak;
    double logkr;
    std::shared_ptr<const Function> Ksurf;
    std::shared_ptr<const Function> Zsurf;
};'''

KExpression3D_cpp = '''
class KExpression3D_cpp : public Expression
{
public:

  KExpression3D_cpp() :
  Expression(),
  alphak(0),
  logkr(0)
  {
  }

    void eval(Array<double>& values, const Array<double>& x) const
    {
        Ksurf->eval(values, x);
        const double Ksurf_val = values[0];

        Zsurf->eval(values, x);
        const double Zsurf_val = values[0];

        values[0] = pow(10, (logkr + ((Ksurf_val - logkr) * pow(1 - (1e-3 * 0.2*(x[2] - Zsurf_val)), -alphak))));

    }

public:
    double alphak;
    double logkr;
    std::shared_ptr<const Function> Ksurf;
    std::shared_ptr<const Function> Zsurf;
};'''

HtopoExpression3D_cpp = '''
class HtopoExpression3D_cpp : public Expression
{
public:

  HtopoExpression3D_cpp() :
  Expression()
  {
  }

    void eval(Array<double>& values, const Array<double>& x) const
    {

        Psurf->eval(values, x);
        const double Psurf_val = values[0];

        values[0] = Psurf_val;
    }

public:
	std::shared_ptr<const Function> Psurf;
};'''


################### FEM solving functions #################################
class Poroelasticity2D:
    def __init__(self, data_file, ndim, mesh, boundaries, plot_figs, event, new_mesh, permeability, days, SSE_days,
		                 sub_cycle_years, sigma_b, xcenter, SS_migrate, u0_EQ, u0_SSE, u0_sub, surface_k, ocean_k, B,
		                 loop, loop_variable, percent_lith, dehydrate_depths, dehydrate_flux, xstick, depth_kappa, k_frack, **kwargs):

        self.start = time.time()
        self.data_file = data_file
        self.mesh = mesh
        self.boundaries = boundaries
        self.plot_figs = plot_figs
        self.event = event
        self.new_mesh = new_mesh
        self.permeability = permeability
        self.days = days
        self.SSE_days = SSE_days
        self.sub_cycle_years = sub_cycle_years
        self.sigma_b = sigma_b
        self.xcenter = xcenter
        self.SS_migrate = SS_migrate
        self.u0_EQ = u0_EQ
        self.u0_SSE = u0_SSE
        self.u0_sub = u0_sub
        self.surface_k = surface_k
        self.ocean_k = ocean_k
        self.loop = loop
        self.loop_variable = loop_variable
        # self.solver = KrylovSolver('bicgstab', 'ilu')
        #self.solver = LUSolver('umfpack')
        self.solver = LinearSolver('mumps')
        self.prm = self.solver.parameters
        #self.prm['reuse_factorization'] = True  # Saves Factorization of LHS

        self.V = VectorElement('Lagrange', self.mesh.ufl_cell(), 3)
        self.R = FiniteElement('DG', self.mesh.ufl_cell(), 1)
        self.W = FiniteElement('RT', self.mesh.ufl_cell(), 2)

        self.ME = FunctionSpace(self.mesh, MixedElement([self.R, self.V, self.W]))
        #self.ME = FunctionSpace(self.mesh, MixedElement([self.R, self.V]))

        self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
        self.Kspace = FunctionSpace(self.mesh, self.Kelement)

        self.Vspace = FunctionSpace(self.mesh, self.V)
        self.Qspace = FunctionSpace(self.mesh, self.Kelement)
        self.Q_gspace = VectorFunctionSpace(self.mesh, "Lagrange", 1)
        self.Wspace = TensorFunctionSpace(self.mesh, "Lagrange", 1)

        (self.p, self.u, self.q) = TrialFunctions(self.ME)
        (self.r, self.v, self.w) = TestFunctions(self.ME)
        self.ds = Measure("ds", domain = self.mesh, subdomain_data = self.boundaries)  #
        self.n = FacetNormal(self.mesh)  # normal vector
        self.dim = self.ME.dim()
        self.kdim = self.Qspace.dim()
        self.pdim = self.ME.sub(0).dim()
        self.udim = self.ME.sub(1).dim()
        self.geom_dim = ndim + 1

        ##############    TESTING

        """
        class Surface(SubDomain):
            def inside(self, x, on_boundary):
                return near(x[1], 0)
        
        # Get cell facet connectivity
        tdim = self.mesh.topology().dim()
        self.mesh.init(tdim, tdim - 1)
        
        # Mark facets on the boundary
        bdry_facets = FacetFunction('size_t', self.mesh, False)
        Surface().mark(bdry_facets, True)
        BDRY_FACETS = bdry_facets.array()


        # Get all dofs which belong to cells with some facet on the boundary
        #order = 3
        #V = FunctionSpace(mesh, 'DG', order)
        dofmap = self.Qspace.dofmap()
        first_dof, last_dof = dofmap.ownership_range()

        bdry_dofs = np.concatenate([dofmap.cell_dofs(cell.index())
                                    for cell in cells(mesh)
                                    if any(BDRY_FACETS[cell.entities(tdim - 1)])])
        
        bdry_dofs.sort()
        
        # Consider parallel
        #bdry_dofs -= first_dof
        
        # Some function to work with
        Pbound = p.vector()
        surf0 = Pbound.get_local()
        surf0[bdry_dofs] = 0
        Pbound.set_local(values)
        quit()
        
        # Check. Project higher order to 0 to make plots same
        if order:
            u = project(u, FunctionSpace(mesh, 'DG', 0))
        
        """


        ##############   TESTING


        self.results_path = 'results/numpy_results/'
        self.var_path = 'results/numpy_variables/'
        self.paraviewpath = 'results/paraview/'
        if not os.path.exists(self.paraviewpath):
            os.makedirs(self.paraviewpath)

        self.extract_coords()
        print "Coordinates extracted", (time.time() - self.start) / 60.0, "minutes"
        self.k_interpolation()
        print "Permeability interpolated", (time.time() - self.start) / 60.0, "minutes"
        self.time_stepping()

        ######################### POROELASTIC PARAMETERS #############################
        self.delta = 1e-6  # Dirichlet control regularization on bottom boundary
        self.delta2 = 1e-10  # Dirichlet control regularization on bottom boundary
        self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
        self.B = .5  # Skempton's coefficient  ...Should this be constant?
        self.nuu = .4  # undrained poissons ratio
        self.nu = 0.25  # Poisson's ratio
        # Young's modulus [MPa]. Fit from Deshon paper (2006?)
        #self.E = Expression('(5.7972e-10*(-(pow(x[1],3)))) - (7.25283e-5*(-(pow(x[1],2)))) + (3.8486e-6*(-x[1])) + 6.94e4', degree = 1)
        #self.E = Expression('347 * pow(-x[1], 0.516) + 5.324e4', degree = 1)
        self.E = Expression('500 * pow(-x[1], 0.516) + 2e4', degree = 1)

        #self.E = Expression('1e3', degree = 1)

        self.mu = Expression('E / (2.0*(1.0 + nu))', E = self.E, nu = self.nu, degree = 1)  # shear modulus
        self.lmbda = Expression('E*nu / ((1.0 + nu)*(1.0 - 2.0*nu))', E = self.E, nu = self.nu,
                                degree = 1)  # Lame's parameter
        self.alpha = 3 * (self.nuu - self.nu) / (self.B * (1 - 2 * self.nu) * (1 + self.nuu))  # Biot's coefficient
        #self.Se = Expression('(9*(nuu-nu)*(1-(2*nuu)))/(2*mu*pow(B,2)*(1-2*nu)*pow((1+nuu),2))', nuu = self.nuu,
        #                     nu = self.nu, mu = self.mu, B = self.B, degree = 1)  # mass specific storage
        self.Se = Expression('(3*alpha*(1-(2*nu))*(1-(alpha*B)))/(2*mu*B*(1+nu))', alpha = self.alpha,
                             nu = self.nu, mu = self.mu, B = self.B, degree = 1)  # mass specific storage
        self.kstick = 8e-5
        self.stick = Expression('1/(1+exp(-kstick*(1.1e5-x[0])))', kstick = self.kstick, degree = 1)
        self.d = self.u.geometric_dimension()  # number of space dimensions
        self.I = Identity(self.d)
        self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
        self.year = 60.0 * 60.0 * 24.0 * 365.0

        self.rhof = Constant('1e-3') # pore fluid density [kg *1e-6]
        self.rho_rock = Constant('3e-3')  # rock density

        print "alpha: ", self.alpha

        ######################### VERTICAL AND HORIZONTAL STRESS  #############################

        self.p_hyrostatic = Expression('rho_f * x[1]', rho_f = self.rhof, degree=1)
        self.sig_v = Expression('rho_rock * x[1]', rho_rock = self.rho_rock, degree=1)
        self.p_frack = Expression('sig_v - ph', sig_v = self.sig_v, ph = self.p_hyrostatic, degree=1)
        self.sig_min = Expression('0', degree=1)


        ################### INITIAL AND BOUNDARY CONDITIONS ######################
        ####### Initial conditions: (('p0','ux0','uy0')) #########
        self.w_init = Expression(('0', '0', '0', '0', '0'), degree = 1)
        self.sol_old = Function(self.ME)
        self.sol_old.interpolate(self.w_init)
        zero_scalar = Constant(0.0)
        zero_vector = Expression(("0.0", "0.0"), degree = 1)
        ub = Expression(('1', '0'), degree = 1)

        bcu1 = DirichletBC(self.ME.sub(1).sub(0), zero_scalar, self.boundaries, 4)  # right/ back side (no slip condition)
        bcu2 = DirichletBC(self.ME.sub(1).sub(1), zero_scalar, self.boundaries, 5)  # mantle (no slip condition)

        bcq1 = DirichletBC(self.ME.sub(2), ('0', '0'), self.boundaries, 1)
        bcq2 = DirichletBC(self.ME.sub(2), ('0', '0'), self.boundaries, 2)
        bcq3 = DirichletBC(self.ME.sub(2), ('0', '0'), self.boundaries, 3)
        bcq4 = DirichletBC(self.ME.sub(2), ('0', '0'), self.boundaries, 4)
        bcq5 = DirichletBC(self.ME.sub(2), ('0', '0'), self.boundaries, 5)

        if self.event == 'topo':
            self.bcs = [bcu1, bcu2, bcq3, bcq4, bcq5]   # These are the BC's that are actually applied
        else:
            self.bcs = [bcu1, bcu2, bcq2, bcq3, bcq4, bcq5]

        self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case
        self.hydro = 1 #ocean bottom
        self.surface = 2

        if self.event == 'SSE':
            self.u_add = self.u0_SSE
        elif self.event == 'sub':  # Amount of slip per day (at center of event)
            self.u_add = self.u0_sub
        elif self.event == 'sub_EQ':  # Amount of slip per day (at center of event)
            self.u_add = self.u0_sub

        ################### SET UP TO SAVE SOLUTION  ######################
        self.Sol_all = np.empty((self.dim, self.nsteps))  # matrix to save entire solution
        self.Sol_surf = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
        self.sea_flux = np.empty((self.ocean_dof_pgrad.shape[0], self.nsteps))  # ocean bottom solution
        self.flow_velocity = np.empty((self.pdim*ndim, self.nsteps))  # ocean bottom solution
        self.Psi = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
        self.vol_strain = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
        self.sea_flux_total = np.empty(self.nsteps)  # matrix to save surface post-seismic solution


        if loop == 'yes':
            if self.loop_variable == 'kappa':
                self.loop_name = "k" + str(-self.surface_k)
                pfilename = self.paraviewpath+"pressure2D_loop_k%s.pvd" % str(-self.surface_k)
                qfilename = self.paraviewpath+"flux2D_loop_k%s.pvd" % str(-self.surface_k)
                qcontfilename = self.paraviewpath+"flux_cont2D_loop_k%s.pvd" % str(-self.surface_k)
                ufilename = self.paraviewpath+"def2D_loop_k%s.pvd" %  str(-self.surface_k)
                sigfilename = self.paraviewpath + "stress2D_k%s.pvd" % str(-self.surface_k)
                effsigfilename = self.paraviewpath + "effstress2D_k%s.pvd" % str(-self.surface_k)
                #eigsigfilename = self.paraviewpath + "eigstress2D_k%s.pvd" % str(-self.surface_k)
            elif self.loop_variable == 'sigma':
                self.loop_name = 'sig' + str(self.sigma_b)[:2]
                pfilename = self.paraviewpath+"pressure2D_loop_%s.pvd" % str(self.sigma_b)[:2]
                qfilename = self.paraviewpath+"pressure_vel2D_loop_%s.pvd" % str(self.sigma_b)[:2]
                ufilename = self.paraviewpath+"def2D_loop_%s.pvd" % str(self.sigma_b)[:2]
                sigfilename = self.paraviewpath + "stress2D_%s.pvd" % str(self.sigma_b)[:2]
                effsigfilename = self.paraviewpath + "effstress2D_%s.pvd" % str(self.sigma_b)[:2]
                eigsigfilename = self.paraviewpath + "eigstress2D_%s.pvd" % str(self.sigma_b)[:2]
            if self.loop_variable == 'B':
                self.loop_name = 'B' + str(-self.B)
                pfilename = self.paraviewpath+"pressure2D_loop_B%s.pvd" % str(-self.B)
                qfilename = self.paraviewpath+"pressure_vel2D_loop_B%s.pvd" % str(-self.B)
                ufilename = self.paraviewpath+"def2D_loop_k%s.pvd" %  str(-self.B)
                sigfilename = self.paraviewpath + "stress2D_B%s.pvd" % str(-self.B)
                effsigfilename = self.paraviewpath + "effstress2D_B%s.pvd" % str(-self.B)
                eigsigfilename = self.paraviewpath + "eigstress2D_B%s.pvd" % str(-self.B)
        else:
            pfilename = self.paraviewpath + "pressure2D.pvd"
            qfilename = self.paraviewpath + "flux2D.pvd"
            qcontfilename = self.paraviewpath + "flux_cont2D.pvd"
            ufilename = self.paraviewpath + "def2D.pvd"
            sigfilename = self.paraviewpath + "stress2D.pvd"
            effsigfilename = self.paraviewpath + "effstress2D.pvd"
            eigsigfilename = self.paraviewpath + "eigstress2D.pvd"
            eigeffsigfilename = self.paraviewpath + "eigeffstress2D.pvd"

        self.pfile = File(pfilename)
        self.qfile = File(qfilename)
        self.qcontfile = File(qcontfilename)
        self.ufile = File(ufilename)
        self.sigfile = File(sigfilename)
        self.effsigfile = File(effsigfilename)
        #self.eigsigfile = File(eigsigfilename)
        #self.eigeffsigfile = File(eigeffsigfilename)


        ################### SET UP, RUN, SOLVE and SAVE  ######################


        self.plot_k()
        self.plot_mu()
        self.plot_Se()
        # quit()

        self.assemble_system()
        print "System assembled", (time.time() - self.start) / 60.0, "minutes"

        self.solve_system()
        #self.plot_initial_cond()
        self.save_solution()

    def time_stepping(self):
        if self.event == 'EQ':  # 15 min for 1 day, 4 hrs for 20 days, 2 days for 30 days, per month for 10 years
            self.dtt = [15, 240, 720, 288e1, 2*432e2]#, 5256e2]  # in minutes
            self.ntime = [48, 60, 60, 30, 60]#, 50]
            # self.dtt = [15, 240, 288e1, 432e2]  # in minutes
            # self.ntime = [96, 120, 15, 122]
            # self.dtt = [15]  # in minutes
            # self.ntime = [2]
        elif self.event == 'SSE':  # 12 hrs for SSE days, 2 hr for 4 days, 12 hrs for total time-4-SSE_days
            self.dtt = [720, 120, 720]  # in minutes
            self.ntime = [self.SSE_days * 2, 48, (self.days - self.SSE_days - 4) * 2]
        elif self.event == 'sub_EQ':  # 1/year for 50 years followed by 'EQ' scheme
            self.dtt = [60 * 24 * 365, 15, 240, 288e1, 432e2, 60 * 24 * 365]  # in minutes
            self.ntime = [self.sub_cycle_years, 96, 120, 15, 122, 40]
        elif self.event == 'sub':  # 1/year for 50 years
            self.dtt = [60 * 24 * 365]  # in minutes
            self.ntime = [self.sub_cycle_years]
        elif self.event == 'topo':  # 1/year for 100 years
            self.dtt = [60 * 24 * 365]  # in minutes
            self.ntime = [self.sub_cycle_years]
            self.dtt = [60 * 24 * 365]  # in minutes
            self.ntime = [self.sub_cycle_years]
        self.dtt = [i * 60 for i in self.dtt]  # in seconds
        self.dtt_repeated = np.repeat(self.dtt, self.ntime)
        self.dtt_comsum = np.cumsum(self.dtt_repeated)
        self.nsteps = self.dtt_comsum.size

    def k_interpolation(self):
        log_kr = -18.0
        alpha_k = 0.8
        bctest1 = DirichletBC(self.Qspace, (1), self.boundaries, 1)  # ocean surface
        bctest2 = DirichletBC(self.Qspace, (1), self.boundaries, 2)  # land surface
        ktest = Function(self.Qspace)
        bctest1.apply(ktest.vector())
        bctest2.apply(ktest.vector())
        self.dof_surfQ = np.where(ktest.vector() == 1)[0]

        if self.permeability == 'mapped':
            self.kappa = KExpression2D(self.surface_k, self.ocean_k, log_kr, alpha_k, self.coords_Q, self.dof_surfQ, degree = 1)

        elif self.permeability == 'constant':
            #self.kappa = Expression('pow(10, kappa)', kappa = self.surface_k, degree = 1)
            kappa_surf = self.surface_k
            self.k_surf_func = Function(self.Kspace)
            self.z_surf_func = Function(self.Kspace)
            self.k_surf_func.vector()[:] = kappa_surf
            # self.z_surf_func.vector()[self.dof_surfK] = self.coords_K[self.dof_surfK, 1]
            self.z_surf_func.vector()[:] = 0.0


            self.kappa = Expression(KExpression2D_cpp, degree=1)
            self.kappa.alphak = alpha_k
            self.kappa.logkr = log_kr
            self.kappa.Ksurf = self.k_surf_func
            self.kappa.Zsurf = self.z_surf_func

    def topo_flow_interp(self):

        wb = pyxl.load_workbook(self.data_file)
        topo_data = wb['topo_data']
        x_topo = np.array([[topo_data.cell(row = i, column = 1).value] for i in \
                           range(2, topo_data.max_row+1)]).reshape(topo_data.max_row-1, )
        h_topo =  np.array([[topo_data.cell(row = i, column = 2).value] for i in \
                            range(2, topo_data.max_row+1)]).reshape(topo_data.max_row-1, )

        x_surf, y_surf = self.coords_Q[self.dof_surfQ, 0], self.coords_Q[self.dof_surfQ, 1]
        self.p_topo = HtopoExpression2D(x_surf, x_topo, h_topo, degree = 1)

    def plot_k(self):
        self.kappafile = File(self.paraviewpath + 'kappa_2D.pvd')
        kappaplot = Function(self.Qspace)
        kappaplot.interpolate(Expression('log10(kappa)', kappa = self.kappa, degree = 1))
        print kappaplot.vector().min(), kappaplot.vector().max()
        self.kappafile << kappaplot

    def plot_mu(self):
        self.Gfile = File(self.paraviewpath + 'Shearmod_2D.pvd')
        muplot = Function(self.Qspace)
        muplot.interpolate(Expression('mu', mu = self.mu, degree = 1))
        print "Mu min max: ",muplot.vector().min(), muplot.vector().max()
        self.Gfile << muplot

    def plot_Se(self):
        self.Sefile = File(self.paraviewpath + 'Se_2D.pvd')
        Seplot = Function(self.Qspace)
        Seplot.interpolate(Expression('mu', mu = self.Se, degree = 1))
        print "Se min max: ", Seplot.vector().min(), Seplot.vector().max()
        self.Sefile << Seplot

    def strain(self, w):  # strain = 1/2 (grad u + grad u^T)
        return sym(nabla_grad(w))

    def sigma(self, w):  # stress = 2 mu strain + lambda tr(strain) I
        return 2.0 * self.mu * self.strain(w) + self.lmbda * div(w) * self.I

    def assemble_system(self):

        dt = self.dtt_comsum[0]

        self.p_Mass = assemble(self.Se * self.p * self.r * dx)  # Mass matrix
        self.u_Div = assemble(nabla_div(self.u) * self.r * dx)  # Divergence matrix
        self.q_Div = assemble(nabla_div(self.q) * self.r * dx)  # Divergence matrix

        self.u_E = assemble(inner(self.sigma(self.u), nabla_grad(self.v)) * dx)  # Elasticity
        self.p_Grad = assemble(-self.p * (nabla_div(self.v)) * dx)  # Gradient

        self.q_K = assemble(-(self.muf / self.kappa) * inner(self.q, self.w) * dx)  # Stiffness matrix
        self.q_Grad = assemble(self.p * (nabla_div(self.w)) * dx)  # Gradient

        if self.event == 'topo':
            self.p_Boundary = assemble((1 / self.delta2) * self.p * dot(self.w, self.n) * self.ds(self.hydro) \
                                       + (1 / self.delta2) * self.p * dot(self.w, self.n) * self.ds(self.surface))

        else:
            self.p_Boundary = assemble((1 / self.delta2) * self.p * dot(self.w, self.n) * self.ds(self.hydro))

        self.u_Boundary = assemble((1 / self.delta) * dot(self.u, self.n) * dot(self.v, self.n) * self.ds(self.slab)
                                   + (1 / self.delta) * inner(self.T * self.u, self.T * self.v) * self.ds(self.slab))

        # Left hand side (LHS)
        self.A = self.p_Mass + self.alpha * self.u_Div + dt * self.q_Div \
                 + self.u_E + self.alpha * self.p_Grad + self.u_Boundary \
                 + dt * (self.q_K + self.q_Grad + self.p_Boundary)

        # RHS
        self.L = self.p_Mass + self.alpha * self.u_Div

    def extract_coords(self):
        if not os.path.exists(self.var_path):
            os.makedirs(self.var_path)
        if self.new_mesh == 'yes':  # create surface/GPS indices and save them

            self.coordinates = self.ME.tabulate_dof_coordinates().reshape((self.dim, -1))
            self.x_all, self.y_all = self.coordinates[:, 0], self.coordinates[:, 1]

            self.coords_Q = self.Qspace.tabulate_dof_coordinates().reshape((self.kdim, -1))

            self.coords_V = self.Vspace.tabulate_dof_coordinates().reshape((self.udim, -1))

            bctest0 = DirichletBC(self.ME, (1, 1, 1,1,1), self.boundaries, 1)  # ocean surface
            bctest1 = DirichletBC(self.ME, (2, 2, 2,2,2), self.boundaries, 2)  # top surface

            ptest = Function(self.ME)
            bctest0.apply(ptest.vector())
            bctest1.apply(ptest.vector())
            self.ocean_dofs = np.where(ptest.vector() == 1)[0]
            self.surface_dofs = np.where(ptest.vector() == 2)[0]

            self.size_p = self.pdim
            self.size_u = self.dim - self.size_p

            self.ocean_dof_p = self.ocean_dofs[np.where(self.ocean_dofs < self.size_p)]
            self.ocean_dof_u = self.ocean_dofs[np.where(self.ocean_dofs >= self.size_p)] - self.size_p
            self.ocean_dof_pgrad = np.hstack((self.ocean_dof_p, (self.ocean_dof_p + self.ocean_dof_p.shape[0])))
            self.surface_dof_p = self.surface_dofs[np.where(self.surface_dofs < self.size_p)]
            self.surface_dof_u = self.surface_dofs[np.where(self.surface_dofs >= self.size_p)] - self.size_p
            self.surface_dof_pgrad = np.hstack((self.surface_dof_p, (self.surface_dof_p + self.surface_dof_p.shape[0])))

            np.save(self.var_path+"x_all_2D", self.x_all)
            np.save(self.var_path+"y_all_2D", self.y_all)
            np.save(self.var_path+"surface_dofs_2D", self.surface_dofs)
            np.save(self.var_path+"ocean_dofs_2D", self.ocean_dofs)
            # np.save(self.var_path+"indices_surf_CR_2D", self.indices_surf)
            # np.save(self.var_path+"size_surf_CR_2D", self.size_surf)
            np.save(self.var_path+"size_p_CR_2D", self.size_p)
        elif self.new_mesh == 'no':  # load saved indices/variables
            # self.x_all = np.load(self.var_path+"x_all_2D.npy")
            # self.y_all = np.load(self.var_path+"y_all_2D.npy")
            self.x_surf = np.load(self.var_path+"x_surf_2D.npy")
            self.y_surf = np.load(self.var_path+"y_surf_2D.npy")
            self.indices_surf = np.load(self.var_path+"indices_surf_CR_2D.npy")
            self.size_surf = np.load(self.var_path+"size_surf_CR_2D.npy")

    def plot_initial_cond(self):  # plot the initial conditions
        self.topofile = File(self.paraviewpath + 'topo_2D.pvd')
        topoplot = Function(self.Qspace)
        topoplot.interpolate(Expression('topo_p', topo_p = self.p_topo, degree = 1))
        print topoplot.vector().min(), topoplot.vector().max()
        self.topofile << topoplot

    def streamfunction(self):
        """Stream function for a given general 2D velocity field.
        The boundary conditions are weakly imposed through the term

            inner(q, grad(psi)*n)*ds,

        where grad(psi) = [-v, u] is set on all boundaries.
        This should work for any collection of boundaries:
        walls, inlets, outlets etc.
        """
        vel = self.q_velocity
        qtest = TestFunction(self.Qspace)
        psi = TrialFunction(self.Qspace)
        a = inner(grad(qtest), grad(psi)) * dx
        # L   = dot(q, u[1].dx(0) - u[0].dx(1))*dx
        L = dot(qtest, curl(vel)) * dx
        bcu = []
        L = L + qtest * (self.n[1] * vel[0] - self.n[0] * vel[1]) * self.ds

        # Compute solution
        psi = Function(self.Qspace)
        A = assemble(a)
        b = assemble(L)
        normalize(b)  # Because we only have Neumann conditions
        [bc.apply(A, b) for bc in bcu]
        solve(A, psi.vector(), b)
        normalize(psi.vector())

        return psi

    def solve_system(self):
        count = 0

        if self.event == 'EQ':  # For an EQ, this can be outside the time loop
            u0slab = Expression(('u0*exp(-(pow((x[0] - xcenter),2)/(pow(sigma,2))))', '0'), \
                                u0 = self.u0_EQ, xcenter = self.xcenter, sigma = self.sigma_b, degree = 1)
            #u0slab = Expression(('0', '0'),degree = 1)
            u_Boundary0 = assemble((1 / self.delta) * inner(u0slab, self.T * self.v) * self.ds(
                self.slab))


        ######## TIME LOOP ########
        for i in range(self.nsteps):

            if i == 0:
                dt = self.dtt_comsum[i]
            else:
                dt = self.dtt_comsum[i] - self.dtt_comsum[i - 1]
            if i > 2:
                dt_old = self.dtt_comsum[i - 1] - self.dtt_comsum[i - 2]
            if i > 2 and dt != dt_old:
                self.A = self.p_Mass + self.alpha * self.u_Div + dt * self.q_Div \
                         + self.u_E + self.alpha * self.p_Grad + self.u_Boundary \
                         + dt * (self.q_K + self.q_Grad + self.p_Boundary)

            if i > 0:
                if self.event == 'SSE' and count <= self.SSE_days:
                    self.u0_SSE += self.u_add
                    self.xcenter += self.SS_migrate
                elif self.event == 'sub':
                    self.u0_sub += self.u_add
                elif self.event == 'sub_EQ':
                    self.u0_sub += self.u_add * (dt / self.year)

            # Slip boundary Expression

            if self.event == 'topo':
                self.topo_flow_interp()
                p_Boundary0 = assemble(
                    dt * (1 / self.delta2) * self.p_topo * dot(self.w, self.n) * self.ds(self.surface))

            elif self.event == 'SSE':
                u0slab = Expression(('u0*exp(-(pow((x[0] - xcenter),2)/(pow(sigma,2))))', '0'), \
                                    u0 = self.u0_SSE, xcenter = self.xcenter, sigma = self.sigma_b, degree = 1)
                u_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) \
                                       * self.ds(self.slab))  # Prescribed slip boundary term

            elif self.event == 'sub':
                u0slab = Expression(('u0_sub*stick', '0'), u0_sub = self.u0_sub, stick = self.stick, degree = 1)
                u_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) \
                                       * self.ds(self.slab))

            if self.event == 'sub_EQ' and i < self.sub_cycle_years:
                u0slab = Expression(('u0_sub*stick', '0'), u0_sub = self.u0_sub, stick = self.stick, degree = 1)
                u_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) \
                                       * self.ds(self.slab))

            if self.event == 'sub_EQ' and i >= self.sub_cycle_years:  # EQ event
                u0slab = Expression(('u0_sub*stick + u0d*exp(-(pow((x[0] - xcenter),2)/(pow(sigma,2))))', '0'), \
                                    u0d = self.u0_EQ, xcenter = self.xcenter, sigma = self.sigma_b, \
                                    u0_sub = self.u0_sub, stick = self.stick, degree = 1)
                u_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slab, self.T * self.v) \
                                       * self.ds(self.slab))

            b = Vector()

            if self.event == 'topo':
                self.p_Mass.init_vector(b, 0)
                self.L.mult(self.sol_old.vector(), b)
                b += p_Boundary0
            else:
                self.p_Mass.init_vector(b, 0)
                self.L.mult(self.sol_old.vector(), b)
                b += u_Boundary0
            [bc.apply(b) for bc in self.bcs]
            [bc.apply(self.A) for bc in self.bcs]
            print "BC's applied", (time.time() - self.start) / 60.0, "minutes"

            self.sol = Function(self.ME)
            self.solver.set_operator(self.A)
            self.solver.solve(self.sol.vector(), b)
            print "System solved", (time.time() - self.start) / 60.0, "minutes"

            p, u, q = self.sol.split()


            Vq_cont = VectorFunctionSpace(self.mesh, "Lagrange", 1)
            q_cont = project(q,Vq_cont)

            # u_post = self.w.sub(1) - self.w0.sub(1)
            # Deformation at each timestep. For an EQ, this is the postseismic def.
            # For a SSE, this is muddled with the imposed slip at each timestep

            #self.domain_flux = (-self.kappa / self.muf) * grad(p)  # element-wise flux vector
            self.bound_flux = assemble(dot(q,self.n) * self.ds(1))
            print "outward flux through seafloor: ", self.bound_flux
            self.bound_flux2 = assemble(dot(q,self.n) * self.ds(2))
            print "outward flux through land surface: ", self.bound_flux2

            #self.q_velocity = Function(self.Q_gspace)
            #self.q_velocity.assign(project(self.domain_flux, self.Q_gspace))
            #print "Computed velocity field", (time.time() - self.start) / 60.0, "minutes"

            #self.v_strain = Function(self.Qspace)
            #self.v_strain.assign(project(nabla_div(u), self.Qspace))
            #self.psi = self.streamfunction()  # compute numerical streamfunction

            def get_eig(self, hes):
                mesh = hes.function_space().mesh()
                [eigL, eigR] = np.linalg.eig(hes.vector().array().reshape([4, self.pdim]).transpose().reshape([self.pdim, 2, 2]))

                eig = Function(self.Q_gspace)

                eig.vector().set_local(eigL.reshape([self.pdim*2,1], order = 'F').astype(float).flatten())

                return eig

            self.stress = Function(self.Wspace)
            self.stress.assign(project(self.sigma(u), self.Wspace))
            self.effstress = Function(self.Wspace)
            self.effstress.assign(project(self.sigma(u)-p*self.I, self.Wspace))
            #self.eigstress = get_eig(self, self.stress)
            #self.eigeffstress = get_eig(self, self.effstress)

            ####################### SAVE SOLUTION ###########################
            p.rename('p', 'p')
            self.pfile << p
            u.rename('u', 'u')
            self.ufile << u
            q.rename('q', 'q')
            self.qfile << q
            q_cont.rename('q_cont', 'q_cont')
            self.qcontfile << q_cont
            #self.qfile << self.q_velocity
            self.sigfile << self.stress
            self.effsigfile << self.effstress
            #self.eigsigfile << self.eigstress
            #self.eigeffsigfile << self. eigeffstress

            self.sea_flux_total[count] = self.bound_flux
            #self.sea_flux[:, count] = self.q_velocity.vector()[self.ocean_dof_pgrad]
            #self.flow_velocity[:, count] = self.q_velocity.vector()
            #self.Psi[:, count] = self.psi.vector()
            #self.vol_strain[:, count] = self.v_strain.vector()
            self.Sol_surf[:, count] = self.sol.vector()[self.surface_dofs]
            self.Sol_all[:, count] = self.sol.vector()


            ####################### UPDATE SOLUTION ###########################

            self.sol_old = self.sol
            count += 1
            print "Timestep:", count, "dt:", dt
            #plt.colorbar(contour)

    def save_solution(self):
        if not os.path.exists(self.results_path):
            os.makedirs(self.results_path)
        if self.event == 'SSE':
            np.save(self.results_path+"Sol_surf_2D_SSE.npy", self.Sol_surf)
            np.save(self.results_path+"Sol_all_2D_SSE.npy", self.Sol_all)
            np.save(self.results_path+"sea_flux_2D_SSE.npy", self.sea_flux)
            np.save(self.var_path+"dtt_comsum_SSE.npy", self.dtt_comsum)
        elif self.event == 'sub':
            np.save(self.results_path+"Sol_surf_2D_sub.npy", self.Sol_surf)
            np.save(self.results_path+"Sol_all_2D_sub.npy", self.Sol_all)
            np.save(self.results_path+"sea_flux_2D_sub.npy", self.sea_flux)
            np.save(self.var_path+"dtt_comsum_sub.npy", self.dtt_comsum)
            np.save(self.var_path+"sub_cycle_sub.npy", self.sub_cycle_years)
        elif self.event == 'sub_EQ':
            np.save(self.results_path+"Sol_surf_2D_sub_EQ.npy", self.Sol_surf)
            np.save(self.results_path+"Sol_all_2D_sub_EQ.npy", self.Sol_all)
            np.save(self.results_path+"sea_flux_2D_sub_EQ.npy", self.sea_flux)
            np.save(self.var_path+"dtt_comsum_sub_EQ.npy", self.dtt_comsum)
            np.save(self.var_path+"sub_cycle_sub_EQ.npy", self.sub_cycle_years)
        elif self.event == 'EQ':
            if self.loop == 'yes':
                np.save(self.results_path+"Sol_surf_2D_EQ_loop_%s.npy" % self.loop_name, self.Sol_surf)
                np.save(self.results_path+"sea_flux_total_2D_EQ_loop_%s.npy" % self.loop_name,self.sea_flux_total)
                #np.save(self.results_path+"sea_flux_2D_EQ_loop_%s.npy" % self.loop_name, self.sea_flux)
                #np.save(self.results_path+"vol_strain_2D_EQ_loop_%s.npy" % self.loop_name, self.vol_strain)
                #np.save(self.results_path+"flow_velocity_2D_EQ_loop_%s.npy" % self.loop_name, self.flow_velocity)
                #np.save(self.results_path+"Streamfun_2D_EQ_loop_%s.npy" % self.loop_name, self.Psi)
                np.save(self.results_path+"Sol_all_2D_EQ_loop_%s.npy" % self.loop_name, self.Sol_all)
                np.save(self.var_path+"dtt_comsum_EQ.npy", self.dtt_comsum)
            if self.loop == 'no':
                np.save(self.results_path+"Sol_surf_2D_EQ.npy", self.Sol_surf)
                np.save(self.results_path+"Sol_all_2D_EQ.npy", self.Sol_all)
                #np.save(self.results_path+"sea_flux_2D_EQ.npy", self.sea_flux)
                np.save(self.var_path+"dtt_comsum_EQ.npy", self.dtt_comsum)
                #np.save(self.results_path+"flow_velocity_2D_EQ.npy", self.flow_velocity)
                #np.save(self.results_path+"vol_strain_2D_EQ.npy", self.vol_strain)
                np.save(self.results_path+"Streamfun_2D_EQ.npy", self.Psi)
                np.save(self.results_path+"sea_flux_total_2D_EQ.npy", self.sea_flux_total)
        elif self.event == 'topo':
            if self.loop == 'yes':
                np.save(self.results_path+"Sol_surf_2D_topo_loop_%s.npy" % self.loop_name, self.Sol_surf)
                np.save(self.results_path+"Streamfun_2D_topo_loop_%s.npy" % self.loop_name, self.Psi)
                np.save(self.results_path+"sea_flux_total_2D_topo_loop_%s.npy" % self.loop_name,
                        self.sea_flux_total)
                np.save(self.results_path+"sea_flux_2D_topo_loop_%s.npy" % self.loop_name, self.sea_flux)
                np.save(self.results_path+"flow_velocity_2D_topo_loop_%s.npy" % self.loop_name,
                        self.flow_velocity)
                np.save(self.results_path+"vol_strain_2D_topo_loop_%s.npy" % self.loop_name, self.vol_strain)
                np.save(self.results_path+"Sol_all_2D_topo_loop_%s.npy" % self.loop_name, self.Sol_all)
                np.save(self.var_path+"dtt_comsum_topo.npy", self.dtt_comsum)
            if self.loop == 'no':
                np.save(self.results_path+"Sol_surf_2D_topo.npy", self.Sol_surf)
                np.save(self.results_path+"Sol_all_2D_topo.npy", self.Sol_all)
                np.save(self.results_path+"sea_flux_2D_topo.npy", self.sea_flux)
                np.save(self.var_path+"dtt_comsum_topo.npy", self.dtt_comsum)
                np.save(self.results_path+"flow_velocity_2D_topo.npy", self.flow_velocity)
                np.save(self.results_path+"vol_strain_2D_topo.npy", self.vol_strain)
                np.save(self.results_path+"Streamfun_2D_topo.npy", self.Psi)
                np.save(self.results_path+"sea_flux_total_2D_topo.npy", self.sea_flux_total)

class Poroelasticity3D:
    def __init__(self, data_file, origin, theta, ndim, mesh, boundaries, plot_figs, EQtype, event, new_mesh, permeability, sub_cycle_years, u0_subdip, u0_substrike, surface_k, ocean_k, loop, **kwargs):

        print "Running 3D RT EQ model..."

        self.import_solution = False
        print "import solution: ", self.import_solution


        self.start = time.time()
        self.data_file = data_file
        self.origin = origin
        self.theta = theta
        self.mesh = mesh
        self.boundaries = boundaries
        self.plot_figs = plot_figs
        self.EQtype = EQtype
        self.event = event
        self.new_mesh = new_mesh
        self.permeability = permeability
        self.sub_cycle_years = sub_cycle_years
        self.u0_subx = u0_subdip
        self.u0_suby = u0_substrike
        self.surface_k = surface_k
        self.ocean_k = ocean_k

        #self.solver = PETScKrylovSolver("bicgstab","ilu")
        #self.solver = KrylovSolver("tfqmr","ilu")
        #self.solver = PETScKrylovSolver("cg")

        #self.solver = PETScKrylovSolver("gmres")
        #self.solver.parameters.nonzero_initial_guess = True

        self.solver = LUSolver('mumps')
        self.prm = self.solver.parameters
        #self.prm["relative_tolerance"] = 1e-6
        #self.prm["maximum_iterations"] = 1000
        #self.prm["monitor_convergence"] = True
        self.prm['reuse_factorization'] = True  # Saves Factorization of LHS

        self.V = VectorElement('Lagrange', self.mesh.ufl_cell(), 2)
        self.R = FiniteElement('DG', self.mesh.ufl_cell(), 0)
        self.W = FiniteElement('RT', self.mesh.ufl_cell(), 1)
        self.Vq_cont = VectorFunctionSpace(self.mesh, "CG", 1)

        self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
        self.Kspace = FunctionSpace(self.mesh, self.Kelement)

        self.ME = FunctionSpace(self.mesh, MixedElement([self.R, self.V, self.W]))

        self.Vspace = FunctionSpace(self.mesh, self.V)
        self.V1space = VectorFunctionSpace(self.mesh, 'CG', 1)
        self.Qspace = FunctionSpace(self.mesh, self.Kelement)
        self.Q_gspace = VectorFunctionSpace(self.mesh, "Lagrange", 1)
        self.Wspace = TensorFunctionSpace(self.mesh, "Lagrange", 1)

        (self.p, self.u, self.q) = TrialFunctions(self.ME)
        (self.r, self.v, self.w) = TestFunctions(self.ME)
        self.ds = Measure("ds", domain = self.mesh, subdomain_data = self.boundaries)  #
        self.n = FacetNormal(self.mesh)  # normal vector
        self.dim = self.ME.dim()
        self.kdim = self.Qspace.dim()
        self.pdim = self.ME.sub(0).dim()
        self.udim = self.ME.sub(1).dim()
        self.geom_dim = ndim + 1

        ##################### GPS STATION LOCATIONS ##############################
        wb = pyxl.load_workbook(self.data_file)
        gps_data = wb['GPS_data']

        self.x_gps, self.y_gps = latlon2XY(gps_data, self.origin, self.theta)
        self.z_gps = np.zeros(len(self.x_gps))

        #np.savetxt('XY_GPS.csv', (self.x_gps, self.y_gps), delimiter = ',')

        self.results_path = 'results/numpy_results/'
        self.var_path = 'results/numpy_variables/'
        self.paraviewpath = 'results/paraview3D/'
        self.inverse_results = '/home/fenics/shared/inverse_codes/results_deterministic/numpy_results/'
        self.inverse_var = '/home/fenics/shared/inverse_codes/results_deterministic/numpy_variables/'


        if not os.path.exists(self.results_path):
            os.makedirs(self.results_path)
        if not os.path.exists(self.var_path):
            os.makedirs(self.var_path)
        if not os.path.exists(self.paraviewpath):
            os.makedirs(self.paraviewpath)

        self.extract_coords()
        print "Coordinates extracted", (time.time() - self.start) / 60.0, "minutes"
        self.k_interpolation()
        print "Permeability interpolated", (time.time() - self.start) / 60.0, "minutes"
        self.timestepping()

        ######################### POROELASTIC PARAMETERS #############################
        self.delta = 1e-6  # Dirichlet control regularization on bottom boundary
        self.delta2 = 1e-6  # Dirichlet control regularization on bottom boundary
        self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
        self.B = 0.6  # Skempton's coefficient  ...Should this be constant?
        self.nuu = 0.4  # undrained poissons ratio
        self.nu = 0.25  # Poisson's ratio
        # self.E = Constant('5e4') # Young's modulus [MPa]
        # self.E = Expression('(5.7972e-10*(-(pow(x[2],3)))) - (7.25283e-5*(-(pow(x[2],2)))) + (3.8486e-6*(-x[2])) + 6.94e4', degree = 1)  # Young's modulus [MPa]. Fit from Deshon paper


        #self.E = Expression('5.324e4', degree = 1)

        #self.E = Expression('347 * pow(-x[2], 0.516) + 5.324e4', degree = 1)
        self.E = Expression('500 * pow(-x[2], 0.516) + 2e4', degree = 1)
        self.mu = Expression('E / (2.0*(1.0 + nu))', E = self.E, nu = self.nu, degree = 1)  # shear modulus
        self.lmbda = Expression('E*nu / ((1.0 + nu)*(1.0 - 2.0*nu))', E = self.E, nu = self.nu,
                                degree = 1)  # Lame's parameter
        self.alpha = Expression('3 * (nuu - nu) / (B * (1 - 2 * nu) * (1 + nuu))', nuu = self.nuu,
                             nu = self.nu, B = self.B, degree = 1)  # Biot's coefficient
        self.Se = Expression('(9*(nuu-nu)*(1-(2*nuu)))/(2*mu*pow(B,2)*(1-2*nu)*pow((1+nuu),2))', nuu = self.nuu,
                             nu = self.nu, mu = self.mu, B = self.B, degree = 1)  # mass specific storage

        self.kstick = 8e-5
        self.stick = Expression('1/(1+exp(-kstick*(1.1e5-x[0])))', kstick = self.kstick, degree = 1)
        self.d = self.u.geometric_dimension()  # number of space dimensions
        self.I = Identity(self.d)
        self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
        self.year = 60 * 60 * 24 * 365

        ################### INITIAL AND BOUNDARY CONDITIONS ######################
        ####### Initial conditions: (('p0','ux0','uy0')) ##########
        # w_init = Expression(('-1*x[1]/1e3', '0', '0'), u0d=u0d, sigma=sigma_b,xcenter=xcenter, ycenter=ycenter)
        self.w_init = Expression(('0.', '0.', '0.', '0.', '0.', '0.', '0.'), degree = 1)

        if self.import_solution:

            #output_file = HDF5File(self.mesh.mpi_comm(), self.results_path + '3d_results.h5', 'w')
            #output_file.close()
            input_file = HDF5File(self.mesh.mpi_comm(), self.results_path + '3d_results.h5',  "r")
            #input_file.read(invmesh, 'mesh')
            self.sol_old = Function(self.ME)

            #input_file.read(self.sol_old, 'solution')
            #input_file.close()

            # Read functions back in

            attr = input_file.attributes("solution")
            nsteps = attr['count']
            print nsteps
            quit()
            for i in range(nsteps):
                dataset = "fun/vector_%d" % i
                attr = hdf.attributes(dataset)
                print 'Retrieving time step:', attr['timestamp']
                hdf.read(F, dataset)
            print self.sol_old.vector().array().shape
        elif not self.import_solution:
            self.sol_old = Function(self.ME)
            self.sol_old.interpolate(self.w_init)

        zero_scalar = Constant(0.0)
        zero_vector = Expression(("0.0", "0.0", "0.0"), degree = 1)
        ub = Expression(('1', '0', "0"), degree = 1)

        bcu1 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 6)  # right/ back side (no slip condition)
        bcu2 = DirichletBC(self.ME.sub(1).sub(2), zero_scalar, self.boundaries, 7)  # mantle (no slip condition)

        bcq1 = DirichletBC(self.ME.sub(2), ('0', '0', '0'), self.boundaries, 1)
        bcq2 = DirichletBC(self.ME.sub(2), ('0', '0', '0'), self.boundaries, 2)
        bcq3 = DirichletBC(self.ME.sub(2), ('0', '0', '0'), self.boundaries, 3)
        bcq4 = DirichletBC(self.ME.sub(2), ('0', '0', '0'), self.boundaries, 4)
        bcq5 = DirichletBC(self.ME.sub(2), ('0', '0', '0'), self.boundaries, 5)
        bcq6 = DirichletBC(self.ME.sub(2), ('0', '0', '0'), self.boundaries, 6)
        bcq7 = DirichletBC(self.ME.sub(2), ('0', '0', '0'), self.boundaries, 7)

        if self.event == 'topo':
            self.bcs = [bcu1, bcu2, bcq3, bcq4, bcq5, bcq6, bcq7]   # These are the BC's that are actually applied
        else:
            self.bcs = [bcu1, bcu2, bcq2, bcq3, bcq4, bcq5, bcq6, bcq7]

        self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case
        self.hydro = 1 #ocean bottom
        self.surface = 2

        if self.event == 'SSE':
            self.u_addx = self.u0_SSEx
            self.u_addy = self.u0_SSEy
        elif self.event == 'sub':  # Amount of slip per day (at center of event)
            self.u_addx = self.u0_subx
            self.u_addy = self.u0_suby
        elif self.event == 'sub_EQ':  # Amount of slip per day (at center of event)
            self.u_addx = self.u0_subx
            self.u_addy = self.u0_suby


        ################### SET UP TO SAVE SOLUTION  ######################

        self.Sol_all = np.empty((self.dim, self.nsteps))  # matrix to save entire solution
        self.Sol_surf = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
        self.Sol_gps = np.empty((self.x_gps.shape[0]*ndim, self.nsteps))  # matrix to save surface solution
        self.sea_flux = np.empty((self.ocean_dof_pgrad.shape[0], self.nsteps))  # ocean bottom solution
        self.flow_velocity = np.empty((self.pdim*ndim, self.nsteps))  # ocean bottom solution
        self.Psi = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
        self.vol_strain = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
        self.sea_flux_total = np.empty(self.nsteps)  # matrix to save surface post-seismic solution

        if loop == 'yes':
            if self.loop_variable == 'kappa':
                self.loop_name = "k" + str(-self.surface_k)
                pfilename = self.paraviewpath+"pressure2D_loop_k%s.pvd" % str(-self.surface_k)
                qfilename = self.paraviewpath+"pressure_vel2D_loop_k%s.pvd" % str(-self.surface_k)
                ufilename = self.paraviewpath+"def2D_loop_k%s.pvd" %  str(-self.surface_k)
                sigfilename = self.paraviewpath + "stress2D_k%s.pvd" % str(-self.surface_k)
                effsigfilename = self.paraviewpath + "effstress2D_k%s.pvd" % str(-self.surface_k)
                eigsigfilename = self.paraviewpath + "eigstress2D_k%s.pvd" % str(-self.surface_k)
            elif self.loop_variable == 'sigma':
                self.loop_name = 'sig' + str(self.sigma_b)[:2]
                pfilename = self.paraviewpath+"pressure2D_loop_%s.pvd" % str(self.sigma_b)[:2]
                qfilename = self.paraviewpath+"pressure_vel2D_loop_%s.pvd" % str(self.sigma_b)[:2]
                ufilename = self.paraviewpath+"def2D_loop_%s.pvd" % str(self.sigma_b)[:2]
                sigfilename = self.paraviewpath + "stress2D_%s.pvd" % str(self.sigma_b)[:2]
                effsigfilename = self.paraviewpath + "effstress2D_%s.pvd" % str(self.sigma_b)[:2]
                eigsigfilename = self.paraviewpath + "eigstress2D_%s.pvd" % str(self.sigma_b)[:2]
            if self.loop_variable == 'B':
                self.loop_name = 'B' + str(-self.B)
                pfilename = self.paraviewpath+"pressure2D_loop_B%s.pvd" % str(-self.B)
                qfilename = self.paraviewpath+"pressure_vel2D_loop_B%s.pvd" % str(-self.B)
                ufilename = self.paraviewpath+"def2D_loop_k%s.pvd" %  str(-self.B)
                sigfilename = self.paraviewpath + "stress2D_B%s.pvd" % str(-self.B)
                effsigfilename = self.paraviewpath + "effstress2D_B%s.pvd" % str(-self.B)
                eigsigfilename = self.paraviewpath + "eigstress2D_B%s.pvd" % str(-self.B)
        else:
            pfilename = self.paraviewpath + "pressure3D.pvd"
            qfilename = self.paraviewpath + "flux3D.pvd"
            ufilename = self.paraviewpath + "def3D.pvd"
            sigfilename = self.paraviewpath + "stress3D.pvd"
            effsigfilename = self.paraviewpath + "effstress3D.pvd"
            eigsigfilename = self.paraviewpath + "eigstress3D.pvd"
            qcontfilename = self.paraviewpath + "flux_project3D.pvd"

        self.pfile = File(pfilename)
        self.qfile = File(qfilename)
        self.ufile = File(ufilename)
        self.sigfile = File(sigfilename)
        self.effsigfile = File(effsigfilename)
        self.eigsigfile = File(eigsigfilename)
        self.qcontfile = File(qcontfilename)
        print "Solution matrices created", (time.time() - self.start) / 60.0, "minutes"

        ################### SET UP, RUN, SOLVE and SAVE  ######################
        self.slip_interpolation()
        print "Slip data interpolated", (time.time() - self.start) / 60.0, "minutes"
        #self.plot_slip()
        #self.plot_k()


        # self.plot_initial_cond()
        dtE = Expression('dt', dt = self.dtt_comsum[0], degree = 1)
        self.assemble_system(dtE)
        print "System assembled", (time.time() - self.start) / 60.0, "minutes"

        self.solve_system()
        self.save_solution()
        #if self.event =='topo':
        #self.plot_topo()

    def extract_coords(self):
        if not os.path.exists(self.var_path):
            os.makedirs(self.var_path)
        if self.new_mesh == 'yes':  # create surface/GPS indices and save them

            self.coordinates = self.ME.tabulate_dof_coordinates().reshape((self.dim, -1))
            self.x_all, self.y_all, self.z_all = self.coordinates[:, 0], self.coordinates[:, 1], self.coordinates[:, 2]

            self.coords_Q = self.Qspace.tabulate_dof_coordinates().reshape((self.kdim, -1))
            self.coords_V = self.Vspace.tabulate_dof_coordinates().reshape((self.udim, -1))

            bctest0 = DirichletBC(self.ME, (0, 1, 1, 1, 0, 0, 0), self.boundaries, 1)  # ocean surface
            bctest1 = DirichletBC(self.ME, (0, 2, 2, 2, 0, 0, 0), self.boundaries, 2)  # top surface
            bctest2 = DirichletBC(self.ME, (0, 3, 3, 3, 0, 0, 0), self.boundaries, 3)  # slab surface

            ptest = Function(self.ME)
            bctest0.apply(ptest.vector())
            bctest1.apply(ptest.vector())
            self.ocean_dofs = np.where(ptest.vector() == 1)[0]
            self.surface_dofs = np.where(ptest.vector() == 2)[0]
            bctest2.apply(ptest.vector())
            self.slab_dofs = np.where(ptest.vector() == 3)[0]

            self.size_p = self.pdim
            self.size_u = self.udim

            self.ocean_dof_p = self.ocean_dofs[np.where(self.ocean_dofs < self.size_p)]
            self.ocean_dof_u = self.ocean_dofs[np.where(self.ocean_dofs >= self.size_p)] - self.size_p
            self.ocean_dof_pgrad = np.hstack((self.ocean_dof_p, (self.ocean_dof_p + self.ocean_dof_p.shape[0]),
                                              (self.ocean_dof_p + 2 * self.ocean_dof_p.shape[0])))
            self.surface_dof_p = self.surface_dofs[np.where(self.surface_dofs < self.size_p)]
            self.surface_dof_u = self.surface_dofs[np.where(self.surface_dofs >= self.size_p)] - self.size_p
            self.surface_dof_pgrad = np.hstack((self.surface_dof_p, (self.surface_dof_p + self.surface_dof_p.shape[0]),
                                                (self.surface_dof_p + 2 * self.surface_dof_p.shape[0])))

            #extract dofs of GPS stations - only for 'u'
            self.size_gps = self.x_gps.shape[0]
            size_w_gps = 3 * self.size_gps
            indices_gps = np.empty((self.size_gps, 3))
            for j in range(0, self.size_gps):
                indice_gps = np.where(np.sqrt(pow(self.x_all[:] - self.x_gps[j],2) + pow(self.y_all[:] - self.y_gps[j],2)+ pow(self.z_all[:] - 0.0,2)) < 0.1)[0]
                indices_gps[j, :] = indice_gps
                self.GPS_dofs = indices_gps.reshape((size_w_gps), order = 'F')
            #print indice_gps


            np.save(self.var_path + "x_all_3D", self.x_all)
            np.save(self.var_path + "y_all_3D", self.y_all)
            np.save(self.var_path + "z_all_3D", self.z_all)
            np.save(self.var_path + "GPS_dofs_u_3D", self.GPS_dofs)
            np.save(self.var_path + "surface_dofs_u_3D", self.surface_dofs)
            np.save(self.var_path + "ocean_dofs_u_3D", self.ocean_dofs)
            np.save(self.var_path + "slab_dofs_u_3D", self.slab_dofs)
            np.save(self.var_path + "size_p_3D", self.size_p)
            np.save(self.var_path + "size_u_3D", self.size_u)

        elif self.new_mesh == 'no':  # load saved indices/variables
            self.x_all = np.load(self.var_path + "x_all_3D.npy")
            self.y_all = np.load(self.var_path + "y_all_3D.npy")
            self.z_all = np.load(self.var_path + "z_all_3D.npy")
            self.GPS_dofs = np.load(self.var_path + "GPS_dofs_3D.npy")
            self.surface_dofs = np.load(self.var_path + "surface_dofs_3D.npy")
            self.ocean_dofs = np.load(self.var_path + "ocean_dofs_3D.npy")
            self.size_p = np.load(self.var_path + "size_p_3D.npy")
            self.size_u = np.load(self.var_path + "size_u_3D.npy")

            self.ocean_dof_p = self.ocean_dofs[np.where(self.ocean_dofs < self.size_p)]
            self.ocean_dof_u = self.ocean_dofs[np.where(self.ocean_dofs >= self.size_p)] - self.size_p
            self.ocean_dof_pgrad = np.hstack((self.ocean_dof_p, (self.ocean_dof_p + self.ocean_dof_p.shape[0]),
                                              (self.ocean_dof_p + 2 * self.ocean_dof_p.shape[0])))
            self.surface_dof_p = self.surface_dofs[np.where(self.surface_dofs < self.size_p)]
            self.surface_dof_u = self.surface_dofs[np.where(self.surface_dofs >= self.size_p)] - self.size_p
            self.surface_dof_pgrad = np.hstack((self.surface_dof_p, (self.surface_dof_p + self.surface_dof_p.shape[0]),
                                                (self.surface_dof_p + 2 * self.surface_dof_p.shape[0])))

    def timestepping(self):
        if self.event == 'EQ':  # 15 min for 1/2 day, 4 hrs for 10 days, 12 hours for 30 days, 2 days for 2 months, per month for 10 years, per year for 50 years
            self.dtt = [15, 240, 720, 288e1, 2*432e2]#, 5256e2]  # in minutes
            self.ntime = [48, 60, 60, 30, 60]#, 50]
            #self.dtt = [15, 240, 720, 288e1, 432e2]#, 5256e2]  # in minutes
            #self.ntime = [3, 3, 3, 3, 3]#, 50]
            #self.dtt = [15, 240, 288e1, 432e2]#, 5256e2]  # in minutes
            #self.ntime = [2, 2, 2, 2]#, 50]
            #self.dtt = [15]
            #self.ntime = [1]
        elif self.event == 'SSE':  # 12 hrs for SSE days, 2 hr for 4 days, 12 hrs for total time-4-SSE_days
            self.dtt = [720, 120, 720]  # in minutes
            self.ntime = [self.SSE_days * 2, 48, (self.days - self.SSE_days - 4) * 2]
            self.dtt = [1]
            self.ntime = [2]
        elif self.event == 'sub_EQ':  # 1/year for 50 years followed by 'EQ' scheme
            self.dtt = [60 * 24 * 365, 15, 240, 288e1, 432e2]  # in minutes
            self.ntime = [self.sub_cycle_years, 96, 120, 15, 122]
            self.dtt = [1]
            self.ntime = [2]
        elif self.event == 'sub':  # 1/year for 50 years
            self.dtt = [60 * 24 * 365]  # in minutes
            self.ntime = [self.sub_cycle_years]
        #self.dtt = [1]
        #self.ntime = [2]
        elif self.event == 'topo':  # 1/year for 100 years
            self.dtt = [60 * 24 * 365]  # in minutes
            self.ntime = [self.sub_cycle_years]
        self.dtt = [i * 60 for i in self.dtt]  # in seconds
        self.dtt_repeated = np.repeat(self.dtt, self.ntime)
        self.dtt_comsum = np.cumsum(self.dtt_repeated)
        self.nsteps = self.dtt_comsum.size

    def k_interpolation(self):
        log_kr = -18.0
        alpha_k = 0.9

        self.coords_K = self.Kspace.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
        bctest1 = DirichletBC(self.Kspace, (1), self.boundaries, 1)  # ocean surface
        bctest2 = DirichletBC(self.Kspace, (1), self.boundaries, 2)  # land surface
        ktest = Function(self.Kspace)
        bctest1.apply(ktest.vector())
        bctest2.apply(ktest.vector())
        self.dof_surfK = np.where(ktest.vector() == 1)[0]

        if self.permeability == 'mapped':

            kappa_surf = kappakriging(self.ocean_k, self.data_file, self.origin, self.theta, log_kr, alpha_k,
                                      self.coords_K, self.dof_surfK)

            #
            # print "trying distance calculation"
            #
            # #kdist = sp.spatial.distance.cdist(self.coords_K[:,0:2], self.coords_K[self.dof_surfK][:, 0:2])
            # self.ksurfall = np.empty(self.coords_K.shape[0])
            # self.zsurfall = np.empty(self.coords_K.shape[0])
            #
            # for i in range(0, self.coords_K.shape[0]):
            # 	idx = (np.abs(np.sqrt(pow((self.coords_K[self.dof_surfK][:, 0] - self.coords_K[i,0]), 2)
            # 	                + pow((self.coords_K[self.dof_surfK][:, 1] - self.coords_K[i,1]), 2)))).argmin()
            # 	#idx = kdist[i,:].argmin()
            # 	self.ksurfall[i] = kappa_surf[idx]
            # 	self.zsurfall[i] = self.coords_K[self.dof_surfK][idx,2]
            #
            # oceanz = np.where(self.zsurfall < 0.0 )[0]
            # self.ksurfall[oceanz] = self.ocean_k

            k_surf_func = Function(self.Kspace)
            z_surf_func = Function(self.Kspace)
            k_surf_func.vector()[:] = kappa_surf
            # z_surf_func.vector()[self.dof_surfK] = self.coords_K[self.dof_surfK, 2]

            z_surf_func.vector()[:] = 0.0
            # z_kappa.assign(project(z_surf_func, self.Kspace, solver_type='mumps'))

            self.kappa = Expression(KExpression3D_cpp, degree=1)
            self.kappa.alphak = alpha_k
            self.kappa.logkr = log_kr
            self.kappa.Ksurf = k_surf_func
            self.kappa.Zsurf = z_surf_func
            # self.kappa = KExpression3DInterp(self.ocean_k, self.data_file, self.origin, self.theta, log_kr, alpha_k,
            #                                 self.coords_K, self.dof_surfK, degree = 1)

        elif self.permeability == 'constant':
            # self.kappa = KExpression3DConstant(self.surface_k, self.ocean_k,
            #                                   log_kr, alpha_k, self.coords_K, self.dof_surfK, degree = 1)

            self.kappa = Expression('1e-12', degree=1)

        # self.kappa = Expression('pow(10, (((k_surf - log_kr)/20) * (1e-3 * x[2]) + k_surf))',
        #                        log_kr=log_kr, k_surf = self.surface_k, alpha_k=alpha_k, degree = 2)

        # self.kappa = Expression('pow(10, (log_kr + ((k_surf - log_kr) * pow((1 - (1e-3 * (x[2]))), -alpha_k))))',
        #                        log_kr=log_kr, k_surf = self.surface_k, alpha_k=alpha_k, degree = 1)

    def slip_interpolation(self):
        tic = time.time()

        if self.EQtype == 'inversion':

            ## import inverse solution and mesh ##
            mesh_size_inv = 'med_inv'  # fine, med, coarse
            path = "/home/fenics/shared/meshes"
            invmesh = Mesh(path + '/CR3D_' + mesh_size_inv + '.xml')

            #invmesh = Mesh()
            input_file = HDF5File(self.mesh.mpi_comm(), self.inverse_results + "u0slab.h5", "r")
            #input_file.read(invmesh, 'mesh')
            Vh = VectorFunctionSpace(invmesh, 'Lagrange', 1)
            u0load = Function(Vh)
            ufwdload = Function(Vh)
            input_file.read(u0load, 'slip')
            input_file.read(ufwdload, 'ufwd')
            input_file.close()

            ## import dofs of slab boundary from inverse solution
            dofslab_inv = np.load(self.inverse_var + 'slab_dofs_u_3D.npy')

            u0loadslab = Function(Vh)
            #u0loadslab.vector()[:] = 0.0
            u0loadslab.vector()[dofslab_inv] = u0load.vector()[dofslab_inv]

            self.u0slabV = Function(self.V1space)
            #self.u0slabV.assign(project(u0loadslab, self.V1space, solver_type='mumps'))
            self.u0slabV.assign(project(u0loadslab, self.V1space, solver_type='gmres'))


            uxdim = self.u0slabV.vector()[:].array().shape[0]/3

            ux = self.u0slabV.vector().array()[0:uxdim]
            uy = self.u0slabV.vector().array()[uxdim:2*uxdim]
            uz = self.u0slabV.vector().array()[2*uxdim:3*uxdim]

            u0x_func = Function(self.Qspace)
            u0y_func = Function(self.Qspace)
            u0z_func = Function(self.Qspace)

            u0x_func.vector()[:] = ux
            u0y_func.vector()[:] = uy
            u0z_func.vector()[:] = uz

            self.u0slab = Expression(('ux', 'uy', 'uz'), ux = u0x_func, uy = u0y_func, uz = u0z_func, degree=1)


        elif self.EQtype == 'data':

            # import from excel file ##
            wb = pyxl.load_workbook(self.data_file)
            slip_data = wb['slip_data']
            self.u0slab = SlipExpression(slip_data, self.origin, self.theta, self.coords_V, dof_slab, self.var_path, degree = 1)

        print "slip intepolation tictoc", (time.time() - tic) / 60.0, "minutes"

    def topo_flow_interp(self):
        self.coords_K = self.Kspace.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
        bctest1 = DirichletBC(self.Kspace, (1), self.boundaries, 1)  # ocean surface
        bctest2 = DirichletBC(self.Kspace, (1), self.boundaries, 2)  # land surface
        qtest = Function(self.Kspace)
        bctest1.apply(qtest.vector())
        bctest2.apply(qtest.vector())
        self.dof_surfK = np.where(qtest.vector() == 1)[0]

        wb = pyxl.load_workbook(self.data_file)
        topo_data = wb['topo_data3D']

        topo_p_surf = topokriging(topo_data, self.origin, self.theta,
                                  self.coords_K, self.dof_surfK, self.var_path)

        # pdist = sp.spatial.distance.cdist(self.coords_K[:, 0:2], self.coords_K[self.dof_surfK][:, 0:2])
        # self.psurfall = np.empty(pdist.shape[0])
        # self.zsurfall_p = np.empty(pdist.shape[0])
        #
        # for i in range(0, pdist.shape[0]):
        # 	idx = pdist[i, :].argmin()
        # 	self.psurfall[i] = topo_p_surf[idx]
        # 	self.zsurfall_p[i] = self.coords_K[self.dof_surfK][idx, 2]
        #
        # oceanz = np.where(self.zsurfall_p < 0.0)[0]
        low_p = np.where(topo_p_surf < 0.0)[0]
        # self.psurfall[oceanz] = 0.0
        topo_p_surf[low_p] = 0.01
        #
        # p_surf_func.vector()[:] = self.psurfall
        # self.p_topo = Expression(HtopoExpression3D_cpp, degree = 1)
        # self.p_topo.Psurf = p_surf_func

        # NEW method
        p_surf_func = Function(self.Kspace)
        p_surf_func.vector()[self.dof_surfK] = topo_p_surf
        self.p_topo = Expression(('ptopo'), ptopo=p_surf_func, degree=1)

    def plot_k(self):
        self.kappafile = File(self.paraviewpath + 'kappa_3D.pvd')
        kappaplot = Function(self.Kspace)
        kappaplot.interpolate(Expression('log10(kappa)', kappa=self.kappa, degree=1))
        # print kappaplot.vector().min(), kappaplot.vector().max()
        self.kappafile << kappaplot

    def plot_topo(self):
        self.topofile = File(self.paraviewpath + 'topo_3D.pvd')
        topoplot = Function(self.Kspace)
        topoplot.interpolate(Expression('p', p=self.p_topo, degree=1))
        # print kappaplot.vector().min(), kappaplot.vector().max()
        self.topofile << topoplot

    def plot_slip(self):
        self.slipfile = File(self.paraviewpath + 'slip_3D.pvd')
        if self.EQtype == 'inversion':
            #slipplot = self.u0slabV
            #slipplot = self.ufwdV
            slipplot = Function(self.Vspace)
            slipplot.interpolate(self.u0slab)
        else:
            slipplot = Function(self.Vspace)
            slipplot.interpolate(self.u0slab)
        # print kappaplot.vector().min(), kappaplot.vector().max()
        self.slipfile << slipplot

    def SSE_interpolation(self, timestep):

        bctest1 = DirichletBC(self.Vspace, (1, 1, 1), self.boundaries, 3)  # slab surface
        stest = Function(self.Vspace)
        bctest1.apply(stest.vector())
        dof_slab = np.where(stest.vector() == 1)[0]

        self.u0_SSE = SSEExpression(self.coords_V, dof_slab, degree = 1)

    def strain(self, w):  # strain = 1/2 (grad u + grad u^T)
        return sym(nabla_grad(w))

    def sigma(self, w):  # stress = 2 mu strain + lambda tr(strain) I
        return 2.0 * self.mu * self.strain(w) + self.lmbda * div(w) * self.I

    def assemble_system(self, dt, assembleRHS = True):
        if assembleRHS:
            self.p_Mass = assemble(self.Se * self.p * self.r * dx)  # Mass matrix
            self.u_Div = assemble(self.alpha * nabla_div(self.u) * self.r * dx)  # Divergence matrix
            self.u_E = assemble(inner(self.sigma(self.u), nabla_grad(self.v)) * dx)  # Elasticity
            self.p_Grad = assemble(-self.alpha * self.p * (nabla_div(self.v)) * dx)  # Gradient
            self.u_Boundary = assemble((1 / self.delta) * dot(self.u, self.n) * dot(self.v, self.n) * self.ds(self.slab)
                                           + (1 / self.delta) * inner(self.T * self.u, self.T * self.v) * self.ds(self.slab))

        self.q_Div = assemble(dt * nabla_div(self.q) * self.r * dx)  # Divergence matrix
        self.q_K = assemble(-(dt * self.muf / self.kappa) * inner(self.q, self.w) * dx)  # Stiffness matrix
        self.q_Grad = assemble(dt * self.p * (nabla_div(self.w)) * dx)  # Gradient

        print "all LHS matrices assembled", (time.time() - self.start) / 60.0, "minutes"

        if self.event == 'topo':
            self.p_Boundary = assemble((dt / self.delta2) * self.p * dot(self.w, self.n) * self.ds(self.hydro) \
                                       + (dt / self.delta2) * self.p * dot(self.w, self.n) * self.ds(self.surface))
        else:
            self.p_Boundary = assemble((dt / self.delta2) * self.p * dot(self.w, self.n) * self.ds(self.hydro))

        # Left hand side (LHS)
        # self.A = self.p_Mass + self.alpha * self.u_Div + dt * self.q_Div + self.u_E +  \
        #          self.p_Grad + dt * (self.q_K + self.q_Grad + self.p_Boundary) + self.u_Boundary

        self.A = self.p_Mass + self.u_Div + self.q_Div + self.u_E +  \
                 self.p_Grad + self.q_K + self.q_Grad + self.p_Boundary + self.u_Boundary

        print "LHS assembled", (time.time() - self.start) / 60.0, "minutes"

        if assembleRHS:
            #  self.L = self.p_Mass + self.alpha * self.u_Div
            self.L = self.p_Mass + self.u_Div
            print "RHS assembled", (time.time() - self.start) / 60.0, "minutes"

    def streamfunction3D(self, constrained_domain = None):
        """Stream function for a given 3D velocity field.
        The boundary conditions are weakly imposed through the term

            inner(q, grad(psi)*n)*ds,

        where u = curl(psi) is used to fill in for grad(psi) and set
        boundary conditions. This should work for walls, inlets, outlets, etc.
        """

        vel = self.q_velocity
        qtest = TestFunction(self.Qspace)
        psi = TrialFunction(self.Qspace)
        a = inner(grad(qtest), grad(psi)) * dx - inner(qtest, dot(self.n, grad(psi))) * self.ds
        L = inner(grad(qtest), curl(vel)) * dx - dot(grad(qtest), cross(self.n, vel)) * self.ds
        psi_ = Function(self.Qspace)
        A = assemble(a)
        b = assemble(L)
        solver = LUSolver('petsc')
        b = assemble(L, tensor = b)
        solver.set_operator(A)
        solver.solve(psi_.vector(), b)

        return psi_

    def test_PSD(self):

        def is_pd(x):
            return np.all(np.linalg.eigvals(x) > 0)

        def is_psd(x, tol = 1e-5):
            E, V = np.linalg.eigh(x)
            return np.all(E > -tol)

        def is_symmetric(x):
            return (x.transpose() == x).all()

        A = self.A.array()

        print('A is {}'.format('symmetric' if is_symmetric(A) else ('not symmetric')))

        np.linalg.cholesky(A)

    def solve_system(self):

        #sol_file = open(self.results_path+"sol_all.txt", "w")

        if self.import_solution:
            count = self.import_step
        else:
            count = 0

        sea_change = 1.0

        if self.event == 'EQ':# and self.EQtype == 'data':  # For an EQ, this can be outside the time loop since we only use it for the first time step
            # u0slab = Expression(('u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + '
            #                      '(pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
            #                      'u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + '
            #                      '(pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))','0'),
            #                     u0d = -4, u0s = -1, xcenter = 75e3,
            #                     ycenter = 130e3,sigmax = 1.5e4,
            #                     sigmay = 2e4, degree = 1)
            # self.u0slab = Expression(('0','0','0'), degree = 1)
            u_Boundary0 = assemble((1 / self.delta) * inner(self.T * self.u0slab, self.T * self.v) * self.ds(
                self.slab))  # Prescribed slip boundary term

        for i in range(self.nsteps):

            if i == 0:
                dt = self.dtt_comsum[i]
                self.sol = Function(self.ME)
            else:
                dt = self.dtt_comsum[i] - self.dtt_comsum[i - 1]
            if i > 1:
                dt_old = self.dtt_comsum[i - 1] - self.dtt_comsum[i - 2]
                sea_change = np.abs((self.sea_flux_total[count-1] - self.sea_flux_total[count - 2])
                                    /self.sea_flux_total[count-1])
            if i > 1 and dt != dt_old: # Rebuild Left hand side when 'dt' changes
                print "re-assembling LHS", (time.time() - self.start) / 60.0, "minutes"
                dtE = Expression('dt', dt=dt, degree=1)
                self.assemble_system(dtE, assembleRHS=False)
                self.solver = LUSolver('mumps')
                self.prm = self.solver.parameters
                self.prm['reuse_factorization'] = True  # Saves Factorization of LHS

            #     self.solver.parameters.nonzero_initial_guess = False
            #     print "Using MUMPS solver"
            # elif sea_change > 0.3:
            #     self.solver.parameters.nonzero_initial_guess = False
            #     #self.solver = LUSolver('mumps')
            #     print "Using MUMPS solver"
            # else:
            #     self.solver.parameters.nonzero_initial_guess = True
            #     print "Using PETSC GMRES solver"


            print "SGD change fraction: ", sea_change
            print "dt: ", dt/60, 'minutes'

            if i > 0:
                # Add slip to the boundary for a SSE and move center of slip along boundary
                if self.event == 'SSE' and count <= self.SSE_days:
                    print "This doesn't work yet..."
                    quit()
                    self.u0_SSE = SSE_interpolation(i)
                elif self.event == 'sub':
                    self.u0_subx += self.u_addx * (dt / self.year)
                    self.u0_suby += self.u_addy * (dt / self.year)
                elif self.event == 'sub_EQ':
                    self.u0_subx += self.u_addx * (dt / self.year)
                    self.u0_suby += self.u_addy * (dt / self.year)

            # Slip boundary Expression

            if self.event == 'topo':
                self.topo_flow_interp()
                p_Boundary0 = assemble(dt * (1 / self.delta2) * self.p_topo * dot(self.w, self.n) * self.ds(self.surface))

            elif self.event == 'SSE':
                u_Boundary0 = assemble((1 / self.delta) * inner(self.T * self.u0_SSE, self.T * self.v) * self.ds(
                    self.slab))

            elif self.event == 'sub':
                self.u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx = self.u0_subx,
                                         u0_suby = self.u0_suby, stick = self.stick, degree = 1)
                u_Boundary0 = assemble((1 / self.delta) * inner(self.T * self.u0slab, self.T * self.v) * self.ds(
                    self.slab))

            elif self.event == 'sub_EQ':
                u0slabsub = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx = self.u0_subx,
                                       u0_suby = self.u0_suby, stick = self.stick, degree = 1)

                if i < self.sub_cycle_years:
                    u_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slabsub, self.T * self.v) * self.ds(self.slab))

                if i >= self.sub_cycle_years:
                    u0slabtotal = self.u0slabsub + self.u0slab
                    u_Boundary0 = assemble((1 / self.delta) * inner(self.T * u0slabtotal, self.T * self.v) * self.ds(
                        self.slab))

            print "Boundary condition set", (time.time() - self.start) / 60.0, "minutes"
            b = Vector()

            if self.event == 'topo':
                self.p_Mass.init_vector(b, 0)
                self.L.mult(self.sol_old.vector(), b)
                b += p_Boundary0
            else:
                self.p_Mass.init_vector(b, 0)
                self.L.mult(self.sol_old.vector(), b)
                b += u_Boundary0
            [bc.apply(b) for bc in self.bcs]
            [bc.apply(self.A) for bc in self.bcs]
            print "BC's applied", (time.time() - self.start) / 60.0, "minutes"

            self.solver.set_operator(self.A)
            self.solver.solve(self.sol.vector(), b)
            print "System solved", (time.time() - self.start) / 60.0, "minutes"

            p, u, q = self.sol.split()

            q_cont = Function(self.Vq_cont)
            q_cont.assign(project(q, self.Vq_cont, solver_type='mumps'))
            #q_cont.assign(project(q, self.Vq_cont, solver_type='gmres'))


            self.bound_flux = assemble(dot(q,self.n) * self.ds(1))
            print "outward flux through seafloor: ", self.bound_flux
            self.bound_flux2 = assemble(dot(q,self.n) * self.ds(2))
            print "outward flux through land surface: ", self.bound_flux2

            #self.v_strain = Function(self.Qspace)
            #self.v_strain.assign(project(nabla_div(u), self.Qspace))
            #self.psi = self.streamfunction3D()  # compute numerical streamfunction

            #self.stress = Function(self.Wspace)
            #self.stress.assign(project(self.sigma(u), self.Wspace))

            ####################### SAVE SOLUTION ###########################
            p.rename("p", "p")
            self.pfile << p
            u.rename("u", "u")
            self.ufile << u
            q_cont.rename("q_cont", "q_cont")
            self.qcontfile << q_cont
            q.rename("q", "q")
            self.qfile << q

            #print "Paraview files updated", (time.time() - self.start) / 60.0, "minutes"


            #self.sigfile << self.stress

            self.sea_flux_total[count] = self.bound_flux
            #self.sea_flux[:, count] = self.q_velocity.vector()[self.ocean_dof_pgrad]
            #self.flow_velocity[:, count] = self.q_velocity.vector()
            #self.Psi[:, count] = self.psi.vector()
            #self.vol_strain[:, count] = self.v_strain.vector()
            #self.Sol_surf[:, count] = self.sol.vector()[self.surface_dofs]
            #self.Sol_all[:, count] = self.sol.vector()
            self.Sol_gps[:, count] = self.sol.vector()[self.GPS_dofs]

            tic = time.time()
            np.save(self.results_path + "Sol_all_3D_latest.npy", self.sol.vector())
            np.save(self.results_path + "Sol_gps_3D_EQ_mapped.npy", self.Sol_gps)

            #sol_file.write("%d\n" % self.sol.vector())

            print "numpy save tictoc: ", (tic - self.start) / 60.0, "minutes"

            print "Numpy Vectors saved", (time.time() - self.start) / 60.0, "minutes"

            if count == 0:
                output_file = HDF5File(self.mesh.mpi_comm(), self.results_path + '3d_results.h5', 'w')
            output_file.write(self.sol, 'solution', count)
            #output_file.write(u, 'def', count)
            #output_file.write(q, 'vel', count)
            #output_file.write(q_cont, 'vel_CG1', count)

            print "Solution saved as HDF5 file ", (time.time() - self.start) / 60.0, "minutes"


            # np.save(self.results_path + "Sol_surf_3D_EQ_mapped.npy", self.Sol_surf)
            # np.save(self.results_path + "Sol_gps_3D_EQ_mapped.npy", self.Sol_gps)
            # np.save(self.results_path + "Sol_all_3D_EQ_mapped.npy", self.Sol_all)
            # np.save(self.results_path + "sea_flux_total_3D_EQ_mapped.npy", self.sea_flux_total)

            ######################## PLOT SOLUTION ##########################
            if self.plot_figs == 'yes':
                HTML(X3DOM().html(p))
            # contour = plot(u)
            # contour = plot(p)

            ####################### UPDATE SOLUTION ###########################
            self.sol_old = self.sol
            print "Solution updated", (time.time() - self.start) / 60.0, "minutes"

            count += 1
            print "Timestep:", count

        output_file.close()
        #sol_file.close()
        # plt.colorbar(contour)

    def save_solution(self):
        if not os.path.exists(self.results_path):
            os.makedirs(self.results_path)
        if self.permeability == 'mapped':

            if self.event == 'SSE':
                np.save(self.results_path + "Sol_surf_3D_SSE_mapped.npy", self.Sol_surf)
                np.save(self.results_path + "Sol_gps_3D_SSE_mapped.npy", self.Sol_gps)
                np.save(self.results_path + "Sol_all_3D_SSE_mapped.npy", self.Sol_all)
            elif self.event == 'sub':
                np.save(self.results_path + "Sol_surf_3D_sub_mapped.npy", self.Sol_surf)
                np.save(self.results_path + "Sol_gps_3D_sub_mapped.npy", self.Sol_gps)
                np.save(self.results_path + "Sol_all_3D_sub_mapped.npy", self.Sol_all)
            elif self.event == 'sub_EQ':
                np.save(self.results_path + "Sol_surf_3D_sub_EQ_mapped.npy", self.Sol_surf)
                np.save(self.results_path + "Sol_gps_3D_sub_EQ_mapped.npy", self.Sol_gps)
                np.save(self.results_path + "Sol_all_3D_sub_EQ_mapped.npy", self.Sol_all)
            elif self.event == 'EQ':
                np.save(self.var_path+"dtt_comsum_EQ.npy", self.dtt_comsum)
                #np.save(self.results_path + "Sol_surf_3D_EQ_mapped.npy", self.Sol_surf)
                np.save(self.results_path + "Sol_gps_3D_EQ_mapped.npy", self.Sol_gps)
                #np.save(self.results_path + "Sol_all_3D_EQ_mapped.npy", self.Sol_all)
                #np.save(self.results_path + "vol_strain_3D_EQ_mapped.npy", self.vol_strain)
                #np.save(self.results_path + "flow_velocity_3D_EQ_mapped.npy", self.flow_velocity)
                #np.save(self.results_path + "psi_3D_EQ_mapped.npy", self.Psi)
                #np.save(self.results_path + "sea_flux_3D_EQ_mapped.npy", self.sea_flux)
                np.save(self.results_path + "sea_flux_total_3D_EQ_mapped.npy", self.sea_flux_total)
            elif self.event == 'topo':
                np.save(self.results_path + "Sol_surf_3D_topo_mapped.npy", self.Sol_surf)
                np.save(self.results_path + "Sol_all_3D_topo_mapped.npy", self.Sol_all)
                np.save(self.results_path + "sea_flux_3D_topo_mapped.npy", self.sea_flux)
                np.save(self.var_path + "dtt_comsum_topo_mapped.npy", self.dtt_comsum)
                np.save(self.results_path + "flow_velocity_3D_topo_mapped.npy", self.flow_velocity)
                np.save(self.results_path + "vol_strain_3D_topo_mapped.npy", self.vol_strain)
                np.save(self.results_path + "Streamfun_3D_topo_mapped.npy", self.Psi)
                np.save(self.results_path + "sea_flux_total_3D_topo_mapped.npy", self.sea_flux_total)
        elif self.permeability == 'constant':
            if self.event == 'SSE':
                np.save(self.results_path + "Sol_surf_3D_SSE_constant.npy", self.Sol_surf)
                np.save(self.results_path + "Sol_gps_3D_SSE_constant.npy", self.Sol_gps)
                np.save(self.results_path + "Sol_all_3D_SSE_constant.npy", self.Sol_all)
            elif self.event == 'sub':
                np.save(self.results_path + "Sol_surf_3D_sub_constant.npy", self.Sol_surf)
                np.save(self.results_path + "Sol_gps_3D_sub_constant.npy", self.Sol_gps)
                np.save(self.results_path + "Sol_all_3D_sub_constant.npy", self.Sol_all)
            elif self.event == 'sub_EQ':
                np.save(self.results_path + "Sol_surf_3D_sub_EQ_constant.npy", self.Sol_surf)
                np.save(self.results_path + "Sol_gps_3D_sub_EQ_constant.npy", self.Sol_gps)
                np.save(self.results_path + "Sol_all_3D_sub_EQ_constant.npy", self.Sol_all)
                np.save(self.results_path + "vol_strain_3D_sub_EQ_constant.npy", self.vol_strain)
                np.save(self.results_path + "flow_velocity_3D_sub_EQ_constant.npy", self.flow_velocity)
                np.save(self.results_path + "psi_3D_EQ_mapped.npy", self.Psi)
                np.save(self.results_path + "sea_flux_3D_sub_EQ_constant.npy", self.sea_flux)
                np.save(self.results_path + "sea_flux_total_3D_sub_EQ_constant.npy", self.sea_flux_total)
            elif self.event == 'EQ':
                np.save(self.results_path + "Sol_surf_3D_EQ_constant.npy", self.Sol_surf)
                np.save(self.results_path + "Sol_gps_3D_EQ_constant.npy", self.Sol_gps)
                np.save(self.results_path + "Sol_all_3D_EQ_constant.npy", self.Sol_all)
                np.save(self.results_path + "vol_strain_3D_EQ_constant.npy", self.vol_strain)
                np.save(self.results_path + "flow_velocity_3D_EQ_constant.npy", self.flow_velocity)
                np.save(self.results_path + "sea_flux_3D_EQ_constant.npy", self.sea_flux)
                np.save(self.results_path + "sea_flux_total_3D_EQ_constant.npy", self.sea_flux_total)

class Laplace3D:
    def __init__(self, data_file, origin, theta, ndim, mesh, boundaries, plot_figs, event, new_mesh, permeability,
                 years, surface_k, ocean_k, loop, **kwargs):

        self.start = time.time()
        self.data_file = data_file
        self.origin = origin
        self.theta = theta
        self.mesh = mesh
        self.boundaries = boundaries
        self.plot_figs = plot_figs
        self.event = event
        self.new_mesh = new_mesh
        self.permeability = permeability
        self.years = years
        self.surface_k = surface_k
        self.ocean_k = ocean_k
        # self.solver = PETScKrylovSolver("bicgstab","ilu")
        # self.solver = KrylovSolver("tfqmr","ilu")

        self.solver = LUSolver('mumps')
        self.prm = self.solver.parameters
        # self.prm["relative_tolerance"] = 5e-6
        # self.prm["maximum_iterations"] = 1000
        # self.prm["monitor_convergence"] = True
        # self.prm["form_compiler"]["representation"] = "uflacs"
        # self.prm['reuse_factorization'] = True  # Saves Factorization of LHS

        self.R = FiniteElement('DG', self.mesh.ufl_cell(), 1)
        self.W = FiniteElement('RT', self.mesh.ufl_cell(), 2)

        self.ME = FunctionSpace(self.mesh, MixedElement([self.R, self.W]))

        self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
        self.Kspace = FunctionSpace(self.mesh, self.Kelement)

        self.Qspace = FunctionSpace(self.mesh, self.Kelement)
        self.Q_gspace = VectorFunctionSpace(self.mesh, "Lagrange", 1)
        self.Wspace = TensorFunctionSpace(self.mesh, "Lagrange", 1)

        (self.p, self.q) = TrialFunctions(self.ME)
        (self.r, self.w) = TestFunctions(self.ME)
        self.ds = Measure("ds", domain = self.mesh, subdomain_data = self.boundaries)  #
        self.n = FacetNormal(self.mesh)  # normal vector
        self.dim = self.ME.dim()
        self.kdim = self.Qspace.dim()
        self.pdim = self.ME.sub(0).dim()
        self.qdim = self.ME.sub(1).dim()
        self.geom_dim = ndim + 1

        ##################### GPS STATION LOCATIONS ##############################

        self.x_gps = np.array([65308.5, 78441.2, 65707.9, 73555.6, 61257.8, 65291.6, 82112.3,
                               58804.1, 106822.8, 60152.4, 144960.9, 88660.1, 134829.2, 82409,
                               65214.4, 67249.7, 91916.7, 111195.9, 92905.4, 77176.5, 90809.3,
                               107806.1, 104439.5, 137051.8, 118095, 82238.4, 84123.2, 98197.5,
                               105163, 154799.8, 133655.9, 111467, 125498.1, 106995.1, 149450.7,
                               192061, 133903.2])

        self.y_gps = np.array([173604.8, 185529.2, 168017.6, 174095.8, 157386.9, 154352.5,
                               172814.3, 138685.6, 190893.8, 136426.2, 230041.9, 163931.9,
                               208246.3, 149993.8, 127836, 122418.9, 141284.3, 162013.4,
                               131283.2, 109747.4, 124059.5, 141615.9, 126126.3, 160541,
                               132800.4, 90851.5, 92291, 105309.4, 104713.8, 153672.9,
                               122823.1, 94557.3, 99698.5, 77971.8, 123239.8, 157764, 45232])

        self.z_gps = np.zeros(self.x_gps.shape)

        self.results_path = 'results/numpy_results/'
        self.var_path = 'results/numpy_variables/'
        self.paraviewpath = 'results/paraview3D/'
        if not os.path.exists(self.paraviewpath):
            os.makedirs(self.paraviewpath)

        self.extract_coords()
        print "Coordinates extracted", (time.time() - self.start) / 60.0, "minutes"
        self.k_interpolation()
        print "Permeability interpolated", (time.time() - self.start) / 60.0, "minutes"
        self.timestepping()

        ######################### POROELASTIC PARAMETERS #############################
        self.delta2 = 1e-2  # Dirichlet control regularization on bottom boundary
        self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
        self.B = 0.7  # Skempton's coefficient  ...Should this be constant?
        self.nuu = .38  # undrained poissons ratio
        self.nu = 0.27  # Poisson's ratio
        # self.E = Constant('5e4') # Young's modulus [MPa]
        # self.E = Expression('(5.7972e-10*(-(pow(x[2],3)))) - (7.25283e-5*(-(pow(x[2],2)))) + (3.8486e-6*(-x[2])) + 6.94e4', degree = 1)  # Young's modulus [MPa]. Fit from Deshon paper

        self.E = Expression('347 * pow(-x[2], 0.516) + 5.324e4', degree = 1)
        self.mu = Expression('E / (2.0*(1.0 + nu))', E = self.E, nu = self.nu, degree = 1)  # shear modulus
        self.lmbda = Expression('E*nu / ((1.0 + nu)*(1.0 - 2.0*nu))', E = self.E, nu = self.nu,
                                degree = 1)  # Lame's parameter
        self.alpha = 3 * (self.nuu - self.nu) / (self.B * (1 - 2 * self.nu) * (1 + self.nuu))  # Biot's coefficient
        self.Se = Expression('(9*(nuu-nu)*(1-(2*nuu)))/(2*mu*pow(B,2)*(1-2*nu)*pow((1+nuu),2))', nuu = self.nuu,
                             nu = self.nu, mu = self.mu, B = self.B, degree = 1)  # mass specific storage

        self.d = self.q.geometric_dimension()  # number of space dimensions
        self.I = Identity(self.d)
        self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
        self.year = 60 * 60 * 24 * 365

        ################### INITIAL AND BOUNDARY CONDITIONS ######################
        ####### Initial conditions: (('p0','ux0','uy0')) ##########
        # w_init = Expression(('-1*x[1]/1e3', '0', '0'), u0d=u0d, sigma=sigma_b,xcenter=xcenter, ycenter=ycenter)
        self.w_init = Expression(('0', '0', '0', '0'), degree = 1)
        self.sol_old = Function(self.ME)
        self.sol_old.interpolate(self.w_init)
        zero_scalar = Constant(0.0)
        zero_vector = Expression(("0.0", "0.0", "0.0"), degree = 1)


        bcq1 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 1)
        bcq2 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 2)
        bcq3 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 3)
        bcq4 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 4)
        bcq5 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 5)
        bcq6 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 6)
        bcq7 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 7)

        if self.event == 'topo':
            self.bcs = [bcq3, bcq4, bcq5, bcq6, bcq7]  # These are the BC's that are actually applied
        else:
            self.bcs = [bcq2, bcq3, bcq4, bcq5, bcq6, bcq7]

        self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case
        self.hydro = 1  # ocean bottom
        self.surface = 2


        ################### SET UP TO SAVE SOLUTION  ######################

        self.Sol_all = np.empty((self.dim, self.nsteps))  # matrix to save entire solution
        self.Sol_surf = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
        #self.sea_flux = np.empty((self.ocean_dof_pgrad.shape[0], self.nsteps))  # ocean bottom solution
        self.flow_velocity = np.empty((self.pdim * ndim, self.nsteps))  # ocean bottom solution
        self.sea_flux_total = np.empty(self.nsteps)  # matrix to save surface post-seismic solution

        if loop == 'yes':
            if self.loop_variable == 'kappa':
                self.loop_name = "k" + str(-self.surface_k)
                pfilename = self.paraviewpath + "pressure2D_loop_k%s.pvd" % str(-self.surface_k)
                qfilename = self.paraviewpath + "pressure_vel2D_loop_k%s.pvd" % str(-self.surface_k)
                ufilename = self.paraviewpath + "def2D_loop_k%s.pvd" % str(-self.surface_k)
                sigfilename = self.paraviewpath + "stress2D_k%s.pvd" % str(-self.surface_k)
                effsigfilename = self.paraviewpath + "effstress2D_k%s.pvd" % str(-self.surface_k)
                eigsigfilename = self.paraviewpath + "eigstress2D_k%s.pvd" % str(-self.surface_k)
            elif self.loop_variable == 'sigma':
                self.loop_name = 'sig' + str(self.sigma_b)[:2]
                pfilename = self.paraviewpath + "pressure2D_loop_%s.pvd" % str(self.sigma_b)[:2]
                qfilename = self.paraviewpath + "pressure_vel2D_loop_%s.pvd" % str(self.sigma_b)[:2]
                ufilename = self.paraviewpath + "def2D_loop_%s.pvd" % str(self.sigma_b)[:2]
                sigfilename = self.paraviewpath + "stress2D_%s.pvd" % str(self.sigma_b)[:2]
                effsigfilename = self.paraviewpath + "effstress2D_%s.pvd" % str(self.sigma_b)[:2]
                eigsigfilename = self.paraviewpath + "eigstress2D_%s.pvd" % str(self.sigma_b)[:2]
            if self.loop_variable == 'B':
                self.loop_name = 'B' + str(-self.B)
                pfilename = self.paraviewpath + "pressure2D_loop_B%s.pvd" % str(-self.B)
                qfilename = self.paraviewpath + "pressure_vel2D_loop_B%s.pvd" % str(-self.B)
                ufilename = self.paraviewpath + "def2D_loop_k%s.pvd" % str(-self.B)
                sigfilename = self.paraviewpath + "stress2D_B%s.pvd" % str(-self.B)
                effsigfilename = self.paraviewpath + "effstress2D_B%s.pvd" % str(-self.B)
                eigsigfilename = self.paraviewpath + "eigstress2D_B%s.pvd" % str(-self.B)
        else:
            pfilename = self.paraviewpath + "pressure3D.pvd"
            qfilename = self.paraviewpath + "flux3D.pvd"
            ufilename = self.paraviewpath + "def3D.pvd"
            sigfilename = self.paraviewpath + "stress3D.pvd"
            effsigfilename = self.paraviewpath + "effstress3D.pvd"
            eigsigfilename = self.paraviewpath + "eigstress3D.pvd"
            qcontfilename = self.paraviewpath + "flux_project3D.pvd"


        self.qcontfile = File(qcontfilename)
        self.pfile = File(pfilename)
        self.qfile = File(qfilename)

        print "Solution matrices created", (time.time() - self.start) / 60.0, "minutes"

        ################### SET UP, RUN, SOLVE and SAVE  ######################

        self.plot_k()

        # self.plot_initial_cond()
        self.assemble_system()
        print "System assembled", (time.time() - self.start) / 60.0, "minutes"
        self.solve_system()
    #self.save_solution()
    #if self.event == 'topo':
    #	self.plot_topo()

    def extract_coords(self):
        if not os.path.exists(self.var_path):
            os.makedirs(self.var_path)
        if self.new_mesh == 'yes':  # create surface/GPS indices and save them

            self.coordinates = self.ME.tabulate_dof_coordinates().reshape((self.dim, -1))
            self.x_all, self.y_all, self.z_all = self.coordinates[:, 0], self.coordinates[:, 1], self.coordinates[:, 2]

            self.coords_Q = self.Qspace.tabulate_dof_coordinates().reshape((self.kdim, -1))

            bctest0 = DirichletBC(self.ME, (1, 1, 1, 1), self.boundaries, 1)  # ocean surface
            bctest1 = DirichletBC(self.ME, (2, 2, 2, 2), self.boundaries, 2)  # top surface

            ptest = Function(self.ME)
            bctest0.apply(ptest.vector())
            bctest1.apply(ptest.vector())
            self.ocean_dofs = np.where(ptest.vector() == 1)[0]
            self.surface_dofs = np.where(ptest.vector() == 2)[0]

            self.size_p = self.pdim

            np.save(self.var_path + "x_all_3D", self.x_all)
            np.save(self.var_path + "y_all_3D", self.y_all)
            np.save(self.var_path + "z_all_3D", self.z_all)
            # np.save(self.var_path + "GPS_dofs_3D", self.GPS_dofs)
            np.save(self.var_path + "surface_dofs_3D", self.surface_dofs)
            np.save(self.var_path + "ocean_dofs_3D", self.ocean_dofs)
            np.save(self.var_path + "size_p_3D", self.size_p)


        elif self.new_mesh == 'no':  # load saved indices/variables
            self.x_all = np.load(self.var_path + "x_all_3D.npy")
            self.y_all = np.load(self.var_path + "y_all_3D.npy")
            self.z_all = np.load(self.var_path + "z_all_3D.npy")
            self.GPS_dofs = np.load(self.var_path + "GPS_dofs_3D.npy")
            self.surface_dofs = np.load(self.var_path + "surface_dofs_3D.npy")
            self.ocean_dofs = np.load(self.var_path + "ocean_dofs_3D.npy")
            self.size_p = np.load(self.var_path + "size_p_3D.npy")
            self.size_u = np.load(self.var_path + "size_u_3D.npy")

            self.ocean_dof_p = self.ocean_dofs[np.where(self.ocean_dofs < self.size_p)]
            self.ocean_dof_u = self.ocean_dofs[np.where(self.ocean_dofs >= self.size_p)] - self.size_p
            self.ocean_dof_pgrad = np.hstack((self.ocean_dof_p, (self.ocean_dof_p + self.ocean_dof_p.shape[0]),
                                              (self.ocean_dof_p + 2 * self.ocean_dof_p.shape[0])))
            self.surface_dof_p = self.surface_dofs[np.where(self.surface_dofs < self.size_p)]
            self.surface_dof_u = self.surface_dofs[np.where(self.surface_dofs >= self.size_p)] - self.size_p
            self.surface_dof_pgrad = np.hstack((self.surface_dof_p, (self.surface_dof_p + self.surface_dof_p.shape[0]),
                                                (self.surface_dof_p + 2 * self.surface_dof_p.shape[0])))

    def timestepping(self):

        self.dtt = [60 * 24 * 365]  # in minutes
        self.ntime = [self.years]
        self.dtt = [i * 60 for i in self.dtt]  # in seconds
        self.dtt_repeated = np.repeat(self.dtt, self.ntime)
        self.dtt_comsum = np.cumsum(self.dtt_repeated)
        self.nsteps = self.dtt_comsum.size

    def k_interpolation(self):
        log_kr = -20.0
        alpha_k = 0.6
        bctest1 = DirichletBC(self.Qspace, (1), self.boundaries, 1)  # ocean surface
        bctest2 = DirichletBC(self.Qspace, (1), self.boundaries, 2)  # land surface
        ktest = Function(self.Qspace)
        bctest1.apply(ktest.vector())
        bctest2.apply(ktest.vector())
        self.dof_surfQ = np.where(ktest.vector() == 1)[0]

        if self.permeability == 'mapped':
            self.kappa = KExpression3DInterp(self.ocean_k, self.data_file, self.origin, self.theta, log_kr, alpha_k,
                                             self.coords_Q, self.dof_surfQ, degree = 1)

        elif self.permeability == 'constant':
            self.kappa = KExpression3DConstant(self.surface_k, self.ocean_k, log_kr, alpha_k, self.coords_Q,
                                               self.dof_surfQ,
                                               degree = 1)
        # self.kappa = Expression('1e-9')  #

    def topo_flow_interp(self):
        bctest1 = DirichletBC(self.Qspace, (1), self.boundaries, 2)  # land surface
        htest = Function(self.Qspace)
        bctest1.apply(htest.vector())
        self.dof_landQ = np.where(htest.vector() == 1)[0]

        topo_max = 100
        xcenter = 130e3
        ycenter = 130e3
        sigma_bx = 1.5e4
        sigma_by = 2e4

        self.p_topo = Expression(( '0.7 * (9.81 / 1e3) * pmax * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + \
		(pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))'),pmax = topo_max, xcenter = xcenter, ycenter = ycenter, \
                                 sigmax = sigma_bx,sigmay = sigma_by, degree = 1)

    # self.p_topo = HtopoExpression3D(self.data_file, self.origin, self.theta, self.coords_Q, self.dof_landQ, degree=1)

    def plot_k(self):
        self.kappafile = File(self.paraviewpath + 'kappa_3D.pvd')
        kappaplot = Function(self.Qspace)
        kappaplot.interpolate(Expression('log10(kappa)', kappa = self.kappa, degree = 1))
        # print kappaplot.vector().min(), kappaplot.vector().max()
        self.kappafile << kappaplot

    def plot_topo(self):
        self.topofile = File(self.paraviewpath + 'topo_3D.pvd')
        topoplot = Function(self.Qspace)
        topoplot.interpolate(Expression('p', p = self.p_topo, degree = 1))
        # print kappaplot.vector().min(), kappaplot.vector().max()
        self.topofile << topoplot

    def assemble_system(self):

        dt = self.dtt_comsum[0]

        self.p_Mass = assemble(self.Se * self.p * self.r * dx)  # Mass matrix
        self.q_Div = assemble(nabla_div(self.q) * self.r * dx)  # Divergence matrix

        self.q_K = assemble(-(self.muf / self.kappa) * inner(self.q, self.w) * dx)  # Stiffness matrix
        self.q_Grad = assemble(self.p * (nabla_div(self.w)) * dx)  # Gradient

        if self.event == 'topo':
            self.p_Boundary = assemble((1 / self.delta2) * self.p * dot(self.w, self.n) * self.ds(self.hydro) \
                                       + (1 / self.delta2) * self.p * dot(self.w, self.n) * self.ds(self.surface))

        else:
            self.p_Boundary = assemble((1 / self.delta2) * self.p * dot(self.w, self.n) * self.ds(self.hydro))


        # Left hand side (LHS)
        self.A = self.p_Mass + dt * self.q_Div + dt * (self.q_K + self.q_Grad + self.p_Boundary)

        # RHS
        self.L = self.p_Mass

    def streamfunction3D(self, constrained_domain = None):
        """Stream function for a given 3D velocity field.
        The boundary conditions are weakly imposed through the term

            inner(q, grad(psi)*n)*ds,

        where u = curl(psi) is used to fill in for grad(psi) and set
        boundary conditions. This should work for walls, inlets, outlets, etc.
        """

        vel = self.q_velocity
        qtest = TestFunction(self.Qspace)
        psi = TrialFunction(self.Qspace)
        a = inner(grad(qtest), grad(psi)) * dx - inner(qtest, dot(self.n, grad(psi))) * self.ds
        L = inner(grad(qtest), curl(vel)) * dx - dot(grad(qtest), cross(self.n, vel)) * self.ds
        psi_ = Function(self.Qspace)
        A = assemble(a)
        b = assemble(L)
        solver = LUSolver('petsc')
        b = assemble(L, tensor = b)
        solver.set_operator(A)
        solver.solve(psi_.vector(), b)

        return psi_

    def test_PSD(self):

        def is_pd(x):
            return np.all(np.linalg.eigvals(x) > 0)

        def is_psd(x, tol = 1e-5):
            E, V = np.linalg.eigh(x)
            return np.all(E > -tol)

        def is_symmetric(x):
            return (x.transpose() == x).all()

        A = self.A.array()

        print('A is {}'.format('symmetric' if is_symmetric(A) else ('not symmetric')))

        np.linalg.cholesky(A)

    def solve_system(self):
        count = 0

        for i in range(self.nsteps):

            if i == 0:
                dt = self.dtt_comsum[i]
            else:
                dt = self.dtt_comsum[i] - self.dtt_comsum[i - 1]
            if i > 1:
                dt_old = self.dtt_comsum[i - 1] - self.dtt_comsum[i - 2]
            if i > 1 and dt != dt_old:  # Rebuild Left hand side when 'dt' changes
                self.A = self.p_Mass + dt * self.q_Div + dt * (self.q_K + self.q_Grad + self.p_Boundary)

            print "timstep: ", dt / 60, 'minutes'

            # Slip boundary Expression

            if self.event == 'topo':
                self.topo_flow_interp()
                p_Boundary0 = assemble(
                    dt * (1 / self.delta2) * self.p_topo * dot(self.w, self.n) * self.ds(self.surface))

            print "Boundary condition set", (time.time() - self.start) / 60.0, "minutes"

            b = Vector()

            if self.event == 'topo':
                self.p_Mass.init_vector(b, 0)
                self.L.mult(self.sol_old.vector(), b)
                b += p_Boundary0
            else:
                self.p_Mass.init_vector(b, 0)
                self.L.mult(self.sol_old.vector(), b)

            [bc.apply(b) for bc in self.bcs]
            [bc.apply(self.A) for bc in self.bcs]
            print "BC's applied", (time.time() - self.start) / 60.0, "minutes"

            self.sol = Function(self.ME)
            self.solver.set_operator(self.A)
            self.solver.solve(self.sol.vector(), b)
            print "System solved", (time.time() - self.start) / 60.0, "minutes"

            p, q = self.sol.split()

            Vq_cont = VectorFunctionSpace(self.mesh, "CG", 2)
            q_cont = project(q, Vq_cont)

            self.bound_flux = assemble(dot(q, self.n) * self.ds(1))
            print "outward flux through seafloor: ", self.bound_flux
            self.bound_flux2 = assemble(dot(q, self.n) * self.ds(2))
            print "outward flux through land surface: ", self.bound_flux2

            # self.psi = self.streamfunction3D()  # compute numerical streamfunction

            ####################### SAVE SOLUTION ###########################
            p.rename('p', 'p')
            self.pfile << p
            q_cont.rename('q_cont', 'q_cont')
            self.qcontfile << q_cont
            q.rename('q', 'q')
            self.qfile << q

            # self.sigfile << self.stress

            self.sea_flux_total[count] = self.bound_flux
            # self.sea_flux[:, count] = self.q_velocity.vector()[self.ocean_dof_pgrad]
            # self.flow_velocity[:, count] = self.q_velocity.vector()
            # self.Psi[:, count] = self.psi.vector()
            # self.vol_strain[:, count] = self.v_strain.vector()
            # self.Sol_surf[:, count] = self.sol.vector()[self.surface_dofs]
            self.Sol_all[:, count] = self.sol.vector()
            # self.Sol_gps[:, count] = self.sol.vector()[self.GPS_dofs]


            ######################## PLOT SOLUTION ##########################
            if self.plot_figs == 'yes':
                HTML(X3DOM().html(p))


            ####################### UPDATE SOLUTION ###########################
            self.sol_old = self.sol
            count += 1
            print "Timestep:", count

    def save_solution(self):
        if not os.path.exists(self.results_path):
            os.makedirs(self.results_path)
        if self.permeability == 'mapped':
            np.save(self.results_path + "Sol_surf_3D_laplace_mapped.npy", self.Sol_surf)
            np.save(self.results_path + "Sol_all_3D_laplace_mapped.npy", self.Sol_all)
            #np.save(self.results_path + "sea_flux_3D_laplace_mapped.npy", self.sea_flux)
            np.save(self.var_path + "dtt_comsum_laplace_mapped.npy", self.dtt_comsum)
            np.save(self.results_path + "flow_velocity_3D_laplace_mapped.npy", self.flow_velocity)
            #np.save(self.results_path + "Streamfun_3D_laplace_mapped.npy", self.Psi)
            np.save(self.results_path + "sea_flux_total_3D_laplace_mapped.npy", self.sea_flux_total)
        elif self.permeability == 'constant':
            np.save(self.results_path + "Sol_surf_3D_laplace_mapped.npy", self.Sol_surf)
            np.save(self.results_path + "Sol_all_3D_laplace_mapped.npy", self.Sol_all)
            #np.save(self.results_path + "sea_flux_3D_laplace_mapped.npy", self.sea_flux)
            np.save(self.var_path + "dtt_comsum_laplace_mapped.npy", self.dtt_comsum)
            np.save(self.results_path + "flow_velocity_3D_laplace_mapped.npy", self.flow_velocity)
            #np.save(self.results_path + "Streamfun_3D_laplace_mapped.npy", self.Psi)
            np.save(self.results_path + "sea_flux_total_3D_laplace_mapped.npy", self.sea_flux_total)

class Laplace2D:
    def __init__(self, data_file, origin, theta, ndim, mesh, boundaries, plot_figs, event, new_mesh, permeability,
                 years, surface_k, ocean_k, loop, i, **kwargs):

        self.start = time.time()
        self.data_file = data_file
        self.origin = origin
        self.theta = theta
        self.mesh = mesh
        self.boundaries = boundaries
        self.plot_figs = plot_figs
        self.event = event
        self.new_mesh = new_mesh
        self.permeability = permeability
        self.years = years
        self.surface_k = surface_k
        self.ocean_k = ocean_k
        # self.solver = PETScKrylovSolver("bicgstab","ilu")
        # self.solver = KrylovSolver("tfqmr","ilu")

        self.solver = LUSolver('mumps')
        self.prm = self.solver.parameters
        # self.prm["relative_tolerance"] = 5e-6
        # self.prm["maximum_iterations"] = 1000
        # self.prm["monitor_convergence"] = True
        # self.prm["form_compiler"]["representation"] = "uflacs"
        # self.prm['reuse_factorization'] = True  # Saves Factorization of LHS

        self.R = FiniteElement('DG', self.mesh.ufl_cell(), 1)
        self.W = FiniteElement('RT', self.mesh.ufl_cell(), 2)

        self.ME = FunctionSpace(self.mesh, MixedElement([self.R, self.W]))

        self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
        self.Kspace = FunctionSpace(self.mesh, self.Kelement)

        self.Qspace = FunctionSpace(self.mesh, self.Kelement)
        self.Q_gspace = VectorFunctionSpace(self.mesh, "Lagrange", 1)
        self.Wspace = TensorFunctionSpace(self.mesh, "Lagrange", 1)

        (self.p, self.q) = TrialFunctions(self.ME)
        (self.r, self.w) = TestFunctions(self.ME)
        self.ds = Measure("ds", domain=self.mesh, subdomain_data=self.boundaries)  #
        self.n = FacetNormal(self.mesh)  # normal vector
        self.dim = self.ME.dim()
        self.kdim = self.Qspace.dim()
        self.pdim = self.ME.sub(0).dim()
        self.qdim = self.ME.sub(1).dim()
        self.geom_dim = ndim + 1

        ##################### GPS STATION LOCATIONS ##############################

        self.x_gps = np.array([65308.5, 78441.2, 65707.9, 73555.6, 61257.8, 65291.6, 82112.3,
                               58804.1, 106822.8, 60152.4, 144960.9, 88660.1, 134829.2, 82409,
                               65214.4, 67249.7, 91916.7, 111195.9, 92905.4, 77176.5, 90809.3,
                               107806.1, 104439.5, 137051.8, 118095, 82238.4, 84123.2, 98197.5,
                               105163, 154799.8, 133655.9, 111467, 125498.1, 106995.1, 149450.7,
                               192061, 133903.2])

        self.y_gps = np.array([173604.8, 185529.2, 168017.6, 174095.8, 157386.9, 154352.5,
                               172814.3, 138685.6, 190893.8, 136426.2, 230041.9, 163931.9,
                               208246.3, 149993.8, 127836, 122418.9, 141284.3, 162013.4,
                               131283.2, 109747.4, 124059.5, 141615.9, 126126.3, 160541,
                               132800.4, 90851.5, 92291, 105309.4, 104713.8, 153672.9,
                               122823.1, 94557.3, 99698.5, 77971.8, 123239.8, 157764, 45232])

        self.z_gps = np.zeros(self.x_gps.shape)

        self.results_path = 'results/numpy_results/'
        self.var_path = 'results/numpy_variables/'
        self.paraviewpath = 'results/paraview3D/'
        if not os.path.exists(self.paraviewpath):
            os.makedirs(self.paraviewpath)

        self.extract_coords()
        print "Coordinates extracted", (time.time() - self.start) / 60.0, "minutes"
        self.k_interpolation()
        print "Permeability interpolated", (time.time() - self.start) / 60.0, "minutes"
        self.timestepping()

        ######################### POROELASTIC PARAMETERS #############################
        self.delta2 = 1e-2  # Dirichlet control regularization on bottom boundary
        self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
        self.B = 0.7  # Skempton's coefficient  ...Should this be constant?
        self.nuu = .38  # undrained poissons ratio
        self.nu = 0.27  # Poisson's ratio
        # self.E = Constant('5e4') # Young's modulus [MPa]
        # self.E = Expression('(5.7972e-10*(-(pow(x[2],3)))) - (7.25283e-5*(-(pow(x[2],2)))) + (3.8486e-6*(-x[2])) + 6.94e4', degree = 1)  # Young's modulus [MPa]. Fit from Deshon paper

        self.E = Expression('347 * pow(-x[2], 0.516) + 5.324e4', degree=1)
        self.mu = Expression('E / (2.0*(1.0 + nu))', E=self.E, nu=self.nu, degree=1)  # shear modulus
        self.lmbda = Expression('E*nu / ((1.0 + nu)*(1.0 - 2.0*nu))', E=self.E, nu=self.nu,
                                degree=1)  # Lame's parameter
        self.alpha = 3 * (self.nuu - self.nu) / (self.B * (1 - 2 * self.nu) * (1 + self.nuu))  # Biot's coefficient
        self.Se = Expression('(9*(nuu-nu)*(1-(2*nuu)))/(2*mu*pow(B,2)*(1-2*nu)*pow((1+nuu),2))', nuu=self.nuu,
                             nu=self.nu, mu=self.mu, B=self.B, degree=1)  # mass specific storage

        self.d = self.q.geometric_dimension()  # number of space dimensions
        self.I = Identity(self.d)
        self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
        self.year = 60 * 60 * 24 * 365

        ################### INITIAL AND BOUNDARY CONDITIONS ######################
        ####### Initial conditions: (('p0','ux0','uy0')) ##########
        # w_init = Expression(('-1*x[1]/1e3', '0', '0'), u0d=u0d, sigma=sigma_b,xcenter=xcenter, ycenter=ycenter)
        self.w_init = Expression(('0', '0', '0', '0'), degree=1)
        self.sol_old = Function(self.ME)
        self.sol_old.interpolate(self.w_init)
        zero_scalar = Constant(0.0)
        zero_vector = Expression(("0.0", "0.0", "0.0"), degree=1)

        bcq1 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 1)
        bcq2 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 2)
        bcq3 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 3)
        bcq4 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 4)
        bcq5 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 5)
        bcq6 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 6)
        bcq7 = DirichletBC(self.ME.sub(1), zero_vector, self.boundaries, 7)

        if self.event == 'topo':
            self.bcs = [bcq3, bcq4, bcq5, bcq6, bcq7]  # These are the BC's that are actually applied
        else:
            self.bcs = [bcq2, bcq3, bcq4, bcq5, bcq6, bcq7]

        self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case
        self.hydro = 1  # ocean bottom
        self.surface = 2

        ################### SET UP TO SAVE SOLUTION  ######################

        self.Sol_all = np.empty((self.dim, self.nsteps))  # matrix to save entire solution
        self.Sol_surf = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
        # self.sea_flux = np.empty((self.ocean_dof_pgrad.shape[0], self.nsteps))  # ocean bottom solution
        self.flow_velocity = np.empty((self.pdim * ndim, self.nsteps))  # ocean bottom solution
        self.sea_flux_total = np.empty(self.nsteps)  # matrix to save surface post-seismic solution

        if loop == 'yes':
            if self.loop_variable == 'kappa':
                self.loop_name = "k" + str(-self.surface_k)
                pfilename = self.paraviewpath + "pressure2D_loop_k%s.pvd" % str(-self.surface_k)
                qfilename = self.paraviewpath + "pressure_vel2D_loop_k%s.pvd" % str(-self.surface_k)
                ufilename = self.paraviewpath + "def2D_loop_k%s.pvd" % str(-self.surface_k)
                sigfilename = self.paraviewpath + "stress2D_k%s.pvd" % str(-self.surface_k)
                effsigfilename = self.paraviewpath + "effstress2D_k%s.pvd" % str(-self.surface_k)
                eigsigfilename = self.paraviewpath + "eigstress2D_k%s.pvd" % str(-self.surface_k)
            elif self.loop_variable == 'sigma':
                self.loop_name = 'sig' + str(self.sigma_b)[:2]
                pfilename = self.paraviewpath + "pressure2D_loop_%s.pvd" % str(self.sigma_b)[:2]
                qfilename = self.paraviewpath + "pressure_vel2D_loop_%s.pvd" % str(self.sigma_b)[:2]
                ufilename = self.paraviewpath + "def2D_loop_%s.pvd" % str(self.sigma_b)[:2]
                sigfilename = self.paraviewpath + "stress2D_%s.pvd" % str(self.sigma_b)[:2]
                effsigfilename = self.paraviewpath + "effstress2D_%s.pvd" % str(self.sigma_b)[:2]
                eigsigfilename = self.paraviewpath + "eigstress2D_%s.pvd" % str(self.sigma_b)[:2]
            if self.loop_variable == 'B':
                self.loop_name = 'B' + str(-self.B)
                pfilename = self.paraviewpath + "pressure2D_loop_B%s.pvd" % str(-self.B)
                qfilename = self.paraviewpath + "pressure_vel2D_loop_B%s.pvd" % str(-self.B)
                ufilename = self.paraviewpath + "def2D_loop_k%s.pvd" % str(-self.B)
                sigfilename = self.paraviewpath + "stress2D_B%s.pvd" % str(-self.B)
                effsigfilename = self.paraviewpath + "effstress2D_B%s.pvd" % str(-self.B)
                eigsigfilename = self.paraviewpath + "eigstress2D_B%s.pvd" % str(-self.B)
        else:
            pfilename = self.paraviewpath + "pressure3D.pvd"
            qfilename = self.paraviewpath + "flux3D.pvd"
            ufilename = self.paraviewpath + "def3D.pvd"
            sigfilename = self.paraviewpath + "stress3D.pvd"
            effsigfilename = self.paraviewpath + "effstress3D.pvd"
            eigsigfilename = self.paraviewpath + "eigstress3D.pvd"
            qcontfilename = self.paraviewpath + "flux_project3D.pvd"

        self.qcontfile = File(qcontfilename)
        self.pfile = File(pfilename)
        self.qfile = File(qfilename)

        print "Solution matrices created", (time.time() - self.start) / 60.0, "minutes"

        ################### SET UP, RUN, SOLVE and SAVE  ######################

        self.plot_k()

        # self.plot_initial_cond()
        self.assemble_system()
        print "System assembled", (time.time() - self.start) / 60.0, "minutes"
        self.solve_system()

    # self.save_solution()
    # if self.event == 'topo':
    #	self.plot_topo()

    def extract_coords(self):
        if not os.path.exists(self.var_path):
            os.makedirs(self.var_path)
        if self.new_mesh == 'yes':  # create surface/GPS indices and save them

            self.coordinates = self.ME.tabulate_dof_coordinates().reshape((self.dim, -1))
            self.x_all, self.y_all = self.coordinates[:, 0], self.coordinates[:, 1]

            self.coords_Q = self.Qspace.tabulate_dof_coordinates().reshape((self.kdim, -1))

            bctest0 = DirichletBC(self.ME, (1, 1, 1), self.boundaries, 1)  # ocean surface
            bctest1 = DirichletBC(self.ME, (2, 2, 2), self.boundaries, 2)  # top surface

            ptest = Function(self.ME)
            bctest0.apply(ptest.vector())
            bctest1.apply(ptest.vector())
            self.ocean_dofs = np.where(ptest.vector() == 1)[0]
            self.surface_dofs = np.where(ptest.vector() == 2)[0]

            self.size_p = self.pdim

            np.save(self.var_path + "x_all_2D", self.x_all)
            np.save(self.var_path + "y_all_2D", self.y_all)
            # np.save(self.var_path + "GPS_dofs_3D", self.GPS_dofs)
            np.save(self.var_path + "surface_dofs_2D", self.surface_dofs)
            np.save(self.var_path + "ocean_dofs_2D", self.ocean_dofs)
            np.save(self.var_path + "size_p_2D", self.size_p)


        elif self.new_mesh == 'no':  # load saved indices/variables
            self.x_all = np.load(self.var_path + "x_all_3D.npy")
            self.y_all = np.load(self.var_path + "y_all_3D.npy")
            self.z_all = np.load(self.var_path + "z_all_3D.npy")
            self.GPS_dofs = np.load(self.var_path + "GPS_dofs_3D.npy")
            self.surface_dofs = np.load(self.var_path + "surface_dofs_3D.npy")
            self.ocean_dofs = np.load(self.var_path + "ocean_dofs_3D.npy")
            self.size_p = np.load(self.var_path + "size_p_3D.npy")
            self.size_u = np.load(self.var_path + "size_u_3D.npy")

            self.ocean_dof_p = self.ocean_dofs[np.where(self.ocean_dofs < self.size_p)]
            self.ocean_dof_u = self.ocean_dofs[np.where(self.ocean_dofs >= self.size_p)] - self.size_p
            self.ocean_dof_pgrad = np.hstack((self.ocean_dof_p, (self.ocean_dof_p + self.ocean_dof_p.shape[0]),
                                              (self.ocean_dof_p + 2 * self.ocean_dof_p.shape[0])))
            self.surface_dof_p = self.surface_dofs[np.where(self.surface_dofs < self.size_p)]
            self.surface_dof_u = self.surface_dofs[np.where(self.surface_dofs >= self.size_p)] - self.size_p
            self.surface_dof_pgrad = np.hstack((self.surface_dof_p, (self.surface_dof_p + self.surface_dof_p.shape[0]),
                                                (self.surface_dof_p + 2 * self.surface_dof_p.shape[0])))

    def k_interpolation(self):
        log_kr = -15
        alpha_k = 0.6

        self.coords_K = self.Kspace.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
        bctest1 = DirichletBC(self.Kspace, (1), self.boundaries, 1)  # ocean surface
        bctest2 = DirichletBC(self.Kspace, (1), self.boundaries, 2)  # land surface
        ktest = Function(self.Kspace)
        bctest1.apply(ktest.vector())
        bctest2.apply(ktest.vector())
        self.dof_surfK = np.where(ktest.vector() == 1)[0]

        if self.permeability == 'mapped':

            kappa_2D = kappakriging2D(self.data_file, self.coords_K, self.dof_surfK)
            self.k_2D_func = Function(self.Kspace)
            self.k_2D_func.vector()[:] = 1e-18
            self.kappa = Expression('kappa', kappa = self.k_2D_func, degree = 1)
        # self.kappa = Expression('1e-18', degree=1)

        elif self.permeability == 'constant':
            # self.kappa = Expression('pow(10, kappa)', kappa=self.surface_k, degree=1)
            # self.kappa = KExpression2D(self.surface_k, self.ocean_k, log_kr, alpha_k, self.coords_K, self.dof_surfK, degree=1)
            kappa_surf = self.surface_k
            self.k_surf_func = Function(self.Kspace)
            self.z_surf_func = Function(self.Kspace)
            self.k_surf_func.vector()[:] = kappa_surf
            self.z_surf_func.vector()[self.dof_surfK] = self.coords_K[self.dof_surfK, 1]

            self.kappa = Expression(KExpression2D_cpp, degree = 1)
            self.kappa.alphak = alpha_k
            self.kappa.logkr = log_kr
            self.kappa.Ksurf = self.k_surf_func
            self.kappa.Zsurf = self.z_surf_func

            self.k_2D_func = Function(self.Kspace)
            self.k_2D_func.interpolate(self.kappa)

    def topo_flow_interp(self):  # debug

        wb = pyxl.load_workbook(self.data_file)
        topo_data = wb['topo_data_2D']
        x_topo = np.array(
            [[topo_data.cell(row = i, column = 1).value] for i in range(2, topo_data.max_row + 1)]).reshape(
            topo_data.max_row - 1, )
        p_topo = 0.7 * (9.81 / 1e3) * np.array(
            [[topo_data.cell(row = i, column = 2).value] for i in range(2, topo_data.max_row + 1)]).reshape(
            topo_data.max_row - 1, )
        self.coords_K = self.Kspace.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
        bctest1 = DirichletBC(self.Kspace, (1), self.boundaries, 1)  # ocean surface
        bctest2 = DirichletBC(self.Kspace, (1), self.boundaries, 2)  # land surface
        qtest = Function(self.Kspace)
        bctest1.apply(qtest.vector())
        bctest2.apply(qtest.vector())
        self.dof_surfK = np.where(qtest.vector() == 1)[0]

        topo_surf_p = griddata(x_topo, p_topo, self.coords_K[self.dof_surfK, 0], method = 'cubic')


        p_surf_func = Function(self.Kspace)
        p_surf_func.vector()[self.dof_surfK] = topo_surf_p
        self.p_topo = Expression(('ptopo'), ptopo = p_surf_func, degree = 1)

    def plot_k(self):
        self.kappafile = File(self.paraviewpath + 'kappa_2D.pvd')
        kappaplot = Function(self.Qspace)
        kappaplot.interpolate(Expression('log10(kappa)', kappa=self.kappa, degree=1))
        # print kappaplot.vector().min(), kappaplot.vector().max()
        self.kappafile << kappaplot

    def plot_topo(self):
        self.topofile = File(self.paraviewpath + 'topo_2D.pvd')
        topoplot = Function(self.Qspace)
        topoplot.interpolate(Expression('p', p=self.p_topo, degree=1))
        # print kappaplot.vector().min(), kappaplot.vector().max()
        self.topofile << topoplot

    def assemble_system(self):

        dt = self.dtt_comsum[0]

        self.p_Mass = assemble(self.Se * self.p * self.r * dx)  # Mass matrix
        self.q_Div = assemble(nabla_div(self.q) * self.r * dx)  # Divergence matrix

        self.q_K = assemble(-(self.muf / self.kappa) * inner(self.q, self.w) * dx)  # Stiffness matrix
        self.q_Grad = assemble(self.p * (nabla_div(self.w)) * dx)  # Gradient

        self.f = assemble(Constant('0')*self.r * dx)

        if self.event == 'topo':
            self.p_Boundary = assemble((1 / self.delta2) * self.p * dot(self.w, self.n) * self.ds(self.hydro) \
                                       + (1 / self.delta2) * self.p * dot(self.w, self.n) * self.ds(self.surface))

        else:
            self.p_Boundary = assemble((1 / self.delta2) * self.p * dot(self.w, self.n) * self.ds(self.hydro))

        # Left hand side (LHS)
        # self.A = self.p_Mass + dt * self.q_Div + dt * (self.q_K + self.q_Grad + self.p_Boundary)
        self.A = self.q_Div + self.q_K + self.q_Grad + self.p_Boundary

        # RHS
        # self.L = self.p_Mass
        self.L = self.f

    def streamfunction(self, vel):
        """Stream function for a given general 2D velocity field.
        The boundary conditions are weakly imposed through the term

            inner(q, grad(psi)*n)*ds,

        where grad(psi) = [-v, u] is set on all boundaries.
        This should work for any collection of boundaries:
        walls, inlets, outlets etc.
        """
        #vel = self.q_velocity
        qtest = TestFunction(self.Qspace)
        psi = TrialFunction(self.Qspace)
        a = inner(grad(qtest), grad(psi)) * dx
        # L   = dot(qtest, u[1].dx(0) - u[0].dx(1))*dx
        L = dot(qtest, curl(vel)) * dx
        bcu = []
        L = L + qtest * (self.n[1] * vel[0] - self.n[0] * vel[1]) * self.ds

        # Compute solution
        psi = Function(self.Qspace)
        A = assemble(a)
        b = assemble(L)
        normalize(b)  # Because we only have Neumann conditions
        [bc.apply(A, b) for bc in bcu]
        solve(A, psi.vector(), b)
        normalize(psi.vector())

        return psi

    def test_PSD(self):

        def is_pd(x):
            return np.all(np.linalg.eigvals(x) > 0)

        def is_psd(x, tol=1e-5):
            E, V = np.linalg.eigh(x)
            return np.all(E > -tol)

        def is_symmetric(x):
            return (x.transpose() == x).all()

        A = self.A.array()

        print('A is {}'.format('symmetric' if is_symmetric(A) else ('not symmetric')))

        np.linalg.cholesky(A)

    def solve_system(self):
        count = 0

        if self.event == 'topo':
            self.topo_flow_interp()
            p_Boundary0 = assemble(
                dt * (1 / self.delta2) * self.p_topo * dot(self.w, self.n) * self.ds(self.surface))

        print "Boundary condition set", (time.time() - self.start) / 60.0, "minutes"

        b = Vector()

        if self.event == 'topo':
            self.p_Mass.init_vector(b, 0)
            self.L.mult(self.sol_old.vector(), b)
            b += p_Boundary0
        else:
            self.p_Mass.init_vector(b, 0)
            self.L.mult(self.sol_old.vector(), b)

        [bc.apply(b) for bc in self.bcs]
        [bc.apply(self.A) for bc in self.bcs]
        print "BC's applied", (time.time() - self.start) / 60.0, "minutes"

        self.sol = Function(self.ME)
        self.solver.set_operator(self.A)
        self.solver.solve(self.sol.vector(), b)
        print "System solved", (time.time() - self.start) / 60.0, "minutes"

        p, q = self.sol.split()

        Vq_cont = VectorFunctionSpace(self.mesh, "CG", 2)
        q_cont = project(q, Vq_cont)

        self.bound_flux = assemble(dot(q, self.n) * self.ds(1))
        print "outward flux through seafloor: ", self.bound_flux
        self.bound_flux2 = assemble(dot(q, self.n) * self.ds(2))
        print "outward flux through land surface: ", self.bound_flux2

        # self.psi = self.streamfunction3D()  # compute numerical streamfunction

        ####################### SAVE SOLUTION ###########################
        p.rename('p', 'p')
        self.pfile << p
        q_cont.rename('q_cont', 'q_cont')
        self.qcontfile << q_cont
        q.rename('q', 'q')
        self.qfile << q

        # self.sigfile << self.stress

        self.sea_flux_total[count] = self.bound_flux
        # self.sea_flux[:, count] = self.q_velocity.vector()[self.ocean_dof_pgrad]
        # self.flow_velocity[:, count] = self.q_velocity.vector()
        # self.Psi[:, count] = self.psi.vector()
        # self.vol_strain[:, count] = self.v_strain.vector()
        # self.Sol_surf[:, count] = self.sol.vector()[self.surface_dofs]
        self.Sol_all[:, count] = self.sol.vector()
        # self.Sol_gps[:, count] = self.sol.vector()[self.GPS_dofs]


        ######################## PLOT SOLUTION ##########################
        if self.plot_figs == 'yes':
            HTML(X3DOM().html(p))

        ####################### UPDATE SOLUTION ###########################
        self.sol_old = self.sol
        count += 1
        print "Timestep:", count

    def save_solution(self):
        if not os.path.exists(self.results_path):
            os.makedirs(self.results_path)
        if self.permeability == 'mapped':
            np.save(self.results_path + "Sol_surf_3D_laplace_mapped.npy", self.Sol_surf)
            np.save(self.results_path + "Sol_all_3D_laplace_mapped.npy", self.Sol_all)
            # np.save(self.results_path + "sea_flux_3D_laplace_mapped.npy", self.sea_flux)
            np.save(self.var_path + "dtt_comsum_laplace_mapped.npy", self.dtt_comsum)
            np.save(self.results_path + "flow_velocity_3D_laplace_mapped.npy", self.flow_velocity)
            # np.save(self.results_path + "Streamfun_3D_laplace_mapped.npy", self.Psi)
            np.save(self.results_path + "sea_flux_total_3D_laplace_mapped.npy", self.sea_flux_total)
        elif self.permeability == 'constant':
            np.save(self.results_path + "Sol_surf_3D_laplace_mapped.npy", self.Sol_surf)
            np.save(self.results_path + "Sol_all_3D_laplace_mapped.npy", self.Sol_all)
            # np.save(self.results_path + "sea_flux_3D_laplace_mapped.npy", self.sea_flux)
            np.save(self.var_path + "dtt_comsum_laplace_mapped.npy", self.dtt_comsum)
            np.save(self.results_path + "flow_velocity_3D_laplace_mapped.npy", self.flow_velocity)
            # np.save(self.results_path + "Streamfun_3D_laplace_mapped.npy", self.Psi)
            np.save(self.results_path + "sea_flux_total_3D_laplace_mapped.npy", self.sea_flux_total)
