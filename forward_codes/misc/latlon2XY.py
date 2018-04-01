"""
latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data iinto numerical models.

INPUTS:
[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
[theta] = rotation amount/angle in radians

Assumes first column is longitude and second column is latitiude

returns transformed coordinates as numpy vectors X, Y

Kimmy McCormack 2017
"""

import numpy as np
import openpyxl as pyxl
import matplotlib.pyplot as plt


def latlon2XY(data_sheet, origin, theta):

	lon = np.array([[data_sheet.cell(row = i, column = 1).value] for i in range(2, data_sheet.max_row+1)])
	lat = np.array([[data_sheet.cell(row = i, column = 2).value] for i in range(2, data_sheet.max_row+1)])

	lon_in_km = (lon - origin[0])*111*np.cos(lat*np.pi/180)
	lat_in_km = (lat - origin[1])*111
	
	theta = np.arctan2(lat_in_km[-1],lon_in_km[-1])
	
	rho = np.sqrt(np.power(lon_in_km,2) + np.power(lat_in_km,2))
	theta_new = np.arctan2(lat_in_km,lon_in_km) - theta

	X, Y = rho*np.cos(theta_new), rho*np.sin(theta_new)

	return X, Y

def XY2latlon(data_sheet, origin, theta):

	X = 1e-3*np.array([[data_sheet.cell(row = i, column = 1).value] for i in range(2, data_sheet.max_row+1)])
	Y = 1e-3*np.array([[data_sheet.cell(row = i, column = 2).value] for i in range(2, data_sheet.max_row+1)])


	rho = np.sqrt(np.power(X,2) + np.power(Y,2))
	theta_new = np.arctan2(Y, X) + theta

	lat = rho*np.sin(theta_new)/111 + origin[1]
	lon = rho*np.cos(theta_new)/(111*np.cos(lat*np.pi/180)) + origin[0]

	return lon, lat

wb = pyxl.load_workbook('data/slab_CAS.xlsx')
data = wb['Sheet1']
origin = np.array([-85.1688, 8.6925]) # origin of top surface of mesh - used as point of rotation
theta = 0.816 # angle of rotation between latitude line and fault trace (42 deg)


X, Y = latlon2XY(data, origin, theta)
np.savetxt('XY_CAS.csv', (X, Y), delimiter=',')



#lon, lat = XY2latlon(data, origin, theta)

#plt.scatter(lon_wells_add,lat_wells_add)

#np.savetxt('lonlat.csv', (lon,lat), delimiter=',')
quit()

fig = plt.figure()
ax1 = fig.add_subplot(111)
# cp = ax1.contour(X, Y, Z_x, levels_x, linewidth=2, colors = 'black', linestyles = 'solid')
#ax1.plot(coast_x, coast_y, '-k')
wells2 = ax1.scatter(X[0:25], Y[0:25], facecolors='red')
wells = ax1.scatter(x_wells, y_wells)
plt.xlim((0, 250))
plt.ylim((0, 250))

fig.savefig('figures/well_test.png', dpi = 300)

