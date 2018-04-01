
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import norm
import openpyxl as pyxl
from scipy.spatial.distance import pdist, squareform
from scipy.optimize import curve_fit
import sympy as sym



def exponential_model(x, n, s, r):
    return n + (s - n ) *(1 - np.exp(-( 3 * x ) /r))

def spherical_model(x ,n ,s ,r):
	return n + (s - n) * (( 3 *x ) /( 2 *r) + np.power( x /( 2 *r), 3))

def latlon2XY(data_sheet, origin, theta, start):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data into numerical models.

	IN:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. svdataositive theta rotates
	counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUT:
	Returns transformed coordinates as numpy vectors X, Y
	"""

	lon = np.array \
		([[data_sheet.cell(row = i, column = 1).value] for i in range(start, data_sheet. max_row +1)]).reshape \
		(data_sheet. max_row - start +1, )
	lat = np.array \
		([[data_sheet.cell(row = i, column = 2).value] for i in range(start, data_sheet. max_row +1)]).reshape \
		(data_sheet. max_row - start +1, )


	lon_in_km = (lon - origin[0] ) * 111 *np.cos( lat *np. pi /180)
	lat_in_km = (lat - origin[1] ) *111

	rho = np.sqrt(np.power(lon_in_km ,2) + np.power(lat_in_km ,2))
	theta_new = np.arctan2(lat_in_km ,lon_in_km) - theta

	X, Y = rho* np.cos(theta_new), rho * np.sin(theta_new)

	return 1e3 * X, 1e3 * Y


def SV_single(svdata, bin, bw, angle):
	'''
	semivariogram for a single lag range
	'''

	X, Y, Z = svdata[:, 0], svdata[:, 1], svdata[:, 2]
	rho = squareform(pdist(svdata[:, :2], 'euclidean'))
	N = rho.shape[0]
	var_out = np.zeros((N,))
	n_var = np.zeros((N,))
	point_var = np.zeros((N,))

	for i in range(N):  # for each point

		theta_var = np.zeros(np.int(np.pi / angle), )
		ind_bin = np.where(np.abs(rho[i, :] - bin) <= bw)[0][1::]

		if len(ind_bin) > 0:

			values = Z[ind_bin]
			# point_var[i] = np.sum(np.power((values[1::] - Z[i]), 2))/(len(ind_bin)*2)
			theta = np.arctan2(Y[ind_bin] - Y[i], X[ind_bin] - X[i])

			for j in range(0, np.int(np.pi / angle)):
				ind_theta = np.where(np.abs(theta - (-np.pi + angle + j * 2 * angle)) <= angle)[0]
				val = values[ind_theta]

				if len(val) > 0:
					theta_var[j] = np.sum(np.power((val[:] - Z[i]), 2)) / (len(ind_theta) * 2)

			nonzero_theta = np.where(theta_var > 0)[0]
			point_var[i] = np.average(theta_var[nonzero_theta])

	nonzero = np.where(point_var > 0.0)[0]
	return np.average(point_var[nonzero])


def SV_all(svdata, bins, bws, angle):
	'''
	variogram for a collection of lags
	'''
	sv = list()
	count=0
	for bin in bins:
		sv.append(SV_single(svdata, bin, bws[count], angle))
		count += 1
	sv = [[bins[i], sv[i]] for i in range(len(bins)) if sv[i] > 0]
	return np.array(sv).T


def C(svdata, bin, bw):
	'''
	Calculate the sill
	'''
	c0 = np.var(svdata[:, 2])
	if bin == 0:
		return c0
	return c0 - SVh(svdata, bin, bw)


origin = np.array([-85.21, 8.64])  # origin of top of mesh
theta = 0.733  # angle of rotation between latitude line and fault race

data_file = 'data/Costa_Rica_data.xlsx'
wb = pyxl.load_workbook(data_file)
data = wb['Sheet5']
#data = wb['storativity']
data_start = 2

x_wells, y_wells = latlon2XY(data, origin, theta, data_start)
#kappa = np.log10(np.array([[data.cell(row = i, column = 3).value] for i in range(data_start, data.max_row + 1)]).reshape(data.max_row - data_start + 1, ))

#storage = np.log10(np.array([[data.cell(row = i, column = 3).value] for i in range(data_start, data.max_row + 1)]).reshape(data.max_row - data_start + 1, ))

elevation = np.array([[data.cell(row = i, column = 3).value] for i in range(data_start, data.max_row + 1)]).reshape(data.max_row - data_start + 1, )

# fig = plt.figure()
# ax = fig.add_subplot(111)
# ax.scatter( x_wells, y_wells, c=kappa, cmap='jet' )
# ax.set_aspect(1)
# xlim(0,22000)
# ylim(0,17500)
# plt.xlabel('Easting [m]')
# plt.ylabel('Northing [m]')
# plt.title('Kappa')

# fig.savefig('figures/variogram.png', dpi = 300)


XY = np.array((x_wells, y_wells)).T


#svdata = np.array((x_wells, y_wells, kappa)).T
svdata = np.array((x_wells, y_wells, elevation)).T

bw = 1500  # bandwidth of bins
max = 36e3
bw_increase = 10e3
svtheta = np.pi/4   # bin angle in radians

bws = [500, 500]
ntime = [bw_increase/(bws[0]*2), (max-bw_increase)/(bws[1]*2)]
bands = np.repeat(2*np.array(bws), ntime)
bins = np.cumsum(bands)


#bins = np.arange(bw, max, bw * 2)  # array of bins

sv = SV_all(svdata, bins, bands, svtheta)

add = np.array([0, .2]).reshape(2, 1)

# sv = np.hstack((add, sv))

nsr_exp, cov_exp = curve_fit(exponential_model, sv[0], sv[1], p0 = (0,1,10000))
#nsr_sp, cov_sp = curve_fit(spherical_model, sv[0], sv[1], p0 = (0,1,5000))

print nsr_exp

fig2 = plt.figure()
ax2 = fig2.add_subplot(111)
ax2.plot(sv[0], sv[1], '.', label = 'Data')
fit = ax2.plot(sv[0], exponential_model(sv[0], *nsr_exp), label = "Exponential model")
#fit2 = ax2.plot(sv[0], exponential_model(sv[0], *nsr_sp), "--", label = "Fitted Curve - spherical")
plt.xlim(0, max)
#plt.ylim(0, 1.5)
plt.xlabel('Lag [m]')
plt.ylabel('Semivariance')
plt.title('Semivariogram')
plt.legend(loc = 'upper left')
fig2.savefig('semivariogram.png', fmt = 'png', dpi = 200)
