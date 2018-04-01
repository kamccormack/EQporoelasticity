import sys
sys.path.append('/home/fenics/shared/local_lib')
from dolfin import *
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator, MultipleLocator, FormatStrFormatter
from matplotlib import cm
from mpl_toolkits.mplot3d import Axes3D
import numpy as np
import scipy as sp
from scipy.interpolate import griddata
from tempfile import TemporaryFile
import openpyxl as pyxl
import matplotlib.tri as tri
dolfin.parameters.reorder_dofs_serial = False

font = {'weight' : 'normal',
        'size'   : 10}

mpl.rc('font', **font)


def GPSlatlon2XY(data_sheet, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = np.array([[data_sheet.cell(row = i, column = 1).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )
	lat = np.array([[data_sheet.cell(row = i, column = 2).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )

	lon_u = np.array([[data_sheet.cell(row = i, column = 5).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )
	lat_u = np.array([[data_sheet.cell(row = i, column = 6).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )
	Uz = np.array([[data_sheet.cell(row = i, column = 4).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )

	lon_in_km = (lon - origin[0])*111*np.cos(lat*np.pi/180)
	lat_in_km = (lat - origin[1])*111
	
	rho_u = np.sqrt(np.power(lon_u,2) + np.power(lat_u,2))
	theta_new_u = np.arctan2(lat_u,lon_u) - theta

	rho = np.sqrt(np.power(lon_in_km,2) + np.power(lat_in_km,2))
	theta_new = np.arctan2(lat_in_km,lon_in_km) - theta

	X, Y = rho*np.cos(theta_new), rho*np.sin(theta_new)
	Ux, Uy = rho_u*np.cos(theta_new_u), rho_u*np.sin(theta_new_u)

	return 1e3*X, 1e3*Y, 1e-3*Ux, 1e-3*Uy, 1e-3*Uz

def GPSlatlon2XY_time(lat_u, lon_u, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	rho_u = np.sqrt(np.power(lon_u, 2) + np.power(lat_u, 2))
	theta_new_u = np.arctan2(lat_u, lon_u) - theta

	Ux, Uy = rho_u * np.cos(theta_new_u), rho_u * np.sin(theta_new_u)

	return Ux, Uy

def latlon2XY_points(lat, lon, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	#lon = point[0]
	#lat = point[1]

	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)

	return 1e3 * X, 1e3 * Y

def GPSlatlon2XY_other(data_sheet, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = np.array([[data_sheet.cell(row = i, column = 1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	Z = np.array([[data_sheet.cell(row = i, column = 3).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat = np.array([[data_sheet.cell(row = i, column = 2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)

	return 1e3*X, 1e3*Y, Z

def XY2latlon(X, Y, origin, theta):

	X = 1e-3*X
	Y = 1e-3*Y


	rho = np.sqrt(np.power(X, 2) + np.power(Y,2))
	theta_new = np.arctan2(Y, X) + theta

	lat = rho*np.sin(theta_new)/111 + origin[1]
	lon = rho*np.cos(theta_new)/(111*np.cos(lat*np.pi/180)) + origin[0]

	return lon, lat
##########################################################################################
path = "/home/fenics/shared/"


## Load in GPS data from txt files

#gps_loc = {}

with open(path + 'data/GD/station_list.txt', 'r') as f:
	stations = f.readlines()

names = [x.split()[0] for x in stations]
gps_lon = np.asarray([x.split()[1] for x in stations], dtype=np.float32)
gps_lat = np.asarray([x.split()[2] for x in stations], dtype=np.float32)



plot_figs = False
# print names
# print gps_lat
# print gps_lon

# gps_data = np.array(len(names))
data = {}

for name in names:

	if plot_figs:
		fig = plt.figure()

	with open(path + 'data/GD/' + name + '.lat', 'r') as latt:
		latread = latt.readlines()

	lat = np.asarray([x.split()[1] for x in latread], dtype=np.float32)
	time = np.asarray([x.split()[0] for x in latread], dtype=np.float32)
	t2_ind = np.where((time > 2012.5))[0][0] - 1
	slope = (np.average(lat[t2_ind-5:t2_ind]) - np.average(lat[0:5])) / (time[t2_ind] - time[0])


	plate_motion = (time - time[0]) * slope
	latnorm = lat - plate_motion
	latnorm -= np.average(latnorm[0:t2_ind])

	data[name + '_time'] = time
	data[name + '_lat'] = latnorm

	if plot_figs:
		ax1 = fig.add_subplot(311)
		ax1.plot(time, plate_motion, '.r')
		ax1.plot(time, latnorm, '.b')
		ax1.plot(time, lat, '.k')
		ax1.ticklabel_format(useOffset=False)


	with open(path + 'data/GD/' + name + '.lon', 'r') as lont:
		lonread = lont.readlines()

	lon = np.asarray([x.split()[1] for x in lonread], dtype=np.float32)

	t2_ind = np.where((time > 2012.5))[0][0] - 1

	slope = (np.average(lon[t2_ind-5:t2_ind]) - np.average(lon[0:5])) / (time[t2_ind] - time[0])


	plate_motion = (time - time[0]) * slope
	lonnorm = lon - plate_motion
	lonnorm -= np.average(lonnorm[0:t2_ind])

	data[name + '_lon'] = lonnorm

	if plot_figs:
		ax1 = fig.add_subplot(312)
		ax1.plot(time, plate_motion, '.r')
		ax1.plot(time, lonnorm, '.b')
		ax1.plot(time, lon, '.k')
		ax1.ticklabel_format(useOffset=False)


	with open(path + 'data/GD/' + name + '.rad', 'r') as radt:
		radread = radt.readlines()

	rad = np.asarray([x.split()[1] for x in radread], dtype=np.float32)

	t2_ind = np.where((time > 2012.5))[0][0] - 1

	slope = (np.average(rad[t2_ind - 5:t2_ind]) - np.average(rad[0:5])) / (time[t2_ind] - time[0])

	plate_motion = (time - time[0]) * slope
	radnorm = rad - plate_motion
	radnorm -= np.average(radnorm[0:t2_ind])

	data[name + '_rad'] = radnorm

	if plot_figs:
		ax1 = fig.add_subplot(313)
		ax1.plot(time, plate_motion, '.r')
		ax1.plot(time, radnorm, '.b')
		ax1.plot(time, rad, '.k')
		ax1.ticklabel_format(useOffset=False)
		fig.savefig('figures/gpsdata_%s.png' %name, dpi=400)

#for name in names:
#	print name, data[name + '_time'].shape

meshpath = "/home/fenics/shared/meshes"
mesh_size = 'medcoarse'  # fine, med, coarse
mesh = Mesh(meshpath+'/CR3D_'+mesh_size+'.xml')
boundaries = MeshFunction("size_t", mesh, meshpath+'/CR3D_'+mesh_size+'_facet_region.xml')


boundarymesh = BoundaryMesh(mesh, 'exterior')
slab_mesh = SubMesh(boundarymesh, boundaries, 3)


results_path = '/home/fenics/shared/forward_codes/results/numpy_results/'
var_path = '/home/fenics/shared/forward_codes/results/numpy_variables/'
x_all = np.load(var_path+"x_all_3D.npy")
y_all = np.load(var_path+"y_all_3D.npy")
z_all = np.load(var_path+"z_all_3D.npy")
size_p = np.load(var_path+"size_p_3D.npy")
#surface_dofs = np.load(var_path+"surface_dofs_3D.npy")
#ocean_dofs = np.load(var_path+"ocean_dofs_3D.npy")
gps_dofs = np.load(var_path + "GPS_dofs_u_3D.npy")
#slab_dofs = np.load(var_path + "slab_dofs_u_3D.npy")
#Sol_all_mapped = np.load(results_path+"Sol_all_3D_EQ_mapped.npy")
Sol_gps_mapped = np.load(results_path+"Sol_gps_3D_EQ_mapped.npy")


#Load GPS data
#origin = np.array([-85.21, 8.64]) # origin of top surface of mesh - used as point of rotation
#theta = 0.733 # angle of rotation between latitude line and fault trace
origin = np.array([-85.1688, 8.6925]) # origin of top surface of mesh - used as point of rotation
theta = 0.816 # angle of rotation between latitude line and fault trace (42 deg)

data_file = path+'data/Data.xlsx'
wb = pyxl.load_workbook(data_file)
gps_data = wb['GPS_data']

X_gps, Y_gps, Ux_gps, Uy_gps, Uz_gps = GPSlatlon2XY(gps_data, origin, theta)
Z_gps = np.zeros(len(X_gps))



#solution at GPS stations
gps_dofs_xy = gps_dofs[0:gps_dofs.shape[0]/3].astype(int)

#Sol_gps_mapped = Sol_all_mapped[gps_dofs.astype(int), :]

X_gps_model, Y_gps_model = x_all[gps_dofs_xy], y_all[gps_dofs_xy]

Ux_gps_model = Sol_gps_mapped[0:X_gps.shape[0],0]
Uy_gps_model = Sol_gps_mapped[X_gps.shape[0]:2*X_gps.shape[0],0]
Uz_gps_model = Sol_gps_mapped[2*X_gps.shape[0]:3*X_gps.shape[0],0]

Ux_gps_model_all = 1e2*Sol_gps_mapped[0:X_gps.shape[0],:]
Uy_gps_model_all = 1e2*Sol_gps_mapped[X_gps.shape[0]:2*X_gps.shape[0],:]
Uz_gps_model_all = 1e2*Sol_gps_mapped[2*X_gps.shape[0]:3*X_gps.shape[0],:]



# #Solution on the slab interface
# slab_dofs_xy = slab_dofs[0:slab_dofs.shape[0]/3].astype(int)
# Sol_slab = Sol_all_mapped[slab_dofs.astype(int), :]
# X_slab, Y_slab, Z_slab = x_all[slab_dofs_xy], y_all[slab_dofs_xy], z_all[slab_dofs_xy]
#
# Ux_slab = Sol_slab[0:X_slab.shape[0],0]
# Uy_slab = Sol_slab[X_slab.shape[0]:2*X_slab.shape[0],0]
# Uz_slab = Sol_slab[2*X_slab.shape[0]:3*X_slab.shape[0],0]
#
# U_mag_slab = np.sqrt(Ux_slab**2 + Uy_slab**2 + Uz_slab**2)
#
# print U_mag_slab.min(), U_mag_slab.max()
# #quit()
# nonzero = np.where(U_mag_slab > 0.1)[0]
#
#
# lon_slab, lat_slab = XY2latlon(X_slab, Y_slab, origin, theta)
#
# mag_all = np.concatenate((lon_slab.reshape(lon_slab.shape[0], 1), lat_slab.reshape(lon_slab.shape[0], 1), U_mag_slab.reshape(lon_slab.shape[0], 1)), axis=1)
#
# np.savetxt('inversion_mag.xyz', (mag_all), delimiter=' ')
#quit()


# data_file3 = 'data/other_slip_data.xlsx'
# wb3 = pyxl.load_workbook(data_file3)
# newman_data = wb3['newman']
#X_slab, Y_slab, Z_slab = GPSlatlon2XY_other(newman_data, origin, theta)
#X_slab_model, Y_slab_model, Z_slab_model = x_all[slab_dofs], y_all[slab_dofs], z_all[slab_dofs]

# print 'ux model: ', Ux_gps_model.min(), Ux_gps_model.max()
# print 'uy model: ', Uy_gps_model.min(), Uy_gps_model.max()
# print 'uz model: ', Uz_gps_model.min(), Uz_gps_model.max()
#
# print 'ux gps: ', Ux_gps.min(), Ux_gps.max()
# print 'uy gps: ', Uy_gps.min(), Uy_gps.max()
# print 'uz gps: ', Uz_gps.min(), Uz_gps.max()
# quit()

#print 1e3*Uz_gps_model.astype(float)

# for j in range(0,20):
# 	#print "Model XYZ:", x_all[gps_dofs.astype(int)[j]],',', y_all[gps_dofs.astype(int)[j]],',', z_all[gps_dofs.astype(int)[j]]
# 	# print "model def xyz:", Ux_gps_model[j],',', Uy_gps_model[j],',', Uz_gps_model[j]
# 	print Ux_gps_model[j], Ux_gps[j], Ux_gps_model[j]/Ux_gps[j]
# 	print Uy_gps_model[j], Uy_gps[j], Uy_gps_model[j] / Uy_gps[j]
#
# quit()

#GPS timeseries data
X_gps_post, Y_gps_post = latlon2XY_points(gps_lat, gps_lon, origin, theta)

#check for match between instaneous and timeseries
ind_gps_time = []
for i in range(0,len(names)):

	#load and rotate timeseries

	gps_time = data[names[i] + '_time']
	lon_time = data[names[i] + '_lon']
	lat_time = data[names[i] + '_lat']
	Uz_time = data[names[i] + '_rad']

	Ux_time, Uy_time = GPSlatlon2XY_time(lat_time, lon_time, theta)

	ind_time = np.where(gps_time[:]>2012.68)[0][0]
	ind = (np.abs(np.sqrt(pow((X_gps[:] - X_gps_post[i]), 2) +
	                      pow((Y_gps[:] - Y_gps_post[i]), 2)))).argmin()

	distance = np.sqrt(pow((X_gps[ind] - X_gps_post[i]), 2) +
	        pow((Y_gps[ind] - Y_gps_post[i]), 2))



	if distance <200:
		#print "##############instant vs. timeseries disp ################## "
		#Ui_gps are in meters
		# need to find earthquake timestep for U_time... not

		#print names[i], "distance: ", distance

		#print "Ux: ", Ux_gps[ind], Ux_time[ind_time]#-Ux_time[ind_time-1]
		#print "Uy: ", Uy_gps[ind], Uy_time[ind_time]#-Uy_time[ind_time-1]
		#print "Uz: ", Uz_gps[ind], Uz_time[ind_time]#-Uz_time[ind_time-1]

		ind_gps_time.append(i)

	#else:
		#print names[i], "distance: ", distance




# print Uz.min(), Uz.max(), Uz[0]
#
# print Uz_gps_post.min(), Uz_gps_post.max()
# #quit()

#gps_time_days = (gps_time - 2012.68)*365
#gps_time_days = gps_time


#indice_gps = np.abs(np.sqrt(pow(X_gps[:]-X_gps_post,2) + pow(Y_gps[:]-Y_gps_post,2))).argmin()
#indice_gps_model = np.abs(np.sqrt(pow(X_gps_model[:]-X_gps_post,2) + pow(Y_gps_model[:]-Y_gps_post,2))).argmin()





# Ux_gps_post += Ux_gps[indice_gps]
# Uy_gps_post += Uy_gps[indice_gps]
# Uz_gps_post += Uz_gps[indice_gps]
#sea_flux = np.load(results_path+"sea_flux_total_3D_EQ_mapped.npy")
#sea_flux[108] = (sea_flux[107]+sea_flux[109])/2

dtt = [15, 240, 720, 288e1, 2 * 432e2]  # , 5256e2]  # in minutes
ntime = [48, 60, 60, 30, 60]  # , 5
#dtt = [15, 240, 720, 288e1, 432e2]  # , 5256e2]  # in minutes
#ntime = [48, 60, 60, 30, 122]  # , 50]
dtt = [i * 60 for i in dtt]  # in seconds
dt = np.repeat(dtt, ntime) #timesteps
dtt_all = np.cumsum(dt) #- 5256e2*50*60 #timeline
dtt_days = dtt_all/(3600*24)

# print np.where(dtt_days > 60)[0]
# quit()

#print np.where(dtt_days >= 500.)


################################################################################################
# print (Ux_gps_model-Ux_gps)/Ux_gps, np.average(np.abs((Ux_gps_model-Ux_gps)/Ux_gps))
# print (Uy_gps_model-Uy_gps)/Uy_gps, np.average(np.abs((Uy_gps_model-Uy_gps)/Uy_gps))
# print (Uz_gps_model-Uz_gps)/Uz_gps, np.average(np.abs((Uz_gps_model-Uz_gps)/Uz_gps))
#
# print "########################################################"
#
# print (Ux_gps_model-Ux_gps), np.average(np.abs(Ux_gps_model-Ux_gps))/np.average(np.abs(Ux_gps))
# print (Uy_gps_model-Uy_gps), np.average(np.abs(Uy_gps_model-Uy_gps))/np.average(np.abs(Uy_gps))
# print (Uz_gps_model-Uz_gps), np.average(np.abs(Uz_gps_model-Uz_gps))/np.average(np.abs(Uz_gps))
#


#quit()


################################################################################################
def shiftedColorMap(cmap, start = 0, midpoint = 0.5, stop = 1.0, name = 'shiftedcmap'):
	'''
	Function to offset the "center" of a colormap. Useful for
	data with a negative min and positive max and you want the
	middle of the colormap's dynamic range to be at zero

	Input
	-----
	  cmap : The matplotlib colormap to be altered
	  start : Offset from lowest point in the colormap's range.
		  Defaults to 0.0 (no lower ofset). Should be between
		  0.0 and `midpoint`.
	  midpoint : The new center of the colormap. Defaults to
		  0.5 (no shift). Should be between 0.0 and 1.0. In
		  general, this should be  1 - vmax/(vmax + abs(vmin))
		  For example if your data range from -15.0 to +5.0 and
		  you want the center of the colormap at 0.0, `midpoint`
		  should be set to  1 - 5/(5 + 15)) or 0.75
	  stop : Offset from highets point in the colormap's range.
		  Defaults to 1.0 (no upper ofset). Should be between
		  `midpoint` and 1.0.
	'''
	cdict = {
		'red':[],
		'green':[],
		'blue':[],
		'alpha':[]
	}

	# regular index to compute the colors
	reg_index = np.linspace(start, stop, 257)

	# shifted index to match the data
	shift_index = np.hstack([
		np.linspace(0.0, midpoint, 128, endpoint = False),
		np.linspace(midpoint, 1.0, 129, endpoint = True)
	])

	for ri, si in zip(reg_index, shift_index):
		r, g, b, a = cmap(ri)

		cdict['red'].append((si, r, r))
		cdict['green'].append((si, g, g))
		cdict['blue'].append((si, b, b))
		cdict['alpha'].append((si, a, a))

	newcmap = mpl.colors.LinearSegmentedColormap(name, cdict)
	plt.register_cmap(cmap = newcmap)

	return newcmap

def plot_gps_u():

	for ind in ind_gps_time:
		# load and rotate timeseries

		gps_time = data[names[ind] + '_time']
		lon_time = data[names[ind] + '_lon']
		lat_time = data[names[ind] + '_lat']
		Uz_time = data[names[ind] + '_rad']

		Ux_time, Uy_time = GPSlatlon2XY_time(lat_time, lon_time, theta)

		ind_time = np.where(gps_time[:] > 2012.68)[0][0] + 1
		end_time = np.where(gps_time[:] > 2013.0)[0][0]

		time_post = 365*(gps_time[ind_time:end_time] - gps_time[ind_time])



		Ux_post, Uy_post, Uz_post = Ux_time[ind_time:end_time], Uy_time[ind_time:end_time], Uz_time[ind_time:end_time]

		ind_model = (np.abs(np.sqrt(pow((X_gps_model[:] - X_gps_post[ind]), 2) +
		                      pow((Y_gps_model[:] - Y_gps_post[ind]), 2)))).argmin()

		#print Ux_gps_model_all[ind_model, 0], Ux_post[0]
		#print Uy_gps_model_all[ind_model, 0], Uy_post[0]
		#print Uz_gps_model_all[ind_model, 0], Uz_post[0]

		#Ux_post += Ux_gps_model_all[ind_model, 0] - Ux_post[0]
		#Uy_post += Uy_gps_model_all[ind_model, 0] - Uy_post[0]
		#Uz_post += Uz_gps_model_all[ind_model, 0] - Uz_post[0]

		distance = np.sqrt(pow((X_gps_model[ind_model] - X_gps_post[ind]), 2) +
		                   pow((Y_gps_model[ind_model] - Y_gps_post[ind]), 2))

		#print 'distance: ', distance

		#
		fig, sol_plotgps = plt.subplots(6, sharex=True)


		#sol_plotgps[0].plot(Ux_gps_model_all[idx_gps, 0:endtime], '--k', label='constant k', linewidth=3, fillstyle='none') #pressure

		# sol_plotgps[1].plot(dtt_days[0:endtime], Uy_gps_model[idx_gps, 0:endtime], '--r', linewidth=3, fillstyle='none') # u_x
		#
		# sol_plotgps[2].plot(dtt_days[0:endtime], Uz_gps_model[idx_gps, 0:endtime], '--g', linewidth=3, fillstyle='none') # u_y

		sol_plotgps[0].plot(dtt_days,Ux_gps_model_all[ind_model, :], 'k', linewidth=3, fillstyle='none')

		sol_plotgps[2].plot(dtt_days, Uy_gps_model_all[ind_model, :], 'r', linewidth=3, fillstyle='none')

		sol_plotgps[4].plot(dtt_days, Uz_gps_model_all[ind_model, :], 'g', linewidth=3, fillstyle='none')


		sol_plotgps[0].plot(time_post , Ux_post, '.k', linewidth=3, fillstyle='none')

		sol_plotgps[2].plot(time_post , Uy_post, '.r', linewidth=3, fillstyle='none')

		sol_plotgps[4].plot(time_post , Uz_post, '.g', linewidth=3, fillstyle='none')

		sol_plotgps[1].plot(dtt_days,Ux_gps_model_all[ind_model, :], 'k', linewidth=3, fillstyle='none')

		sol_plotgps[3].plot(dtt_days, Uy_gps_model_all[ind_model, :], 'r', linewidth=3, fillstyle='none')

		sol_plotgps[5].plot(dtt_days, Uz_gps_model_all[ind_model, :], 'g', linewidth=3, fillstyle='none')

		#sol_plotgps[0].plot(gps_time_days, Ux_gps_post, '.k', linewidth=3, fillstyle='none')

		#sol_plotgps[1].plot(gps_time_days, Uy_gps_post, '.r', linewidth=3, fillstyle='none')

		#sol_plotgps[2].plot(gps_time_days, Uz_gps_post, '.g', linewidth=3, fillstyle='none')

		# sol_plotgps[0].plot(Ux_gps_post, '*k', linewidth=3, fillstyle='none')
		#
		# sol_plotgps[1].plot(Uy_gps_post, '*r', linewidth=3, fillstyle='none')
		# sol_plotgps[2].plot(Uz_gps_post, '*g', linewidth=3, fillstyle='none')


		sol_plotgps[0].set_ylabel('X disp (cm)')
		sol_plotgps[2].set_ylabel('Y disp (cm)')
		sol_plotgps[4].set_ylabel('Z disp (cm)')

		sol_plotgps[0].set_ylim([Ux_post.min()-2. , Ux_post.max()+2.])
		sol_plotgps[2].set_ylim([Uy_post.min()-2. , Uy_post.max()+2.])
		sol_plotgps[4].set_ylim([Uz_post.min()-2. , Uz_post.max()+2.])

		sol_plotgps[1].set_ylim([Ux_gps_model_all[ind_model,:].min()-.2 , Ux_gps_model_all[ind_model,:].max()+.2])
		sol_plotgps[3].set_ylim([Uy_gps_model_all[ind_model,:].min()-.2 , Uy_gps_model_all[ind_model,:].max()+.2])
		sol_plotgps[5].set_ylim([Uz_gps_model_all[ind_model,:].min()-.2 , Uz_gps_model_all[ind_model,:].max()+.2])

		sol_plotgps[0].set_xlim([0, 50])
		sol_plotgps[1].set_xlim([0, 50])
		sol_plotgps[2].set_xlim([0, 50])
		sol_plotgps[3].set_xlim([0, 50])
		sol_plotgps[4].set_xlim([0, 50])
		sol_plotgps[5].set_xlim([0, 50])

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		fig.set_size_inches(6, 8)



		#fig.savefig('figures/3d_surf_GPS_model timeseries_%s.png' % station, dpi = 400)
		fig.savefig('figures/3d_surf_GPS_model timeseries_%s.png' % ind, dpi = 400)

def plot_gps(size_gps, n):

	for i in range(0,n):
		idx_gps = i
		fig, sol_plotgps = plt.subplots(4, sharex=True)


		sol_plotgps[0].plot(Sol_gps[idx_gps, :], '--k', label='constant k', linewidth=3, fillstyle='none') #pressure

		sol_plotgps[1].plot(Sol_gps[(size_gps + idx_gps), :], '--r', linewidth=3, fillstyle='none') # u_x

		sol_plotgps[2].plot(Sol_gps[2*size_gps + idx_gps, :], '--g', linewidth=3, fillstyle='none') # u_y

		sol_plotgps[3].plot(Sol_gps[3*size_gps + idx_gps, :], '--b', linewidth=3, fillstyle='none') # u_z



		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_surf():

	points = 1e3*np.array([100 , 130])

	for i in range(0, points.shape[1]):
		idx_p = i
		fig = plt.figure()
		ax1 = fig.add_subplot(111)

		ax1.plot(Sol_surf[idx, :], 'k*', label='mapped k', linewidth=3, fillstyle='none') #pressure

		#sol_plot_surf[1].plot(Sol_surf[(size_surf + idx), :], '--r', linewidth=3, fillstyle='none') # u_x

		#sol_plot_surf[2].plot(Sol_surf[2*size_surf + idx, :], '--g', linewidth=3, fillstyle='none') # u_y

		#sol_plot_surf[3].plot(Sol_surf[3*size_surf + idx, :], '--b', linewidth=3, fillstyle='none') # u_z

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_gps_k():

	for i in range(0,2):
		idx_gps = i
		fig = plt.figure()
		ax1 = fig.add_subplot(111)


		ax1.plot(Sol_gps_k[idx_gps, :], '--k', label='mapped k', linewidth=3, fillstyle='none') #pressure

		#sol_plotgps_k[1].plot(Sol_gps_k[(size_gps + idx_gps), :], '--r', linewidth=3, fillstyle='none') # u_x

		#sol_plotgps_k[2].plot(Sol_gps_k[2*size_gps + idx_gps, :], '--g', linewidth=3, fillstyle='none') # u_y

		#sol_plotgps_k[3].plot(Sol_gps_k[3*size_gps + idx_gps, :], '--b', linewidth=3, fillstyle='none') # u_z

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		#legend1 = sol_plotgps[0].legend()
		#for label in legend1.get_texts():
		#    label.set_fontsize('medium')

		#for label in legend1.get_lines():
		#    label.set_linewidth(1.5)  # the legend line width

def plot_surf_k(points, splice):


	for i in range(0, points.shape[0]):

		fig = plt.figure()
		ax1 = fig.add_subplot(111)
		#ax2 = fig.add_subplot(212)
		#ax3 = fig.add_subplot(413)
		#ax4 = fig.add_subplot(414)

		

		ax1.plot(dtt_days[0:endtime]/splice, Sol_wells[i, 0:endtime]*1e3/(9.81), '-k', linewidth = 3, fillstyle = 'none')  # p

		#ax2.plot(dtt_days[0:endtime]/splice, Sol_surf[p_surf_ind + idx_u, 0:endtime]*100, '-r', linewidth=3, fillstyle='none') #ux

		#ax3.plot(dtt_days[0:endtime]/splice, Sol_surf[p_surf_ind+u_surf_ind + idx_u, 0:endtime]*100, '-g', linewidth = 3, fillstyle = 'none')  # uy

		#ax2.plot(dtt_days[0:endtime]/splice, Sol_surf[p_surf_ind + 2*u_surf_ind + idx_u, 0:endtime]*100, '-k', linewidth = 3, fillstyle = 'none')  # uz

		ax1.set_ylabel('head change (m)')
		#ax2.set_ylabel('X disp (cm)')
		#ax3.set_ylabel('Y disp (cm)')
		#ax2.set_ylabel('Uplift (cm)')
		#plt.ylabel('Depth (km)')
		#ax1.set_xlim((0, 3))
		#ax2.set_xlim((0, 3))
		#ax3.set_xlim((0, 3))
		plt.xlim((0, 60))
		fig.set_size_inches(8, 2.5)
		plt.xlabel('Days after EQ')
		fig.savefig('figures/3d_surf_timeseries_%s_60day_well.png'% str(points[i,0])[:3], dpi = 400)

		#legend1 = sol_plotgps[0].legend()
		#for label in legend1.get_texts():
		#    label.set_fontsize('medium')

		#for label in legend1.get_lines():
		#    label.set_linewidth(1.5)  # the legend line width

def plot_gps_both(Sol_gps, Sol_gps_k, size_gps, n):

	for i in range(0,n):
		idx_gps = i
		fig, sol_plotgps_both = plt.subplots(4, sharex=True)

		sol_plotgps_both[0].plot(Sol_gps[idx_gps, :], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure
		sol_plotgps_both[0].plot(Sol_gps_k[idx_gps, :], '--k', label='mapped k', linewidth=3, fillstyle='none') #pressure

		sol_plotgps_both[1].plot(Sol_gps[(size_gps + idx_gps), :], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		sol_plotgps_both[1].plot(Sol_gps_k[(size_gps + idx_gps), :], '--r', linewidth=3, fillstyle='none') # u_x

		sol_plotgps_both[2].plot(Sol_gps[2*size_gps + idx_gps, :], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		sol_plotgps_both[2].plot(Sol_gps_k[2*size_gps + idx_gps, :], '--g', linewidth=3, fillstyle='none') # u_y

		sol_plotgps_both[3].plot(Sol_gps[3*size_gps + idx_gps, :], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		sol_plotgps_both[3].plot(Sol_gps_k[3*size_gps + idx_gps, :], '--b', linewidth=3, fillstyle='none') # u_z

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_surf_both(points):

	for i in range(0, points.shape[1]):
		idx_p = np.abs(np.sqrt(np.power((x_surf_p[:] - points[i,0]),2) + np.power((y_surf_p[:] - points[i,1]),2))).argmin()
		idx_u = np.abs(np.sqrt(np.power((x_surf_ux[:] - points[i,0]),2) + np.power((y_surf_ux[:] - points[i,1]),2))).argmin()
		fig, sol_plot_surf_both = plt.subplots(4, sharex=True)

		sol_plot_surf_both[0].plot(dtt[0:endtime], Sol_surf[idx_p, 0:endtime], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure
		#sol_plot_surf_both[0].plot(Sol_surf_k[idx, :], '--k', label='mapped k', linewidth=3, fillstyle='none') #pressure

		sol_plot_surf_both[1].plot(dtt[0:endtime], Sol_surf[(p_surf_ind + idx_u), 0:endtime], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		#sol_plot_surf_both[1].plot(Sol_surf_k[(size_surf + idx), :], '--r', linewidth=3, fillstyle='none') # u_x

		sol_plot_surf_both[2].plot(dtt[0:endtime], Sol_surf[p_surf_ind+u_surf_ind + idx_u, 0:endtime], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		#sol_plot_surf_both[2].plot(Sol_surf_k[2*size_surf + idx, :], '--g', linewidth=3, fillstyle='none') # u_y

		sol_plot_surf_both[3].plot(dtt[0:endtime], Sol_surf[p_surf_ind+2*u_surf_ind + idx_u, 0:endtime], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		#sol_plot_surf_both[3].plot(Sol_surf_k[3*size_surf + idx, :], '--b', linewidth=3, fillstyle='none') # u_z

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plot_surf_both[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

		fig.savefig('figures/3d_surf_timeseries_%s.png'% str(points[i,1])[:2], dpi = 400)

def plot_gps_elastic_compare(nstations):

	for i in range(0,nstations):
		idx_gps = i
		fig, sol_plotgps_compare = plt.subplots(4, sharex=True)

		sol_plotgps_compare[0].plot(Sol_gps[idx_gps, :], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure

		sol_plotgps_compare[1].plot(Sol_gps[(size_gps + idx_gps), :], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		sol_plotgps_compare[1].plot(x_elastic, elastic_gps[idx_gps, :],'ro', fillstyle='full') # u_x

		sol_plotgps_compare[2].plot(Sol_gps[2*size_gps + idx_gps, :], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		sol_plotgps_compare[2].plot(x_elastic, elastic_gps[size_gps + idx_gps, :],'go', fillstyle='full') # u_x

		sol_plotgps_compare[3].plot(Sol_gps[3*size_gps + idx_gps, :], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		sol_plotgps_compare[3].plot(x_elastic, elastic_gps[2*size_gps+idx_gps, :],'bo', fillstyle='full') # u_x

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps_compare[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_surf_elastic_compare(points):


	for i in range(0, points.shape[1]):

		idx = (np.abs(np.sqrt(pow((x_surf - plotpoint[0]),2) + pow((y_surf - plotpoint[1]),2)))).argmin()

		fig, sol_plot_surf_compare = plt.subplots(4, sharex=True)

		sol_plot_surf_compare[0].plot(Sol_surf[idx, :], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure


		sol_plot_surf_compare[1].plot(Sol_surf[(size_surf + idx_gps), :], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		sol_plot_surf_compare[1].plot(x_elastic, elastic_surf[idx_gps, :],'ro', fillstyle='full') # u_x

		sol_plot_surf_compare[2].plot(Sol_surf[2*size_surf + idx_gps, :], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		sol_plot_surf_compare[2].plot(x_elastic, elastic_surf[size_gps + idx_gps, :],'go', fillstyle='full') # u_x

		sol_plot_surf_compare[3].plot(Sol_surf[3*size_surf + idx_gps, :], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		sol_plot_surf_compare[3].plot(x_elastic, elastic_surf[2*size_gps+idx_gps, :],'bo', fillstyle='full') # u_x

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_gps_all(nstations):

	for i in range(0,nstations):
		idx_gps = i
		fig, sol_plotgps_all = plt.subplots(4, sharex=True)

		sol_plotgps_all[0].plot(Sol_gps[idx_gps, :], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure
		sol_plotgps_all[0].plot(Sol_gps_k[idx_gps, :], '--k', label='mapped k', linewidth=3, fillstyle='none') #pressure

		sol_plotgps_all[1].plot(Sol_gps[(size_gps + idx_gps), :], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		sol_plotgps_all[1].plot(Sol_gps_k[(size_gps + idx_gps), :], '--r', linewidth=3, fillstyle='none') # u_x
		sol_plotgps_all[1].plot(x_elastic, elastic_gps_k[idx_gps, :],'ro', fillstyle='full') # u_x

		sol_plotgps_all[2].plot(Sol_gps[2*size_gps + idx_gps, :], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		sol_plotgps_all[2].plot(Sol_gps_k[2*size_gps + idx_gps, :], '--g', linewidth=3, fillstyle='none') # u_y
		sol_plotgps_all[2].plot(x_elastic, elastic_gps_k[size_gps + idx_gps, :],'go', fillstyle='full') # u_x

		sol_plotgps_all[3].plot(Sol_gps[3*size_gps + idx_gps, :], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		sol_plotgps_all[3].plot(Sol_gps_k[3*size_gps + idx_gps, :], '--b', linewidth=3, fillstyle='none') # u_z
		sol_plotgps_all[3].plot(x_elastic, elastic_gps_k[2*size_gps+idx_gps, :],'bo', fillstyle='full') # u_x

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_surf_all(Sol_surf, Sol_surf_k, elastic_surf, size_surf, x_surf, y_surf):

	points = 1e3*np.array([40 , 100])

	nsteps = Sol_surf.shape[1]
	x_elastic = np.array([0,nsteps-1])


	for i in range(0, points.shape[1]):

		idx = (np.abs(np.sqrt(pow((x_surf - plotpoint[0]),2) + pow((y_surf - plotpoint[1]),2)))).argmin()

		fig, sol_plot_surf_all = plt.subplots(4, sharex=True)

		sol_plot_surf_all[0].plot(Sol_surf[idx, :], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure
		sol_plot_surf_all[0].plot(Sol_surf_k[idx, :], '--k', label='mapped k', linewidth=3, fillstyle='none') #pressure

		sol_plot_surf_all[1].plot(Sol_surf[(size_surf + idx_gps), :], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		sol_plot_surf_all[1].plot(Sol_surf_k[(size_surf + idx_gps), :], '--r', linewidth=3, fillstyle='none') # u_x
		sol_plot_surf_all[1].plot(x_elastic, elastic_surf[idx_gps, :],'ro', fillstyle='full') # u_x

		sol_plot_surf_all[2].plot(Sol_surf[2*size_surf + idx_gps, :], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		sol_plot_surf_all[2].plot(Sol_surf_k[2*size_surf + idx_gps, :], '--g', linewidth=3, fillstyle='none') # u_y
		sol_plot_surf_all[2].plot(x_elastic, elastic_surf[size_gps + idx_gps, :],'go', fillstyle='full') # u_x

		sol_plot_surf_all[3].plot(Sol_surf[3*size_surf + idx_gps, :], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		sol_plot_surf_all[3].plot(Sol_surf_k[3*size_surf + idx_gps, :], '--b', linewidth=3, fillstyle='none') # u_z
		sol_plot_surf_all[3].plot(x_elastic, elastic_surf[2*size_gps+idx_gps, :],'bo', fillstyle='full') # u_x

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_pressure_contour():
	fig = plt.figure()
	ax1 = fig.add_subplot(111)


	i = 1

	#stream = streamfun[:, i]
	Z = mag_flux_grid[:,:, i]



	triang = tri.Triangulation(x_surf_all, y_surf_all)
	triang_u = tri.Triangulation(x_ocean_u, y_ocean_u)
	contour_steps = 10
	num_lines = 10
	orig_cmap = cm.coolwarm
	midpoint = 1 - Z.max() / (Z.max() + abs(Z.min()))
	shifted_cmap = shiftedColorMap(orig_cmap, midpoint = midpoint)

	zero = np.array([0])



	# levels_10 = np.array([Z_10.min(), Z_10.min()/2, Z_10.min()/4, Z_10.min()/8, Z_10.min()/16,
	#                      0, Z_10.max()/16, Z_10.max()/8, Z_10.max()/4, Z_10.max()/2, Z_10.max()])
	levels = np.linspace(seaflux_mag[:,i].min(), seaflux_mag[:,i].max(), num = contour_steps, endpoint = True)
	lines = np.linspace(seaflux_mag[:,i].min(), seaflux_mag[:,i].max(),  num = num_lines, endpoint = True)
	#lines_stream = np.linspace(stream.min(), stream.max(), num = num_lines, endpoint = True)


	skip = (slice(None, None, 10))

	contourf = ax1.contourf(x_grid,y_grid, Z, levels, cmap = orig_cmap)
	contour = ax1.contour(x_grid,y_grid, Z, lines, linewidth = 1, colors = 'black', linestyles = 'solid')


	#quiver = ax1.quiver(x_ocean_u[skip], y_ocean_u[skip], seaflux_x[skip][:, i], seaflux_y[skip][:, i], units = 'inches', scale_units = 'width', width = 1.7e-2, scale = .3/1e5, color = 'darkcyan')
	# quiver = ax1.quiver(x_ux[skip], y_ux[skip], vel_x[skip][:, i], vel_y[skip][:, i], scale_units = 'width', scale = 1.0 / 2e5)
	# streamline10 = ax1.streamplot(x_stream, y_stream, Z_x_stream[:, :, i], Z_y_stream[:, :, i], density = [1.0, 3.0], linewidth = lw)
	#streamfun_contour = ax1.tricontour(triang, stream, lines_stream, linewidth =1, colors = 'blue',linestyles = 'solid')



	fig.colorbar(contourf)
	#ax1.set_ylim([-2e3, 1e3])
	#ax1.set_xlim([55e3, 80e3])

	#fig.set_size_inches(8, 2)


	fig.savefig('figures/3d_contour.png', dpi = 400)

def plot_sea_flux():
	fig = plt.figure()
	ax1 = fig.add_subplot(111)
	# ax2 = fig.add_subplot(312)
	# ax3 = fig.add_subplot(313)
	# ax1.set_title('Flux through seafloor')
	
	end = 300
	
	# Co-seismic displacement
	ax1.plot(dtt_days/30, np.zeros(dtt_days.shape[0]), '--k', linewidth = 1)
	#ax1.plot(dtt_days[48:51+end]/30, flux_topo[48:51+end], '-b',label='Topographic flow',  linewidth = 2)
	#ax1.plot(dtt_all[51:51+250], flux_topo[51:51+250], '-b',  linewidth = 3)

	#ax1.plot(dtt_all, flux_subEQ, '-r',  linewidth = 3)
	#ax1.plot(dtt_all[48:50], flux_EQ[48:50], '-r', linewidth = 3)
	#ax1.plot(dtt_days[48:51+end]/30, flux_subEQ[48:51+end], '-r', label='Tectonic flow', linewidth = 2)


	ax1.plot(dtt_days/30, sea_flux, '-r', label='Tectonic flow', linewidth = 2)



		# ax3.plot(dtt[50:-1], sea_flux[50:-1], '-b*', ms=2, mew=2, fillstyle='none', color = 'red')
	# ax1.set_ylabel('Darcy flux (meters/year)')
	# ax2.set_ylabel('Darcy flux (meters/sec)')
	# ax3.set_ylabel('Darcy flux (meters/year)')
	# ax1.set_xlabel('Years')
	
	#ax1.set_yscale('log')
	#ax1.set_ylim([-5e5, 1.5e6])
	ax1.set_ylim([-5e1, 1.5e3])
	ax1.set_ylim([-5, 5])
	ax1.set_xlim([-.1 , 24])
	ax1.set_xlabel('Months after earthquake')
	#ax1.legend(fontsize='small', loc=1)
	
	#fig.set_size_inches(8, 2.5)
	
	#becomes negative at 8 days, positive again at 215
	#above topo for 5 days
	
	fig.savefig('figures/3d_sea_flux.png', dpi = 400)

def plot_gps_surface():
	
	fig = plt.figure()
	ax1 = fig.add_subplot(111)

	
	# quiver = ax1.quiver(x_ocean_u[skip], y_ocean_u[skip], seaflux_x[skip][:, i], seaflux_y[skip][:, i],
	#                    units='inches',scale_units='width', width=1.7e-2, scale=.3 / 1e5, color='darkcyan')
	#
	quiver = ax1.quiver(X_gps/1e3, Y_gps/1e3, Ux_gps, Uy_gps, color = 'blue', label='GPS data',units='inches',
	                    scale_units='width', width=1e-2, scale=5e0)
	quiver2 = ax1.quiver(X_gps/1e3, Y_gps/1e3, Ux_gps_model, Uy_gps_model, color = 'red', label='model', units='inches',
	                     scale_units='width', width=1e-2, scale=5e0)

	# scale = 4.0
	# quiver = ax1.quiver(X_gps/1e3, Y_gps/1e3, Z_gps, Uz_gps, color = 'blue', label='GPS data',units='inches',
	#                     scale_units='width', width=2e-2, scale=10)
	# quiver2 = ax1.quiver(X_gps/1e3, Y_gps/1e3, Z_gps, Uz_gps_model, color = 'red', label='model',
	#                      units='inches',scale_units='width', width=2e-2, scale=10)
	# #
	# ax1.plot(Uz_gps, 'b*', color = 'blue', label='GPS data')
	# ax1.plot(Uz_gps_model, 'ro', color = 'red', label='model')
	#ax1.plot(Uz_gps_model/Uz_gps, 'ro', color = 'red', label = 'model')

	ax1.legend(fontsize='small', loc=1)
	ax1.set_ylim([50, 250])
	ax1.set_xlim([30 , 200])

	#fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_gps_surface.png', dpi = 400)
	#fig.savefig('figures/3d_gps_up_scatter.png', dpi = 400)

def plot_gps_surface_xyz():
	fig = plt.figure()
	fig, plot = plt.subplots(3, sharex = True)

	#
	plot[0].plot(Ux_gps, 'b*', color = 'blue', label='GPS data')
	plot[0].plot(Ux_gps_model, 'ro', color = 'red', label='model')
	plot[0].plot(Z_gps, '--k', label = 'model')

	plot[1].plot(Uy_gps, 'b*', color = 'blue', label='GPS data')
	plot[1].plot(Uy_gps_model, 'ro', color = 'red', label='model')
	plot[1].plot(Z_gps, '--k', label = 'model')

	plot[2].plot(Uz_gps, 'b*', color = 'blue', label='GPS data')
	plot[2].plot(Uz_gps_model, 'ro', color = 'red', label='model')
	plot[2].plot(Z_gps, '--k', label = 'model')
	#ax1.plot(Uz_gps_model/Uz_gps, 'ro', color = 'red', label = 'model')

	plot[0].legend(fontsize = 'small', loc = 4)
	# ax1.set_ylim([0, 250])
	# ax1.set_xlim([0 , 250])

	# fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_gps_surface_compare.png', dpi = 400)

	fig = plt.figure()
	fig, plot = plt.subplots(3, sharex = True)

	plot[0].plot((Ux_gps_model-Ux_gps)/Ux_gps, '.b', color = 'blue', label='GPS data')
	plot[0].plot(Z_gps, '--k', label = 'model')

	plot[1].plot((Uy_gps_model-Uy_gps)/Uy_gps, '.b', color = 'blue', label='GPS data')
	plot[1].plot(Z_gps, '--k', label = 'model')

	plot[2].plot((Uz_gps_model-Uz_gps)/Uz_gps, '.b', color = 'blue', label='GPS data')
	plot[2].plot(Z_gps, '--k', label = 'model')
	#ax1.plot(Uz_gps_model/Uz_gps, 'ro', color = 'red', label = 'model')

	#plot[0].legend(fontsize = 'small', loc = 1)
	# ax1.set_ylim([0, 250])
	# ax1.set_xlim([0 , 250])

	# fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_gps_surface_compare_ratio.png', dpi = 400)

def plot_slab_depth():
	fig = plt.figure()
	ax1 = fig.add_subplot(111)


	ax1.plot(X_slab / 1e3, Z_slab / 1e3, '.k', label = 'GPS data')
	ax1.plot(X_slab_model / 1e3, Z_slab_model / 1e3, '.r', label = 'model',)

	ax1.legend(fontsize = 'small', loc = 1)
	# ax1.set_ylim([0, 250])
	# ax1.set_xlim([0 , 250])

	# fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_slab.png', dpi = 400)

def plot_slab_contour():
	fig = plt.figure()
	ax1 = fig.add_subplot(111)

	#triang = tri.Triangulation(X_slab, Y_slab, slab_mesh.cells())
	triang = tri.Triangulation(1e-3*X_slab, 1e-3*Y_slab)
	#triang = tri.Triangulation(lon_slab, lat_slab)


	Z = U_mag_slab[:]
	Z_slab_plot = 1e-3*Z_slab
	Zmin, Zmax = Z.min(), Z.max()

	contour_steps = 10
	num_lines = 10
	#orig_cmap = cm.coolwarm
	#midpoint = 1 - Zmax / (Zmax + abs(Zmin))
	#shifted_cmap = shiftedColorMap(orig_cmap, midpoint = midpoint)

	zero = np.array([0])

	# levels_10 = np.array([Z_10.min(), Z_10.min()/2, Z_10.min()/4, Z_10.min()/8, Z_10.min()/16,
	#                      0, Z_10.max()/16, Z_10.max()/8, Z_10.max()/4, Z_10.max()/2, Z_10.max()])
	levels = np.linspace(Zmin, Zmax, num = contour_steps, endpoint = True)
	#levels = np.linspace(0.0, 3., num = 13, endpoint = True)

	lines = np.linspace(Zmin, Zmax, num = num_lines, endpoint = True)
	#lines_slab = np.linspace(Z_slab.min, Z_slab.max(), num = 5, endpoint = True)
	lines_slab = -1.0*np.array([50,40,30,20,10])


	skip = (slice(None, None, 7))


	cpf = ax1.tricontourf(triang, Z, levels, cmap = cm.jet, alpha=.5)
	#cp = ax1.tricontour(triang, Z, lines, linewidth = 10 ,colors = 'black', linestyles = 'solid')
	quiver10 = ax1.quiver(1e-3*X_slab[skip], 1e-3*Y_slab[skip], Ux_slab[skip], Uy_slab[skip],  units='inches',
	                     scale_units='width', width=1.5e-2, scale=30e0)

	cp = ax1.tricontour(triang, Z_slab_plot, lines_slab, linewidth = 10 ,colors = 'black', linestyles = 'dashed')


	#ax1.legend(fontsize = 'small', loc = 1)
	plt.axis('scaled')
	ax1.set_ylim([0, 250])
	ax1.set_xlim([0, 130])

	#set tickmarks
	majorLocator = MultipleLocator(25)
	majorFormatter = FormatStrFormatter('%d')
	minorLocator = MultipleLocator(5)
	ax1.xaxis.set_major_locator(majorLocator)
	ax1.xaxis.set_major_formatter(majorFormatter)
	# for the minor ticks, use no labels; default NullFormatter
	ax1.xaxis.set_minor_locator(minorLocator)
	ax1.tick_params(which = 'both', direction = 'out')

	fig.colorbar(cpf)


	# fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_slab.png', dpi = 400)


# ax3.set_yscale('log')

#Z = mag_flux_grid[:,:, 1]
#print Z.max(), Z.min()
#points = 1e3*np.array([[40, 140], [50, 140], [60.5, 140], [70, 140], [80, 140], [90, 140], [100, 140], [110, 140], [120, 140], [130, 140]])
#points = 1e3*np.array([[78, 140], [96, 129], [61, 160], [114,124]])
points = 1e3*np.array([[83.5, 141]])


splice = 1
endtime = 300

#plot_surf_k()
#plot_gps_elastic_compare(20)
#plot_surf_k(wellpoints, splice)
#plot_pressure_contour()
#plot_sea_flux()
#plot_gps_surface()
#plot_gps_surface_xyz()
plot_gps_u()
#plot_slab_depth()
#plot_slab_contour()

#fig3.set_size_inches(7, 4)
#plt.xlim((0, 100))
#plt.ylim((5e-8, 10))
#fig2.savefig('figures/SGD_cumulative.png', dpi = 400)

