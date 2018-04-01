from dolfin import *
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from matplotlib import cm
from mpl_toolkits.mplot3d import Axes3D
import numpy as np
import scipy as sp
from scipy.interpolate import griddata
from tempfile import TemporaryFile
import openpyxl as pyxl
import matplotlib.tri as tri
dolfin.parameters.reorder_dofs_serial = False

font = {'weight' : 'normal',
        'size'   : 17}

mpl.rc('font', **font)

mesh_size = 'med'  # fine, med, coarse
# path = "/home/fenics/shared"
# mesh = Mesh(path+'/meshes/CR3D_'+mesh_size+'.xml')
# boundaries = MeshFunction("size_t", mesh, path+'/meshes/CR3D_'+mesh_size+'_facet_region.xml')

def GPSlatlon2XY(data_sheet, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = np.array([[data_sheet.cell(row = i, column = 1).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )
	lat = np.array([[data_sheet.cell(row = i, column = 2).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )

	lon_u = np.array([[data_sheet.cell(row = i, column = 5).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )
	lat_u = np.array([[data_sheet.cell(row = i, column = 6).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )
	Uz = np.array([[data_sheet.cell(row = i, column = 4).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )

	lon_in_km = (lon - origin[0])*111*np.cos(lat*np.pi/180)
	lat_in_km = (lat - origin[1])*111
	
	rho_u = np.sqrt(np.power(lon_u,2) + np.power(lat_u,2))
	theta_new_u = np.arctan2(lat_u,lon_u) - theta

	rho = np.sqrt(np.power(lon_in_km,2) + np.power(lat_in_km,2))
	theta_new = np.arctan2(lat_in_km,lon_in_km) - theta

	X, Y = rho*np.cos(theta_new), rho*np.sin(theta_new)
	Ux, Uy = rho_u*np.cos(theta_new_u), rho_u*np.sin(theta_new_u)

	return 1e3*X, 1e3*Y, 1e-3*Ux, 1e-3*Uy, 1e-3*Uz


def GPSlatlon2XY_time(data_sheet, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""


	lon_u = np.array([[data_sheet.cell(row = i, column = 5).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat_u = np.array([[data_sheet.cell(row = i, column = 6).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	Uz = np.array([[data_sheet.cell(row = i, column = 4).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	rho_u = np.sqrt(np.power(lon_u, 2) + np.power(lat_u, 2))
	theta_new_u = np.arctan2(lat_u, lon_u) - theta

	Ux, Uy = rho_u * np.cos(theta_new_u), rho_u * np.sin(theta_new_u)

	return 1e-3 * Ux, 1e-3 * Uy, 1e-3 * Uz

def latlon2XY_point(point, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = point[0]
	lat = point[1]

	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)

	return 1e3 * X, 1e3 * Y

station = 'CABA'

results_path = 'results/numpy_results_oct17/'
var_path = 'results/numpy_variables_oct17/'
x_all = np.load(var_path+"x_all_3D.npy")
y_all = np.load(var_path+"y_all_3D.npy")
z_all = np.load(var_path+"z_all_3D.npy")
size_p = np.load(var_path+"size_p_3D.npy")
#surface_dofs = np.load(var_path+"surface_dofs_3D.npy")
#ocean_dofs = np.load(var_path+"ocean_dofs_3D.npy")
gps_dofs = np.load(var_path + "GPS_dofs_u_3D.npy")


with open('data/station_list.txt', 'r') as document:
    stations = {}
    for line in document:
        line = line.split()
        if not line:  # empty line?
            continue
        stations[line[0]] = line[1:]

latlongps = np.array(stations[station]).astype(np.float)


#results_path = "/home/fenics/shared/CR_forward/results/numpy_results/"
# Sol_all = np.load(results_path+"Sol_all_3D_EQ_mapped.npy")
# Sol_gps_k = np.load(results_path+"Sol_gps_3D_EQ_mapped.npy")
# Sol_surf_k = np.load(results_path+"Sol_surf_3D_EQ_mapped.npy")
# #Sol_gps_constant = np.load(results_path+"Sol_gps_3D_EQ_constant.npy")
# #Sol_surf = np.load(results_path+"Sol_surf_3D_EQ_constant.npy")
# #Sol_gps_mapped = np.load(results_path+"Sol_gps_3D_EQ_mapped.npy")
#
# print Sol_all.shape, Sol_all[gps_dofs[1],0:10]
# print Sol_gps_k.shape, Sol_gps_k[1,0:10]
# print Sol_surf_k.shape, Sol_surf_k[100000,0:10]
# quit()
#
#
# early_def = Sol_surf[:,0] - Sol_surf[:,156]
# late_def = Sol_surf[:,156] - Sol_surf[:,-1]
#
# print early_def.min(), early_def.max()
# print late_def.min(), late_def.max()
# quit()


Sol_all_mapped = np.load(results_path+"Sol_all_3D_EQ_mapped.npy")
Sol_gps_mapped = Sol_all_mapped[gps_dofs.astype(int), :]

# wellpoints = 1e3*np.array([[71, 150], [88, 125], [93, 149], [113,140]])
# well_dofs = np.empty(len(wellpoints))
#
# for i in range(len(well_dofs)):
# 	well_dofs[i] = np.abs(np.sqrt(np.power((x_all[0:size_p] - wellpoints[i, 0]), 2)
#                        + np.power((y_all[0:size_p] - wellpoints[i, 1]), 2)
# 	                    + np.power((z_all[0:size_p] - 0), 2))).argmin()
#
# Sol_wells = Sol_all_mapped[well_dofs.astype(int), :]
# np.save(results_path + "Sol_wells_3D.npy", Sol_wells)
#Sol_wells = np.load(results_path+"Sol_wells_3D.npy")

#print Sol_wells.min(), Sol_wells.max()


origin = np.array([-85.21, 8.64]) # origin of top surface of mesh - used as point of rotation
theta = 0.733 # angle of rotation between latitude line and fault trace
data_file = 'data/Data.xlsx'
wb = pyxl.load_workbook(data_file)
gps_data = wb['GPS_data']

X_gps, Y_gps, Ux_gps, Uy_gps, Uz_gps = GPSlatlon2XY(gps_data, origin, theta)
Z_gps = np.zeros(len(X_gps))


Ux_gps_model = Sol_gps_mapped[0:X_gps.shape[0],0]
Uy_gps_model = Sol_gps_mapped[X_gps.shape[0]:2*X_gps.shape[0],0]
Uz_gps_model = Sol_gps_mapped[2*X_gps.shape[0]:3*X_gps.shape[0],0]

Ux_gps_model_all = Sol_gps_mapped[0:X_gps.shape[0],:]
Uy_gps_model_all = Sol_gps_mapped[X_gps.shape[0]:2*X_gps.shape[0],:]
Uz_gps_model_all = Sol_gps_mapped[2*X_gps.shape[0]:3*X_gps.shape[0],:]


XYgps = latlon2XY_point(latlongps, origin, theta)


#print Sol_gps_mapped.shape
#print Ux_gps - Sol_gps_mapped[0:X_gps.shape[0],0]
#print X_gps.shape
#quit()

#elastic_drained_gps = np.load(results_path+"Sol_gps_drained_3D.npy")
#elastic_drained_surf = np.load(results_path+"Sol_surf_drained_3D.npy")
#elastic_undrained_gps = np.load(results_path+"Sol_gps_undrained_3D.npy")
#elastic_undrained_surf = np.load(results_path+"Sol_surf_undrained_3D.npy")

#streamfun = np.load(results_path+"psi_3D_EQ_mapped.npy")
#velocity = np.load(results_path+"flow_velocity_3D_EQ_mapped.npy")

#sea_flux_topo = np.load(results_path+"sea_flux_total_3D_topo_mapped.npy")
sea_flux = np.load(results_path+"sea_flux_total_3D_EQ_mapped.npy")
#Sol_surf = np.load(results_path+"Sol_surf_3D_EQ_mapped.npy")
#Sol_all = np.load(results_path+"Sol_all_3D_EQ_mapped.npy")
# sub_cycle = np.load("saved_variables_poroelastic/sub_cycle_EQ.npy")
#dtt = np.load(var_path+"dtt_comsum_EQ.npy")

#print sea_flux.min()*(3600*24), sea_flux[1]*(3600*24)
#quit()


dtt = [15, 240, 288e1, 432e2]#, 5256e2]  # in minutes
ntime = [96, 120, 15, 122]#, 50]
#dtt = [15, 240, 288e1, 432e2]#, 5256e2]  # in minutes
#ntime = [96, 120, 15, 122]#, 50]
dtt = [i * 60 for i in dtt]  # in seconds
dt = np.repeat(dtt, ntime) #timesteps
dtt_all = np.cumsum(dt) #- 5256e2*50*60 #timeline
dtt_days= dtt_all/(3600*24)


#
# print np.abs(dtt_days[:] - 8.0).argmin()
#
#
# sea_flux[0] = sea_flux[1]
# flux_EQ = np.hstack([np.zeros(50),sea_flux])*(3600*24)
# topoflux = 11.829*(3600*24)
# flux_topo = topoflux*np.ones(len(dtt_all))
# zero_line = np.zeros(len(dtt_all))
# topo_10days = flux_topo[0]*(8)
# sub_flux = 0.5*(3600*24)
# sea_flux_sub = sub_flux*np.ones(len(dtt_all))
# flux_subEQ = flux_EQ+sea_flux_sub
# EQflux_cum = np.sum(flux_subEQ[50:187]*dt[50:187])
# dttsum = np.sum(dt[50:50+96])/(3600*24)
# eqper_s = EQflux_cum/(3600*24*30)
# print "subEQ max, min [m3/day]: ",flux_subEQ.max(),', ', flux_subEQ.min()
# print "total EQ flux over 8 days: ", EQflux_cum
# print "total topo flow over 8 days: ", topo_10days
# print "EQ/topo ratio over 8 days: ", EQflux_cum/topo_10days
# quit()




#size_ux = (Sol_all.shape[0] - size_p) / 3
#size_flux = sea_flux.shape[0] / 3

#p_surf_ind = np.argmax(surface_dofs >= size_p)
#u_surf_ind = (surface_dofs.shape[0] - p_surf_ind) / 3

#size_ux = (u_surf_ind - p_surf_ind) / 3

#ux_ind = p_ind + size_ux



#Z_p = Sol_all[0:size_p, :]
#Z_ux = Sol_all[size_p:size_p + size_ux, :]
#Z_uy = Sol_all[size_p + size_ux::, :]

#norm_sea = np.array([.06652, 0.997785])

#p_ind_ocean = np.argmax(ocean_dofs >= size_p)
#ux_ind_ocean = np.argmax(ocean_dofs >= size_p + size_ux)
#ocean_dof_p = ocean_dofs[0:p_ind]#[np.argsort(x_all[ocean_dofs[0:p_ind]])]
#ocean_dof_ux = ocean_dofs[p_ind::]#[np.argsort(x_all[ocean_dofs[p_ind:ux_ind]])]
#ocean_dof_uy = ocean_dofs[ux_ind::]#[np.argsort(x_all[ocean_dofs[ux_ind::]])]

#
# surf_dof_p = surface_dofs[0:p_surf_ind]
# surf_dof_ux = surface_dofs[p_surf_ind:p_surf_ind + u_surf_ind]
# surf_dof_uy = surface_dofs[p_surf_ind:p_surf_ind + 2*u_surf_ind]
# surf_dof_uz = surface_dofs[p_surf_ind + 2*u_surf_ind::]
#
# x_surf_p, y_surf_p = x_all[surf_dof_p], y_all[surf_dof_p]
# x_surf_ux, y_surf_ux = x_all[surf_dof_ux], y_all[surf_dof_ux]

#all_surf_dof_p = np.hstack((ocean_dof_p,surface_dof_p))

"""
seaflux_x = norm_sea[0] * velocity[ocean_dof_p]
seaflux_y = norm_sea[1] * velocity[ocean_dof_p + size_p]

seaflux_mag = np.sqrt((pow(seaflux_x,2) + pow(seaflux_y,2)))

surfflux_x_all = norm_sea[0] * velocity[all_surf_dof_p]
surfflux_y_all = norm_sea[1] * velocity[all_surf_dof_p + size_p]

surfflux_mag = np.sqrt((pow(surfflux_x_all,2) + pow(surfflux_y_all,2)))


elastic_gps  = np.empty((elastic_drained_gps.shape[0],2))
elastic_surf  = np.empty((elastic_drained_surf.shape[0],2))
x_elastic = np.array([0,Sol_gps.shape[1]-1])


#convert displacements from meters to cm
for j in range(0,elastic_undrained_gps.shape[0]):
	elastic_gps[j,:] = np.array([elastic_undrained_gps[j], elastic_drained_gps[j]])

for j in range(0, elastic_undrained_surf.shape[0]):
	elastic_surf[j,:] = np.array([elastic_undrained_surf[j], elastic_drained_surf[j]])

"""

#convert displacements from meters to mm
#Sol_gps[size_gps:-1,:] = 1e2*Sol_gps[size_gps:-1,:]
#Sol_surf[size_gps:-1,:]  = 1e2*Sol_surf[size_gps:-1,:]
#Sol_gps_k[size_gps:-1,:] = 1e2*Sol_gps_k[size_gps:-1,:]
#Sol_surf_k[size_gps:-1, :]  = 1e2*Sol_surf_k[size_gps:-1,:]

#convert MPa to well head change (meters)

#Sol_gps[0:size_gps, :] = Sol_gps[0:size_gps,:]*1e3/(9.81)
#Sol_surf[0:size_gps, :] = Sol_surf[0:size_gps,:]*1e3/(9.81)
#Sol_gps_k[0:size_gps, :] = Sol_gps_k[0:size_gps,:]*335.130*12/39
#Sol_surf_k[0:size_gps, :] = Sol_surf_k[0:size_gps,:]*335.130*12/39


#nsteps = dtt.size

"""
x_p = x_all[0:size_p]
y_p = y_all[0:size_p]
x_u = x_all[size_p:size_p+size_ux]
y_u = y_all[size_p:size_p+size_ux]
x_ux = x_u[0:size_ux]
y_ux = y_u[0:size_ux]
x_ocean_p = x_all[ocean_dof_p]
y_ocean_p = y_all[ocean_dof_p]
x_ocean_u = x_all[ocean_dof_ux]
y_ocean_u = y_all[ocean_dof_ux]
x_surf_all = x_all[all_surf_dof_p]
y_surf_all = y_all[all_surf_dof_p]




X_slip_fine = np.linspace(0.0, 2.5e5, 250, endpoint = True)
Y_slip_fine = np.linspace(0.0, 2.5e5, 250, endpoint = True)
x_grid, y_grid = np.meshgrid(X_slip_fine, Y_slip_fine)
# slip_data_UX_fine = griddata((X_slip, Y_slip), U_X_slip, (x_bottom, y_bottom), method='linear')

mag_flux_grid = griddata((x_surf_all, y_surf_all), surfflux_mag, (x_grid, y_grid), method = 'linear')
"""


def shiftedColorMap(cmap, start = 0, midpoint = 0.5, stop = 1.0, name = 'shiftedcmap'):
	'''
	Function to offset the "center" of a colormap. Useful for
	data with a negative min and positive max and you want the
	middle of the colormap's dynamic range to be at zero

	Input
	-----
	  cmap : The matplotlib colormap to be altered
	  start : Offset from lowest point in the colormap's range.
		  Defaults to 0.0 (no lower ofset). Should be between
		  0.0 and `midpoint`.
	  midpoint : The new center of the colormap. Defaults to
		  0.5 (no shift). Should be between 0.0 and 1.0. In
		  general, this should be  1 - vmax/(vmax + abs(vmin))
		  For example if your data range from -15.0 to +5.0 and
		  you want the center of the colormap at 0.0, `midpoint`
		  should be set to  1 - 5/(5 + 15)) or 0.75
	  stop : Offset from highets point in the colormap's range.
		  Defaults to 1.0 (no upper ofset). Should be between
		  `midpoint` and 1.0.
	'''
	cdict = {
		'red':[],
		'green':[],
		'blue':[],
		'alpha':[]
	}

	# regular index to compute the colors
	reg_index = np.linspace(start, stop, 257)

	# shifted index to match the data
	shift_index = np.hstack([
		np.linspace(0.0, midpoint, 128, endpoint = False),
		np.linspace(midpoint, 1.0, 129, endpoint = True)
	])

	for ri, si in zip(reg_index, shift_index):
		r, g, b, a = cmap(ri)

		cdict['red'].append((si, r, r))
		cdict['green'].append((si, g, g))
		cdict['blue'].append((si, b, b))
		cdict['alpha'].append((si, a, a))

	newcmap = mpl.colors.LinearSegmentedColormap(name, cdict)
	plt.register_cmap(cmap = newcmap)

	return newcmap

def plot_gps_u(endtime, ngps):

	for i in range(0,ngps):
		idx_gps = i
		fig, sol_plotgps = plt.subplots(3, sharex=True)


		#sol_plotgps[0].plot(Ux_gps_model_all[idx_gps, 0:endtime], '--k', label='constant k', linewidth=3, fillstyle='none') #pressure

		# sol_plotgps[1].plot(dtt_days[0:endtime], Uy_gps_model[idx_gps, 0:endtime], '--r', linewidth=3, fillstyle='none') # u_x
		#
		# sol_plotgps[2].plot(dtt_days[0:endtime], Uz_gps_model[idx_gps, 0:endtime], '--g', linewidth=3, fillstyle='none') # u_y

		sol_plotgps[0].plot(dtt_days[0:endtime],1e2*Ux_gps_model_all[idx_gps, 0:endtime], 'k', linewidth=3, fillstyle='none') #pressure

		sol_plotgps[1].plot(dtt_days[0:endtime], 1e2*Uy_gps_model_all[idx_gps, 0:endtime], 'r', linewidth=3, fillstyle='none') # u_x

		sol_plotgps[2].plot(dtt_days[0:endtime], 1e2*Uz_gps_model_all[idx_gps, 0:endtime], 'g', linewidth=3, fillstyle='none') # u_y

		sol_plotgps[0].set_ylabel('X disp (cm)')
		sol_plotgps[1].set_ylabel('Y disp (cm)')
		sol_plotgps[2].set_ylabel('Z disp (cm)')



		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')


		fig.savefig('figures/3d_surf_GPS_timeseries_%s.png' % str(i), dpi = 400)


def plot_gps(size_gps, n):

	for i in range(0,n):
		idx_gps = i
		fig, sol_plotgps = plt.subplots(4, sharex=True)


		sol_plotgps[0].plot(Sol_gps[idx_gps, :], '--k', label='constant k', linewidth=3, fillstyle='none') #pressure

		sol_plotgps[1].plot(Sol_gps[(size_gps + idx_gps), :], '--r', linewidth=3, fillstyle='none') # u_x

		sol_plotgps[2].plot(Sol_gps[2*size_gps + idx_gps, :], '--g', linewidth=3, fillstyle='none') # u_y

		sol_plotgps[3].plot(Sol_gps[3*size_gps + idx_gps, :], '--b', linewidth=3, fillstyle='none') # u_z



		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_surf():

	points = 1e3*np.array([100 , 130])

	for i in range(0, points.shape[1]):
		idx_p = i
		fig = plt.figure()
		ax1 = fig.add_subplot(111)

		ax1.plot(Sol_surf[idx, :], 'k*', label='mapped k', linewidth=3, fillstyle='none') #pressure

		#sol_plot_surf[1].plot(Sol_surf[(size_surf + idx), :], '--r', linewidth=3, fillstyle='none') # u_x

		#sol_plot_surf[2].plot(Sol_surf[2*size_surf + idx, :], '--g', linewidth=3, fillstyle='none') # u_y

		#sol_plot_surf[3].plot(Sol_surf[3*size_surf + idx, :], '--b', linewidth=3, fillstyle='none') # u_z

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_gps_k():

	for i in range(0,2):
		idx_gps = i
		fig = plt.figure()
		ax1 = fig.add_subplot(111)


		ax1.plot(Sol_gps_k[idx_gps, :], '--k', label='mapped k', linewidth=3, fillstyle='none') #pressure

		#sol_plotgps_k[1].plot(Sol_gps_k[(size_gps + idx_gps), :], '--r', linewidth=3, fillstyle='none') # u_x

		#sol_plotgps_k[2].plot(Sol_gps_k[2*size_gps + idx_gps, :], '--g', linewidth=3, fillstyle='none') # u_y

		#sol_plotgps_k[3].plot(Sol_gps_k[3*size_gps + idx_gps, :], '--b', linewidth=3, fillstyle='none') # u_z

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		#legend1 = sol_plotgps[0].legend()
		#for label in legend1.get_texts():
		#    label.set_fontsize('medium')

		#for label in legend1.get_lines():
		#    label.set_linewidth(1.5)  # the legend line width

def plot_surf_k(points, splice):


	for i in range(0, points.shape[0]):

		fig = plt.figure()
		ax1 = fig.add_subplot(111)
		#ax2 = fig.add_subplot(212)
		#ax3 = fig.add_subplot(413)
		#ax4 = fig.add_subplot(414)

		

		ax1.plot(dtt_days[0:endtime]/splice, Sol_wells[i, 0:endtime]*1e3/(9.81), '-k', linewidth = 3, fillstyle = 'none')  # p

		#ax2.plot(dtt_days[0:endtime]/splice, Sol_surf[p_surf_ind + idx_u, 0:endtime]*100, '-r', linewidth=3, fillstyle='none') #ux

		#ax3.plot(dtt_days[0:endtime]/splice, Sol_surf[p_surf_ind+u_surf_ind + idx_u, 0:endtime]*100, '-g', linewidth = 3, fillstyle = 'none')  # uy

		#ax2.plot(dtt_days[0:endtime]/splice, Sol_surf[p_surf_ind + 2*u_surf_ind + idx_u, 0:endtime]*100, '-k', linewidth = 3, fillstyle = 'none')  # uz

		ax1.set_ylabel('head change (m)')
		#ax2.set_ylabel('X disp (cm)')
		#ax3.set_ylabel('Y disp (cm)')
		#ax2.set_ylabel('Uplift (cm)')
		#plt.ylabel('Depth (km)')
		#ax1.set_xlim((0, 3))
		#ax2.set_xlim((0, 3))
		#ax3.set_xlim((0, 3))
		plt.xlim((0, 60))
		fig.set_size_inches(8, 2.5)
		plt.xlabel('Days after EQ')
		fig.savefig('figures/3d_surf_timeseries_%s_60day_well.png'% str(points[i,0])[:3], dpi = 400)

		#legend1 = sol_plotgps[0].legend()
		#for label in legend1.get_texts():
		#    label.set_fontsize('medium')

		#for label in legend1.get_lines():
		#    label.set_linewidth(1.5)  # the legend line width

def plot_gps_both(Sol_gps, Sol_gps_k, size_gps, n):

	for i in range(0,n):
		idx_gps = i
		fig, sol_plotgps_both = plt.subplots(4, sharex=True)

		sol_plotgps_both[0].plot(Sol_gps[idx_gps, :], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure
		sol_plotgps_both[0].plot(Sol_gps_k[idx_gps, :], '--k', label='mapped k', linewidth=3, fillstyle='none') #pressure

		sol_plotgps_both[1].plot(Sol_gps[(size_gps + idx_gps), :], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		sol_plotgps_both[1].plot(Sol_gps_k[(size_gps + idx_gps), :], '--r', linewidth=3, fillstyle='none') # u_x

		sol_plotgps_both[2].plot(Sol_gps[2*size_gps + idx_gps, :], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		sol_plotgps_both[2].plot(Sol_gps_k[2*size_gps + idx_gps, :], '--g', linewidth=3, fillstyle='none') # u_y

		sol_plotgps_both[3].plot(Sol_gps[3*size_gps + idx_gps, :], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		sol_plotgps_both[3].plot(Sol_gps_k[3*size_gps + idx_gps, :], '--b', linewidth=3, fillstyle='none') # u_z

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_surf_both(points):

	for i in range(0, points.shape[1]):
		idx_p = np.abs(np.sqrt(np.power((x_surf_p[:] - points[i,0]),2) + np.power((y_surf_p[:] - points[i,1]),2))).argmin()
		idx_u = np.abs(np.sqrt(np.power((x_surf_ux[:] - points[i,0]),2) + np.power((y_surf_ux[:] - points[i,1]),2))).argmin()
		fig, sol_plot_surf_both = plt.subplots(4, sharex=True)

		sol_plot_surf_both[0].plot(dtt[0:endtime], Sol_surf[idx_p, 0:endtime], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure
		#sol_plot_surf_both[0].plot(Sol_surf_k[idx, :], '--k', label='mapped k', linewidth=3, fillstyle='none') #pressure

		sol_plot_surf_both[1].plot(dtt[0:endtime], Sol_surf[(p_surf_ind + idx_u), 0:endtime], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		#sol_plot_surf_both[1].plot(Sol_surf_k[(size_surf + idx), :], '--r', linewidth=3, fillstyle='none') # u_x

		sol_plot_surf_both[2].plot(dtt[0:endtime], Sol_surf[p_surf_ind+u_surf_ind + idx_u, 0:endtime], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		#sol_plot_surf_both[2].plot(Sol_surf_k[2*size_surf + idx, :], '--g', linewidth=3, fillstyle='none') # u_y

		sol_plot_surf_both[3].plot(dtt[0:endtime], Sol_surf[p_surf_ind+2*u_surf_ind + idx_u, 0:endtime], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		#sol_plot_surf_both[3].plot(Sol_surf_k[3*size_surf + idx, :], '--b', linewidth=3, fillstyle='none') # u_z

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plot_surf_both[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

		fig.savefig('figures/3d_surf_timeseries_%s.png'% str(points[i,1])[:2], dpi = 400)

def plot_gps_elastic_compare(nstations):

	for i in range(0,nstations):
		idx_gps = i
		fig, sol_plotgps_compare = plt.subplots(4, sharex=True)

		sol_plotgps_compare[0].plot(Sol_gps[idx_gps, :], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure

		sol_plotgps_compare[1].plot(Sol_gps[(size_gps + idx_gps), :], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		sol_plotgps_compare[1].plot(x_elastic, elastic_gps[idx_gps, :],'ro', fillstyle='full') # u_x

		sol_plotgps_compare[2].plot(Sol_gps[2*size_gps + idx_gps, :], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		sol_plotgps_compare[2].plot(x_elastic, elastic_gps[size_gps + idx_gps, :],'go', fillstyle='full') # u_x

		sol_plotgps_compare[3].plot(Sol_gps[3*size_gps + idx_gps, :], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		sol_plotgps_compare[3].plot(x_elastic, elastic_gps[2*size_gps+idx_gps, :],'bo', fillstyle='full') # u_x

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps_compare[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_surf_elastic_compare(points):


	for i in range(0, points.shape[1]):

		idx = (np.abs(np.sqrt(pow((x_surf - plotpoint[0]),2) + pow((y_surf - plotpoint[1]),2)))).argmin()

		fig, sol_plot_surf_compare = plt.subplots(4, sharex=True)

		sol_plot_surf_compare[0].plot(Sol_surf[idx, :], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure


		sol_plot_surf_compare[1].plot(Sol_surf[(size_surf + idx_gps), :], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		sol_plot_surf_compare[1].plot(x_elastic, elastic_surf[idx_gps, :],'ro', fillstyle='full') # u_x

		sol_plot_surf_compare[2].plot(Sol_surf[2*size_surf + idx_gps, :], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		sol_plot_surf_compare[2].plot(x_elastic, elastic_surf[size_gps + idx_gps, :],'go', fillstyle='full') # u_x

		sol_plot_surf_compare[3].plot(Sol_surf[3*size_surf + idx_gps, :], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		sol_plot_surf_compare[3].plot(x_elastic, elastic_surf[2*size_gps+idx_gps, :],'bo', fillstyle='full') # u_x

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_gps_all(nstations):

	for i in range(0,nstations):
		idx_gps = i
		fig, sol_plotgps_all = plt.subplots(4, sharex=True)

		sol_plotgps_all[0].plot(Sol_gps[idx_gps, :], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure
		sol_plotgps_all[0].plot(Sol_gps_k[idx_gps, :], '--k', label='mapped k', linewidth=3, fillstyle='none') #pressure

		sol_plotgps_all[1].plot(Sol_gps[(size_gps + idx_gps), :], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		sol_plotgps_all[1].plot(Sol_gps_k[(size_gps + idx_gps), :], '--r', linewidth=3, fillstyle='none') # u_x
		sol_plotgps_all[1].plot(x_elastic, elastic_gps_k[idx_gps, :],'ro', fillstyle='full') # u_x

		sol_plotgps_all[2].plot(Sol_gps[2*size_gps + idx_gps, :], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		sol_plotgps_all[2].plot(Sol_gps_k[2*size_gps + idx_gps, :], '--g', linewidth=3, fillstyle='none') # u_y
		sol_plotgps_all[2].plot(x_elastic, elastic_gps_k[size_gps + idx_gps, :],'go', fillstyle='full') # u_x

		sol_plotgps_all[3].plot(Sol_gps[3*size_gps + idx_gps, :], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		sol_plotgps_all[3].plot(Sol_gps_k[3*size_gps + idx_gps, :], '--b', linewidth=3, fillstyle='none') # u_z
		sol_plotgps_all[3].plot(x_elastic, elastic_gps_k[2*size_gps+idx_gps, :],'bo', fillstyle='full') # u_x

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_surf_all(Sol_surf, Sol_surf_k, elastic_surf, size_surf, x_surf, y_surf):

	points = 1e3*np.array([40 , 100])

	nsteps = Sol_surf.shape[1]
	x_elastic = np.array([0,nsteps-1])


	for i in range(0, points.shape[1]):

		idx = (np.abs(np.sqrt(pow((x_surf - plotpoint[0]),2) + pow((y_surf - plotpoint[1]),2)))).argmin()

		fig, sol_plot_surf_all = plt.subplots(4, sharex=True)

		sol_plot_surf_all[0].plot(Sol_surf[idx, :], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure
		sol_plot_surf_all[0].plot(Sol_surf_k[idx, :], '--k', label='mapped k', linewidth=3, fillstyle='none') #pressure

		sol_plot_surf_all[1].plot(Sol_surf[(size_surf + idx_gps), :], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		sol_plot_surf_all[1].plot(Sol_surf_k[(size_surf + idx_gps), :], '--r', linewidth=3, fillstyle='none') # u_x
		sol_plot_surf_all[1].plot(x_elastic, elastic_surf[idx_gps, :],'ro', fillstyle='full') # u_x

		sol_plot_surf_all[2].plot(Sol_surf[2*size_surf + idx_gps, :], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		sol_plot_surf_all[2].plot(Sol_surf_k[2*size_surf + idx_gps, :], '--g', linewidth=3, fillstyle='none') # u_y
		sol_plot_surf_all[2].plot(x_elastic, elastic_surf[size_gps + idx_gps, :],'go', fillstyle='full') # u_x

		sol_plot_surf_all[3].plot(Sol_surf[3*size_surf + idx_gps, :], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		sol_plot_surf_all[3].plot(Sol_surf_k[3*size_surf + idx_gps, :], '--b', linewidth=3, fillstyle='none') # u_z
		sol_plot_surf_all[3].plot(x_elastic, elastic_surf[2*size_gps+idx_gps, :],'bo', fillstyle='full') # u_x

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_pressure_contour():
	fig = plt.figure()
	ax1 = fig.add_subplot(111)


	i = 1

	#stream = streamfun[:, i]
	Z = mag_flux_grid[:,:, i]



	triang = tri.Triangulation(x_surf_all, y_surf_all)
	triang_u = tri.Triangulation(x_ocean_u, y_ocean_u)
	contour_steps = 10
	num_lines = 10
	orig_cmap = cm.coolwarm
	midpoint = 1 - Z.max() / (Z.max() + abs(Z.min()))
	shifted_cmap = shiftedColorMap(orig_cmap, midpoint = midpoint)

	zero = np.array([0])



	# levels_10 = np.array([Z_10.min(), Z_10.min()/2, Z_10.min()/4, Z_10.min()/8, Z_10.min()/16,
	#                      0, Z_10.max()/16, Z_10.max()/8, Z_10.max()/4, Z_10.max()/2, Z_10.max()])
	levels = np.linspace(seaflux_mag[:,i].min(), seaflux_mag[:,i].max(), num = contour_steps, endpoint = True)
	lines = np.linspace(seaflux_mag[:,i].min(), seaflux_mag[:,i].max(),  num = num_lines, endpoint = True)
	#lines_stream = np.linspace(stream.min(), stream.max(), num = num_lines, endpoint = True)


	skip = (slice(None, None, 10))

	contourf = ax1.contourf(x_grid,y_grid, Z, levels, cmap = orig_cmap)
	contour = ax1.contour(x_grid,y_grid, Z, lines, linewidth = 1, colors = 'black', linestyles = 'solid')


	#quiver = ax1.quiver(x_ocean_u[skip], y_ocean_u[skip], seaflux_x[skip][:, i], seaflux_y[skip][:, i], units = 'inches', scale_units = 'width', width = 1.7e-2, scale = .3/1e5, color = 'darkcyan')
	# quiver = ax1.quiver(x_ux[skip], y_ux[skip], vel_x[skip][:, i], vel_y[skip][:, i], scale_units = 'width', scale = 1.0 / 2e5)
	# streamline10 = ax1.streamplot(x_stream, y_stream, Z_x_stream[:, :, i], Z_y_stream[:, :, i], density = [1.0, 3.0], linewidth = lw)
	#streamfun_contour = ax1.tricontour(triang, stream, lines_stream, linewidth =1, colors = 'blue',linestyles = 'solid')



	fig.colorbar(contourf)
	#ax1.set_ylim([-2e3, 1e3])
	#ax1.set_xlim([55e3, 80e3])

	#fig.set_size_inches(8, 2)


	fig.savefig('figures/3d_contour.png', dpi = 400)

def plot_sea_flux():
	fig = plt.figure()
	ax1 = fig.add_subplot(111)
	# ax2 = fig.add_subplot(312)
	# ax3 = fig.add_subplot(313)
	# ax1.set_title('Flux through seafloor')
	
	end = 350
	
	# Co-seismic displacement
	ax1.plot(dtt_all, zero_line, '--k', linewidth = 1)
	ax1.plot(dtt_days[48:51+end]/30, flux_topo[48:51+end], '-b',label='Topographic flow',  linewidth = 2)
	#ax1.plot(dtt_all[51:51+250], flux_topo[51:51+250], '-b',  linewidth = 3)

	#ax1.plot(dtt_all, flux_subEQ, '-r',  linewidth = 3)
	#ax1.plot(dtt_all[48:50], flux_EQ[48:50], '-r', linewidth = 3)
	ax1.plot(dtt_days[48:51+end]/30, flux_subEQ[48:51+end], '-r', label='Tectonic flow', linewidth = 2)

	# ax3.plot(dtt[50:-1], sea_flux[50:-1], '-b*', ms=2, mew=2, fillstyle='none', color = 'red')
	# ax1.set_ylabel('Darcy flux (meters/year)')
	# ax2.set_ylabel('Darcy flux (meters/sec)')
	# ax3.set_ylabel('Darcy flux (meters/year)')
	# ax1.set_xlabel('Years')
	
	#ax1.set_yscale('log')
	ax1.set_ylim([-5e5, 1.5e6])
	ax1.set_xlim([-.5 , 12])
	ax1.set_xlabel('Months after earthquake')
	#ax1.legend(fontsize='small', loc=1)
	
	#fig.set_size_inches(8, 2.5)
	
	#becomes negative at 8 days, positive again at 215
	#above topo for 5 days
	
	fig.savefig('figures/3d_sea_flux.png', dpi = 400)

def plot_gps_surface():
	
	fig = plt.figure()
	ax1 = fig.add_subplot(111)
	
	#quiver = ax1.quiver(x_ocean_u[skip], y_ocean_u[skip], seaflux_x[skip][:, i], seaflux_y[skip][:, i],
	#                    units='inches',scale_units='width', width=1.7e-2, scale=.3 / 1e5, color='darkcyan')
	
	# quiver = ax1.quiver(X_gps/1e3, Y_gps/1e3, Ux_gps, Uy_gps, color = 'blue', label='GPS data',units='inches',
	#                     scale_units='width', width=2e-2)#, scale=1e1,)
	# quiver2 = ax1.quiver(X_gps/1e3, Y_gps/1e3, Ux_gps_model, Uy_gps_model, color = 'red', label='model', units='inches',
	#                      scale_units='width', width=2e-2)#, scale=1,)

	# scale = 4.0
	quiver = ax1.quiver(X_gps/1e3, Y_gps/1e3, Z_gps, Uz_gps, color = 'blue', label='GPS data',units='inches',
	                    scale_units='width', width=2e-2, scale=10)
	quiver2 = ax1.quiver(X_gps/1e3, Y_gps/1e3, Z_gps, Uz_gps_model, color = 'red', label='model',
	                     units='inches',scale_units='width', width=2e-2, scale=10)
	#
	# ax1.plot(Uz_gps, 'b*', color = 'blue', label='GPS data')
	# ax1.plot(Uz_gps_model, 'ro', color = 'red', label='model')
	#ax1.plot(Uz_gps_model/Uz_gps, 'ro', color = 'red', label = 'model')

	ax1.legend(fontsize='small', loc=1)
	# ax1.set_ylim([0, 250])
	# ax1.set_xlim([0 , 250])

	#fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_gps_surface_up.png', dpi = 400)
	#fig.savefig('figures/3d_gps_up_scatter.png', dpi = 400)



# ax3.set_yscale('log')

#Z = mag_flux_grid[:,:, 1]
#print Z.max(), Z.min()
#points = 1e3*np.array([[40, 140], [50, 140], [60.5, 140], [70, 140], [80, 140], [90, 140], [100, 140], [110, 140], [120, 140], [130, 140]])
#points = 1e3*np.array([[78, 140], [96, 129], [61, 160], [114,124]])
points = 1e3*np.array([[83.5, 141]])


splice = 1
endtime = 200

#plot_surf_k()
#plot_gps_elastic_compare(20)
#plot_surf_k(wellpoints, splice)
#plot_pressure_contour()
#plot_sea_flux()
#plot_gps_surface()
plot_gps_u(endtime, 2)
#plt.show()

#fig3.set_size_inches(7, 4)
#plt.xlim((0, 100))
#plt.ylim((5e-8, 10))
#fig2.savefig('figures/SGD_cumulative.png', dpi = 400)

