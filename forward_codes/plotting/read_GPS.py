import sys
sys.path.append('/home/fenics/shared/local_lib')

from dolfin import *
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator, MultipleLocator, FormatStrFormatter
from matplotlib import cm
from mpl_toolkits.mplot3d import Axes3D
import numpy as np
import scipy as sp
from scipy.interpolate import griddata
from tempfile import TemporaryFile
import openpyxl as pyxl
import matplotlib.tri as tri

font = {'weight': 'normal',
        'size': 10}

mpl.rc('font', **font)

mesh_size = 'medcoarse'  # fine, med, coarse
path = "/home/fenics/shared/"


# mesh = Mesh(path+'/meshes/CR3D_'+mesh_size+'.xml')
# boundaries = MeshFunction("size_t", mesh, path+'/meshes/CR3D_'+mesh_size+'_facet_region.xml')

def GPSlatlon2XY(data_sheet, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = np.array([[data_sheet.cell(row=i, column=1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat = np.array([[data_sheet.cell(row=i, column=2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	lon_u = np.array([[data_sheet.cell(row=i, column=5).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat_u = np.array([[data_sheet.cell(row=i, column=6).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	Uz = np.array([[data_sheet.cell(row=i, column=4).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho_u = np.sqrt(np.power(lon_u, 2) + np.power(lat_u, 2))
	theta_new_u = np.arctan2(lat_u, lon_u) - theta

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)
	Ux, Uy = rho_u * np.cos(theta_new_u), rho_u * np.sin(theta_new_u)

	return 1e3 * X, 1e3 * Y, 1e-3 * Ux, 1e-3 * Uy, 1e-3 * Uz


def GPSlatlon2XY_time(data_sheet, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	gps_time = np.array([[data_sheet.cell(row=i, column=1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	lon_u = np.array([[data_sheet.cell(row=i, column=3).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat_u = np.array([[data_sheet.cell(row=i, column=2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	Uz = np.array([[data_sheet.cell(row=i, column=4).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	rho_u = np.sqrt(np.power(lon_u, 2) + np.power(lat_u, 2))
	theta_new_u = np.arctan2(lat_u, lon_u) - theta

	Ux, Uy = rho_u * np.cos(theta_new_u), rho_u * np.sin(theta_new_u)

	return gps_time, Ux, Uy, Uz


def latlon2XY_point(point, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = point[0]
	lat = point[1]

	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)

	return 1e3 * X, 1e3 * Y


def GPSlatlon2XY_other(data_sheet, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = np.array([[data_sheet.cell(row=i, column=1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	Z = np.array([[data_sheet.cell(row=i, column=3).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat = np.array([[data_sheet.cell(row=i, column=2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)

	return 1e3 * X, 1e3 * Y, Z


def XY2latlon(X, Y, origin, theta):
	X = 1e-3 * X
	Y = 1e-3 * Y

	rho = np.sqrt(np.power(X, 2) + np.power(Y, 2))
	theta_new = np.arctan2(Y, X) + theta

	lat = rho * np.sin(theta_new) / 111 + origin[1]
	lon = rho * np.cos(theta_new) / (111 * np.cos(lat * np.pi / 180)) + origin[0]

	return lon, lat


##########################################################################################


with open(path + 'data/GD/station_list.txt', 'r') as f:
	stations = f.readlines()

names = [x.split()[0] for x in stations]
gps_lon = np.asarray([x.split()[1] for x in stations], dtype=np.float32)
gps_lat = np.asarray([x.split()[2] for x in stations], dtype=np.float32)

plot_data = False
# print names
# print gps_lat
# print gps_lon

# gps_data = np.array(len(names))
data = {}

for name in names:

	if plot_data:
		fig = plt.figure()

	with open(path + 'data/GD/' + name + '.lat', 'r') as latt:
		latread = latt.readlines()

	lat = np.asarray([x.split()[1] for x in latread], dtype=np.float32)
	time = np.asarray([x.split()[0] for x in latread], dtype=np.float32)
	t2_ind = np.where((time > 2012.5))[0][0] - 1
	slope = (np.average(lat[t2_ind-5:t2_ind]) - np.average(lat[0:5])) / (time[t2_ind] - time[0])
	plate_motion = (time - time[0]) * slope
	latnorm = lat - plate_motion
	latnorm -= np.average(latnorm[0:t2_ind])

	data[name + '_time'] = time
	data[name + '_lat'] = latnorm

	if plot_data:
		ax1 = fig.add_subplot(311)
		ax1.plot(time, plate_motion, '.r')
		ax1.plot(time, latnorm, '.b')
		ax1.plot(time, lat, '.k')
		ax1.ticklabel_format(useOffset=False)


	with open(path + 'data/GD/' + name + '.lon', 'r') as lont:
		lonread = lont.readlines()

	lon = np.asarray([x.split()[1] for x in lonread], dtype=np.float32)

	t2_ind = np.where((time > 2012.5))[0][0] - 1

	slope = (np.average(lon[t2_ind-5:t2_ind]) - np.average(lon[0:5])) / (time[t2_ind] - time[0])
	plate_motion = (time - time[0]) * slope
	lonnorm = lon - plate_motion
	lonnorm -= np.average(lonnorm[0:t2_ind])

	data[name + '_lon'] = lonnorm

	if plot_data:
		ax1 = fig.add_subplot(312)
		ax1.plot(time, plate_motion, '.r')
		ax1.plot(time, lonnorm, '.b')
		ax1.plot(time, lon, '.k')
		ax1.ticklabel_format(useOffset=False)


	with open(path + 'data/GD/' + name + '.rad', 'r') as radt:
		radread = radt.readlines()

	rad = np.asarray([x.split()[1] for x in radread], dtype=np.float32)

	t2_ind = np.where((time > 2012.5))[0][0] - 1

	slope = (np.average(rad[t2_ind - 5:t2_ind]) - np.average(rad[0:5])) / (time[t2_ind] - time[0])
	plate_motion = (time - time[0]) * slope
	radnorm = rad - plate_motion
	radnorm -= np.average(radnorm[0:t2_ind])

	data[name + '_rad'] = radnorm

	if plot_data:
		ax1 = fig.add_subplot(313)
		ax1.plot(time, plate_motion, '.r')
		ax1.plot(time, radnorm, '.b')
		ax1.plot(time, rad, '.k')
		ax1.ticklabel_format(useOffset=False)
		fig.savefig('figures/gpsdata_%s.png' %name, dpi=400)


for name in names:
	print name, data[name + '_time'].shape

# #latlongps = np.array(stations[station]).astype(np.float)
# latlongps = np.array(stations).astype(np.float)
#
# quit()
#
# #Load GPS data
# origin = np.array([-85.1688, 8.6925]) # origin of top surface of mesh - used as point of rotation
# theta = 0.816 # angle of rotation between latitude line and fault trace (42 deg)
#
# data_file = path+'data/Data.xlsx'
# wb = pyxl.load_workbook(data_file)
# gps_data = wb['GPS_data']
#
# X_gps, Y_gps, Ux_gps, Uy_gps, Uz_gps = GPSlatlon2XY(gps_data, origin, theta)
# Z_gps = np.zeros(len(X_gps))
#
#
# data_file2 = path+'data/GPS_post.xlsx'
# wb2 = pyxl.load_workbook(data_file2)
# gps_data_post = wb2['CABA_val']
#
# gps_time, Ux_gps_post, Uy_gps_post, Uz_gps_post = GPSlatlon2XY_time(gps_data_post, theta)
#
#
# Uz = np.array([[gps_data_post.cell(row = i, column = 3).value] for i in range(2, gps_data_post.max_row + 1)]).reshape(
# 	gps_data_post.max_row - 1, )
#
# gps_time_days = (gps_time - 2012.68)*365
# #gps_time_days = gps_time



################################################################################################
def shiftedColorMap(cmap, start=0, midpoint=0.5, stop=1.0, name='shiftedcmap'):
	'''
	Function to offset the "center" of a colormap. Useful for
	data with a negative min and positive max and you want the
	middle of the colormap's dynamic range to be at zero

	Input
	-----
	  cmap : The matplotlib colormap to be altered
	  start : Offset from lowest point in the colormap's range.
		  Defaults to 0.0 (no lower ofset). Should be between
		  0.0 and `midpoint`.
	  midpoint : The new center of the colormap. Defaults to
		  0.5 (no shift). Should be between 0.0 and 1.0. In
		  general, this should be  1 - vmax/(vmax + abs(vmin))
		  For example if your data range from -15.0 to +5.0 and
		  you want the center of the colormap at 0.0, `midpoint`
		  should be set to  1 - 5/(5 + 15)) or 0.75
	  stop : Offset from highets point in the colormap's range.
		  Defaults to 1.0 (no upper ofset). Should be between
		  `midpoint` and 1.0.
	'''
	cdict = {
		'red': [],
		'green': [],
		'blue': [],
		'alpha': []
	}

	# regular index to compute the colors
	reg_index = np.linspace(start, stop, 257)

	# shifted index to match the data
	shift_index = np.hstack([
		np.linspace(0.0, midpoint, 128, endpoint=False),
		np.linspace(midpoint, 1.0, 129, endpoint=True)
	])

	for ri, si in zip(reg_index, shift_index):
		r, g, b, a = cmap(ri)

		cdict['red'].append((si, r, r))
		cdict['green'].append((si, g, g))
		cdict['blue'].append((si, b, b))
		cdict['alpha'].append((si, a, a))

	newcmap = mpl.colors.LinearSegmentedColormap(name, cdict)
	plt.register_cmap(cmap=newcmap)

	return newcmap


def plot_gps_u(endtime, ngps):
	for i in range(0, ngps):
		idx_gps = indice_gps
		fig, sol_plotgps = plt.subplots(3, sharex=True)

		# sol_plotgps[0].plot(Ux_gps_model_all[idx_gps, 0:endtime], '--k', label='constant k', linewidth=3, fillstyle='none') #pressure

		# sol_plotgps[1].plot(dtt_days[0:endtime], Uy_gps_model[idx_gps, 0:endtime], '--r', linewidth=3, fillstyle='none') # u_x
		#
		# sol_plotgps[2].plot(dtt_days[0:endtime], Uz_gps_model[idx_gps, 0:endtime], '--g', linewidth=3, fillstyle='none') # u_y

		sol_plotgps[0].plot(dtt_days[0:endtime], 1e2 * Ux_gps_model_all[idx_gps, 0:endtime], 'k', linewidth=3,
		                    fillstyle='none')

		sol_plotgps[1].plot(dtt_days[0:endtime], 1e2 * Uy_gps_model_all[idx_gps, 0:endtime], 'r', linewidth=3,
		                    fillstyle='none')

		sol_plotgps[2].plot(dtt_days[0:endtime], 1e2 * Uz_gps_model_all[idx_gps, 0:endtime], 'g', linewidth=3,
		                    fillstyle='none')

		sol_plotgps[0].plot(gps_time_days, Ux_gps_post, '.k', linewidth=3, fillstyle='none')

		sol_plotgps[1].plot(gps_time_days, Uy_gps_post, '.r', linewidth=3, fillstyle='none')

		sol_plotgps[2].plot(gps_time_days, Uz_gps_post, '.g', linewidth=3, fillstyle='none')

		# sol_plotgps[0].plot(Ux_gps_post, '*k', linewidth=3, fillstyle='none')
		#
		# sol_plotgps[1].plot(Uy_gps_post, '*r', linewidth=3, fillstyle='none')
		# sol_plotgps[2].plot(Uz_gps_post, '*g', linewidth=3, fillstyle='none')


		sol_plotgps[0].set_ylabel('X disp (cm)')
		sol_plotgps[1].set_ylabel('Y disp (cm)')
		sol_plotgps[2].set_ylabel('Z disp (cm)')

		sol_plotgps[0].set_xlim([0, 100])
		sol_plotgps[1].set_xlim([0, 100])
		sol_plotgps[2].set_xlim([0, 100])

		# plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')

		fig.savefig('figures/3d_surf_GPS_model timeseries_%s.png' % station, dpi=400)


def plot_gps(size_gps, n):
	for i in range(0, n):
		idx_gps = i
		fig, sol_plotgps = plt.subplots(4, sharex=True)

		sol_plotgps[0].plot(Sol_gps[idx_gps, :], '--k', label='constant k', linewidth=3, fillstyle='none')  # pressure

		sol_plotgps[1].plot(Sol_gps[(size_gps + idx_gps), :], '--r', linewidth=3, fillstyle='none')  # u_x

		sol_plotgps[2].plot(Sol_gps[2 * size_gps + idx_gps, :], '--g', linewidth=3, fillstyle='none')  # u_y

		sol_plotgps[3].plot(Sol_gps[3 * size_gps + idx_gps, :], '--b', linewidth=3, fillstyle='none')  # u_z

		# plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width


def plot_gps_surface():
	fig = plt.figure()
	ax1 = fig.add_subplot(111)

	# quiver = ax1.quiver(x_ocean_u[skip], y_ocean_u[skip], seaflux_x[skip][:, i], seaflux_y[skip][:, i],
	#                    units='inches',scale_units='width', width=1.7e-2, scale=.3 / 1e5, color='darkcyan')
	#
	quiver = ax1.quiver(X_gps / 1e3, Y_gps / 1e3, Ux_gps, Uy_gps, color='blue', label='GPS data', units='inches',
	                    scale_units='width', width=1e-2, scale=5e0)
	quiver2 = ax1.quiver(X_gps / 1e3, Y_gps / 1e3, Ux_gps_model, Uy_gps_model, color='red', label='model',
	                     units='inches',
	                     scale_units='width', width=1e-2, scale=5e0)

	# scale = 4.0
	# quiver = ax1.quiver(X_gps/1e3, Y_gps/1e3, Z_gps, Uz_gps, color = 'blue', label='GPS data',units='inches',
	#                     scale_units='width', width=2e-2, scale=10)
	# quiver2 = ax1.quiver(X_gps/1e3, Y_gps/1e3, Z_gps, Uz_gps_model, color = 'red', label='model',
	#                      units='inches',scale_units='width', width=2e-2, scale=10)
	# #
	# ax1.plot(Uz_gps, 'b*', color = 'blue', label='GPS data')
	# ax1.plot(Uz_gps_model, 'ro', color = 'red', label='model')
	# ax1.plot(Uz_gps_model/Uz_gps, 'ro', color = 'red', label = 'model')

	ax1.legend(fontsize='small', loc=1)
	ax1.set_ylim([50, 250])
	ax1.set_xlim([30, 200])

	# fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_gps_surface_up.png', dpi=400)


def plot_gps_surface_xyz():
	fig = plt.figure()
	fig, plot = plt.subplots(3, sharex=True)

	#
	plot[0].plot(Ux_gps, 'b*', color='blue', label='GPS data')
	plot[0].plot(Ux_gps_model, 'ro', color='red', label='model')
	plot[0].plot(Z_gps, '--k', label='model')

	plot[1].plot(Uy_gps, 'b*', color='blue', label='GPS data')
	plot[1].plot(Uy_gps_model, 'ro', color='red', label='model')
	plot[1].plot(Z_gps, '--k', label='model')

	plot[2].plot(Uz_gps, 'b*', color='blue', label='GPS data')
	plot[2].plot(Uz_gps_model, 'ro', color='red', label='model')
	plot[2].plot(Z_gps, '--k', label='model')
	# ax1.plot(Uz_gps_model/Uz_gps, 'ro', color = 'red', label = 'model')

	plot[0].legend(fontsize='small', loc=4)
	# ax1.set_ylim([0, 250])
	# ax1.set_xlim([0 , 250])

	# fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_gps_surface_compare.png', dpi=400)

	fig = plt.figure()
	fig, plot = plt.subplots(3, sharex=True)

	plot[0].plot((Ux_gps_model - Ux_gps) / Ux_gps, '.b', color='blue', label='GPS data')
	plot[0].plot(Z_gps, '--k', label='model')

	plot[1].plot((Uy_gps_model - Uy_gps) / Uy_gps, '.b', color='blue', label='GPS data')
	plot[1].plot(Z_gps, '--k', label='model')

	plot[2].plot((Uz_gps_model - Uz_gps) / Uz_gps, '.b', color='blue', label='GPS data')
	plot[2].plot(Z_gps, '--k', label='model')
	# ax1.plot(Uz_gps_model/Uz_gps, 'ro', color = 'red', label = 'model')

	# plot[0].legend(fontsize = 'small', loc = 1)
	# ax1.set_ylim([0, 250])
	# ax1.set_xlim([0 , 250])

	# fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_gps_surface_compare_ratio.png', dpi=400)


splice = 1
endtime = 300


# plot_gps_surface()
# plot_gps_surface_xyz()
# plot_gps_u(endtime, 1)
