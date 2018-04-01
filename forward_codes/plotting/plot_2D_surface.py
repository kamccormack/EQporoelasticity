import sys

sys.path.append('/home/fenics/shared/local_lib')
sys.path.append('/home/fenics/shared/forward_codes/')
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from matplotlib import cm
from mpl_toolkits.mplot3d import Axes3D
import numpy as np
import scipy as sp
from scipy.interpolate import griddata
from tempfile import TemporaryFile
import openpyxl as pyxl
import matplotlib.tri as tri


def GPSlatlon2XY(data_sheet, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = np.array([[data_sheet.cell(row=i, column=1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat = np.array([[data_sheet.cell(row=i, column=2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	lon_u = np.array([[data_sheet.cell(row=i, column=5).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat_u = np.array([[data_sheet.cell(row=i, column=6).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	Uz = np.array([[data_sheet.cell(row=i, column=4).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho_u = np.sqrt(np.power(lon_u, 2) + np.power(lat_u, 2))
	theta_new_u = np.arctan2(lat_u, lon_u) - theta

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)
	Ux, Uy = rho_u * np.cos(theta_new_u), rho_u * np.sin(theta_new_u)

	return 1e3 * X, 1e3 * Y, 1e-3 * Ux, 1e-3 * Uy, 1e-3 * Uz

def GPSlatlon2XY_time(lat_u, lon_u, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	rho_u = np.sqrt(np.power(lon_u, 2) + np.power(lat_u, 2))
	theta_new_u = np.arctan2(lat_u, lon_u) - theta

	Ux, Uy = rho_u * np.cos(theta_new_u), rho_u * np.sin(theta_new_u)

	return Ux, Uy

def latlon2XY_points(lat, lon, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	#lon = point[0]
	#lat = point[1]

	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)

	return 1e3 * X, 1e3 * Y



event = 'EQ'  # syn = synthetic event, data = real EQ, sub = subduction, sub_EQ
loop = 'yes'
variable = 'cross'  # sigma or kappa
origin = np.array([-85.1688, 8.6925]) # origin of top surface of mesh - used as point of rotation
theta = 0.816 # angle of rotation between latitude line and fault trace (42 deg)

loop_through = np.arange(110., 160., 5.)
##########################################################################
############################ LOAD GPS DATA ##############################
##########################################################################

path = "/home/fenics/shared/"
with open(path + 'data/GD/station_list.txt', 'r') as f:
	stations = f.readlines()

names = [x.split()[0] for x in stations]
gps_lon = np.asarray([x.split()[1] for x in stations], dtype=np.float32)
gps_lat = np.asarray([x.split()[2] for x in stations], dtype=np.float32)

X_gps_post, Y_gps_post = latlon2XY_points(gps_lat, gps_lon, origin, theta)


plot_figs = False
data = {}
for name in names:

	if plot_figs:
		fig = plt.figure()

	with open(path + 'data/GD/' + name + '.lat', 'r') as latt:
		latread = latt.readlines()

	lat = np.asarray([x.split()[1] for x in latread], dtype=np.float32)
	time = np.asarray([x.split()[0] for x in latread], dtype=np.float32)
	t2_ind = np.where((time > 2012.5))[0][0] - 1
	slope = (np.average(lat[t2_ind-5:t2_ind]) - np.average(lat[0:5])) / (time[t2_ind] - time[0])


	plate_motion = (time - time[0]) * slope
	latnorm = lat - plate_motion
	latnorm -= np.average(latnorm[0:t2_ind])

	data[name + '_time'] = time
	data[name + '_lat'] = latnorm

	if plot_figs:
		ax1 = fig.add_subplot(311)
		ax1.plot(time, plate_motion, '.r')
		ax1.plot(time, latnorm, '.b')
		ax1.plot(time, lat, '.k')
		ax1.ticklabel_format(useOffset=False)


	with open(path + 'data/GD/' + name + '.lon', 'r') as lont:
		lonread = lont.readlines()

	lon = np.asarray([x.split()[1] for x in lonread], dtype=np.float32)

	t2_ind = np.where((time > 2012.5))[0][0] - 1

	slope = (np.average(lon[t2_ind-5:t2_ind]) - np.average(lon[0:5])) / (time[t2_ind] - time[0])


	plate_motion = (time - time[0]) * slope
	lonnorm = lon - plate_motion
	lonnorm -= np.average(lonnorm[0:t2_ind])

	data[name + '_lon'] = lonnorm

	if plot_figs:
		ax1 = fig.add_subplot(312)
		ax1.plot(time, plate_motion, '.r')
		ax1.plot(time, lonnorm, '.b')
		ax1.plot(time, lon, '.k')
		ax1.ticklabel_format(useOffset=False)


	with open(path + 'data/GD/' + name + '.rad', 'r') as radt:
		radread = radt.readlines()

	rad = np.asarray([x.split()[1] for x in radread], dtype=np.float32)

	t2_ind = np.where((time > 2012.5))[0][0] - 1

	slope = (np.average(rad[t2_ind - 5:t2_ind]) - np.average(rad[0:5])) / (time[t2_ind] - time[0])

	plate_motion = (time - time[0]) * slope
	radnorm = rad - plate_motion
	radnorm -= np.average(radnorm[0:t2_ind])

	data[name + '_rad'] = radnorm

	if plot_figs:
		ax1 = fig.add_subplot(313)
		ax1.plot(time, plate_motion, '.r')
		ax1.plot(time, radnorm, '.b')
		ax1.plot(time, rad, '.k')
		ax1.ticklabel_format(useOffset=False)
		fig.savefig('figures/gpsdata_%s.png' %name, dpi=400)



data_file = path+'data/Data.xlsx'
wb = pyxl.load_workbook(data_file)
gps_data = wb['GPS_data']

X_gps, Y_gps, Ux_gps, Uy_gps, Uz_gps = GPSlatlon2XY(gps_data, origin, theta)
Z_gps = np.zeros(len(X_gps))


# cross_ind_time = []
# ind_starttime = []
# Ux_cosies = []
# Uz_cosies = []
# for i in cross_ind:
#
# 	#load and rotate timeseries
# 	gps_time = data[names[i] + '_time']
# 	lon_time = data[names[i] + '_lon']
# 	lat_time = data[names[i] + '_lat']
# 	Uz_time = data[names[i] + '_rad']
#
# 	Ux_time, Uy_time = GPSlatlon2XY_time(lat_time, lon_time, theta)
#
# 	ind_time = np.where(gps_time[:]>2012.68)[0][0]
# 	print names[i], gps_time[ind_time]
#
# 	if gps_time[ind_time] < 2012.9:
# 		cross_ind_time.append(i)
# 		ind_starttime.append(ind_time)
# 		Ux_cosies.append(Ux_time[ind_time])
# 		Uz_cosies.append(Uz_time[ind_time])

##########################################################################
############################ LOAD VARIABLES ##############################
##########################################################################

resultspath = '/home/fenics/shared/forward_codes/results/numpy_results/'
variablespath = '/home/fenics/shared/forward_codes/results/numpy_variables/'

x_all = np.load(variablespath + "x_all_2D.npy")
y_all = np.load(variablespath + "y_all_2D.npy")
surface_dofs = np.load(variablespath + "surface_dofs_2D.npy")
ocean_dofs = np.load(variablespath + "ocean_dofs_2D.npy")
# elastic_drained = np.load(resultspath+"Sol_surf_drained_2D.npy")
# elastic_undrained = np.load(resultspath+"Sol_surf_undrained_2D.npy")
size_p = np.load(variablespath + "size_p_CR_2D.npy")

if event == 'EQ':
	if loop == 'yes':
		if variable == 'kappa':
			# sea_flux = np.load(resultspath+"sea_flux_total_2D_EQ.npy")
			# dtt = np.load(variablespath+"dtt_comsum_EQ.npy")
			# Sol_surfk10 = np.load(resultspath + "Sol_surf_2D_EQ_loop_k10.npy")
			Sol_surfk10 = np.load(resultspath + "Sol_surf_2D_EQ_loop_k10.npy")
			Sol_surf = np.load(resultspath + "Sol_surf_2D_EQ_loop_k11.npy")
			Sol_surfk12 = np.load(resultspath + "Sol_surf_2D_EQ_loop_k12.npy")
			Sol_surfk13 = np.load(resultspath + "Sol_surf_2D_EQ_loop_k13.npy")
			
			#Sol_all_k10 = np.load(resultspath + "Sol_all_2D_EQ_loop_k10.npy")
			
			seaflux_k10 = np.load(resultspath + "sea_flux_total_2D_EQ_loop_k10.npy")
			seaflux_k11 = np.load(resultspath + "sea_flux_total_2D_EQ_loop_k11.npy")
			seaflux_k12 = np.load(resultspath + "sea_flux_total_2D_EQ_loop_k12.npy")
			seaflux_k13 = np.load(resultspath + "sea_flux_total_2D_EQ_loop_k13.npy")
		elif variable == 'sigma':
			# sea_flux = np.load(resultspath+"sea_flux_total_2D_EQ.npy")
			Sol_surf10 = np.load(resultspath + "Sol_surf_2D_EQ_loop_sig10.npy")
			Sol_surf = np.load(resultspath + "Sol_surf_2D_EQ_loop_sig15.npy")
			Sol_surf20 = np.load(resultspath + "Sol_surf_2D_EQ_loop_sig20.npy")
			Sol_surf25 = np.load(resultspath + "Sol_surf_2D_EQ_loop_sig25.npy")
			# dtt = np.load(variablespath+"dtt_comsum_EQ.npy")
			
			seaflux_10 = np.load(resultspath + "sea_flux_total_2D_EQ_loop_sig10.npy")
			seaflux_15 = np.load(resultspath + "sea_flux_total_2D_EQ_loop_sig15.npy")
			seaflux_20 = np.load(resultspath + "sea_flux_total_2D_EQ_loop_sig20.npy")
			seaflux_25 = np.load(resultspath + "sea_flux_total_2D_EQ_loop_sig25.npy")
		elif variable == 'cross':

			Sol_surf110 = np.load(resultspath + "Sol_surf_2D_EQ_loop_cross110.npy")
			Sol_surf115 = np.load(resultspath + "Sol_surf_2D_EQ_loop_cross115.npy")
			Sol_surf120 = np.load(resultspath + "Sol_surf_2D_EQ_loop_cross120.npy")
			Sol_surf125 = np.load(resultspath + "Sol_surf_2D_EQ_loop_cross125.npy")
			Sol_surf130 = np.load(resultspath + "Sol_surf_2D_EQ_loop_cross130.npy")
			Sol_surf135 = np.load(resultspath + "Sol_surf_2D_EQ_loop_cross135.npy")
			Sol_surf140 = np.load(resultspath + "Sol_surf_2D_EQ_loop_cross140.npy")
			Sol_surf145 = np.load(resultspath + "Sol_surf_2D_EQ_loop_cross145.npy")
			Sol_surf150 = np.load(resultspath + "Sol_surf_2D_EQ_loop_cross150.npy")
			Sol_surf155 = np.load(resultspath + "Sol_surf_2D_EQ_loop_cross155.npy")

	else:
		sea_flux = np.load(resultspath + "sea_flux_total_2D_EQ.npy")
		Sol_surf = np.load(resultspath + "Sol_surf_2D_EQ.npy")
		sub_cycle = np.load(variablespath + "sub_cycle_EQ.npy")
		dtt = np.load(variablespath + "dtt_comsum_EQ.npy")
	dtt = [60, 240, 720, 288e1, 2 * 432e2]  # , 5256e2]  # in minutes
	ntime = [24, 60, 60, 30, 60]  # , 50]
	#dtt = [15, 240, 288e1, 432e2]  # in minutes
	#ntime = [96, 120, 15, 122]
	dtt = [i * 60 for i in dtt]  # in seconds
	dtt_repeated = np.repeat(dtt, ntime)
	dtt = np.cumsum(dtt_repeated)

elif event == 'SSE':
	sea_flux = np.load(resultspath + "sea_flux_2D_SSE.npy")
	Sol_surf = np.load(resultspath + "Sol_surf_2D_SSE.npy")
	dtt = np.load(variablespath + "dtt_comsum_SSE.npy")

elif event == 'sub':
	sea_flux = np.load(resultspath + "sea_flux_2D_sub.npy")
	Sol_surf = np.load(resultspath + "Sol_surf_2D_sub.npy")
	sub_cycle = np.load(variablespath + "sub_cycle_sub.npy")
	dtt = np.load(variablespath + "dtt_comsum_sub.npy")

elif event == 'sub_EQ':
	sea_flux = np.load(resultspath + "sea_flux_2D_sub_EQ.npy")
	Sol_surf = np.load(resultspath + "Sol_surf_2D_sub_EQ.npy")
	sub_cycle = np.load(variablespath + "sub_cycle_sub_EQ.npy")
	dtt = np.load(variablespath + "dtt_comsum_sub_EQ.npy")

elif event == 'topo':
	sea_flux = np.load(resultspath + "sea_flux_2D_sub_EQ.npy")
	sea_flux_topo = np.load(resultspath + "sea_flux_2D_topo.npy")
	Sol_surf = np.load(resultspath + "Sol_surf_2D_topo.npy")

dttnew = [60, 240, 720, 288e1, 2 * 432e2]  # , 5256e2]  # in minutes
ntimenew = [24, 60, 60, 30, 60]  # , 50]
#dttnew = [15, 240, 720, 288e1, 2 * 432e2]  # , 5256e2]  # in minutes
#ntimenew = [48, 60, 60, 30, 60]  # , 50]
dttnew = [i * 60 for i in dttnew]  # in seconds
dttnew_repeated = np.repeat(dttnew, ntimenew)
dttnew = np.cumsum(dttnew_repeated)

# # sub_cycle = 50
# sub_cycle = np.load(variablespath + "sub_cycle_sub_EQ.npy")
# dtt = np.load(variablespath + "dtt_comsum_topo.npy")
# dtt_subEQ = np.load(variablespath + "dtt_comsum_sub_EQ.npy")  # GPS data

gps_file = '/home/fenics/shared/data/GPS_displacement.xlsx'

gps_wb = pyxl.load_workbook(gps_file)
gps_data = gps_wb['Sheet2']
gps_X = np.array([[gps_data.cell(row=i, column=1).value]
                  for i in range(2, gps_data.max_row + 1)]).reshape(gps_data.max_row - 1, )
gps_Ux = 0.1 * np.array([[gps_data.cell(row=i, column=3).value]
                         for i in range(2, gps_data.max_row + 1)]).reshape(gps_data.max_row - 1, )
gps_Uz = 0.1 * np.array([[gps_data.cell(row=i, column=5).value]
                         for i in range(2, gps_data.max_row + 1)]).reshape(gps_data.max_row - 1, )

##########################################################################
###################### PARSE AND SORT VARIABLES ##########################
##########################################################################

p_ind = np.argmax(surface_dofs >= size_p)
size_u_surf = x_all[surface_dofs].shape[0]
size_ux = (size_u_surf - p_ind) / 2
ux_ind = p_ind + size_ux

p_ind_o = np.argmax(ocean_dofs >= size_p)
size_u_ocean = x_all[ocean_dofs].shape[0]
size_ux_o = (size_u_ocean - p_ind_o) / 2
ux_ind_o = p_ind_o + size_ux_o
# print surface_dofs.shape
# print p_ind, ux_ind, ux_ind-p_ind


"""
ocean_dof_p = ocean_dofs[0:p_ind_ocean]
ocean_dof_ux = ocean_dofs[p_ind:ux_ind_ocean]
ocean_dof_uy = ocean_dofs[ux_ind_ocean::]
"""
surface_dof_p = surface_dofs[0:p_ind][np.argsort(x_all[surface_dofs[0:p_ind]])]
surface_dof_ux = surface_dofs[p_ind:ux_ind][np.argsort(x_all[surface_dofs[p_ind:ux_ind]])]
surface_dof_uy = surface_dofs[ux_ind::][np.argsort(x_all[surface_dofs[ux_ind::]])]
sort_p, sort_u = np.argsort(x_all[surface_dofs[0:p_ind]]), np.argsort(x_all[surface_dofs[ux_ind::]])

ocean_dof_p = ocean_dofs[0:p_ind_o][np.argsort(x_all[ocean_dofs[0:p_ind_o]])]
ocean_dof_ux = ocean_dofs[p_ind_o:ux_ind_o][np.argsort(x_all[ocean_dofs[p_ind_o:ux_ind_o]])]
ocean_dof_uy = ocean_dofs[ux_ind_o::][np.argsort(x_all[ocean_dofs[ux_ind_o::]])]
sort_po, sort_uo = np.argsort(x_all[ocean_dofs[0:p_ind_o]]), np.argsort(x_all[ocean_dofs[ux_ind_o::]])

sort_p_all, sort_u_all = np.concatenate((sort_po, sort_p)), np.concatenate((sort_uo, sort_u))

# print '#################'
# print surface_dof_ux[0:10]
# print x_all[surface_dof_ux][0:10]
# print '#################'
#
# #print np.argsort(x_all[surface_dofs[p_ind:ux_ind]])[0:20]
# print sort_u[0:10]
# print x_all[surface_dofs[p_ind:ux_ind]][sort_u][0:10]
# print '#################'
# print np.argsort(x_all[surface_dofs[p_ind:ux_ind]])[0:10]
# print x_all[surface_dofs[np.argsort(x_all[surface_dofs[p_ind:ux_ind]])]][0:10]

x_surf_p = x_all[surface_dof_p]
x_surf_ux = x_all[surface_dof_ux]
x_surf_uy = x_all[surface_dof_uy]

y_surf_p = y_all[surface_dof_p]
y_surf_ux = y_all[surface_dof_ux]
y_surf_uy = y_all[surface_dof_uy]

if loop == 'yes':
	if variable == 'kappa':
		Sol_surf_pk10 = Sol_surfk10[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_pk11 = Sol_surf[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_pk12 = Sol_surfk12[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_pk13 = Sol_surfk13[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		
		Sol_surf_xk10 = Sol_surfk10[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_xk11 = Sol_surf[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_xk12 = Sol_surfk12[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_xk13 = Sol_surfk13[p_ind:ux_ind, :][sort_u, :] * 1e2
		
		Sol_surf_yk10 = Sol_surfk10[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_yk11 = Sol_surf[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_yk12 = Sol_surfk12[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_yk13 = Sol_surfk13[ux_ind::, :][sort_u, :] * 1e2
	
	
	if variable == 'sigma':
		Sol_surf_p10 = Sol_surf10[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_p15 = Sol_surf[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_p20 = Sol_surf20[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_p25 = Sol_surf25[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		
		Sol_surf_x10 = Sol_surf10[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_x15 = Sol_surf[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_x20 = Sol_surf20[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_x25 = Sol_surf25[p_ind:ux_ind, :][sort_u, :] * 1e2
		
		Sol_surf_y10 = Sol_surf10[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_y15 = Sol_surf[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_y20 = Sol_surf20[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_y25 = Sol_surf25[ux_ind::, :][sort_u, :] * 1e2
	
	if variable == 'cross':
		Sol_surf_p110 = Sol_surf110[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_p115 = Sol_surf115[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_p120 = Sol_surf120[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_p125 = Sol_surf125[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_p130 = Sol_surf130[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_p135 = Sol_surf135[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_p140 = Sol_surf140[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_p145 = Sol_surf145[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_p150 = Sol_surf150[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_p155 = Sol_surf155[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		
		Sol_surf_x110 = Sol_surf110[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_x115 = Sol_surf115[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_x120 = Sol_surf120[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_x125 = Sol_surf125[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_x130 = Sol_surf130[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_x135 = Sol_surf135[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_x140 = Sol_surf140[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_x145 = Sol_surf145[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_x150 = Sol_surf150[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_x155 = Sol_surf155[p_ind:ux_ind, :][sort_u, :] * 1e2
		
		Sol_surf_y110 = Sol_surf110[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_y115 = Sol_surf115[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_y120 = Sol_surf120[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_y125 = Sol_surf125[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_y130 = Sol_surf130[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_y135 = Sol_surf135[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_y140 = Sol_surf140[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_y145 = Sol_surf145[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_y150 = Sol_surf150[ux_ind::, :][sort_u, :] * 1e2
		Sol_surf_y155 = Sol_surf155[ux_ind::, :][sort_u, :] * 1e2



else:
	
	Sol_surf_p = Sol_surf[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
	Sol_surf_x = Sol_surf[p_ind:ux_ind, :][sort_u, :]
	Sol_surf_y = Sol_surf[ux_ind::, :][sort_u, :]

nsteps = dtt.size
y_surf_zero = y_surf_ux - y_surf_ux

"""e_drained_x = elastic_drained[0:size_ux][sort_u]
e_drained_y = elastic_drained[size_ux::][sort_u]
e_undrained_x = elastic_undrained[0:size_ux][sort_u]
e_undrained_y = elastic_undrained[size_ux::][sort_u]"""

# Sol_surf_x_post = Sol_surf_x - np.transpose(np.tile(Sol_surf_x[:,0], (nsteps,1)))
# Sol_surf_y_post = Sol_surf_y - np.transpose(np.tile(Sol_surf_y[:,0], (nsteps,1)))
# Sol_surf_p_post = Sol_surf_p - np.transpose(np.tile(Sol_surf_p[:,0], (nsteps,1)))

# print Sol_surf_x_post.max(), Sol_surf_x_post.min()
# print Sol_surf_x.max(), Sol_surf_x.min()

"""
undrained_diff_x, undrained_diff_y = Sol_surf_x[:,0] - e_undrained_x, Sol_surf_y[:,0] - e_undrained_y
drained_diff_x, drained_diff_y = Sol_surf_x[:, -1] - e_drained_x, Sol_surf_y[:,-1] - e_drained_y

elastic_results_x  = np.empty((e_drained_x.shape[0],2))
elastic_results_y  = np.empty((e_drained_x.shape[0],2))

if event == 'sub_EQ':
    for j in range(0,e_drained_x.shape[0]):
        elastic_results_x[j,:] = np.array([e_undrained_x[j], e_drained_x[j]]) + np.array([Sol_surf_x[j,sub_cycle-1],Sol_surf_x[j,sub_cycle-1]])
        elastic_results_y[j,:] = np.array([e_undrained_y[j], e_drained_y[j]]) + np.array([Sol_surf_y[j,sub_cycle-1],Sol_surf_y[j,sub_cycle-1]])
else:
    for j in range(0,e_drained_x.shape[0]):
        elastic_results_x[j,:] = np.array([e_undrained_x[j], e_drained_x[j]])
        elastic_results_y[j,:] = np.array([e_undrained_y[j], e_drained_y[j]])
"""

##########################################################################
######################### CONVERT UNITS ##################################
##########################################################################

if event == 'topo':
	dt_minus = np.concatenate([[0], dtt_subEQ[0:-1]])
	dt = dtt_subEQ - dt_minus
	
	m3_h20 = 4e4 * sea_flux * dt
	m3_h20_topo = 4e4 * sea_flux_topo[sub_cycle] * dt
	
	# Cumulative cubic kilometers SGD
	cum_h20_topo = np.cumsum(m3_h20_topo) * 1e-9
	cum_h20_sub = np.cumsum(m3_h20) * 1e-9
	cum_h20_both = cum_h20_sub + cum_h20_topo
	
	# month 1 - [216] - 21 days
	# month 2 - [231] - 30 days
	# month 3 - [232]
	# month 4 - [233]
	# month 5 - [234]
	# month 6- [235]
	
	nummonths = 24
	EQmonthsadd = np.arange(281, 281 + (nummonths - 1), 1)
	EQmonths = np.concatenate(
		[[sub_cycle - 1, sub_cycle, 266.0], EQmonthsadd])  # indice intervals for months following the EQ
	ratio = np.empty((EQmonths.shape[0] - 1))
	# percentage = np.empty((EQmonths.shape[0]-1))
	# dt_percent = np.empty((EQmonths.shape[0]-1))
	
	for i in range(0, EQmonths.shape[0] - 1):
		sub_water = np.sum(m3_h20[EQmonths[i]:EQmonths[i + 1]])
		topo_water = np.sum(m3_h20_topo[EQmonths[i]:EQmonths[i + 1]])
		# dt_percent[i] = np.sum(dt[EQmonths[i]:EQmonths[i+1]])/(3600*24*30)
		ratio[i] = sub_water / topo_water
	# percentage[i] = 100*sub_water/(sub_water+topo_water)
	
	percentage = (sea_flux * dt) / (sea_flux * dt + sea_flux_topo[sub_cycle] * dt)
	
	total_h20_topo = np.sum(m3_h20_topo)
	total_h20_b4 = np.sum(m3_h20[0:sub_cycle])
	total_h20_EQ = np.sum(m3_h20[sub_cycle::1])
	
	sixmo_h20_topo = np.sum(m3_h20_topo[sub_cycle:sub_cycle + 1]) / 2
	sixmo_h20_EQ = np.sum(m3_h20[sub_cycle:285])
	
	print 'cubic meters expelled from topographic flow', total_h20_topo
	print 'cubic meters expelled through seafloor before EQ', total_h20_b4
	print 'cubic meters expelled through after EQ', total_h20_EQ
	print '#####################################################'
	print 'cubic meters expelled from topographic flow - 6 months', sixmo_h20_topo
	print 'cubic meters expelled through seafloor in 6 months after EQ', sixmo_h20_EQ
	print 'EQ/topo in 6 months after EQ', sixmo_h20_EQ / sixmo_h20_topo
	
	dtt_subEQ = dtt_subEQ / (3600 * 24 * 365) - 50

# convert dtt from seconds to days
dtt = dtt / (3600 * 24)
dttnew = dttnew/ (3600 * 24)
if event == 'sub_EQ':
	x_elastic = np.array([dtt[sub_cycle], dtt[-1]])  # timesteps to plot elastic solution
else:
	x_elastic = np.array([dtt[0], dtt[-1]])


# convert sea flux from m/s to m/day
# sea_flux = sea_flux*(3600*24)

# convert MPa to well head change (meters)
# Sol_surf_p9[:] = Sol_surf_pk9[:]*1e3/(9.81)


##########################################################################
######################### PLOTTING FUNCTIONS #############################
##########################################################################
def shiftedColorMap(cmap, start=0, midpoint=0.5, stop=1.0, name='shiftedcmap'):
	'''
    Function to offset the "center" of a colormap. Useful for
    data with a negative min and positive max and you want the
    middle of the colormap's dynamic range to be at zero

    Input
    -----
      cmap : The matplotlib colormap to be altered
      start : Offset from lowest point in the colormap's range.
          Defaults to 0.0 (no lower ofset). Should be between
          0.0 and `midpoint`.
      midpoint : The new center of the colormap. Defaults to
          0.5 (no shift). Should be between 0.0 and 1.0. In
          general, this should be  1 - vmax/(vmax + abs(vmin))
          For example if your data range from -15.0 to +5.0 and
          you want the center of the colormap at 0.0, `midpoint`
          should be set to  1 - 5/(5 + 15)) or 0.75
      stop : Offset from highets point in the colormap's range.
          Defaults to 1.0 (no upper ofset). Should be between
          `midpoint` and 1.0.
    '''
	cdict = {
		'red': [],
		'green': [],
		'blue': [],
		'alpha': []
	}
	
	# regular index to compute the colors
	reg_index = np.linspace(start, stop, 257)
	
	# shifted index to match the data
	shift_index = np.hstack([
		np.linspace(0.0, midpoint, 128, endpoint=False),
		np.linspace(midpoint, 1.0, 129, endpoint=True)
	])
	
	for ri, si in zip(reg_index, shift_index):
		r, g, b, a = cmap(ri)
		
		cdict['red'].append((si, r, r))
		cdict['green'].append((si, g, g))
		cdict['blue'].append((si, b, b))
		cdict['alpha'].append((si, a, a))
	
	newcmap = mpl.colors.LinearSegmentedColormap(name, cdict)
	plt.register_cmap(cmap=newcmap)
	
	return newcmap


cmap = mpl.cm.winter


def plot_timeseries(points):
	for i in range(0, points.shape[0]):
		idx = np.abs(x_surf[:] - points[i]).argmin()
		fname = points[i] / 1e3
		fig = plt.figure()
		ax1 = fig.add_subplot(311)
		ax2 = fig.add_subplot(312)
		ax3 = fig.add_subplot(313)
		ax1.set_title('Modeled poroelastic timeseries %3d km from trench' % fname)
		
		ax1.plot(dtt, Sol_surf_p[idx, :], '-k.', label='constant k', ms=10, mew=1.5, fillstyle='none')  # pressure
		
		ax2.plot(dtt, Sol_surf_x[idx, :], '-r.', ms=10, mew=1.5, fillstyle='none')  # u_x
		# ax2.plot(x_elastic, elastic_results_x[idx, :],'ro', ms = 13, fillstyle='full', label = 'elastic solution') # u_x
		
		ax3.plot(dtt, Sol_surf_y[idx, :], '-g.', ms=10, mew=1.5, fillstyle='none')  # u_y
		# ax3.plot(x_elastic, elastic_results_y[idx, :],'go', ms=13, fillstyle='full', label = 'elastic solution') # u_x
		
		
		# plt.ylabel('Depth (km)')
		ax1.set_ylabel('meters of head change')
		ax2.set_ylabel('Displacement (cm)')
		ax3.set_ylabel('Displacement (cm)')
		plt.xlabel('Days after EQ')
		"""
        legend1 = ax1.legend()
        for label in legend1.get_texts():
            label.set_fontsize('medium')
        legend2 = ax2.legend()
        for label in legend2.get_texts():
            label.set_fontsize('medium')
        legend3 = ax3.legend()
        for label in legend3.get_texts():
            label.set_fontsize('medium')
        """
		fig.savefig('figures/2D_timeseries_%s.png' % str(points[i])[:2], dpi=400)


def plot_timeseries_head(points):
	
	days = 30.
	x_elastic = np.array([0, dtt[endtime - 1]])
	
	for i in range(0, points.shape[0]):
		idx_p = np.abs(x_surf_p[:] - points[i]).argmin()
		idx_u = np.abs(x_surf_ux[:] - points[i]).argmin()
		
		print idx_p, idx_u
		
		fname = points[i] / 1e3
		fig = plt.figure()
		ax1 = fig.add_subplot(311)
		ax2 = fig.add_subplot(312)
		ax3 = fig.add_subplot(313)
		# ax1.set_title('Modeled poroelastic timeseries %3d km from trench' % fname)
		
		ax1.plot(dttnew/days, Sol_surf_pk10[idx_p, :], 'red', label='log(k) = -10', linewidth=5)
		ax1.plot(dttnew/days, Sol_surf_pk11[idx_p, 0:endtime], 'k', label='log(k) = -11', linewidth=5)
		ax1.plot(dttnew/days, Sol_surf_pk12[idx_p, 0:endtime], 'lightblue', label='log(k) = -12', linewidth=5)
		ax1.plot(dttnew/days, Sol_surf_pk13[idx_p, 0:endtime], 'b', label='log(k) = -13', linewidth=5)
		
		ax2.plot(dttnew/days, Sol_surf_xk10[idx_u, :], 'red', label='log(k) = -10',
		         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none', color = 'pink') #pressure
		ax2.plot(dttnew/days, Sol_surf_xk11[idx_u, 0:endtime], 'k', label='log(k) = -11',
		         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none') #pressure
		ax2.plot(dttnew/days, Sol_surf_xk12[idx_u, 0:endtime], 'lightblue', label='log(k) = -12',
		         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none', color = 'lightblue') #pressure
		ax2.plot(dttnew/days, Sol_surf_xk13[idx_u, 0:endtime], 'b', label='log(k) = -13',
		         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none') #pressure
		
		ax3.plot(dttnew/days, Sol_surf_yk10[idx_u, :], 'red', label='log(k) = -10',
		         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none', color = 'pink') #pressure
		ax3.plot(dttnew/days, Sol_surf_yk11[idx_u, 0:endtime], 'k', label='log(k) = -11',
		         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none') #pressure
		ax3.plot(dttnew/days, Sol_surf_yk12[idx_u, 0:endtime], 'lightblue', label='log(k) = -12',
		         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none', color = 'lightblue') #pressure
		ax3.plot(dttnew/days, Sol_surf_yk13[idx_u, 0:endtime], 'b', label='log(k) = -13',
		         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none') #pressure
		
		# plt.ylabel('Depth (km)')
		# ax1.set_ylabel('meters of head change')
		# ax2.set_ylabel('X Displacement (cm)')
		# ax3.set_ylabel('Uplift (cm)')
		# plt.xlabel('Days after EQ')
		# ax1.legend()
		
		ax1.set_xlim((0, 4))
		ax2.set_xlim((0, 4))
		ax3.set_xlim((0, 4))

		
		# fig.set_size_inches(8, 10)
		# plt.xlim((0, 11000))
		# plt.ylim((5e-8, 10))
		
		# fig1.savefig('figures/SGD_sea_flux.png', dpi = 400)
		# fig3.savefig('figures/SGD_percent.png', dpi=400)#
		
		legend1 = ax1.legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')
		# legend2 = ax2.legend()
		# for label in legend2.get_texts():
		#     label.set_fontsize('medium')
		# legend3 = ax3.legend()
		# for label in legend3.get_texts():
		#     label.set_fontsize('medium')
		
		fig.savefig('figures/2D_timeseries_%s.png' % str(points[i])[:2], dpi=400)


def plot_cross_time():
	days = 1.
	endday = 40
	end = np.where(dttnew / days > endday)[0][0]
	
	for i in range(0, loop_through.shape[0]):
		
		print "###################", "cross section", str(loop_through[i])[:3], "#####################"
		
		Sol_surf = np.load(resultspath + "Sol_surf_2D_EQ_loop_cross"+str(loop_through[i])[:3]+".npy")
		Sol_surf_p = Sol_surf[0:p_ind, :][sort_p, :] * (1e3 / (9.81))
		Sol_surf_x = Sol_surf[p_ind:ux_ind, :][sort_u, :] * 1e2
		Sol_surf_y = Sol_surf[ux_ind::, :][sort_u, :] * 1e2

		cross_ind = np.where(np.sqrt((Y_gps_post - 1e3*loop_through[i]) ** 2) <= 5e3)[0]
		

		
		cross_ind_time = []
		ind_starttime = []
		Ux_cosies = []
		Uz_cosies = []
		for ind in cross_ind:
			
			if len(cross_ind) == 0:
				break

			# load and rotate timeseries
			gps_time = data[names[ind] + '_time']
			lon_time = data[names[ind] + '_lon']
			lat_time = data[names[ind] + '_lat']
			Uz_time = data[names[ind] + '_rad']

			Ux_time, Uy_time = GPSlatlon2XY_time(lat_time, lon_time, theta)

			ind_time = np.where(gps_time[:] > 2012.68)[0][0]


			print names[ind], gps_time[ind_time]

			if gps_time[ind_time] < 2012.9:
				cross_ind_time.append(ind)
				ind_starttime.append(ind_time)
				Ux_cosies.append(Ux_time[ind_time])
				Uz_cosies.append(Uz_time[ind_time])
				
			else:
				break


			idx_p = np.abs(x_surf_p[:] - X_gps_post[ind]).argmin()
			idx_u = np.abs(x_surf_ux[:] - X_gps_post[ind]).argmin()
			

			ind_time = np.where(gps_time[:] > 2012.68)[0][0] + 1
			end_time = np.where(gps_time[:] > 2012.68 + (endday/365.))[0][0]
			end_time = np.where(gps_time[:] > 2013.2)[0][0]


			time_post = 365 * (gps_time[ind_time:end_time] - gps_time[ind_time])  # /days
			Ux_post, Uz_post = Ux_time[ind_time:end_time], Uz_time[ind_time:end_time]

			Ux_post += Sol_surf_x[idx_u, 0] - Ux_post[0]
			Uz_post += Sol_surf_y[idx_u, 0] - Uz_post[0]
			
			Ux_post_norm = []
			Uz_post_norm = []
			Uxpercent = [0.]
			Uzpercent = [0.]

			for j in range(0,time_post.shape[0]):
				norm_i = np.abs(np.sqrt((dttnew - time_post[j]) ** 2)).argmin()
				Ux_post_norm.append(Ux_post[j] - (Sol_surf_x[idx_u, norm_i] - Sol_surf_x[idx_u, 0]))
				Uz_post_norm.append(Uz_post[j] - (Sol_surf_y[idx_u, norm_i] - Sol_surf_y[idx_u, 0]))
				if j>0:
					xper = 1e2*(Sol_surf_x[idx_u, norm_i] - Sol_surf_x[idx_u, 0])/(Ux_post[j] - Ux_post[0])
					if xper >= 100.:
						#print "percent error x", names[ind], xper
						#print "numer = ",Sol_surf_x[idx_u, norm_i] - Sol_surf_x[idx_u, 0]
						#print "denominator = ", Ux_post[j] - Ux_post[0]
						Uxpercent.append(100.)
					elif xper <= -100.:
						#print "percent error x", names[ind], xper
						#print "numer = ",Sol_surf_x[idx_u, norm_i] - Sol_surf_x[idx_u, 0]
						#print "denominator = ", Ux_post[j] - Ux_post[0]
						Uxpercent.append(-100.)
					else:
						Uxpercent.append(xper)
						
					zper = 1e2*(Sol_surf_y[idx_u, norm_i] - Sol_surf_y[idx_u, 0]) / (Uz_post[j] - Uz_post[0])
					if zper >= 100.:
						#print "percent error z", names[ind], zper
						#print "numer = ",Sol_surf_y[idx_u, norm_i] - Sol_surf_y[idx_u, 0]
						#print "denominator = ", Uz_post[j] - Uz_post[0]
						Uzpercent.append(100.)
					elif zper <= -100.:
						#print "percent error z", names[ind], zper
						#print "numer = ",Sol_surf_y[idx_u, norm_i] - Sol_surf_y[idx_u, 0]
						#print "denominator = ", Uz_post[j] - Uz_post[0]
						Uzpercent.append(-100.)
					else:
						Uzpercent.append(zper)

			
			#fname = points[i] / 1e3
			fig = plt.figure()
			ax1 = fig.add_subplot(311)
			ax2 = fig.add_subplot(312)
			ax3 = fig.add_subplot(313)
			# ax1.set_title('Modeled poroelastic timeseries %3d km from trench' % fname)

			ax1.plot(dttnew / days, Sol_surf_p[idx_p, :], 'blue', linewidth=3)

			ax2.plot(dttnew[0:end] / days, Sol_surf_x[idx_u, 0:end], 'green', label='model',
			         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none') #pressure
			ax2.plot(time_post, Ux_post, '.k', label='data',linewidth=3)#, fillstyle='none')
			ax2.plot(time_post, Ux_post_norm, '.g', label='residual',linewidth=3, fillstyle='none')



			ax3.plot(dttnew[0:end] / days, Sol_surf_y[idx_u, 0:end], 'red', label='model',
			         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none') #pressure
			ax3.plot(time_post, Uz_post, '.k', label='data',linewidth=3)#, fillstyle='none')
			ax3.plot(time_post, Uz_post_norm, '.r', label='residual',linewidth=3, fillstyle='none')


			# plt.ylabel('Depth (km)')
			ax1.set_ylabel('Head change (m)')
			ax2.set_ylabel('Trench-ward (cm)')
			ax3.set_ylabel('Uplift (cm)')
			plt.xlabel('Days after EQ')
			# ax1.legend()

			ax1.set_xlim((0, endday))
			ax2.set_xlim((0, endday))
			ax3.set_xlim((0, endday))
			
			#ax1.set_ylim([Ux_post.min()-2. , Ux_post.max()+2.])
			ax2.set_ylim([Ux_post.min()-1. , Ux_post.max()+3.])
			ax3.set_ylim([Uz_post.min()-1. , Uz_post.max()+2.])

			# fig.set_size_inches(8, 10)
			# plt.xlim((0, 11000))
			# plt.ylim((5e-8, 10))

			# fig1.savefig('figures/SGD_sea_flux.png', dpi = 400)
			# fig3.savefig('figures/SGD_percent.png', dpi=400)#

			#legend1 = ax1.legend()
			#for label in legend1.get_texts():
			#	label.set_fontsize('medium')
			# Shrink current axis by 20%
			box = ax1.get_position()
			ax1.set_position([box.x0, box.y0, box.width * 0.8, box.height])
			box2 = ax2.get_position()
			ax2.set_position([box2.x0, box2.y0, box2.width * 0.8, box2.height])
			box3 = ax3.get_position()
			ax3.set_position([box3.x0, box3.y0, box3.width * 0.8, box3.height])
			
			# Put a legend to the right of the current axis
			#ax2.legend(loc='center left', bbox_to_anchor=(1, 0.5))
			
			legend2 = ax2.legend(loc='center left', bbox_to_anchor=(1, 0.5))
			for label in legend2.get_texts():
			    label.set_fontsize('small')
			legend3 = ax3.legend(loc='center left', bbox_to_anchor=(1, 0.5))
			for label in legend3.get_texts():
			    label.set_fontsize('small')


			fig.savefig('figures/2D_crosstimeseries_cross%s_%s.png' % (str(loop_through[i])[:3], str(X_gps_post[ind])[:3]), dpi=400)
		
			fig = plt.figure()
			ax1 = fig.add_subplot(211)
			ax2 = fig.add_subplot(212)
			
			ax1.plot(time_post, Uxpercent, '.g')
			
			ax2.plot(time_post, Uzpercent, '.r')

			# plt.ylabel('Depth (km)')
			ax1.set_ylabel('Percent trench-ward')
			ax2.set_ylabel('Percent uplift')
			# ax3.set_ylabel('Uplift (cm)')
			# plt.xlabel('Days after EQ')
			# ax1.legend()
			
			ax1.set_xlim((0, endday))
			ax2.set_xlim((0, endday))
			ax3.set_xlim((0, endday))
			
			ax1.set_ylim([np.asarray(Uxpercent).min()-10. , np.asarray(Uxpercent).max()+10.])
			ax2.set_ylim([np.asarray(Uzpercent).min()-10. , np.asarray(Uzpercent).max()+10.])
			
			fig.savefig('figures/2D_cross_percent_%s_%s.png' % (str(loop_through[i])[:3], str(X_gps_post[ind])[:3]), dpi=400)
	
	
		fig = plt.figure()
		ax1 = fig.add_subplot(211)
		ax2 = fig.add_subplot(212)
		
		# ax1.set_title('Nicoya Peninsula Topography')
		# ax1.plot(X_topo, h_topo, '-b*', ms=3, mew=2, fillstyle='none', color='green', linewidth=3)
		ax1.scatter(1e-3*x_surf_ux, Sol_surf_x[:, 0])
		ax2.scatter(1e-3*x_surf_uy, Sol_surf_y[:, 0])
		ax1.plot(1e-3 * X_gps_post[cross_ind_time], Ux_cosies, '.r')
		ax2.plot(1e-3 * X_gps_post[cross_ind_time], Uz_cosies, '.r')
		
		# ax1.plot(X_topo,h_topo2,'-b*', ms=3, mew=2, fillstyle='none', color = 'blue',linewidth = 3)
		# plt.axis('scaled')
		#ax1.set_xlim((60, 200))
		#ax2.set_xlim((60, 200))  # plt.ylim((-1e4, 1e4))
		
		# fig.set_size_inches(7, 4)
		
		fig.savefig('figures/cross_surf_%s.png' % str(loop_through[i])[:3], dpi=400)



def plot_surface_loop():
	fig = plt.figure()
	ax1 = fig.add_subplot(311)
	ax2 = fig.add_subplot(312)
	ax3 = fig.add_subplot(313)
	# ax3 = fig.add_subplot(313)
	# ax1.set_title('Modeled poroelastic timeseries %3d km from trench' % fname)
	
	# ax1.plot(dtt[0:endtime], Sol_surf_p9[idx, 0:endtime], '-r.', label = 'log(k) = -9',ms=10, mew=1.5, fillstyle='none', color = 'red') #pressure
	ax1.plot(x_surf_p, Sol_surf_p10[:, 0], '#F7941D', label='sigma = 10',
	         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none', color = 'pink') #pressure
	ax1.plot(x_surf_p, Sol_surf_p15[:, 0], '#BF1E2E', label='sigma = 15',
	         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none') #pressure
	ax1.plot(x_surf_p, Sol_surf_p20[:, 0], '#0F76BB', label='sigma = 20',
	         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none', color = 'lightblue') #pressure
	ax1.plot(x_surf_p, Sol_surf_p25[:, 0], '#00A54F', label='sigma = 25',
	         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none') #pressure
	
	ax1.set_yticks(ax1.get_yticks()[::2])
	ax1.set_xlim((60.5, 250))
	
	ax2.plot(x_surf_uy, Sol_surf_y10[:, 0], '#F7941D', label='W:D = 2:1',
	         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none', color = 'pink') #pressure
	ax2.plot(x_surf_uy, Sol_surf_y15[:, 0], '#BF1E2E', label='W:D = 3:1',
	         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none') #pressure
	ax2.plot(x_surf_uy, Sol_surf_y20[:, 0], '#0F76BB', label='W:D = 4:1',
	         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none', color = 'lightblue') #pressure
	ax2.plot(x_surf_uy, Sol_surf_y25[:, 0], '#00A54F', label='W:D = 5:1',
	         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none') #pressure
	ax2.plot(gps_X, gps_Uz, 'ok', label='GPS Data')
	
	ax2.set_xlim((60.5, 250))
	
	ax3.plot(x_surf_ux, Sol_surf_x10[:, 0], '#F7941D', label='W:D = 2:1',
	         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none', color = 'pink') #pressure
	ax3.plot(x_surf_ux, Sol_surf_x15[:, 0], '#BF1E2E', label='W:D = 3:1',
	         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none') #pressure
	ax3.plot(x_surf_ux, Sol_surf_x20[:, 0], '#0F76BB', label='W:D = 4:1',
	         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none', color = 'lightblue') #pressure
	ax3.plot(x_surf_ux, Sol_surf_x25[:, 0], '#00A54F', label='W:D = 5:1',
	         linewidth=3)  # ,ms=1, mew=1.5, fillstyle='none') #pressure
	ax3.plot(gps_X, gps_Ux, 'ok', label='GPS Data')
	
	# plt.ylabel('Depth (km)')
	ax1.set_ylabel('meters of head change')
	# ax2.set_ylabel('Displacement (cm)')
	# ax3.set_ylabel('Displacement (cm)')
	plt.xlabel('Meters from trench')
	# ax1.legend()
	
	# fig.set_size_inches(8, 5)
	plt.xlim((60.5, 250))
	# plt.ylim((5e-8, 10))
	
	
	ax1.set_ylabel('head change (m)')
	ax2.set_ylabel('Uplift (cm)')
	ax3.set_ylabel('X disp. (cm)')
	# ax2.set_xlabel('Meters from trench')
	
	# legend1 = ax1.legend()
	# for label in legend1.get_texts():
	#    label.set_fontsize('medium')
	legend2 = ax2.legend(prop={'size': 8})
	for label in legend2.get_texts():
		label.set_fontsize('small')
	
	fig.savefig('figures/sig_loop_surface.png', dpi=400)


def plot_surface_e():
	fig = plt.figure()
	ax1 = fig.add_subplot(211)
	ax2 = fig.add_subplot(212)
	ax1.set_title('Elastic Solution')
	
	scaleplot = 2.5e3
	
	Ux, Uy = e_undrained_x, e_undrained_y
	x_surf_0, y_surf_0 = x_surf + scaleplot * Ux, Uy
	
	Ux_post, Uy_post = e_drained_x - e_undrained_x, e_drained_y - e_undrained_y
	x_surf_post, y_surf_post = x_surf_0 + scaleplot * Ux_post, y_surf_0 + Uy_post
	
	ax1.plot(x_surf_0, y_surf_0, '-r*', ms=2, mew=2, fillstyle='none', color='black', label='pre-seismic land surface')
	
	ax1.plot(x_surf, y_surf_zero, '-k*', ms=2, mew=2, fillstyle='none', color='red', label='co-seismic land surface')
	
	ax1.quiver(x_surf, y_surf_zero, scaleplot * Ux, Uy, scale_units='xy', angles='xy', scale=1, width=2e-3)
	
	ax2.plot(x_surf_0, y_surf_0, '-r*', ms=2, mew=2, fillstyle='none', color='red',
	         label='undrained co-seimic solution')
	
	ax2.plot(x_surf_post, y_surf_post, '-r*', ms=2, mew=2, fillstyle='none', color='magenta',
	         label='drained co-seimic solution')
	
	ax2.quiver(x_surf_0, y_surf_0, scaleplot * Ux_post, Uy_post, scale_units='xy', angles='xy', scale=1, width=1.5e-3)
	
	ax1.set_ylabel('Displacement (cm)')
	ax2.set_ylabel('Displacement (cm)')
	ax2.set_xlabel('Meters from trench')
	plt.xlabel('Meters from trench')
	legend1 = ax1.legend()
	for label in legend1.get_texts():
		label.set_fontsize('medium')
	legend2 = ax2.legend()
	for label in legend2.get_texts():
		label.set_fontsize('medium')


def plot_diff_surf():  # Plot the difference between the elastic and poroelastic solutions
	
	fig = plt.figure()
	ax1 = fig.add_subplot(221)
	ax2 = fig.add_subplot(223)
	# ax1.set_title('Difference between elastic and poroelastic solution')
	ax1.set_title('Co-seismic solution')
	ax2.set_title('Post-seismic solution')
	scaleplot = 2.5e3
	
	Ux_p, Uy_p = Sol_surf_x[:, 0], Sol_surf_y[:, 0]
	Ux_e, Uy_e = e_undrained_x, e_undrained_y
	x_surf_p0, y_surf_p0 = x_surf + Ux_p, Uy_p
	x_surf_e0, y_surf_e0 = x_surf + Ux_e, Uy_e
	Ux_diff0, Uy_diff0 = x_surf + scaleplot * (Ux_p - Ux_e), (Uy_p - Uy_e)
	Ux_surf_diff0, Uy_surf_diff0 = (Ux_p - Ux_e), (Uy_p - Uy_e)
	
	Ux_p_post, Uy_p_post = Sol_surf_x[:, -1], Sol_surf_y[:, -1]
	Ux_e_post, Uy_e_post = e_drained_x, e_drained_y
	x_surf_ppost, y_surf_ppost = x_surf_p0 + scaleplot * Ux_p_post, y_surf_p0 + Uy_p_post
	x_surf_epost, y_surf_epost = x_surf_e0 + scaleplot * Ux_e_post, y_surf_e0 + Uy_e_post
	Ux_diff_post, Uy_diff_post = x_surf_e0 + (Ux_p_post - Ux_e_post), (Uy_p_post - Uy_e_post)
	Ux_surf_diff_post, Uy_surf_diff_post = (Ux_p_post - Ux_e_post), (Uy_p_post - Uy_e_post)
	
	ax1.plot(x_surf_e0, y_surf_zero, '-r*', ms=2, mew=2, fillstyle='none', color='red', label='elastic co-seismic')
	ax1.plot(Ux_diff0, Uy_diff0, '-k*', ms=2, mew=2, fillstyle='none', color='blue', label='poroelastic co-seismic')
	ax1.quiver(x_surf_e0, y_surf_zero, scaleplot * Ux_surf_diff0, Uy_diff0, scale_units='xy', angles='xy', scale=1,
	           width=2e-3)
	
	ax2.plot(x_surf_epost, y_surf_zero, '-r*', ms=2, mew=2, fillstyle='none', color='magenta',
	         label='elastic post-seismic')
	ax2.plot(Ux_diff_post, Uy_diff_post, '-k*', ms=2, mew=2, fillstyle='none', color='blueviolet',
	         label='poroelastic post-seismic')
	ax2.quiver(x_surf_epost, y_surf_zero, scaleplot * Ux_surf_diff_post, Uy_diff_post, scale_units='xy', angles='xy',
	           scale=1, width=2e-3)
	
	plt.xlabel('Meters from trench')
	legend1 = ax1.legend(bbox_to_anchor=(1.05, 1), loc=2, borderaxespad=0.)
	for label in legend1.get_texts():
		label.set_fontsize('medium')
	legend2 = ax2.legend(bbox_to_anchor=(1.05, 1), loc=2, borderaxespad=0.)
	for label in legend2.get_texts():
		label.set_fontsize('medium')


def plot_sea_flux():
	if event == 'sub_EQ':
		
		fig = plt.figure()
		ax1 = fig.add_subplot(111)
		# ax2 = fig.add_subplot(312)
		# ax3 = fig.add_subplot(313)
		# ax1.set_title('Flux through seafloor')
		
		# Co-seismic displacement
		ax1.plot(dtt, sea_flux, '-bo', ms=5, mew=1, fillstyle='none', color='blue')
		# ax1.plot(dtt[0:50], sea_flux[0:50], '-b*', ms=2, mew=2, fillstyle='none', color = 'blue')
		ax1.plot(dtt[50:-1], sea_flux[50:-1], '-b*', ms=2, mew=2, fillstyle='none', color='red')
		# ax2.plot(dtt[0:50], sea_flux[0:50], '-b*', ms=2, mew=2, fillstyle='none', color = 'blue')
		# ax3.plot(dtt[50:-1], sea_flux[50:-1], '-b*', ms=2, mew=2, fillstyle='none', color = 'red')
		# ax1.set_ylabel('Darcy flux (meters/year)')
		# ax2.set_ylabel('Darcy flux (meters/sec)')
		# ax3.set_ylabel('Darcy flux (meters/year)')
		# ax1.set_xlabel('Years')
		
		ax1.set_yscale('log')
	# ax3.set_yscale('log')
	
	elif event == 'topo':
		ind = np.arange(7)  # the x locations for the groups
		# width = np.concatenate([[1],dt_percent[1::]])
		# months = np.concatenate([[-1],(np.cumsum(dt_percent[0:-1]) - np.sum(dt_percent[0:1]))])
		sixmo = 287
		twoyear = 303
		
		fig1 = plt.figure()
		fig2 = plt.figure()
		fig3 = plt.figure()
		ax1 = fig1.add_subplot(211)
		ax2 = fig1.add_subplot(212)
		ax3 = fig2.add_subplot(111)
		# ax4 = fig2.add_subplot(212)
		ax5 = fig3.add_subplot(111)
		
		ax1.plot(dtt, (3600 * 24 * 365) * sea_flux_topo, '-b', label='Topographic flow',
		         ms=2, mew=2, fillstyle='none', color='blue', linewidth=2.0)
		ax1.plot(dtt_subEQ, (3600 * 24 * 365) * sea_flux, '-r', label='Tectonic driven flow',
		         ms=2, mew=2, fillstyle='none', color='red', linewidth=2.0)
		ax1.set_xlim([-50, 50])
		ax1.set_yscale('log')
		ax1.legend(fontsize='small')
		
		ax2.plot(dtt[sub_cycle - 2:sub_cycle + 2], (3600 * 24 * 365) * sea_flux_topo[sub_cycle - 2:sub_cycle + 2],
		         '-b*', label='Topographic flow', ms=2, mew=2,
		         fillstyle='none', color='blue', linewidth=2.0)
		ax2.plot(dtt_subEQ[sub_cycle - 2:twoyear], (3600 * 24 * 365) * sea_flux[sub_cycle - 2:twoyear],
		         '-b*', label='Tectonic driven flow', ms=2, mew=2,
		         fillstyle='none', color='red', linewidth=2.0)
		ax2.set_xlim([-.1, 2])
		ax2.set_xlabel('Years after earthquake')
		ax2.set_yscale('log')
		ax2.legend(fontsize='small')
		
		ax3.plot(dtt_subEQ, cum_h20_sub, '-bo', label='Tectonic flow',
		         ms=2, mew=2, fillstyle='none', color='red', linewidth=2.0)
		ax3.plot(dtt_subEQ, cum_h20_topo, '-bo', label='Topographic flow',
		         ms=2, mew=2, fillstyle='none', color='blue', linewidth=2.0)
		ax3.plot(dtt_subEQ, cum_h20_both, '-bo', label='Total flow',
		         ms=2, mew=2, fillstyle='none', color='purple', linewidth=2.0)
		ax3.set_xlim([-50, 50])
		# ax3.set_xlabel('Years after earthquake')
		# ax3.set_ylabel('Cumulative SGD')
		ax3.legend(loc=2)
		
		print dtt.shape
		print percentage
		
		ax5.plot(dtt_subEQ, 100 * percentage, '-b.', color='blue', linewidth=3)
		ax5.set_xlim([-.2, 2])
		ax5.set_ylim([0, 100])
	# ax5.set_xlabel('Months after earthquake')
	# ax5.set_ylabel('Percent of SGD driven by tectonic processes')
	
	
	# percentagetext = np.round(percentage,1)
	# for i, v in enumerate(percentagetext[0:2]):
	#    ax5.text(i-.95, v+1 , str(v), color='darkblue', fontsize = '10')
	# for i, v in enumerate(percentagetext[-1::]):
	#    ax5.text(i + nummonths-2.5, v + 1.4, str(v), color='darkblue', fontsize = '10')
	
	
	
	else:
		fig = plt.figure()
		ax1 = fig.add_subplot(111)
		ax1.set_title('Flux through seafloor')
		# Co-seismic displacement
		ax1.plot(dtt, sea_flux, '-b*', ms=2, mew=2, fillstyle='none', color='blue')
	
	# fig3.set_size_inches(7, 4)
	# plt.xlim((0, 100))
	# plt.ylim((5e-8, 10))
	# fig2.savefig('figures/SGD_cumulative.png', dpi = 400)
	# fig1.savefig('figures/SGD_sea_flux.png', dpi = 400)
	fig3.savefig('figures/SGD_percent.png', dpi=400)  #


def plot_surf_post():
	fig_x = plt.figure()
	ax1 = fig_x.add_subplot(111, projection='3d')
	fig_y = plt.figure()
	ax2 = fig_y.add_subplot(111, projection='3d')
	fig_p = plt.figure()
	ax3 = fig_p.add_subplot(111, projection='3d')
	
	X, Y = np.meshgrid(dtt, x_surf)
	Z_x = Sol_surf_x_post
	Z_y = Sol_surf_y_post
	Z_p = Sol_surf_p_post
	
	surfx = ax1.plot_surface(X, Y, Z_x, cmap=cm.coolwarm, linewidth=0, antialiased=False)
	
	surfy = ax2.plot_surface(X, Y, Z_y, cmap=cm.coolwarm, linewidth=0, antialiased=False)
	
	surfp = ax3.plot_surface(X, Y, Z_p, cmap=cm.coolwarm, linewidth=0, antialiased=False)
	
	# ax1.zaxis.set_major_locator(LinearLocator(10))
	# ax1.zaxis.set_major_formatter(FormatStrFormatter('%.02f'))
	# plt.title('Surf X')
	
	# plt.xlabel('Meters from trench')
	# plt.ylabel('Meters from trench')
	
	fig_x.colorbar(surfx)
	fig_y.colorbar(surfy)
	fig_p.colorbar(surfp)


def plot_surf_all():
	fig_x = plt.figure()
	ax1 = fig_x.add_subplot(111, projection='3d')
	fig_y = plt.figure()
	ax2 = fig_y.add_subplot(111, projection='3d')
	fig_p = plt.figure()
	ax3 = fig_p.add_subplot(111, projection='3d')
	
	X, Y = np.meshgrid(dtt, x_surf)
	Z_x = Sol_surf_x
	Z_y = Sol_surf_y
	Z_p = Sol_surf_p
	
	surfx = ax1.plot_surface(X, Y, Z_x, cmap=cm.coolwarm, linewidth=0, antialiased=False)
	
	surfy = ax2.plot_surface(X, Y, Z_y, cmap=cm.coolwarm, linewidth=0, antialiased=False)
	
	surfp = ax3.plot_surface(X, Y, Z_p, cmap=cm.coolwarm, linewidth=0, antialiased=False)
	
	# ax1.zaxis.set_major_locator(LinearLocator(10))
	# ax1.zaxis.set_major_formatter(FormatStrFormatter('%.02f'))
	# plt.title('Surf X')
	
	# plt.xlabel('Meters from trench')
	# plt.ylabel('Meters from trench')
	
	fig_x.colorbar(surfx)
	fig_y.colorbar(surfy)
	fig_p.colorbar(surfp)


def plot_surf_early():
	fig_x = plt.figure()
	ax1 = fig_x.add_subplot(111, projection='3d')
	fig_y = plt.figure()
	ax2 = fig_y.add_subplot(111, projection='3d')
	fig_p = plt.figure()
	ax3 = fig_p.add_subplot(111, projection='3d')
	
	X, Y = np.meshgrid(dtt[0:endtime], x_surf)
	Z_x = Sol_surf_x_post[:, 0:endtime]
	Z_y = Sol_surf_y_post[:, 0:endtime]
	Z_p = Sol_surf_p_post[:, 0:endtime]
	
	surfx = ax1.plot_surface(X, Y, Z_x, cmap=cm.coolwarm, linewidth=0, antialiased=False)
	
	surfy = ax2.plot_surface(X, Y, Z_y, cmap=cm.coolwarm, linewidth=0, antialiased=False)
	
	surfp = ax3.plot_surface(X, Y, Z_p, cmap=cm.coolwarm, linewidth=0, antialiased=False)
	
	# ax1.zaxis.set_major_locator(LinearLocator(10))
	# ax1.zaxis.set_major_formatter(FormatStrFormatter('%.02f'))
	# plt.title('Surf X')
	
	# plt.xlabel('Meters from trench')
	# plt.ylabel('Meters from trench')
	
	fig_x.colorbar(surfx)
	fig_y.colorbar(surfy)
	fig_p.colorbar(surfp)


def plot_contour_post():
	fig_x = plt.figure()
	ax1 = fig_x.add_subplot(111)
	fig_y = plt.figure()
	ax2 = fig_y.add_subplot(111)
	fig_p = plt.figure()
	ax3 = fig_p.add_subplot(111)
	
	X, Y = np.meshgrid(dtt, x_surf)
	Z_x = Sol_surf_x_post
	Z_y = Sol_surf_y_post
	Z_p = Sol_surf_p_post
	
	contour_steps = 20
	levels_x = np.linspace(Z_x.min(), Z_x.max(), num=contour_steps, endpoint=True)
	levels_y = np.linspace(Z_y.min(), Z_y.max(), num=contour_steps, endpoint=True)
	levels_p = np.linspace(Z_p.min(), Z_p.max(), num=contour_steps, endpoint=True)
	
	cpxf = ax1.contourf(X, Y, Z_x, levels_x, cmap=cm.coolwarm)
	# cpx = ax1.contour(X, Y, Z_x, levels_x, linewidth=2, colors = 'black', linestyles = 'solid')
	
	cpyf = ax2.contourf(X, Y, Z_y, levels_y, cmap=cm.coolwarm)
	# cpy = ax2.contour(X, Y, Z_y, levels_y, linewidth=2, colors = 'black', linestyles = 'solid')
	
	cppf = ax3.contourf(X, Y, Z_p, levels_p, cmap=cm.coolwarm)
	# cpp = ax3.contour(X, Y, Z_p, levels_p, linewidth=2, colors = 'black', linestyles = 'solid')
	
	
	
	# ax1.zaxis.set_major_locator(LinearLocator(10))
	# ax1.zaxis.set_major_formatter(FormatStrFormatter('%.02f'))
	# plt.title('Surf X')
	
	# plt.xlabel('Meters from trench')
	# plt.ylabel('Meters from trench')
	
	fig_x.colorbar(cpxf)
	fig_y.colorbar(cpyf)
	fig_p.colorbar(cppf)


def plot_topo():
	wb_slip = pyxl.load_workbook('data/water_table.xlsx')
	topo_data = wb_slip['Sheet1']
	X_topo = np.array([[cell.value for cell in col] for col in topo_data['A2':'A21']])
	h_topo = np.array([[cell.value for cell in col] for col in topo_data['B2':'B21']])
	h_topo2 = 0.75 * np.array([[cell.value for cell in col] for col in topo_data['B2':'B21']])
	
	fig = plt.figure()
	ax1 = fig.add_subplot(111)
	ax1.set_title('Nicoya Peninsula Topography')
	ax1.plot(X_topo, h_topo, '-b*', ms=3, mew=2, fillstyle='none', color='green', linewidth=3)
	# ax1.plot(X_topo,h_topo2,'-b*', ms=3, mew=2, fillstyle='none', color = 'blue',linewidth = 3)
	
	plt.axis('scaled')
	plt.xlim((6e4, 25e4))
	plt.ylim((-1e4, 1e4))
	
	fig.set_size_inches(7, 4)
	
	fig.savefig('figures/topo_flow2.png', dpi=400)


# fig1.savefig('figures/SGD_sea_flux.png', dpi = 400)
# fig3.savefig('figures/SGD_percent.png', dpi=400)#


##########################################################################
######################### PLOTTING COMMANDS ##############################
##########################################################################


# Array points along the surface to be plotted as timeseries
# points = np.arange(60e3, 150e3, 5e5)
points = np.array([60,65,70,75,80,85,90,95,100,105,110,120,130,140,150 ])
# points = 1e3*np.array([63, 65, 80, 95, 110, 130])
#points = np.array([65, 95, 130])
endtime = 353

# plot_timeseries(points)
#plot_timeseries_head(points)
plot_cross_time()
# plot_surface_loop()
# plot_surface_p()
# plot_surface_e()
# plot_diff_surf()
# plot_sea_flux()
# plot_topo()


# plot_surf_post()
# plot_surf_all()
# plot_surf_early()
# plot_contour_post()



# plt.show()
