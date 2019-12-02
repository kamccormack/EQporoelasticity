"""
Kimmy McCormack
Functions used by 'model_run.py' to solve coupled set of
poroelastic equations in 3D subduction zone domains

Units are in meters and MPa
"""
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as pyxl
import os
import scipy as sp
import scipy.interpolate
import shutil
from dolfin import *
from tempfile import TemporaryFile
from pykrige.ok import OrdinaryKriging
import pykrige.kriging_tools as kt
import matplotlib.tri as tri
import time
import gc
import sys
sys.path.append('/EQporoelasticity/local_lib')
gc.collect()

"""Optimization options"""
parameters["reorder_meshes"] = False
parameters['form_compiler']['optimize'] = True
parameters["form_compiler"]["cpp_optimize"] = True
ffc_options = {"optimize": True, "eliminate_zeros": True,"precompute_basis_const": True, "precompute_ip_const": True}


"""Data interpolation functions"""
def slabkriging(XYZslab, u0load, xvec, yvec, **kwargs):
	"""
	Interpolates inverted fault slip onto the mesh by by kriging.

		INPUTS:
			XYZslab = coordinates of fault slip solution in the model cartesian space
			u0load = slip vector at each solution node
			xvec - vector of dof x locations
			yvec - vector of dof y locations

		OUPUTS:
			u03 - vector containing inverted slip at all dofs
	"""
	
	x_all, y_all = xvec[:], yvec[:]
	
	nonzero = np.where(np.abs(u0load[:, 0]) > 0.10)[0]
	skip = (slice(None, None, 3))
	
	uxslip_load, uyslip_load, uzslip_load = u0load[:, 0][nonzero][skip], \
	                                        u0load[:, 1][nonzero][skip], u0load[:, 2][nonzero][skip]
	xslip, yslip, zslip = XYZslab[:, 0][nonzero][skip], \
	                      XYZslab[:, 1][nonzero][skip], XYZslab[:, 2][nonzero][skip]
	
	# Interpolate data using pykrige
	OK_uxslip = OrdinaryKriging(xslip, yslip, uxslip_load,
	                            variogram_model='linear',
	                            variogram_parameters=[2e-4, 1e-2])
	uxslip, ss_ux = OK_uxslip.execute('points', x_all, y_all)
	OK_uyslip = OrdinaryKriging(xslip, yslip, uyslip_load,
	                            variogram_model='linear',
	                            variogram_parameters=[2e-4, 1e-2])
	uyslip, ss_uy = OK_uyslip.execute('points', x_all, y_all)
	OK_uzslip = OrdinaryKriging(xslip, yslip, uzslip_load,
	                            variogram_model='linear',
	                            variogram_parameters=[2e-4, 1e-2])
	uzslip, ss_uz = OK_uzslip.execute('points', x_all, y_all)
	u03 = np.concatenate((uxslip.reshape((uxslip.shape[0], 1)),
	                      uyslip.reshape((uxslip.shape[0], 1)),
	                      uzslip.reshape((uxslip.shape[0], 1))), axis=1)
	
	return u03

def surfkriging(xvec, yvec, zvec, dof_surf, **kwargs):
	"""Interpolates surface z value data by kriging.

		INPUTS:
			[x,y,z] = coordinates of degrees of freedom onto which the z value is mapped
			[dof_surf] = the degrees of freedom of the surface of the mesh

			OrdinaryKriging function takes variogram parameters as:
				linear - [slope, nugget]
				power - [scale, exponent, nugget]
			`   gaussian - [sill, range, nugget]
				spherical - [sill, range, nugget]
				exponential - [sill, range, nugget]
				hole-effect - [sill, range, nugget]

		OUPUTS:
			z_surf - surface height on all mesh nodes as a numpy array
	"""

	x_surf, y_surf, z_surf = xvec[dof_surf], yvec[dof_surf], zvec[dof_surf]
	x_all, y_all = xvec[:], yvec[:]

	"""PyKrige interpolation"""
	OK_surf = OrdinaryKriging(x_surf, y_surf, z_surf,
	                          variogram_model='gaussian', variogram_parameters=[5e5, 5e4, 1e2])
	z_surf, ss = OK_surf.execute('points', x_all, y_all)
	
	return z_surf.compressed()

def kappakriging(data_file, origin, theta, xvec, yvec, **kwargs):
	"""Interpolates surface permeability data by kriging. Take surface value to interpolate kappa with depth.

		INPUTS:
			[data_sheet] = excel sheet containing pointwise permeability data in (lon, lat, permaebility[m^2]) format.
					Default is to ignore the first row (containing headers).
			[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
			[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
					counter-clockwise from latlon to XY
			[coords_K] = coordinates of degrees of freedom onto which the permeability is mapped
			[dof_surf] = the degrees of freedom of the surface of the mesh

			OrdinaryKriging function takes variogram parameters as:
				linear - [slope, nugget]
				power - [scale, exponent, nugget]
			`   gaussian - [sill, range, nugget]
				spherical - [sill, range, nugget]
				exponential - [sill, range, nugget]

		OUPUTS:
			kappa [m^2] - the interpolated 3D permeabiltiy field in the form of a FEniCS Expression()
	"""

	x_all, y_all = xvec[:], yvec[:]
	wb = pyxl.load_workbook(data_file)
	well_data = wb['well_data']
	x_wells, y_wells = latlon2XY(well_data, origin, theta)
	permeability = np.array(
		[[well_data.cell(row=i, column=5).value] for i in range(2, well_data.max_row + 1)]).reshape(
		well_data.max_row - 1, )
	logk = np.log10(permeability)

	"""PyKrige interpolation"""
	# OK = OrdinaryKriging(x_wells, y_wells, logk, variogram_model = 'power', variogram_parameters = [1, 1, 2e4])
	OK = OrdinaryKriging(x_wells, y_wells, logk, variogram_model='exponential',
	                     variogram_parameters=[1.086, 1.28e4, 0.771])
	k_surf, ss = OK.execute('points', x_all, y_all)
	np.save("results/numpy_variables/k_surf", k_surf.compressed())

	return k_surf.compressed()

def latlon2XY(data_sheet, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = np.array([[data_sheet.cell(row=i, column=1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat = np.array([[data_sheet.cell(row=i, column=2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)

	return 1e3 * X, 1e3 * Y

class SlipExpression(Expression):
	""" Interpolates an arbitrary displacement field to the fault boundary of the mesh.

	INPUTS:
		[data] = an excel sheet with slip vectors recorded as
				(lon, lat, dip-slip [meters] , strike-slip [meters]) in the first four columns
		[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
		[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY
		[coords_V] = coordinates of degrees of freedom onto which the slip vectors are mapped
		[dof_slab] = the degrees of freedom of the fault boundary of the mesh
		[var_path] = the path of stored variables

	OUTPUTS:
		FEniCS Expression() containing the interpolated vector slip displacement on the fault boundary
	"""

	def __init__(self, data, origin, theta, coords_V, dof_slab, var_path, **kwargs):
		self.x_slab, self.y_slab, self.z_slab = coords_V[dof_slab, 0], coords_V[dof_slab, 1], coords_V[dof_slab, 2]

		X_slip, Y_slip = latlon2XY(data, origin, theta)

		Ux_slip = lon = np.array([[data.cell(row=i, column=3).value] for i in range(2, data.max_row + 1)]).reshape(
			data.max_row - 1, )
		Uy_slip = lon = np.array([[data.cell(row=i, column=4).value] for i in range(2, data.max_row + 1)]).reshape(
			data.max_row - 1, )

		# Interpolate data using pykrige
		OK_uxslip = OrdinaryKriging(X_slip, Y_slip, Ux_slip, variogram_model='linear', \
		                            variogram_parameters=[2e-4, 1e-2])
		self.uxslip, ss_ux = OK_uxslip.execute('points', self.x_slab, self.y_slab)

		OK_uyslip = OrdinaryKriging(X_slip, Y_slip, Uy_slip, variogram_model='linear', \
		                            variogram_parameters=[2e-4, 1e-2])
		self.uyslip, ss_ux = OK_uyslip.execute('points', self.x_slab, self.y_slab)
		np.save(var_path + 'uxslip', self.uxslip.compressed())
		np.save(var_path + 'uyslip', self.uyslip.compressed())

	def U_X_interp(self, X, Y):
		indx = (np.abs(np.sqrt(pow((self.x_slab[:] - X), 2) + pow((self.y_slab[:] - Y), 2)))).argmin()
		return self.uxslip[indx]

	def U_Y_interp(self, X, Y):
		indy = (np.abs(np.sqrt(pow((self.x_slab[:] - X), 2) + pow((self.y_slab[:] - Y), 2)))).argmin()
		return self.uyslip[indy]

	def eval(self, values, x, **kwargs):
		values[0] = self.U_X_interp(x[0], x[1])
		values[1] = self.U_Y_interp(x[0], x[1])
		values[2] = 0.0

	def value_shape(self):
		return (3,)
	
		
"""FEM solving functions"""
class Poroelasticity3DSynthetic:
	def __init__(self, data_file, origin, theta, ndim, mesh, plot_figs, event, permeability,
	             SSE_days, sub_cycle_years, sigma_bx, xcenter, sigma_by, ycenter,
	             SS_migratedip, SS_migratestrike, u0_EQdip, u0_EQstrike, u0_SSEdip,
	             u0_SSEstrike, u0_subdip, u0_substrike, surface_k, ocean_k):

		self.start = time.time()
		self.data_file = data_file
		self.origin = origin
		self.theta = theta
		self.mesh = mesh
		self.plot_figs = plot_figs
		self.event = event
		self.permeability = permeability
		self.SSE_days = SSE_days
		self.sub_cycle_years = sub_cycle_years
		self.sigma_bx = sigma_bx
		self.xcenter = xcenter
		self.sigma_by = sigma_by
		self.ycenter = ycenter
		self.SS_migratex = SS_migratedip
		self.SS_migratey = SS_migratestrike
		self.u0_EQx = u0_EQdip
		self.u0_EQy = u0_EQstrike
		self.u0d_SSE = u0_SSEdip
		self.u0s_SSE = u0_SSEstrike
		self.u0_subx = u0_subdip
		self.u0_suby = u0_substrike
		self.surface_k = surface_k
		self.ocean_k = ocean_k

		self.solver = LinearSolver('mumps')
		self.prm = self.solver.parameters
		self.V = VectorElement('Lagrange', self.mesh.ufl_cell(), 2)
		self.Q = FiniteElement('Lagrange', self.mesh.ufl_cell(), 1)
		self.ME = FunctionSpace(self.mesh, MixedElement([self.Q, self.V]))
		self.Vspace = FunctionSpace(self.mesh, self.V)
		self.Qspace = FunctionSpace(self.mesh, self.Q)
		self.Q_gspace = VectorFunctionSpace(self.mesh, "Lagrange", 1)
		self.Wspace = TensorFunctionSpace(self.mesh, "Lagrange", 1)

		self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
		self.Kspace = FunctionSpace(self.mesh, self.Kelement)

		(self.p, self.u) = TrialFunctions(self.ME)
		(self.q, self.v) = TestFunctions(self.ME)
		self.ds = Measure("ds", domain=self.mesh)
		self.dx = Measure("dx", domain=self.mesh)#
		self.n = FacetNormal(self.mesh)  # normal vector
		self.dim = self.ME.dim()
		self.pdim = self.ME.sub(0).dim()
		self.udim = self.ME.sub(1).dim()
		self.geom_dim = ndim + 1

		self.results_path = 'results/numpy_results/'
		self.var_path = 'results/numpy_variables/'
		self.paraviewpath = 'results/paraview/'
		if not os.path.exists(self.paraviewpath):
			os.makedirs(self.paraviewpath)

		self.extract_coords()
		print("Coordinates extracted", (time.time() - self.start) / 60.0, "minutes")
		self.k_interpolation()
		print("Permeability interpolated", (time.time() - self.start) / 60.0, "minutes")
		self.timestepping()
		
		"""POROELASTIC PARAMETERS"""
		self.delta = 1e-6  # Dirichlet control regularization on bottom boundary
		self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
		self.B = .6  # Skempton's coefficient
		self.nuu = .38  # undrained poissons ratio
		self.nu = 0.27  # Poisson's ratio
		x = SpatialCoordinate(self.mesh)
		self.E = interpolate(347 * pow(-x[2], 0.516) + 5.324e4, self.Kspace)
		self.mu = interpolate(self.E / (2.0*(1.0 + self.nu)), self.Kspace)  # shear modulus
		self.lmbda = interpolate(self.E*self.nu / ((1.0 + self.nu)*(1.0 - 2.0*self.nu)), self.Kspace)  # Lame's parameter
		self.alpha = 3 * (self.nuu - self.nu) / (self.B * (1 - 2 * self.nu) * (1 + self.nuu))  # Biot's coefficient
		self.Se = interpolate((9*(self.nuu-self.nu)*(1-(2*self.nuu)))/
		                      (2*self.mu*pow(self.B,2)*(1-2*self.nu)*pow((1+self.nuu),2)), self.Kspace)  # mass specific storage
		self.kstick = 8e-5
		self.stick = Expression('1/(1+exp(-kstick*(1.1e5-x[0])))',
		                        kstick=self.kstick, degree=1)
		self.d = self.u.geometric_dimension()  # number of space dimensions
		self.I = Identity(self.d)
		self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
		self.year = 60 * 60 * 24 * 365

		"""INITIAL AND BOUNDARY CONDITIONS"""
		p0d = 0  # initial pressure
		self.w_init = Expression(('p0d', '0', '0', '0'), p0d=p0d, degree=1)
		self.w0 = Function(self.ME)
		self.w0.interpolate(self.w_init)
		self.ub0 = Expression(('0', '0', '0'), degree=1)
		self.ub = Expression(('1', '1', '1'), degree=1)
		self.pb0 = Expression('1', degree=1)
		self.bc1 = DirichletBC(self.ME.sub(1), self.ub0,  1)  # ocean surface
		self.bc2 = DirichletBC(self.ME.sub(1), self.ub,  2)  # land surface
		self.bc3 = DirichletBC(self.ME.sub(1), self.ub0,  3)  # subducting slab
		self.bc4 = DirichletBC(self.ME.sub(1), self.ub0,  4)  # ocean wedge
		self.bc5 = DirichletBC(self.ME.sub(1), self.ub0,  5)  # north side
		self.bc6 = DirichletBC(self.ME.sub(1), self.ub0,  6)  # back side (no slip)
		self.bc7 = DirichletBC(self.ME.sub(1), self.ub0,  7)  # south side
		self.bc8 = DirichletBC(self.ME.sub(1), self.ub0,  8)  # mantle
		self.bcp1 = DirichletBC(self.ME.sub(0), ('0'),  1)  # ocean surface (free flow condition)
		self.bcp2 = DirichletBC(self.ME.sub(0), ('1'),  2)  # land surface

		self.bcs = [self.bc6, self.bc7, self.bcp1]  # These are the BC's that are actually applied
		self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case

		if self.event == 'SSE':
			self.u_addx = self.u0_SSEx
			self.u_addy = self.u0_SSEy
		elif self.event == 'sub':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby
		elif self.event == 'sub_EQ':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby

		"""SET UP TO SAVE SOLUTION"""
		self.Sol_all = np.empty((self.dim, self.nsteps))  # matrix to save entire solution
		self.Sol_surf = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
		self.Sol_gps = np.empty((self.x_gps.shape[0] * self.geom_dim, self.nsteps))  # matrix to save surface solution
		elf.sea_flux = np.empty((self.ocean_dof_pgrad.shape[0], self.nsteps))  # ocean bottom solution
		self.flow_velocity = np.empty((self.pdim * ndim, self.nsteps))  # ocean bottom solution
		self.vol_strain = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
		self.sea_flux_total = np.empty(self.nsteps)  # matrix to save surface post-seismic solution
		print("Solution matrices created", (time.time() - self.start) / 60.0, "minutes")

		pfilename = self.paraviewpath + "pressure3D_syn.pvd"
		p_velfilename = self.paraviewpath + "pressure_vel3D_syn.pvd"
		ufilename = self.paraviewpath + "def3D_syn.pvd"
		sigfilename = self.paraviewpath + "stress3D_syn.pvd"
	
		self.pfile = File(pfilename)
		self.p_velfile = File(p_velfilename)
		self.ufile = File(ufilename)
		self.sigfile = File(sigfilename)

		"""SET UP, RUN, SOLVE and SAVE"""
		dt = self.dtt_comsum[0]
		self.assemble_system(dt)
		print("System assembled", (time.time() - self.start) / 60.0, "minutes")
		self.solve_system()
		self.save_solution()

	def time_stepping(self):
		if self.event == 'EQ':
			"""15 min for 1/2 day,
			4 hrs for 10 days,
			12 hours for 30 days,
			2 days for 2 months,
			per month for 10 years
			"""
			self.dtt = [60, 240, 720, 288e1, 2 * 432e2]   # in minutes
			self.ntime = [24, 60, 60, 30, 60]  # , 50]
		elif self.event == 'SSE':
			"""
			12 hrs for SSE days,
			12 hrs for total time-SSE_days,
			2 months for 10 years
			"""
			self.dtt = [288e1 / 2, 432e2 * 2]
			self.ntime = [self.SSE_days / 1, (self.days - self.SSE_days) / 60]  # SSE_days, 10 years
		elif self.event == 'sub_EQ':
			"""
			1/year for 50 years followed by 'EQ' scheme
			"""
			self.dtt = [60 * 24 * 365, 15, 240, 288e1, 432e2, 60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years, 96, 120, 15, 122, 40]
		elif self.event == 'sub':  # 1/year for 50 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		elif self.event == 'topo':  # 1/year for 100 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		self.dtt = [i * 60 for i in self.dtt]  # in seconds
		self.dtt_repeated = np.repeat(self.dtt, self.ntime)
		self.dtt_comsum = np.cumsum(self.dtt_repeated)
		self.dt = self.dtt_comsum[0]
		self.nsteps = self.dtt_comsum.size
	
	def k_interpolation(self):
		log_kr = -18.0
		alpha_k = 0.6
		bctest1 = DirichletBC(self.Kspace, (1.0), 1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1.0), 2)  # land surface
		ktest = Function(self.Kspace)
		bctest1.apply(ktest)
		bctest2.apply(ktest)
		self.dof_surfK = np.where(ktest.vector().array() == 1.0)[0]
		
		if self.permeability == 'mapped':
			
			"""interpolate surface permeability onto all nodes"""
			kappa_surf = kappakriging(self.data_file,
			                          self.origin, self.theta,
			                          self.xvec, self.yvec)
			z_surf = surfkriging(self.xvec, self.yvec, self.zvec, self.dof_surfK)
			
			oceandof = np.where(z_surf < -10.)[0]
			kappa_surf[oceandof] = self.ocean_k
			
			self.kappa = Function(self.Kspace)
			self.kappa_vec = pow(10, (log_kr +
			                          (((kappa_surf) - log_kr) *
			                           pow((1 - (1e-3 * 0.25 * (self.zvec - z_surf)))
			                               , -alpha_k))))
			self.kappa.dat.data[:] = self.kappa_vec
		
		elif self.permeability == 'constant':
			"""interpolate surface permeability onto all nodes"""
			kappa_surf = Function(self.Kspace)
			kappa_surf_vec = pow(10, self.surface_k)
			kappa_surf.dat.data[:] = kappa_surf_vec
			z_surf = surfkriging(self.xvec, self.yvec, self.zvec, self.dof_surfK)
			oceandof = np.where(z_surf < -10.)[0]
			kappa_surf.dat.data[oceandof] = self.ocean_k
	
			self.kappa = Function(self.Kspace)
			self.kappa_vec = pow(10, (log_kr +
			                          (((kappa_surf) - log_kr) *
			                           pow((1 - (1e-3 * 0.25 * (self.zvec - z_surf)))
			                               , -alpha_k))))
			self.kappa.dat.data[:] = self.kappa_vec
			
	def strain(self, w):  # strain = 1/2 (grad u + grad u^T)
		return sym(nabla_grad(w))

	def sigma(self, w):  # stress = 2 mu strain + lambda tr(strain) I
		mu = interpolate(self.mu, self.Kspace)
		lmbda = interpolate(self.lmbda, self.Kspace)
		return 2.0 * mu * self.strain(w) + lmbda * div(w) * self.I
	
	def assemble_system(self, dt, assembleRHS=True):
		
		if assembleRHS:
			
			self.a_Mass = self.Se * self.p * self.q * dx  # Mass matrix
			self.a_Div = self.alpha*nabla_div(self.u) * self.q * dx  # Divergence matrix
			self.a_Grad = -self.alpha * self.p * (nabla_div(self.v)) * dx  # Gradient
			self.a_E = inner(self.sigma(self.u), nabla_grad(self.v)) * dx # Elasticity
			self.a_Boundary = (1 / self.delta) * dot(self.u, self.n) * dot(self.v, self.n)\
			                          * self.ds(self.slab) + (1 / self.delta) * inner(self.T * self.u, self.T * self.v) * self.ds(self.slab)
	
	
		self.a_K = dt * (self.kappa / self.muf) * inner(nabla_grad(self.p),
		                                                         nabla_grad(self.q)) * dx  # Stiffness matrix
		
		print("all LHS matrices assembled", (time.time() - self.start) / 60.0, "minutes")
		
		self.A = assemble(self.a_Mass + self.a_Div + self.a_K +
		                  self.a_Grad + self.a_E + self.a_Boundary, mat_type='aij')
		
		print("LHS assembled", (time.time() - self.start) / 60.0, "minutes")
		
		if assembleRHS:
			self.L = assemble(self.a_Mass + self.a_Div)
			print("RHS assembled", (time.time() - self.start) / 60.0, "minutes")

	def extract_coords(self):

		if not os.path.exists(self.var_path):
			os.makedirs(self.var_path)

		self.x_all, self.y_all, self.z_all =  SpatialCoordinate(self.mesh)

		xf = Function(self.Qspace)
		yf = Function(self.Qspace)
		zf = Function(self.Qspace)
		xm, ym, zm = SpatialCoordinate(self.mesh)
		
		self.xvec = xf.interpolate(xm).dat.data_ro
		self.yvec = yf.interpolate(ym).dat.data_ro
		self.zvec = zf.interpolate(zm).dat.data_ro

		bctest0 = DirichletBC(self.Qspace, (1), 1)  # ocean surface
		bctest1 = DirichletBC(self.Qspace, (2), 2)  # top surface

		ptest = Function(self.Qspace)
		bctest0.apply(ptest.vector())
		bctest1.apply(ptest.vector())
		self.ocean_dofs = np.where(ptest.vector().array() == 1)[0]
		self.surface_dofs = np.where(ptest.vector().array() == 2)[0]

		self.size_gps = self.x_gps.shape[0]
		size_w_gps = self.size_gps
		indices_gps = np.empty((self.size_gps, ))
		for j in range(0, self.size_gps):
			indice_gps = (np.abs(np.sqrt(pow((self.xvec[:] - self.x_gps[j]), 2) + pow((self.yvec[:] - self.y_gps[j]), 2)))).argmin()
			indices_gps[j] = indice_gps
			self.GPS_dofs = indices_gps

		np.save(self.var_path + "x_all_3D", self.xvec)
		np.save(self.var_path + "y_all_3D", self.yvec)
		np.save(self.var_path + "z_all_3D", self.zvec)
		np.save(self.var_path + "GPS_dofs_3D", self.GPS_dofs)
		np.save(self.var_path + "surface_dofs_3D", self.surface_dofs)
		np.save(self.var_path + "ocean_dofs_3D", self.ocean_dofs)

	def plot_initial_cond(self):  # plot the initial conditions
		plot(self.w0.sub(1), mode="glyphs", key="def")  # , rescale = False)
		plot(self.w0.sub(0), mode='color', key="pressure")  # , rescale = False)

	def solve_system(self):
		count = 0
		if self.event == 'EQ':
			"""For an EQ, this can be outside the time loop
			since we only use it for the first time step"""
			u0slab = Expression(('u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
			                    'u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
			                    '0'), u0d=self.u0_EQx, u0s=self.u0_EQy, xcenter=self.xcenter, ycenter=self.ycenter,
			                    sigmax=self.sigma_bx, sigmay=self.sigma_by, degree=1)
			u0slabf = Function(self.ME.sub(1))
			u0slabf.interpolate(u0slab)
			bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
				self.slab)
			with assemble(bndform).dat.vec as vec:
				bnd_vec = vec
			
		for i in range(self.nsteps):
			
			if i == 0:
				dt = self.dtt_comsum[i]
				self.sol = Function(self.ME)
			else:
				dt = self.dtt_comsum[i] - self.dtt_comsum[i - 1]
			if i > 1:
				dt_old = self.dtt_comsum[i - 1] - self.dtt_comsum[i - 2]
				
			if i > 1 and dt != dt_old:  # Rebuild Left hand side when 'dt' changes
				print("re-assembling LHS", (time.time() - self.start) / 60.0, "minutes")
				self.assemble_system(dt, assembleRHS=False)
				self.prm['reuse_factorization'] = True  # Saves Factorization of LHS
			
			if i > 0:
				"""Add slip to the boundary for a SSE
				and move center of slip along boundary"""
				if self.event == 'SSE' and count <= self.SSE_days:
					self.u0d_SSE += self.u_addx
					self.u0s_SSE += self.u_addy
					self.xcenter += self.SS_migratex
					self.ycenter += self.SS_migratey
				elif self.event == 'sub':
					self.u0_subx += self.u_addx
					self.u0_suby += self.u_addy
				elif self.event == 'sub_EQ':
					self.u0_subx += self.u_addx * (dt / self.year)
					self.u0_suby += self.u_addy * (dt / self.year)

			"""Slip boundary Expression"""
			if self.event == 'SSE':
				u0slab = Expression((
					'u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
					'u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
					'0'), u0d=self.u0d_SSE, u0s=self.u0s_SSE, xcenter=self.xcenter,
					ycenter=self.ycenter, sigmax=self.sigma_bx, sigmay=self.sigma_by,
					degree=1)
				u0slabf = Function(self.ME.sub(1))
				u0slabf.interpolate(u0slab)
				bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
					self.slab)
				with assemble(bndform).dat.vec as vec:
					bnd_vec = vec
			elif self.event == 'sub':
				u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                    u0_suby=self.u0_suby, stick=self.stick, degree=1)
				u0slabf = Function(self.ME.sub(1))
				u0slabf.interpolate(u0slab)
				bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
					self.slab)
				with assemble(bndform).dat.vec as vec:
					bnd_vec = vec
			elif self.event == 'sub_EQ' and i < self.sub_cycle_years:
				u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                    u0_suby=self.u0_suby, stick=self.stick, degree=1)
				u0slabf = Function(self.ME.sub(1))
				u0slabf.interpolate(u0slab)
				bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
					self.slab)
				with assemble(bndform).dat.vec as vec:
					bnd_vec = vec
			elif self.event == 'sub_EQ' and i >= self.sub_cycle_years:
				u0slab = Expression(('u0_subx*stick + u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + '
				                     '(pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
				                     'u0_suby*stick + u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) '
				                     '+ (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
				                     '0'), u0d=self.u0_subx, u0s=self.u0_suby, xcenter=self.xcenter,
				                    ycenter=self.ycenter, sigmax=self.sigma_bx, sigmay=self.Tsigma_by, degree=1)
				u0slabf = Function(self.ME.sub(1))
				u0slabf.interpolate(u0slab)
				bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
					self.slab)
				with assemble(bndform).dat.vec as vec:
					bnd_vec = vec
		
			"""move the RHS and slip BC to PETSc to multiply in solution from
			# previous timestep and add in slip boundary term"""
			RHS_mat = self.L.M.handle
			b = RHS_mat.getVecs()[0]
			with assemble(self.w0).dat.vec as vec:
				w0vec = vec
			RHS_mat.mult(w0vec, b)

			"""transfer back to Function through a numpy array"""
			b += bnd_vec
			b_func = Function(self.ME)
			b_func.vector().set_local(b.array)
			
			print("Boundary condition set", (time.time() - self.start) / 60.0, "minutes")
			
			L = b_func
			A = self.A
			w = Function(self.ME)
			
			params = {
                            "mat_solver_package": "mumps",
			          }
			
			solve(A, w, b_func, bcs=self.bcs, solver_parameters=params)
			print("System solved", (time.time() - self.start) / 60.0, "minutes")
			p, u = w.split()
			self.domain_flux = (-self.kappa / self.muf) * grad(p)  # element-wise flux vector
			self.bound_flux = assemble(inner(self.domain_flux, self.n) * self.ds(1))
			self.bound_flux2 = assemble(inner(self.domain_flux, self.n) * self.ds(2))
			print("outward flux through seafloor: ", self.bound_flux)
			print("outward flux through land surface: ", self.bound_flux2)
			
			self.q_velocity = Function(self.Q_gspace)
			self.q_velocity.assign(project(self.domain_flux, self.Q_gspace))
			print("Computed velocity field", (time.time() - self.start) / 60.0, "minutes")
			self.v_strain = Function(self.Qspace)
			self.v_strain.assign(project(nabla_div(u), self.Qspace))
			self.stress = Function(self.Wspace)
			self.stress.assign(project(self.sigma(u), self.Wspace))

			"""SAVE SOLUTION"""
			p.rename('p', 'p')
			u.rename('u', 'u')
			q_velocity.rename('q', 'q')
			self.pfile.write(p)
			self.ufile.write(u)
			self.qfile.write(q)
			#self.sigfile.write(self.stress)
			self.sea_flux_total[count] = self.bound_flux
			self.sea_flux[:, count] = self.q_velocity.vector()[self.ocean_dof_pgrad]
			self.flow_velocity[:, count] = self.q_velocity.vector()
			self.vol_strain[:, count] = self.v_strain.vector()
			self.Sol_surf[:, count] = self.w.vector()[self.surface_dofs]
			self.Sol_all[:, count] = self.w.vector()

			self.w0 = w
			count += 1
			print("Timestep:", count)

	def save_solution(self):

		if not os.path.exists(self.results_path):
			os.makedirs(self.results_path)
		if self.permeability == 'mapped':
			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_mapped.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_mapped.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_EQ_mapped.npy", self.Sol_all)
			elif self.event == 'EQ':
				np.save(self.results_path + "Sol_surf_3D_EQ_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_EQ_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_EQ_mapped.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_EQ_mapped.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_EQ_mapped.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_EQ_mapped.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_EQ_mapped.npy", self.sea_flux_total)
		elif self.permeability == 'constant':
			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_constant.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_constant.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_EQ_constant.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_sub_EQ_constant.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_sub_EQ_constant.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_sub_EQ_constant.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_sub_EQ_constant.npy", self.sea_flux_total)
			elif self.event == 'EQ':
				np.save(self.results_path + "Sol_surf_3D_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_EQ_constant.npy", self.Sol_all)
				np.save(self.results_path + "vol_strain_3D_EQ_constant.npy", self.vol_strain)
				np.save(self.results_path + "flow_velocity_3D_EQ_constant.npy", self.flow_velocity)
				np.save(self.results_path + "sea_flux_3D_EQ_constant.npy", self.sea_flux)
				np.save(self.results_path + "sea_flux_total_3D_EQ_constant.npy", self.sea_flux_total)

class Poroelasticity3D:
	def __init__(self, data_file, origin, theta, ndim, mesh, plot_figs, EQtype, event, permeability,
	             sub_cycle_years, u0_subdip, u0_substrike, surface_k, ocean_k, **kwargs):

		self.start = time.time()
		self.data_file = data_file
		self.origin = origin
		self.theta = theta
		self.mesh = mesh
		self.plot_figs = plot_figs
		self.EQtype = EQtype
		self.event = event
		self.permeability = permeability
		self.sub_cycle_years = sub_cycle_years
		self.u0_subx = u0_subdip
		self.u0_suby = u0_substrike
		self.surface_k = surface_k
		self.ocean_k = ocean_k
	
		self.V = VectorElement('Lagrange', self.mesh.ufl_cell(), 2)
		self.Q = FiniteElement('Lagrange', self.mesh.ufl_cell(), 1)
		self.ME = FunctionSpace(self.mesh, MixedElement([self.Q, self.V]))
		self.Vspace = FunctionSpace(self.mesh, self.V)
		self.Qspace = FunctionSpace(self.mesh, self.Q)
		self.Q_gspace = VectorFunctionSpace(self.mesh, "Lagrange", 1)
		self.Wspace = TensorFunctionSpace(self.mesh, "Lagrange", 1)
		self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
		self.Kspace = FunctionSpace(self.mesh, self.Kelement)
		
		(self.p, self.u) = TrialFunctions(self.ME)
		(self.q, self.v) = TestFunctions(self.ME)
		self.ds = Measure("ds", domain=self.mesh)
		self.dx = Measure("dx", domain=self.mesh)#
		self.n = FacetNormal(self.mesh)  # normal vector
		self.dim = self.ME.dim()
		self.pdim = self.ME.sub(0).dim()
		self.udim = self.ME.sub(1).dim()
		self.geom_dim = ndim + 1


		"""GPS STATION LOCATIONS"""
		wb = pyxl.load_workbook(self.data_file)
		gps_data = wb['GPS_data']
		
		self.x_gps, self.y_gps = latlon2XY(gps_data, self.origin, self.theta)
		self.z_gps = np.zeros(len(self.x_gps))
		
		self.results_path = 'results/numpy_results/'
		self.var_path = 'results/numpy_variables/'
		self.paraviewpath = 'results/paraview/'
		self.inverse_results = '/fenics/shared/inverse_codes/results_deterministic/numpy_results/'
		self.inverse_var = '/fenics/shared/inverse_codes/results_deterministic/numpy_variables/'
	
		if not os.path.exists(self.results_path):
			os.makedirs(self.results_path)
		if not os.path.exists(self.var_path):
			os.makedirs(self.var_path)
		if not os.path.exists(self.paraviewpath):
			os.makedirs(self.paraviewpath)

		self.extract_coords()
		print("Coordinates extracted", (time.time() - self.start) / 60.0, "minutes")
		self.k_interpolation()
		print("Permeability interpolated", (time.time() - self.start) / 60.0, "minutes")
		self.timestepping()

		"""POROELASTIC PARAMETERS"""
		self.delta = 1e-6  # Dirichlet control regularization on bottom boundary
		self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
		self.B = .6  # Skempton's coefficient
		self.nuu = .35  # undrained poissons ratio
		self.nu = 0.25  # Poisson's ratio
		x = SpatialCoordinate(self.mesh)
		self.E = interpolate(500 * pow(-x[2], 0.516) + 2e4, self.Kspace)
		#self.E = interpolate(347 * pow(-x[2], 0.516) + 5.324e4, self.Kspace)
		self.mu = interpolate(self.E / (2.0*(1.0 + self.nu)), self.Kspace)  # shear modulus
		self.lmbda = interpolate(self.E*self.nu / ((1.0 + self.nu)*(1.0 - 2.0*self.nu)), self.Kspace)  # Lame's parameter
		self.alpha = 3 * (self.nuu - self.nu) / (self.B * (1 - 2 * self.nu) * (1 + self.nuu))  # Biot's coefficient
		self.Se = interpolate((9*(self.nuu-self.nu)*(1-(2*self.nuu)))/
		                      (2*self.mu*pow(self.B,2)*(1-2*self.nu)*pow((1+self.nuu),2)), self.Kspace)  # mass specific storage

		self.kstick = 8e-5
		self.stick = Expression('1/(1+exp(-kstick*(1.1e5-x[0])))', kstick=self.kstick, degree=1)
		self.d = self.u.geometric_dimension()  # number of space dimensions
		self.I = Identity(self.d)
		self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
		self.year = 60 * 60 * 24 * 365

		""" INITIAL AND BOUNDARY CONDITIONS """
		####### Initial conditions: (('p0','ux0','uy0')) ##########
		p0d = 0  # initial pressure
		self.w_init = Expression(('p0d', '0', '0', '0'), p0d=p0d, degree=1)
		self.sol_old = Function(self.ME)
		self.w0.interpolate(self.w_init)
		self.ub0 = Expression(('0', '0', '0'), degree=1)
		self.ub = Expression(('1', '1', '1'), degree=1)
		self.pb0 = Expression('1', degree=1)
		self.bc1 = DirichletBC(self.ME.sub(1), self.ub0,  1)  # ocean surface
		self.bc2 = DirichletBC(self.ME.sub(1), self.ub,  2)  # land surface
		self.bc3 = DirichletBC(self.ME.sub(1), self.ub0,  3)  # subducting slab
		self.bc5 = DirichletBC(self.ME.sub(1), self.ub0,  5)  # north side
		self.bc6 = DirichletBC(self.ME.sub(1), self.ub0,  6)  # back side (no slip)
		self.bc7 = DirichletBC(self.ME.sub(1), self.ub0,  7)  # south side
		self.bc8 = DirichletBC(self.ME.sub(1), self.ub0,  8)  # mantle

		self.bcp1 = DirichletBC(self.ME.sub(0), ('0.'),  1, method="geometric")  # ocean surface (free flow condition)
		self.bcp2 = DirichletBC(self.ME.sub(0), ('1'),  2)  # land surface

		self.bcs = [self.bc6, self.bc7, self.bcp1]  # These are the BC's that are actually applied

		self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case

		if self.event == 'SSE':
			self.u_addx = self.u0_SSEx
			self.u_addy = self.u0_SSEy
		elif self.event == 'sub':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby
		elif self.event == 'sub_EQ':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby

		"""SET UP TO SAVE SOLUTION"""
		self.Sol_all = np.empty((self.dim, self.nsteps))  # matrix to save entire solution
		self.Sol_surf = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
		self.sea_flux = np.empty((self.ocean_dof_pgrad.shape[0], self.nsteps))  # ocean bottom solution
		self.flow_velocity = np.empty((self.pdim * ndim, self.nsteps))  # ocean bottom solution
		self.vol_strain = np.empty((self.size_p, self.nsteps))  # ocean bottom solution
		self.sea_flux_total = np.empty(self.nsteps)  # matrix to save surface post-seismic solution
		self.Sol_surf_p = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
		self.Sol_gps = np.empty((self.x_gps.shape[0], ndim, self.nsteps))  # matrix to save surface solution

		pfilename = self.paraviewpath + "pressure3D.pvd"
		p_velfilename = self.paraviewpath + "flux0_comp3D.pvd"
		ufilename = self.paraviewpath + "def3D.pvd"
		sigfilename = self.paraviewpath + "stress3D.pvd"

		self.pfile = File(pfilename)
		self.qfile = File(p_velfilename)
		self.ufile = File(ufilename)
		self.sigfile = File(sigfilename)
		print("Solution matrices created", (time.time() - self.start) / 60.0, "minutes")

		"""SET UP, RUN, SOLVE and SAVE"""
		self.slip_interpolation()
		print("Slip data interpolated", (time.time() - self.start) / 60.0, "minutes")
		dt = self.dtt_comsum[0]
		self.assemble_system(dt)
		print("System assembled", (time.time() - self.start) / 60.0, "minutes")
		self.solve_system()
		self.save_solution()
	
	def extract_coords(self):
		
		if not os.path.exists(self.var_path):
			os.makedirs(self.var_path)

		self.x_all, self.y_all, self.z_all = SpatialCoordinate(self.mesh)
		
		xf = Function(self.Qspace)
		yf = Function(self.Qspace)
		zf = Function(self.Qspace)
		xm, ym, zm = SpatialCoordinate(self.mesh)
		
		self.xvec = xf.interpolate(xm).dat.data_ro
		self.yvec = yf.interpolate(ym).dat.data_ro
		self.zvec = zf.interpolate(zm).dat.data_ro
		
		bctest0 = DirichletBC(self.Qspace, (1), 1)  # ocean surface
		bctest1 = DirichletBC(self.Qspace, (2), 2)  # top surface
		bctest2 = DirichletBC(self.Qspace, (3), 3)  # slab surface

		ptest = Function(self.Qspace)
		bctest0.apply(ptest.vector())
		bctest1.apply(ptest.vector())
		bctest2.apply(ptest.vector())
		self.ocean_dofs = np.where(ptest.vector().array() == 1)[0]
		self.surface_dofs = np.where(ptest.vector().array() == 2)[0]
		self.slab_dofs = np.where(ptest.vector().array() == 3)[0]

		self.size_gps = self.x_gps.shape[0]
		size_w_gps =  self.size_gps
		indices_gps = np.empty((self.size_gps, ))
		for j in range(0, self.size_gps):
			indice_gps = (np.abs(np.sqrt(pow((self.xvec[:] - self.x_gps[j]), 2)
			                             + pow((self.yvec[:] - self.y_gps[j]), 2)
			                             + pow((self.zvec[:] - 0.), 2)))).argmin()
			indices_gps[j] = indice_gps
		self.GPS_dofs = indices_gps.astype('int')
		
		np.save(self.var_path + "x_all_3D", self.xvec)
		np.save(self.var_path + "y_all_3D", self.yvec)
		np.save(self.var_path + "z_all_3D", self.zvec)
		np.save(self.var_path + "GPS_dofs_3D", self.GPS_dofs)
		np.save(self.var_path + "surface_dofs_3D", self.surface_dofs)
		np.save(self.var_path + "ocean_dofs_3D", self.ocean_dofs)
	
	def time_stepping(self):
		if self.event == 'EQ':
			"""15 min for 1/2 day,
			4 hrs for 10 days,
			12 hours for 30 days,
			2 days for 2 months,
			per month for 10 years
			"""
			self.dtt = [60, 240, 720, 288e1, 2 * 432e2]   # in minutes
			self.ntime = [24, 60, 60, 30, 60]  # , 50]
		elif self.event == 'SSE':
			"""
			12 hrs for SSE days,
			12 hrs for total time-SSE_days,
			2 months for 10 years
			"""
			self.dtt = [288e1 / 2, 432e2 * 2]
			self.ntime = [self.SSE_days / 1, (self.days - self.SSE_days) / 60]  # SSE_days, 10 years
		elif self.event == 'sub_EQ':
			"""
			1/year for 50 years followed by 'EQ' scheme
			"""
			self.dtt = [60 * 24 * 365, 15, 240, 288e1, 432e2, 60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years, 96, 120, 15, 122, 40]
		elif self.event == 'sub':  # 1/year for 50 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		elif self.event == 'topo':  # 1/year for 100 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		self.dtt = [i * 60 for i in self.dtt]  # in seconds
		self.dtt_repeated = np.repeat(self.dtt, self.ntime)
		self.dtt_comsum = np.cumsum(self.dtt_repeated)
		self.dt = self.dtt_comsum[0]
		self.nsteps = self.dtt_comsum.size
	
	def k_interpolation(self):
		log_kr = -18.0
		alpha_k = 0.6
		bctest1 = DirichletBC(self.Kspace, (1.0), 1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1.0), 2)  # land surface
		ktest = Function(self.Kspace)
		bctest1.apply(ktest)
		bctest2.apply(ktest)
		self.dof_surfK = np.where(ktest.vector().array() == 1.0)[0]
		
		if self.permeability == 'mapped':
			
			"""interpolate surface permeability onto all nodes"""
			kappa_surf = kappakriging(self.data_file,
			                          self.origin, self.theta,
			                          self.xvec, self.yvec)
			z_surf = surfkriging(self.xvec, self.yvec, self.zvec, self.dof_surfK)
			
			oceandof = np.where(z_surf < -10.)[0]
			kappa_surf[oceandof] = self.ocean_k
			
			self.kappa = Function(self.Kspace)
			self.kappa_vec = pow(10, (log_kr +
			                          (((kappa_surf) - log_kr) *
			                           pow((1 - (1e-3 * 0.25 * (self.zvec - z_surf)))
			                               , -alpha_k))))
			self.kappa.dat.data[:] = self.kappa_vec
		
		elif self.permeability == 'constant':
			"""interpolate surface permeability onto all nodes"""
			kappa_surf = Function(self.Kspace)
			kappa_surf_vec = self.surface_k
			kappa_surf.dat.data[:] = kappa_surf_vec
			z_surf = surfkriging(self.xvec, self.yvec, self.zvec, self.dof_surfK)
			oceandof = np.where(z_surf < -10.)[0]
			kappa_surf.dat.data[oceandof] = self.ocean_k
			
			self.kappa = Function(self.Kspace)
			self.kappa_vec = pow(10, (log_kr +
			                          (((kappa_surf) - log_kr) *
			                           pow((1 - (1e-3 * 0.25 * (self.zvec - z_surf)))
			                               , -alpha_k))))
			self.kappa.dat.data[:] = self.kappa_vece
	
	def slip_interpolation(self):
		
		if self.EQtype == 'inversion':
			
			u0load = np.load(self.inverse_results + 'u0_slab_array.npy')
			size_u = int(u0load.shape[0]/3)
			u0load = u0load.reshape((size_u, 3), order = 'F')
			XYZslab = np.load(self.inverse_var + 'XYZ_slab.npy')[0:size_u, :]
			
			u0slab_interp = slabkriging(XYZslab, u0load, self.xvec, self.yvec)

			self.u0slab = Function(self.Q_gspace)
			self.u0slab.dat.data[:] = u0slab_interp

		elif self.EQtype == 'data':
			
			"""import from excel file"""
			wb = pyxl.load_workbook(self.data_file)
			slip_data = wb['slip_data']
			self.u0slab = SlipExpression(slip_data, self.origin, self.theta, self.coords_V, dof_slab, self.var_path,
			                             degree=1)
	
	def strain(self, w):  # strain = 1/2 (grad u + grad u^T)
		return sym(nabla_grad(w))

	def sigma(self, w):  # stress = 2 mu strain + lambda tr(strain) I
		return 2.0 * self.mu * self.strain(w) + self.lmbda * div(w) * self.I
	
	def assemble_system(self, dt, assembleRHS=True):
		mat_type = 'aij'
		if assembleRHS:
			self.a_Mass = self.Se * self.p * self.q * dx  # Mass matrix
			self.a_Div = self.alpha * nabla_div(self.u) * self.q * dx  # Divergence matrix
			self.a_Grad = -self.alpha * self.p * (nabla_div(self.v)) * dx  # Gradient
			self.a_E = inner(self.sigma(self.u), nabla_grad(self.v)) * dx  # Elasticity
			self.a_Boundary = (1 / self.delta) * dot(self.u, self.n) * dot(self.v, self.n) \
			                  * self.ds(self.slab) + (1 / self.delta) * inner(self.T * self.u,
			                                                                  self.T * self.v) * self.ds(self.slab)
		
		self.a_K = dt * (self.kappa / self.muf) * inner(nabla_grad(self.p),
		                                                nabla_grad(self.q)) * dx  # Stiffness matrix
		
		print("all LHS matrices assembled", (time.time() - self.start) / 60.0, "minutes")
		
		self.A = assemble(self.a_Mass + self.a_Div + self.a_K +
		                  self.a_Grad + self.a_E + self.a_Boundary, bcs=self.bcs, mat_type=mat_type)
		
		print("LHS assembled", (time.time() - self.start) / 60.0, "minutes")
		
		if assembleRHS:
			self.L = assemble(self.a_Mass + self.a_Div, bcs=self.bcs, mat_type=mat_type)
			print("RHS assembled", (time.time() - self.start) / 60.0, "minutes")
			
	def test_PSD(self):

		def is_pd(x):
			return np.all(np.linalg.eigvals(x) > 0)

		def is_psd(x, tol=1e-5):
			E, V = np.linalg.eigh(x)
			return np.all(E > -tol)

		def is_symmetric(x):
			return (x.transpose() == x).all()

		A = self.A.array()

		print('A is {}'.format('symmetric' if is_symmetric(A) else ('not symmetric')))

		np.linalg.cholesky(A)

	def solve_system(self):
		count = 0
		if self.event == 'EQ':  # For an EQ, this can be outside the time loop since we only use it for the first time step
			bndform = (1 / self.delta) * inner(self.u0slab, self.T * self.v) * self.ds(
				self.slab)
			with assemble(bndform).dat.vec as vec:
				bnd_vec = vec
				
		"""et up solvers and preconditioners"""
		params = {
			'ksp_type': 'preonly',
			"ksp_monitor": True,
			'pc_type': 'lu',
			"pc_factor_mat_solver_package": "mumps",
			'ksp_initial_guess_nonzero': True,
				}
		
		A = self.A
		solver = LinearSolver(A, solver_parameters=params)
		self.sol = Function(self.ME)
		dt = self.dtt_comsum[0]
		
		
		for i in range(self.nsteps):
			if i == 0:
				dt = self.dtt_comsum[i]
			else:
				dt = self.dtt_comsum[i] - self.dtt_comsum[i - 1]
			if i > 1:
				dt_old = self.dtt_comsum[i - 1] - self.dtt_comsum[i - 2]
			if i > 1 and dt != dt_old:  # Rebuild Left hand side when 'dt' changes
				print("re-assembling LHS", (time.time() - self.start) / 60.0, "minutes")
				self.assemble_system(dt, assembleRHS=False)
				A = self.A
				solver = LinearSolver(A, solver_parameters=params)
			if i > 0:
				# Add slip to the boundary for a SSE and move center of slip along boundary
				if self.event == 'SSE' and count <= self.SSE_days:
					print("This doesn't work yet...")
					quit()
					self.u0_SSE = SSE_interpolation(i)
				elif self.event == 'sub':
					self.u0_subx += self.u_addx * (dt / self.year)
					self.u0_suby += self.u_addy * (dt / self.year)
				elif self.event == 'sub_EQ':
					self.u0_subx += self.u_addx * (dt / self.year)
					self.u0_suby += self.u_addy * (dt / self.year)
			
			"""Slip boundary Expression"""
			if self.event == 'SSE':
				u0slab = Expression((
					'u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
					'u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
					'0'), u0d=self.u0d_SSE, u0s=self.u0s_SSE, xcenter=self.xcenter,
					ycenter=self.ycenter, sigmax=self.sigma_bx, sigmay=self.sigma_by,
					degree=1)
				u0slabf = Function(self.ME.sub(1))
				u0slabf.interpolate(u0slab)
				bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
					self.slab)
				with assemble(bndform).dat.vec as vec:
					bnd_vec = vec
			elif self.event == 'sub':
				u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                    u0_suby=self.u0_suby, stick=self.stick, degree=1)
				u0slabf = Function(self.ME.sub(1))
				u0slabf.interpolate(u0slab)
				bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
					self.slab)
				with assemble(bndform).dat.vec as vec:
					bnd_vec = vec
			elif self.event == 'sub_EQ' and i < self.sub_cycle_years:
				u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                    u0_suby=self.u0_suby, stick=self.stick, degree=1)
				u0slabf = Function(self.ME.sub(1))
				u0slabf.interpolate(u0slab)
				bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
					self.slab)
				with assemble(bndform).dat.vec as vec:
					bnd_vec = vec
			elif self.event == 'sub_EQ' and i >= self.sub_cycle_years:
				u0slab = Expression(('u0_subx*stick + u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + '
				                     '(pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
				                     'u0_suby*stick + u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) '
				                     '+ (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
				                     '0'), u0d=self.u0_subx, u0s=self.u0_suby, xcenter=self.xcenter,
				                    ycenter=self.ycenter, sigmax=self.sigma_bx, sigmay=self.Tsigma_by, degree=1)
				u0slabf = Function(self.ME.sub(1))
				u0slabf.interpolate(u0slab)
				bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
					self.slab)
				with assemble(bndform).dat.vec as vec:
					bnd_vec = vec
					
			"""move the RHS and slip BC to PETSc to multiply in solution from
			# previous timestep and add in slip boundary term"""
			RHS_mat = self.L.M.handle
			b = RHS_mat.getVecs()[0]
			with assemble(self.sol_old).dat.vec as vec:
				w0vec = vec
			RHS_mat.mult(w0vec, b)
			
			"""transfer back to Function through a numpy array"""
			b += bnd_vec
			b_func = Function(self.ME)
			b_func.vector().set_local(b.array)
			[bc.apply(b_func) for bc in self.bcs]
			
			print("Boundary condition set", (time.time() - self.start) / 60.0, "minutes")
			
			solver.solve(self.sol, b_func)
			print("System solved", (time.time() - self.start) / 60.0, "minutes")
			p, u = self.sol.split()
			self.domain_flux = (-self.kappa / self.muf) * grad(p)  # element-wise flux vector
			self.bound_flux = assemble(inner(self.domain_flux, self.n) * self.ds(1))
			self.bound_flux2 = assemble(inner(self.domain_flux, self.n) * self.ds(2))
			print("outward flux through seafloor: ", self.bound_flux)
			print("outward flux through land surface: ", self.bound_flux2)
			#
			q = Function(self.Q_gspace)
			q.assign(project(self.domain_flux, self.Q_gspace))
			print("Computed velocity field", (time.time() - self.start) / 60.0, "minutes")
			
			u_V1 = Function(self.Q_gspace)
			u_V1.assign(project(u, self.Q_gspace))
			
			# self.v_strain = Function(self.Qspace)
			# self.v_strain.assign(project(nabla_div(u), self.Qspace))
			# self.stress = Function(self.Wspace)
			# self.stress.assign(project(self.sigma(u), self.Wspace))
			
			""" SAVE SOLUTION """
			p.rename('p', 'p')
			u.rename('u', 'u')
			q.rename('q', 'q')
			self.pfile.write(p)
			self.ufile.write(u)
			self.qfile.write(q)
			
			self.sea_flux_total[count] = self.bound_flux
			self.Sol_surf_p[:, count] = p.vector().array()[self.surface_dofs]
			self.Sol_gps[:, :, count] = u_V1.dat.data[self.GPS_dofs, :]#.reshape((size_u,), order = 'F')

			self.sol_old = self.sol
			count += 1
			print("Timestep:", count)

	def save_solution(self):
		if not os.path.exists(self.results_path):
			os.makedirs(self.results_path)
		if self.permeability == 'mapped':

			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_mapped.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_mapped.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_EQ_mapped.npy", self.Sol_all)
			elif self.event == 'EQ':
				np.save(self.results_path + "Sol_surfp_3D_EQ_mapped.npy", self.Sol_surf_p)
				np.save(self.results_path + "Sol_gps_3D_EQ_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "sea_flux_total_3D_EQ_mapped.npy", self.sea_flux_total)
		elif self.permeability == 'constant':
			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_constant.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_constant.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surfp_3D_sub_EQ_constant.npy", self.Sol_surf_p)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "sea_flux_total_3D_sub_EQ_constant.npy", self.sea_flux_total)
			elif self.event == 'EQ':
				np.save(self.results_path + "Sol_surfp_3D_EQ_constant.npy", self.Sol_surf_p)
				np.save(self.results_path + "Sol_gps_3D_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "sea_flux_total_3D_EQ_constant.npy", self.sea_flux_total)

