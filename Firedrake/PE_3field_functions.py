"""
Kimmy McCormack
Functions used by 'model_run.py' to solve coupled set of
poroelastic equations in 3D subduction zone domains

Units are in meters and MPa
"""
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as pyxl
import os
import scipy as sp
import scipy.interpolate
import shutil
from dolfin import *
from tempfile import TemporaryFile
from pykrige.ok import OrdinaryKriging
import pykrige.kriging_tools as kt
import matplotlib.tri as tri
import time
import gc
import sys
sys.path.append('/EQporoelasticity/local_lib')
gc.collect()

"""Optimization options"""
parameters["reorder_meshes"] = False
parameters['form_compiler']['optimize'] = True
parameters["form_compiler"]["cpp_optimize"] = True
ffc_options = {"optimize": True, "eliminate_zeros": True,"precompute_basis_const": True, "precompute_ip_const": True}


"""Data interpolation functions"""
def slabkriging(XYZslab, u0load, xvec, yvec, **kwargs):
	"""
	Interpolates inverted fault slip onto the mesh by by kriging.

		INPUTS:
			XYZslab = coordinates of fault slip solution in the model cartesian space
			u0load = slip vector at each solution node
			xvec - vector of dof x locations
			yvec - vector of dof y locations

		OUPUTS:
			u03 - vector containing inverted slip at all dofs
	"""
	
	x_all, y_all = xvec[:], yvec[:]
	
	nonzero = np.where(np.abs(u0load[:, 0]) > 0.10)[0]
	skip = (slice(None, None, 3))
	
	uxslip_load, uyslip_load, uzslip_load = u0load[:, 0][nonzero][skip], \
	                                        u0load[:, 1][nonzero][skip], u0load[:, 2][nonzero][skip]
	xslip, yslip, zslip = XYZslab[:, 0][nonzero][skip], \
	                      XYZslab[:, 1][nonzero][skip], XYZslab[:, 2][nonzero][skip]
	
	# Interpolate data using pykrige
	OK_uxslip = OrdinaryKriging(xslip, yslip, uxslip_load, variogram_model='linear', \
	                            variogram_parameters=[2e-4, 1e-2])
	uxslip, ss_ux = OK_uxslip.execute('points', x_all, y_all)
	
	OK_uyslip = OrdinaryKriging(xslip, yslip, uyslip_load, variogram_model='linear', \
	                            variogram_parameters=[2e-4, 1e-2])
	uyslip, ss_uy = OK_uyslip.execute('points', x_all, y_all)
	
	OK_uzslip = OrdinaryKriging(xslip, yslip, uzslip_load, variogram_model='linear', \
	                            variogram_parameters=[2e-4, 1e-2])
	uzslip, ss_uz = OK_uzslip.execute('points', x_all, y_all)
	
	u03 = np.concatenate((uxslip.reshape((uxslip.shape[0], 1)),
	                      uyslip.reshape((uxslip.shape[0], 1)),
	                      uzslip.reshape((uxslip.shape[0], 1))), axis=1)
	
	return u03

def surfkriging(xvec, yvec, zvec, dof_surf, **kwargs):
	"""Interpolates surface z value data by kriging.

		INPUTS:
			[x,y,z] = coordinates of degrees of freedom onto which the z value is mapped
			[dof_surf] = the degrees of freedom of the surface of the mesh

			OrdinaryKriging function takes variogram parameters as:
				linear - [slope, nugget]
				power - [scale, exponent, nugget]
			`   gaussian - [sill, range, nugget]
				spherical - [sill, range, nugget]
				exponential - [sill, range, nugget]
				hole-effect - [sill, range, nugget]

		OUPUTS:
			z_surf - surface height on all mesh nodes as a numpy array
	"""
	
	x_surf, y_surf, z_surf = xvec[dof_surf], yvec[dof_surf], zvec[dof_surf]
	x_all, y_all = xvec[:], yvec[:]
	
	# PyKrige interpolation
	OK_surf = OrdinaryKriging(x_surf, y_surf, z_surf,
	                          variogram_model='gaussian', variogram_parameters=[5e5, 5e4, 1e2])
	z_surf, ss = OK_surf.execute('points', x_all, y_all)
	
	return z_surf.compressed()

def kappakriging(data_file, origin, theta, xvec, yvec, **kwargs):
	"""Interpolates surface permeability data by kriging. Take surface value to interpolate kappa with depth.

		INPUTS:
			[data_sheet] = excel sheet containing pointwise permeability data in (lon, lat, permaebility[m^2]) format.
					Default is to ignore the first row (containing headers).
			[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
			[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
					counter-clockwise from latlon to XY
			[coords_K] = coordinates of degrees of freedom onto which the permeability is mapped
			[dof_surf] = the degrees of freedom of the surface of the mesh

			OrdinaryKriging function takes variogram parameters as:
				linear - [slope, nugget]
				power - [scale, exponent, nugget]
			`   gaussian - [sill, range, nugget]
				spherical - [sill, range, nugget]
				exponential - [sill, range, nugget]

		OUPUTS:
			kappa [m^2] - the interpolated 3D permeabiltiy field in the form of a FEniCS Expression()
	"""

	x_all, y_all = xvec[:], yvec[:]
	wb = pyxl.load_workbook(data_file)
	well_data = wb['well_data']
	x_wells, y_wells = latlon2XY(well_data, origin, theta)
	permeability = np.array(
		[[well_data.cell(row=i, column=5).value] for i in range(2, well_data.max_row + 1)]).reshape(
		well_data.max_row - 1, )
	logk = np.log10(permeability)

	"""PyKrige interpolation"""
	# OK = OrdinaryKriging(x_wells, y_wells, logk, variogram_model = 'power', variogram_parameters = [1, 1, 2e4])
	OK = OrdinaryKriging(x_wells, y_wells, logk, variogram_model='exponential',
	                     variogram_parameters=[1.086, 1.28e4, 0.771])
	k_surf, ss = OK.execute('points', x_all, y_all)
	np.save("results/numpy_variables/k_surf", k_surf.compressed())

	return k_surf.compressed()

def latlon2XY(data_sheet, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""
	
	lon = np.array([[data_sheet.cell(row=i, column=1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat = np.array([[data_sheet.cell(row=i, column=2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	
	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111
	
	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta
	
	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)
	
	return 1e3 * X, 1e3 * Y

class SlipExpression(Expression):
	""" Interpolates an arbitrary displacement field to the fault boundary of the mesh.

	INPUTS:
		[data] = an excel sheet with slip vectors recorded as
				(lon, lat, dip-slip [meters] , strike-slip [meters]) in the first four columns
		[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
		[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY
		[coords_V] = coordinates of degrees of freedom onto which the slip vectors are mapped
		[dof_slab] = the degrees of freedom of the fault boundary of the mesh
		[var_path] = the path of stored variables

	OUTPUTS:
		FEniCS Expression() containing the interpolated vector slip displacement on the fault boundary
	"""
	
	def __init__(self, data, origin, theta, coords_V, dof_slab, var_path, **kwargs):
		self.x_slab, self.y_slab, self.z_slab = coords_V[dof_slab, 0], coords_V[dof_slab, 1], coords_V[dof_slab, 2]
		
		X_slip, Y_slip = latlon2XY(data, origin, theta)
		Ux_slip =  np.array([[data.cell(row=i, column=3).value] for i in range(2, data.max_row + 1)]).reshape(
			data.max_row - 1, )
		Uy_slip =  np.array([[data.cell(row=i, column=4).value] for i in range(2, data.max_row + 1)]).reshape(
			data.max_row - 1, )
		
		# Interpolate data using pykrige
		OK_uxslip = OrdinaryKriging(X_slip, Y_slip, Ux_slip, variogram_model='linear', \
		                            variogram_parameters=[2e-4, 1e-2])
		self.uxslip, ss_ux = OK_uxslip.execute('points', self.x_slab, self.y_slab)
		
		OK_uyslip = OrdinaryKriging(X_slip, Y_slip, Uy_slip, variogram_model='linear', \
		                            variogram_parameters=[2e-4, 1e-2])
		self.uyslip, ss_ux = OK_uyslip.execute('points', self.x_slab, self.y_slab)
		
		np.save(var_path + 'uxslip', self.uxslip.compressed())
		np.save(var_path + 'uyslip', self.uyslip.compressed())
		
		# self.uxslip = np.load("saved_variables_poroelastic/uxslip_med.npy")
		# self.uyslip = np.load("saved_variables_poroelastic/uyslip_med.npy")
	
	def U_X_interp(self, X, Y):
		indx = (np.abs(np.sqrt(pow((self.x_slab[:] - X), 2) + pow((self.y_slab[:] - Y), 2)))).argmin()
		return self.uxslip[indx]
	
	def U_Y_interp(self, X, Y):
		indy = (np.abs(np.sqrt(pow((self.x_slab[:] - X), 2) + pow((self.y_slab[:] - Y), 2)))).argmin()
		return self.uyslip[indy]
	
	def eval(self, values, x, **kwargs):
		values[0] = self.U_X_interp(x[0], x[1])
		values[1] = self.U_Y_interp(x[0], x[1])
		values[2] = 0.0
	
	def value_shape(self):
		return (3,)


"""FEM solving functions"""
class Poroelasticity3D:
	def __init__(self, data_file, origin, theta, ndim, mesh, plot_figs, EQtype,
	             event, permeability, sub_cycle_years, u0_subdip, u0_substrike,
	             surface_k, ocean_k, **kwargs):
				
		self.start = time.time()
		self.data_file = data_file
		self.origin = origin
		self.theta = theta
		self.mesh = mesh
		self.plot_figs = plot_figs
		self.EQtype = EQtype
		self.event = event
		self.permeability = permeability
		self.sub_cycle_years = sub_cycle_years
		self.u0_subx = u0_subdip
		self.u0_suby = u0_substrike
		self.surface_k = surface_k
		self.ocean_k = ocean_k
		
		self.V = VectorElement('Lagrange', self.mesh.ufl_cell(), 2)
		self.R = FiniteElement('DG', self.mesh.ufl_cell(), 0)
		self.W = FiniteElement('RT', self.mesh.ufl_cell(), 1)
		self.Vq_cont = VectorFunctionSpace(self.mesh, "CG", 1)
		
		self.Kelement = FiniteElement('CG', self.mesh.ufl_cell(), 1)
		self.Kspace = FunctionSpace(self.mesh, self.Kelement)
		self.Rspace = FunctionSpace(self.mesh, self.R)
		
		self.ME = FunctionSpace(self.mesh, MixedElement([self.R, self.V, self.W]))
		
		self.Vspace = FunctionSpace(self.mesh, self.V)
		self.V1space = VectorFunctionSpace(self.mesh, 'CG', 1)
		self.Qspace = FunctionSpace(self.mesh, self.Kelement)
		self.Q_gspace = VectorFunctionSpace(self.mesh, "Lagrange", 1)
		self.Wspace = TensorFunctionSpace(self.mesh, "Lagrange", 1)
		
		(self.p, self.u, self.q) = TrialFunctions(self.ME)
		(self.r, self.v, self.w) = TestFunctions(self.ME)
		self.ds = Measure("ds", domain=self.mesh)
		self.dx = Measure("dx", domain=self.mesh)  #
		self.n = FacetNormal(self.mesh)  # normal vector
		self.dim = self.ME.dim()
		self.kdim = self.Qspace.dim()
		self.pdim = self.ME.sub(0).dim()
		self.udim = self.ME.sub(1).dim()
		self.geom_dim = ndim + 1
		
		"""GPS STATION LOCATIONS """
		wb = pyxl.load_workbook(self.data_file)
		gps_data = wb['GPS_data']
		
		self.x_gps, self.y_gps = latlon2XY(gps_data, self.origin, self.theta)
		self.z_gps = np.zeros(len(self.x_gps))

		"""Input/Output paths"""
		self.results_path = 'results/numpy_results/'
		self.var_path = 'results/numpy_variables/'
		self.paraviewpath = 'results/paraview3D/'
		self.inverse_results = '/fenics/shared/inverse_codes/results_deterministic/numpy_results/'
		self.inverse_var = '/fenics/shared/inverse_codes/results_deterministic/numpy_variables/'
		
		
		if not os.path.exists(self.results_path):
			os.makedirs(self.results_path)
		if not os.path.exists(self.var_path):
			os.makedirs(self.var_path)
		if not os.path.exists(self.paraviewpath):
			os.makedirs(self.paraviewpath)
		
		self.extract_coords()
		print( "Coordinates extracted", (time.time() - self.start) / 60.0, "minutes")
		self.k_interpolation()
		self.plot_k()
		#quit()
		print( "Permeability interpolated", (time.time() - self.start) / 60.0, "minutes")
		self.timestepping()
		
		"""POROELASTIC PARAMETERS"""
		self.delta = 1e-6  # Dirichlet control regularization on bottom boundary
		self.delta2 = 1e-6  # Dirichlet control regularization on bottom boundary
		self.muf = Constant('1e-9')  # pore fluid viscosity [MPa s]
		self.B = 0.6  # Skempton's coefficient
		self.nuu = 0.4  # undrained poissons ratio
		self.nu = 0.25  # Poisson's ratio
		x = SpatialCoordinate(self.mesh)

		self.E = interpolate(500 * pow(-x[2], 0.516) + 2e4, self.Kspace)
		self.mu = interpolate(self.E / (2.0*(1.0 + self.nu)), self.Kspace)  # shear modulus
		self.lmbda = interpolate(self.E*self.nu / ((1.0 + self.nu)*(1.0 - 2.0*self.nu)), self.Kspace)  # Lame's parameter
		self.alpha = 3 * (self.nuu - self.nu) / (self.B * (1 - 2 * self.nu) * (1 + self.nuu))  # Biot's coefficient
		self.Se = interpolate((9*(self.nuu-self.nu)*(1-(2*self.nuu)))/
		                      (2*self.mu*pow(self.B,2)*(1-2*self.nu)*pow((1+self.nuu),2)), self.Kspace)  # mass specific storage
		
		self.kstick = 8e-5
		self.stick = Expression('1/(1+exp(-kstick*(1.1e5-x[0])))', kstick=self.kstick, degree=1)
		self.d = self.u.geometric_dimension()  # number of space dimensions
		self.I = Identity(self.d)
		self.T = (self.I - outer(self.n, self.n))  # tangent operator for boundary condition
		self.year = 60 * 60 * 24 * 365
		
		"""INITIAL AND BOUNDARY CONDITIONS"""
		####### Initial conditions: (('p0','ux0','uy0')) ##########
		self.sol_old = Function(self.ME)

		zero_scalar = Constant(0.0)
		zero_vector = Expression(('0.0', '0.0', '0.0'), degree = 1)
		ub = Expression(('1', '0', '0'), degree = 1)
		
		bcu1 = DirichletBC(self.ME.sub(1), zero_vector, 6)  # right/ back side (no slip condition)
		bcu2 = DirichletBC(self.ME.sub(1).sub(2), zero_scalar, 8)  # mantle (no slip condition)
		bcq1 = DirichletBC(self.ME.sub(2), ('0.0', '0.0', '0.0'),  1)  # ocean surface
		bcq2 = DirichletBC(self.ME.sub(2), ('0.0', '0.0', '0.0'),  2)  # land surface
		bcq3 = DirichletBC(self.ME.sub(2), ('0.0', '0.0', '0.0'),  3)  # subducting slab
		bcq5 = DirichletBC(self.ME.sub(2), ('0.0', '0.0', '0.0'),  5)  # north side
		bcq6 = DirichletBC(self.ME.sub(2), ('0.0', '0.0', '0.0'),  6)  # back side (no slip)
		bcq7 = DirichletBC(self.ME.sub(2), ('0.0', '0.0', '0.0'),  7)  # south side
		bcq8 = DirichletBC(self.ME.sub(2), ('0.0', '0.0', '0.0'),  8)  # mantle

		bcp1 = DirichletBC(self.ME.sub(0), ('0.'),  1, method="geometric")  # ocean surface (free flow condition)
		
		self.bcs = [bcp1, bcu1, bcu2, bcq2, bcq3, bcq8, bcq5, bcq6, bcq7]
		
		self.slab = 3  # the slab is labeled as boundary 3 for this mesh, but not always the case
		self.hydro = 1 #ocean bottom
		self.surface = 2
		
		if self.event == 'SSE':
			self.u_addx = self.u0_SSEx
			self.u_addy = self.u0_SSEy
		elif self.event == 'sub':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby
		elif self.event == 'sub_EQ':  # Amount of slip per day (at center of event)
			self.u_addx = self.u0_subx
			self.u_addy = self.u0_suby
		
		
		"""SET UP TO SAVE SOLUTION"""
		self.Sol_surf_p = np.empty((self.surface_dofs.shape[0], self.nsteps))  # matrix to save surface solution
		self.Sol_gps = np.empty((self.x_gps.shape[0]*ndim, self.nsteps))  # matrix to save surface solution
		self.sea_flux_total = np.empty(self.nsteps)  # matrix to save surface post-seismic solution
		
		pfilename = self.paraviewpath + "pressure3D.pvd"
		qfilename = self.paraviewpath + "flux3D.pvd"
		ufilename = self.paraviewpath + "def3D.pvd"
		
		self.pfile = File(pfilename)
		self.qfile = File(qfilename)
		self.ufile = File(ufilename)

		print( "Solution matrices created", (time.time() - self.start) / 60.0, "minutes")
		
		"""SET UP, RUN, SOLVE and SAVE """
		self.slip_interpolation()
		print( "Slip data interpolated", (time.time() - self.start) / 60.0, "minutes")
		dt = self.dtt_comsum[0]
		self.assemble_system(dt)
		print( "System assembled", (time.time() - self.start) / 60.0, "minutes")
		self.solve_system()
		self.save_solution()
	
	def extract_coords(self):
		if not os.path.exists(self.var_path):
			os.makedirs(self.var_path)
			
		self.x_all, self.y_all, self.z_all = SpatialCoordinate(self.mesh)
		
		xf = Function(self.Qspace)
		yf = Function(self.Qspace)
		zf = Function(self.Qspace)
		xm, ym, zm = SpatialCoordinate(self.mesh)
		
		self.xvec = xf.interpolate(xm).dat.data_ro
		self.yvec = yf.interpolate(ym).dat.data_ro
		self.zvec = zf.interpolate(zm).dat.data_ro
		
		u1dim = self.xvec.shape[0]

		bctestp0 = DirichletBC(self.Rspace, (2), 2, method="geometric")  # land surface
		ptest = Function(self.Rspace)
		bctestp0.apply(ptest.vector())
		self.surface_dofs_p = np.where(ptest.vector().array() == 2)[0]

		bctest0 = DirichletBC(self.Qspace, (1), 1)  # ocean surface
		bctest1 = DirichletBC(self.Qspace, (2), 2)  # top surface
		bctest3 = DirichletBC(self.Qspace, (3), 3)  # slab surface
		utest = Function(self.Qspace)
		bctest0.apply(utest.vector())
		bctest1.apply(utest.vector())
		bctest3.apply(utest.vector())
		self.ocean_dofs = np.where(utest.vector().array() == 1)[0]
		self.surface_dofs = np.where(utest.vector().array() == 2)[0]
		self.slab_dofs = np.where(utest.vector().array() == 3)[0]
		
		self.size_gps = self.x_gps.shape[0]
		size_w_gps = 3 * self.size_gps
		indices_gps = np.empty((self.size_gps, 3))
		for j in range(0, self.size_gps):
			indice_gps = (np.abs(np.sqrt(
					pow((self.xvec[:] - self.x_gps[j]), 2) + pow((self.yvec[:] - self.y_gps[j]), 2)))).argmin()
			indices_gps[j, :] = [indice_gps, indice_gps+u1dim, indice_gps+(2*u1dim)]
		self.GPS_dofs = indices_gps.reshape((size_w_gps), order = 'F').astype('int')
		
		np.save(self.var_path + "x_all_3D", self.xvec)
		np.save(self.var_path + "y_all_3D", self.yvec)
		np.save(self.var_path + "z_all_3D", self.zvec)
		np.save(self.var_path + "GPS_dofs_3D", self.GPS_dofs)
		np.save(self.var_path + "surface_dofs_3D", self.surface_dofs)
		np.save(self.var_path + "ocean_dofs_3D", self.ocean_dofs)
	
	def time_stepping(self):
		if self.event == 'EQ':
			"""15 min for 1/2 day,
			4 hrs for 10 days,
			12 hours for 30 days,
			2 days for 2 months,
			per month for 10 years
			"""
			self.dtt = [60, 240, 720, 288e1, 2 * 432e2]   # in minutes
			self.ntime = [24, 60, 60, 30, 60]  # , 50]
		elif self.event == 'SSE':
			"""
			12 hrs for SSE days,
			12 hrs for total time-SSE_days,
			2 months for 10 years
			"""
			self.dtt = [288e1 / 2, 432e2 * 2]
			self.ntime = [self.SSE_days / 1, (self.days - self.SSE_days) / 60]  # SSE_days, 10 years
		elif self.event == 'sub_EQ':
			"""
			1/year for 50 years followed by 'EQ' scheme
			"""
			self.dtt = [60 * 24 * 365, 15, 240, 288e1, 432e2, 60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years, 96, 120, 15, 122, 40]
		elif self.event == 'sub':  # 1/year for 50 years
			self.dtt = [60 * 24 * 365]  # in minutes
			self.ntime = [self.sub_cycle_years]
		self.dtt = [i * 60 for i in self.dtt]  # in seconds
		self.dtt_repeated = np.repeat(self.dtt, self.ntime)
		self.dtt_comsum = np.cumsum(self.dtt_repeated)
		self.dt = self.dtt_comsum[0]
		self.nsteps = self.dtt_comsum.size
	
	def k_interpolation(self):
		log_kr = -18.0
		alpha_k = 0.6
		
		bctest1 = DirichletBC(self.Kspace, (1.0), 1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1.0), 2)  # land surface
		ktest = Function(self.Kspace)
		bctest1.apply(ktest)
		bctest2.apply(ktest)
		self.dof_surfK = np.where(ktest.vector().array() == 1.0)[0]
		
		if self.permeability == 'mapped':
			
			"""interpolate surface permeability onto all nodes"""
			kappa_surf = kappakriging(self.data_file, self.origin, self.theta,
			                          self.xvec, self.yvec)
			z_surf = surfkriging(self.xvec, self.yvec, self.zvec, self.dof_surfK)
			
			oceandof = np.where(z_surf < -10.)[0]
			kappa_surf[oceandof] = self.ocean_k
			self.kappa = Function(self.Kspace)
			self.kappa_vec = pow(10, (log_kr + ((kappa_surf - log_kr) *
			                                    pow((1 - (1e-3 * 0.25*(self.zvec - z_surf))), -alpha_k))))
			self.kappa.dat.data[:] = self.kappa_vec
			
		elif self.permeability == 'constant':
			"""interpolate surface permeability onto all nodes"""
			kappa_surf = Function(self.Kspace)
			kappa_surf_vec = pow(10, self.surface_k)
			kappa_surf.dat.data[:] = kappa_surf_vec
			z_surf = surfkriging(self.xvec, self.yvec, self.zvec, self.dof_surfK)
			oceandof = np.where(z_surf < -10.)[0]
			kappa_surf.dat.data[oceandof] = self.ocean_k
			
			self.kappa = Function(self.Kspace)
			self.kappa_vec = pow(10, (log_kr +
			                          (((kappa_surf) - log_kr) *
			                           pow((1 - (1e-3 * 0.25 * (self.zvec - z_surf)))
			                               , -alpha_k))))
			self.kappa.dat.data[:] = self.kappa_vec
	
	
	def slip_interpolation(self):
		
		if self.EQtype == 'inversion':
			
			u0load = np.load(self.inverse_results + 'u0_slab_array.npy')
			size_u = int(u0load.shape[0] / 3)
			u0load = u0load.reshape((size_u, 3), order='F')
			XYZslab = np.load(self.inverse_var + 'XYZ_slab.npy')[0:size_u, :]
			u0slab_interp = slabkriging(XYZslab, u0load, self.xvec, self.yvec)
			self.u0slab = Function(self.Q_gspace)
			self.u0slab.dat.data[:] = u0slab_interp
		
		elif self.EQtype == 'data':
			# import from excel file ##
			wb = pyxl.load_workbook(self.data_file)
			slip_data = wb['slip_data']
			self.u0slab = SlipExpression(slip_data, self.origin, self.theta, self.coords_V, dof_slab, self.var_path,
			                             degree=1)
	
	def topo_flow_interp(self):
		self.coords_K = self.Kspace.tabulate_dof_coordinates().reshape((self.Kspace.dim(), -1))
		bctest1 = DirichletBC(self.Kspace, (1), self.boundaries, 1)  # ocean surface
		bctest2 = DirichletBC(self.Kspace, (1), self.boundaries, 2)  # land surface
		qtest = Function(self.Kspace)
		bctest1.apply(qtest.vector())
		bctest2.apply(qtest.vector())
		self.dof_surfK = np.where(qtest.vector() == 1)[0]
		
		wb = pyxl.load_workbook(self.data_file)
		topo_data = wb['topo_data3D']
		
		topo_p_surf = topokriging(topo_data, self.origin, self.theta,
		                          self.coords_K, self.dof_surfK, self.var_path)
		
		low_p = np.where(topo_p_surf < 0.0)[0]
		topo_p_surf[low_p] = 0.01

		p_surf_func = Function(self.Kspace)
		p_surf_func.vector()[self.dof_surfK] = topo_p_surf
		self.p_topo = Expression(('ptopo'), ptopo=p_surf_func, degree=1)

	def strain(self, w):  # strain = 1/2 (grad u + grad u^T)
		return sym(nabla_grad(w))
	
	def sigma(self, w):  # stress = 2 mu strain + lambda tr(strain) I
		return 2.0 * self.mu * self.strain(w) + self.lmbda * div(w) * self.I
	
	def assemble_system(self, dt, assembleRHS=True):
		mat_type = 'aij'
		if assembleRHS:
			self.p_Mass = self.Se * self.p * self.r * dx  # Mass matrix
			self.u_Div = self.alpha * nabla_div(self.u) * self.r * dx  # Divergence matrix
			self.u_E = inner(self.sigma(self.u), nabla_grad(self.v)) * dx  # Elasticity
			self.p_Grad = -self.alpha * self.p * (nabla_div(self.v)) * dx  # Gradient
			self.u_Boundary = (1 / self.delta) * dot(self.u, self.n) * dot(self.v, self.n) * self.ds(self.slab) \
			                  + (1 / self.delta) * inner(self.T * self.u, self.T * self.v) * self.ds(self.slab)
		
		self.q_Div = dt * nabla_div(self.q) * self.r * dx  # Divergence matrix
		self.q_K = -(dt * self.muf / self.kappa) * inner(self.q, self.w) * dx  # Stiffness matrix
		self.q_Grad = dt * self.p * (nabla_div(self.w)) * dx  # Gradient
		
		print ("all LHS matrices assembled", (time.time() - self.start) / 60.0, "minutes")
		
		self.A = assemble(self.p_Mass + self.u_Div + self.q_Div + self.u_E + \
		                  self.p_Grad + self.q_K + self.q_Grad + self.u_Boundary, bcs=self.bcs, mat_type=mat_type)
		
		print("LHS assembled", (time.time() - self.start) / 60.0, "minutes")
		
		if assembleRHS:
			#  self.L = self.p_Mass + self.alpha * self.u_Div
			self.L = assemble(self.p_Mass + self.u_Div, bcs=self.bcs, mat_type=mat_type)
			print("RHS assembled", (time.time() - self.start) / 60.0, "minutes")

	def solve_system(self):

		if self.event == 'EQ':
			"""For an EQ, this can be outside the time loop since we
			only use it for the first time step"""
			bndform = (1 / self.delta) * inner(self.u0slab, self.T * self.v) * self.ds(
				self.slab)
			with assemble(bndform).dat.vec as vec:
				bnd_vec = vec

		# set up solvers and preconditioners
		params = {
			'ksp_type': 'gmres',
			"ksp_monitor": True,
			'pc_type': 'lu',
			"pc_factor_mat_solver_package": "mumps",
			'pc_factor_reuse_ordering': True,  # PCFactorSetReuseOrdering
			'pc_factor_reuse_fill': True,
			'ksp_initial_guess_nonzero': True
		}

		A = self.A
		solver = LinearSolver(A, solver_parameters=params)
		self.sol = Function(self.ME)
		dt = self.dtt_comsum[0]

		for i in range(self.nsteps):

			if i > 0:
				dt = self.dtt_comsum[i] - self.dtt_comsum[i - 1]
			if i > 1:
				dt_old = self.dtt_comsum[i - 1] - self.dtt_comsum[i - 2]
			if i > 1 and dt != dt_old:  # Rebuild Left hand side when 'dt' changes
				print ("re-assembling LHS", (time.time() - self.start) / 60.0, "minutes")
				self.assemble_system(dt, assembleRHS=False)
				A = self.A
				solver = LinearSolver(A, solver_parameters=params)
			if i > 0:
				# Add slip to the boundary for a SSE and move center of slip along boundary
				if self.event == 'SSE' and count <= self.SSE_days:
					print ("This doesn't work yet...")
					quit()
					self.u0_SSE = SSE_interpolation(i)
				elif self.event == 'sub':
					self.u0_subx += self.u_addx * (dt / self.year)
					self.u0_suby += self.u_addy * (dt / self.year)
				elif self.event == 'sub_EQ':
					self.u0_subx += self.u_addx * (dt / self.year)
					self.u0_suby += self.u_addy * (dt / self.year)

			"""Slip boundary Expression"""
			if self.event == 'SSE':
				u0slab = Expression((
					'u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
					'u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
					'0'), u0d=self.u0d_SSE, u0s=self.u0s_SSE, xcenter=self.xcenter,
					ycenter=self.ycenter, sigmax=self.sigma_bx, sigmay=self.sigma_by,
					degree=1)
				u0slabf = Function(self.ME.sub(1))
				u0slabf.interpolate(u0slab)
				bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
					self.slab)
				with assemble(bndform).dat.vec as vec:
					bnd_vec = vec
			elif self.event == 'sub':
				u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                    u0_suby=self.u0_suby, stick=self.stick, degree=1)
				u0slabf = Function(self.ME.sub(1))
				u0slabf.interpolate(u0slab)
				bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
					self.slab)
				with assemble(bndform).dat.vec as vec:
					bnd_vec = vec
			elif self.event == 'sub_EQ' and i < self.sub_cycle_years:
				u0slab = Expression(('u0_subx*stick', 'u0_suby*stick', '0'), u0_subx=self.u0_subx,
				                    u0_suby=self.u0_suby, stick=self.stick, degree=1)
				u0slabf = Function(self.ME.sub(1))
				u0slabf.interpolate(u0slab)
				bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
					self.slab)
				with assemble(bndform).dat.vec as vec:
					bnd_vec = vec
			elif self.event == 'sub_EQ' and i >= self.sub_cycle_years:
				u0slab = Expression(('u0_subx*stick + u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) + '
				                     '(pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
				                     'u0_suby*stick + u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigmax,2)) '
				                     '+ (pow((x[1] - ycenter),2)/(2*pow(sigmay,2)))))',
				                     '0'), u0d=self.u0_subx, u0s=self.u0_suby, xcenter=self.xcenter,
				                    ycenter=self.ycenter, sigmax=self.sigma_bx, sigmay=self.Tsigma_by, degree=1)
				u0slabf = Function(self.ME.sub(1))
				u0slabf.interpolate(u0slab)
				bndform = (1 / self.delta) * inner(u0slabf, self.T * self.v) * self.ds(
					self.slab)
				with assemble(bndform).dat.vec as vec:
					bnd_vec = vec

			"""move the RHS and slip BC to PETSc to multiply in solution from
			previous timestep and add in slip boundary term"""
			RHS_mat = self.L.M.handle
			b = RHS_mat.getVecs()[0]
			with assemble(self.sol_old).dat.vec as vec:
				sol_oldvec = vec
			RHS_mat.mult(sol_oldvec, b)

			"""transfer back to Function through a numpy array"""
			b += bnd_vec
			b_func = Function(self.ME)
			b_func.vector().set_local(b.array)
			[bc.apply(b_func) for bc in self.bcs]
			print("Boundary condition set", (time.time() - self.start) / 60.0, "minutes")

			solver.solve(self.sol, b_func)
			print("System solved", (time.time() - self.start) / 60.0, "minutes")
			p, u, q = self.sol.split()

			q_cont = Function(self.Vq_cont)
			q_cont.assign(project(q, self.Vq_cont))
			# q_cont.assign(project(q, self.Vq_cont, solver_type='gmres'))

			p_cont = Function(self.Qspace)
			p_cont.assign(project(p, self.Qspace))
			
			u_V1 = Function(self.Q_gspace)
			u_V1.assign(project(u, self.Q_gspace))
			print( "u_v1 shape: ", u_V1.vector().array().shape)

			self.bound_flux = assemble(dot(q, self.n) * self.ds(1))# + dot(q, self.n) * self.ds(4))
			print ("outward flux through seafloor: ", self.bound_flux)
			self.bound_flux2 = assemble(dot(q, self.n) * self.ds(2))
			print ("outward flux through land surface: ", self.bound_flux2)

			"""SAVE SOLUTION """
			p.rename('p', 'p')
			u.rename('u', 'u')
			q.rename('q', 'q')
			q_cont.rename("q_cont", "q_cont")
			p_cont.rename("p_cont", "p_cont")

			self.pfile.write(p)
			self.ufile.write(u)
			self.qfile.write(q)
			self.qcontfile.write(q_cont)
			self.pcontfile.write(p_cont)

			self.sea_flux_total[count] = self.bound_flux
			self.Sol_surf_p[:, count] = p_cont.vector().array()[self.surface_dofs]
			self.Sol_gps[:, count] = u_V1.vector().array()[self.GPS_dofs]

			self.sol_old = self.sol
			print ("Solution updated", (time.time() - self.start) / 60.0, "minutes")

			count += 1
			print ("Timestep:", count)

	def save_solution(self):
		if not os.path.exists(self.results_path):
			os.makedirs(self.results_path)
		if self.permeability == 'mapped':
			if self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_mapped.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_mapped.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_EQ_mapped.npy", self.Sol_all)
			elif self.event == 'EQ':
				np.save(self.results_path + "Sol_surfp_3D_EQ_mapped.npy", self.Sol_surf_p)
				np.save(self.results_path + "Sol_gps_3D_EQ_mapped.npy", self.Sol_gps)
				np.save(self.results_path + "sea_flux_total_3D_EQ_mapped.npy", self.sea_flux_total)
		elif self.permeability == 'constant':
			if self.event == 'SSE':
				np.save(self.results_path + "Sol_surf_3D_SSE_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_SSE_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_SSE_constant.npy", self.Sol_all)
			elif self.event == 'sub':
				np.save(self.results_path + "Sol_surf_3D_sub_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_constant.npy", self.Sol_gps)
				np.save(self.results_path + "Sol_all_3D_sub_constant.npy", self.Sol_all)
			elif self.event == 'sub_EQ':
				np.save(self.results_path + "Sol_surf_3D_sub_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_sub_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "sea_flux_total_3D_sub_EQ_constant.npy", self.sea_flux_total)
			elif self.event == 'EQ':
				np.save(self.results_path + "Sol_surf_3D_EQ_constant.npy", self.Sol_surf)
				np.save(self.results_path + "Sol_gps_3D_EQ_constant.npy", self.Sol_gps)
				np.save(self.results_path + "sea_flux_total_3D_EQ_constant.npy", self.sea_flux_total)


