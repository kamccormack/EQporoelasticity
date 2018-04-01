from dolfin import *
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator, MultipleLocator, FormatStrFormatter
from matplotlib import cm
from mpl_toolkits.mplot3d import Axes3D
import numpy as np
import scipy as sp
from scipy.interpolate import griddata
from tempfile import TemporaryFile
import openpyxl as pyxl
import matplotlib.tri as tri
dolfin.parameters.reorder_dofs_serial = False

font = {'weight' : 'normal',
        'size'   : 10}

mpl.rc('font', **font)

mesh_size = 'med'  # fine, med, coarse
path = "/home/fenics/shared/"

def GPSlatlon2XY(data_sheet, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = np.array([[data_sheet.cell(row = i, column = 1).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )
	lat = np.array([[data_sheet.cell(row = i, column = 2).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )

	lon_u = np.array([[data_sheet.cell(row = i, column = 5).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )
	lat_u = np.array([[data_sheet.cell(row = i, column = 6).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )
	Uz = np.array([[data_sheet.cell(row = i, column = 4).value] for i in range(2, data_sheet.max_row+1)]).reshape(data_sheet.max_row-1, )

	lon_in_km = (lon - origin[0])*111*np.cos(lat*np.pi/180)
	lat_in_km = (lat - origin[1])*111
	
	rho_u = np.sqrt(np.power(lon_u,2) + np.power(lat_u,2))
	theta_new_u = np.arctan2(lat_u,lon_u) - theta

	rho = np.sqrt(np.power(lon_in_km,2) + np.power(lat_in_km,2))
	theta_new = np.arctan2(lat_in_km,lon_in_km) - theta

	X, Y = rho*np.cos(theta_new), rho*np.sin(theta_new)
	Ux, Uy = rho_u*np.cos(theta_new_u), rho_u*np.sin(theta_new_u)

	return 1e3*X, 1e3*Y, 1e-3*Ux, 1e-3*Uy, 1e-3*Uz

def GPSlatlon2XY_time(data_sheet, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	gps_time = np.array([[data_sheet.cell(row = i, column = 1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	lon_u = np.array([[data_sheet.cell(row = i, column = 3).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat_u = np.array([[data_sheet.cell(row = i, column = 2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	Uz = np.array([[data_sheet.cell(row = i, column = 4).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	rho_u = np.sqrt(np.power(lon_u, 2) + np.power(lat_u, 2))
	theta_new_u = np.arctan2(lat_u, lon_u) - theta

	Ux, Uy = rho_u * np.cos(theta_new_u), rho_u * np.sin(theta_new_u)

	return gps_time, Ux, Uy, Uz

def latlon2XY_point(point, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = point[0]
	lat = point[1]

	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)

	return 1e3 * X, 1e3 * Y

def GPSlatlon2XY_other(data_sheet, origin, theta):
	"""
	latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
	into numerical models.

	INPUTS:
	[data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
	[origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
	[theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
				counter-clockwise from latlon to XY

	Assumes first column is longitude and second column is latitude

	OUTPUTS:
	Returns transformed coordinates as numpy vectors X, Y in kilometers
	"""

	lon = np.array([[data_sheet.cell(row = i, column = 1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	Z = np.array([[data_sheet.cell(row = i, column = 3).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )
	lat = np.array([[data_sheet.cell(row = i, column = 2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
		data_sheet.max_row - 1, )

	lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
	lat_in_km = (lat - origin[1]) * 111

	rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
	theta_new = np.arctan2(lat_in_km, lon_in_km) - theta

	X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)

	return 1e3*X, 1e3*Y, Z

def XY2latlon(X, Y, origin, theta):

	X = 1e-3*X
	Y = 1e-3*Y


	rho = np.sqrt(np.power(X, 2) + np.power(Y,2))
	theta_new = np.arctan2(Y, X) + theta

	lat = rho*np.sin(theta_new)/111 + origin[1]
	lon = rho*np.cos(theta_new)/(111*np.cos(lat*np.pi/180)) + origin[0]

	return lon, lat


##########################################################################################

station = 'CABA'

with open(path+'data/station_list.txt', 'r') as document:
    stations = {}
    for line in document:
        line = line.split()
        if not line:  # empty line?
            continue
        stations[line[0]] = line[1:]

latlongps = np.array(stations[station]).astype(np.float)

meshpath = "/fenics/shared/meshes"
mesh_size = 'med_inv'  # fine, med, coarse
mesh = Mesh(meshpath+'/3D_'+mesh_size+'.xml')
boundaries = MeshFunction("size_t", mesh, meshpath+'/3D_'+mesh_size+'_facet_region.xml')


boundarymesh = BoundaryMesh(mesh, 'exterior')
slab_mesh = SubMesh(boundarymesh, boundaries, 3)


results_path = '/fenics/shared/inverse_codes/results_deterministic/numpy_results/'
var_path = '/fenics/shared/inverse_codes/results_deterministic/numpy_variables/'
x_all = np.load(var_path+"x_all_3D.npy")
y_all = np.load(var_path+"y_all_3D.npy")
z_all = np.load(var_path+"z_all_3D.npy")
#surface_dofs = np.load(var_path+"surface_dofs_3D.npy")
#ocean_dofs = np.load(var_path+"ocean_dofs_3D.npy")
gps_dofs = np.load(var_path + "GPS_dofs.npy")
slab_dofs = np.load(var_path + "slab_dofs_u_3D.npy")
u_array = np.load(results_path+"u_array.npy")


#Load GPS data
#origin = np.array([-85.21, 8.64]) # origin of top surface of mesh - used as point of rotation
#theta = 0.733 # angle of rotation between latitude line and fault trace
origin = np.array([-85.1688, 8.6925]) # origin of top surface of mesh - used as point of rotation
theta = 0.816 # angle of rotation between latitude line and fault trace (42 deg)

data_file = path+'data/Data.xlsx'
wb = pyxl.load_workbook(data_file)
gps_data = wb['GPS_data']

X_gps, Y_gps, Ux_gps, Uy_gps, Uz_gps = GPSlatlon2XY(gps_data, origin, theta)
Z_gps = np.zeros(len(X_gps))

#solution at GPS stations
gps_dofs_xy = gps_dofs[0:gps_dofs.shape[0]/3].astype(int)
Sol_gps_mapped = u_array[gps_dofs.astype(int)]
X_gps_model, Y_gps_model = x_all[gps_dofs_xy], y_all[gps_dofs_xy]

Ux_gps_model = Sol_gps_mapped[0:X_gps.shape[0]]
Uy_gps_model = Sol_gps_mapped[X_gps.shape[0]:2*X_gps.shape[0]]
Uz_gps_model = Sol_gps_mapped[2*X_gps.shape[0]:3*X_gps.shape[0]]


#Solution on the slab interface
slab_dofs_xy = slab_dofs[0:slab_dofs.shape[0]/3].astype(int)
Sol_slab = u_array[slab_dofs.astype(int)]
X_slab, Y_slab, Z_slab = x_all[slab_dofs_xy], y_all[slab_dofs_xy], z_all[slab_dofs_xy]

Ux_slab = Sol_slab[0:X_slab.shape[0]]
Uy_slab = Sol_slab[X_slab.shape[0]:2*X_slab.shape[0]]
Uz_slab = Sol_slab[2*X_slab.shape[0]:3*X_slab.shape[0]]


U_mag_slab = np.sqrt(Ux_slab**2 + Uy_slab**2 + Uz_slab**2)

nonzero = np.where(U_mag_slab > 0.1)[0]

lon_slab, lat_slab = XY2latlon(X_slab, Y_slab, origin, theta)

mag_all = np.concatenate((lon_slab.reshape(lon_slab.shape[0], 1), lat_slab.reshape(lon_slab.shape[0], 1), U_mag_slab.reshape(lon_slab.shape[0], 1)), axis=1)

#np.savetxt('inversion_mag.xyz', (mag_all), delimiter=' ')

X_gps_post, Y_gps_post = latlon2XY_point(latlongps, origin, theta)


data_file2 = path+'data/GPS_post.xlsx'
wb2 = pyxl.load_workbook(data_file2)
gps_data_post = wb2['CABA_val']

gps_time, Ux_gps_post, Uy_gps_post, Uz_gps_post = GPSlatlon2XY_time(gps_data_post, theta)


Uz = np.array([[gps_data_post.cell(row = i, column = 3).value] for i in range(2, gps_data_post.max_row + 1)]).reshape(
	gps_data_post.max_row - 1, )

gps_time_days = (gps_time - 2012.68)*365
#gps_time_days = gps_time


indice_gps = np.abs(np.sqrt(pow(X_gps[:]-X_gps_post,2) + pow(Y_gps[:]-Y_gps_post,2))).argmin()
indice_gps_model = np.abs(np.sqrt(pow(X_gps_model[:]-X_gps_post,2) + pow(Y_gps_model[:]-Y_gps_post,2))).argmin()



################################################################################################
def shiftedColorMap(cmap, start = 0, midpoint = 0.5, stop = 1.0, name = 'shiftedcmap'):
	'''
	Function to offset the "center" of a colormap. Useful for
	data with a negative min and positive max and you want the
	middle of the colormap's dynamic range to be at zero

	Input
	-----
	  cmap : The matplotlib colormap to be altered
	  start : Offset from lowest point in the colormap's range.
		  Defaults to 0.0 (no lower ofset). Should be between
		  0.0 and `midpoint`.
	  midpoint : The new center of the colormap. Defaults to
		  0.5 (no shift). Should be between 0.0 and 1.0. In
		  general, this should be  1 - vmax/(vmax + abs(vmin))
		  For example if your data range from -15.0 to +5.0 and
		  you want the center of the colormap at 0.0, `midpoint`
		  should be set to  1 - 5/(5 + 15)) or 0.75
	  stop : Offset from highets point in the colormap's range.
		  Defaults to 1.0 (no upper ofset). Should be between
		  `midpoint` and 1.0.
	'''
	cdict = {
		'red':[],
		'green':[],
		'blue':[],
		'alpha':[]
	}

	# regular index to compute the colors
	reg_index = np.linspace(start, stop, 257)

	# shifted index to match the data
	shift_index = np.hstack([
		np.linspace(0.0, midpoint, 128, endpoint = False),
		np.linspace(midpoint, 1.0, 129, endpoint = True)
	])

	for ri, si in zip(reg_index, shift_index):
		r, g, b, a = cmap(ri)

		cdict['red'].append((si, r, r))
		cdict['green'].append((si, g, g))
		cdict['blue'].append((si, b, b))
		cdict['alpha'].append((si, a, a))

	newcmap = mpl.colors.LinearSegmentedColormap(name, cdict)
	plt.register_cmap(cmap = newcmap)

	return newcmap

def plot_gps_u(endtime, ngps):

	for i in range(0,ngps):
		idx_gps = indice_gps
		fig, sol_plotgps = plt.subplots(3, sharex=True)


		#sol_plotgps[0].plot(Ux_gps_model_all[idx_gps, 0:endtime], '--k', label='constant k', linewidth=3, fillstyle='none') #pressure

		# sol_plotgps[1].plot(dtt_days[0:endtime], Uy_gps_model[idx_gps, 0:endtime], '--r', linewidth=3, fillstyle='none') # u_x
		#
		# sol_plotgps[2].plot(dtt_days[0:endtime], Uz_gps_model[idx_gps, 0:endtime], '--g', linewidth=3, fillstyle='none') # u_y

		sol_plotgps[0].plot(dtt_days[0:endtime],1e2*Ux_gps_model_all[idx_gps, 0:endtime], 'k', linewidth=3, fillstyle='none')

		sol_plotgps[1].plot(dtt_days[0:endtime], 1e2*Uy_gps_model_all[idx_gps, 0:endtime], 'r', linewidth=3, fillstyle='none')

		sol_plotgps[2].plot(dtt_days[0:endtime], 1e2*Uz_gps_model_all[idx_gps, 0:endtime], 'g', linewidth=3, fillstyle='none')

		sol_plotgps[0].plot(gps_time_days, Ux_gps_post, '.k', linewidth=3, fillstyle='none')

		sol_plotgps[1].plot(gps_time_days, Uy_gps_post, '.r', linewidth=3, fillstyle='none')

		sol_plotgps[2].plot(gps_time_days, Uz_gps_post, '.g', linewidth=3, fillstyle='none')

		# sol_plotgps[0].plot(Ux_gps_post, '*k', linewidth=3, fillstyle='none')
		#
		# sol_plotgps[1].plot(Uy_gps_post, '*r', linewidth=3, fillstyle='none')
		# sol_plotgps[2].plot(Uz_gps_post, '*g', linewidth=3, fillstyle='none')


		sol_plotgps[0].set_ylabel('X disp (cm)')
		sol_plotgps[1].set_ylabel('Y disp (cm)')
		sol_plotgps[2].set_ylabel('Z disp (cm)')

		sol_plotgps[0].set_xlim([0, 100])
		sol_plotgps[1].set_xlim([0, 100])
		sol_plotgps[2].set_xlim([0, 100])

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')


		fig.savefig('figures/3d_surf_GPS_model timeseries_%s.png' % station, dpi = 400)

def plot_gps(size_gps, n):

	for i in range(0,n):
		idx_gps = i
		fig, sol_plotgps = plt.subplots(4, sharex=True)


		sol_plotgps[0].plot(Sol_gps[idx_gps, :], '--k', label='constant k', linewidth=3, fillstyle='none') #pressure

		sol_plotgps[1].plot(Sol_gps[(size_gps + idx_gps), :], '--r', linewidth=3, fillstyle='none') # u_x

		sol_plotgps[2].plot(Sol_gps[2*size_gps + idx_gps, :], '--g', linewidth=3, fillstyle='none') # u_y

		sol_plotgps[3].plot(Sol_gps[3*size_gps + idx_gps, :], '--b', linewidth=3, fillstyle='none') # u_z



		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_surf():

	points = 1e3*np.array([100 , 130])

	for i in range(0, points.shape[1]):
		idx_p = i
		fig = plt.figure()
		ax1 = fig.add_subplot(111)

		ax1.plot(Sol_surf[idx, :], 'k*', label='mapped k', linewidth=3, fillstyle='none') #pressure

		#sol_plot_surf[1].plot(Sol_surf[(size_surf + idx), :], '--r', linewidth=3, fillstyle='none') # u_x

		#sol_plot_surf[2].plot(Sol_surf[2*size_surf + idx, :], '--g', linewidth=3, fillstyle='none') # u_y

		#sol_plot_surf[3].plot(Sol_surf[3*size_surf + idx, :], '--b', linewidth=3, fillstyle='none') # u_z

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_gps_all(nstations):

	for i in range(0,nstations):
		idx_gps = i
		fig, sol_plotgps_all = plt.subplots(4, sharex=True)

		sol_plotgps_all[0].plot(Sol_gps[idx_gps, :], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure
		sol_plotgps_all[0].plot(Sol_gps_k[idx_gps, :], '--k', label='mapped k', linewidth=3, fillstyle='none') #pressure

		sol_plotgps_all[1].plot(Sol_gps[(size_gps + idx_gps), :], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		sol_plotgps_all[1].plot(Sol_gps_k[(size_gps + idx_gps), :], '--r', linewidth=3, fillstyle='none') # u_x
		sol_plotgps_all[1].plot(x_elastic, elastic_gps_k[idx_gps, :],'ro', fillstyle='full') # u_x

		sol_plotgps_all[2].plot(Sol_gps[2*size_gps + idx_gps, :], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		sol_plotgps_all[2].plot(Sol_gps_k[2*size_gps + idx_gps, :], '--g', linewidth=3, fillstyle='none') # u_y
		sol_plotgps_all[2].plot(x_elastic, elastic_gps_k[size_gps + idx_gps, :],'go', fillstyle='full') # u_x

		sol_plotgps_all[3].plot(Sol_gps[3*size_gps + idx_gps, :], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		sol_plotgps_all[3].plot(Sol_gps_k[3*size_gps + idx_gps, :], '--b', linewidth=3, fillstyle='none') # u_z
		sol_plotgps_all[3].plot(x_elastic, elastic_gps_k[2*size_gps+idx_gps, :],'bo', fillstyle='full') # u_x

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_surf_all(Sol_surf, Sol_surf_k, elastic_surf, size_surf, x_surf, y_surf):

	points = 1e3*np.array([40 , 100])

	nsteps = Sol_surf.shape[1]
	x_elastic = np.array([0,nsteps-1])


	for i in range(0, points.shape[1]):

		idx = (np.abs(np.sqrt(pow((x_surf - plotpoint[0]),2) + pow((y_surf - plotpoint[1]),2)))).argmin()

		fig, sol_plot_surf_all = plt.subplots(4, sharex=True)

		sol_plot_surf_all[0].plot(Sol_surf[idx, :], '.k', label='constant k',ms=10, mew=1.5, fillstyle='none') #pressure
		sol_plot_surf_all[0].plot(Sol_surf_k[idx, :], '--k', label='mapped k', linewidth=3, fillstyle='none') #pressure

		sol_plot_surf_all[1].plot(Sol_surf[(size_surf + idx_gps), :], '.r', ms=10, mew=1.5, fillstyle='none') # u_x
		sol_plot_surf_all[1].plot(Sol_surf_k[(size_surf + idx_gps), :], '--r', linewidth=3, fillstyle='none') # u_x
		sol_plot_surf_all[1].plot(x_elastic, elastic_surf[idx_gps, :],'ro', fillstyle='full') # u_x

		sol_plot_surf_all[2].plot(Sol_surf[2*size_surf + idx_gps, :], '.g', ms=10, mew=1.5, fillstyle='none') # u_y
		sol_plot_surf_all[2].plot(Sol_surf_k[2*size_surf + idx_gps, :], '--g', linewidth=3, fillstyle='none') # u_y
		sol_plot_surf_all[2].plot(x_elastic, elastic_surf[size_gps + idx_gps, :],'go', fillstyle='full') # u_x

		sol_plot_surf_all[3].plot(Sol_surf[3*size_surf + idx_gps, :], '.b', ms=10, mew=1.5, fillstyle='none') # u_z
		sol_plot_surf_all[3].plot(Sol_surf_k[3*size_surf + idx_gps, :], '--b', linewidth=3, fillstyle='none') # u_z
		sol_plot_surf_all[3].plot(x_elastic, elastic_surf[2*size_gps+idx_gps, :],'bo', fillstyle='full') # u_x

		#plt.ylabel('Depth (km)')
		plt.xlabel('Days after EQ')
		legend1 = sol_plotgps[0].legend()
		for label in legend1.get_texts():
			label.set_fontsize('medium')

		for label in legend1.get_lines():
			label.set_linewidth(1.5)  # the legend line width

def plot_pressure_contour():
	fig = plt.figure()
	ax1 = fig.add_subplot(111)


	i = 1

	#stream = streamfun[:, i]
	Z = mag_flux_grid[:,:, i]



	triang = tri.Triangulation(x_surf_all, y_surf_all)
	triang_u = tri.Triangulation(x_ocean_u, y_ocean_u)
	contour_steps = 10
	num_lines = 10
	orig_cmap = cm.coolwarm
	midpoint = 1 - Z.max() / (Z.max() + abs(Z.min()))
	shifted_cmap = shiftedColorMap(orig_cmap, midpoint = midpoint)

	zero = np.array([0])



	# levels_10 = np.array([Z_10.min(), Z_10.min()/2, Z_10.min()/4, Z_10.min()/8, Z_10.min()/16,
	#                      0, Z_10.max()/16, Z_10.max()/8, Z_10.max()/4, Z_10.max()/2, Z_10.max()])
	levels = np.linspace(seaflux_mag[:,i].min(), seaflux_mag[:,i].max(), num = contour_steps, endpoint = True)
	lines = np.linspace(seaflux_mag[:,i].min(), seaflux_mag[:,i].max(),  num = num_lines, endpoint = True)
	#lines_stream = np.linspace(stream.min(), stream.max(), num = num_lines, endpoint = True)


	skip = (slice(None, None, 10))

	contourf = ax1.contourf(x_grid,y_grid, Z, levels, cmap = orig_cmap)
	contour = ax1.contour(x_grid,y_grid, Z, lines, linewidth = 1, colors = 'black', linestyles = 'solid')


	#quiver = ax1.quiver(x_ocean_u[skip], y_ocean_u[skip], seaflux_x[skip][:, i], seaflux_y[skip][:, i], units = 'inches', scale_units = 'width', width = 1.7e-2, scale = .3/1e5, color = 'darkcyan')
	# quiver = ax1.quiver(x_ux[skip], y_ux[skip], vel_x[skip][:, i], vel_y[skip][:, i], scale_units = 'width', scale = 1.0 / 2e5)
	# streamline10 = ax1.streamplot(x_stream, y_stream, Z_x_stream[:, :, i], Z_y_stream[:, :, i], density = [1.0, 3.0], linewidth = lw)
	#streamfun_contour = ax1.tricontour(triang, stream, lines_stream, linewidth =1, colors = 'blue',linestyles = 'solid')



	fig.colorbar(contourf)
	#ax1.set_ylim([-2e3, 1e3])
	#ax1.set_xlim([55e3, 80e3])

	#fig.set_size_inches(8, 2)


	fig.savefig('figures/3d_contour.png', dpi = 400)

def plot_sea_flux():
	fig = plt.figure()
	ax1 = fig.add_subplot(111)
	# ax2 = fig.add_subplot(312)
	# ax3 = fig.add_subplot(313)
	# ax1.set_title('Flux through seafloor')
	
	end = 350
	
	# Co-seismic displacement
	ax1.plot(dtt_all, zero_line, '--k', linewidth = 1)
	ax1.plot(dtt_days[48:51+end]/30, flux_topo[48:51+end], '-b',label='Topographic flow',  linewidth = 2)
	#ax1.plot(dtt_all[51:51+250], flux_topo[51:51+250], '-b',  linewidth = 3)

	#ax1.plot(dtt_all, flux_subEQ, '-r',  linewidth = 3)
	#ax1.plot(dtt_all[48:50], flux_EQ[48:50], '-r', linewidth = 3)
	ax1.plot(dtt_days[48:51+end]/30, flux_subEQ[48:51+end], '-r', label='Tectonic flow', linewidth = 2)

	# ax3.plot(dtt[50:-1], sea_flux[50:-1], '-b*', ms=2, mew=2, fillstyle='none', color = 'red')
	# ax1.set_ylabel('Darcy flux (meters/year)')
	# ax2.set_ylabel('Darcy flux (meters/sec)')
	# ax3.set_ylabel('Darcy flux (meters/year)')
	# ax1.set_xlabel('Years')
	
	#ax1.set_yscale('log')
	ax1.set_ylim([-5e5, 1.5e6])
	ax1.set_xlim([-.5 , 12])
	ax1.set_xlabel('Months after earthquake')
	#ax1.legend(fontsize='small', loc=1)
	
	#fig.set_size_inches(8, 2.5)
	
	#becomes negative at 8 days, positive again at 215
	#above topo for 5 days
	
	fig.savefig('figures/3d_sea_flux.png', dpi = 400)

def plot_gps_surface():
	
	fig = plt.figure()
	ax1 = fig.add_subplot(111)

	
	# quiver = ax1.quiver(x_ocean_u[skip], y_ocean_u[skip], seaflux_x[skip][:, i], seaflux_y[skip][:, i],
	#                    units='inches',scale_units='width', width=1.7e-2, scale=.3 / 1e5, color='darkcyan')
	#
	quiver = ax1.quiver(X_gps/1e3, Y_gps/1e3, Ux_gps, Uy_gps, color = 'blue', label='GPS data',units='inches',
	                    scale_units='width', width=1e-2, scale=5e0)
	quiver2 = ax1.quiver(X_gps/1e3, Y_gps/1e3, Ux_gps_model, Uy_gps_model, color = 'red', label='model', units='inches',
	                     scale_units='width', width=1e-2, scale=5e0)

	# scale = 4.0
	# quiver = ax1.quiver(X_gps/1e3, Y_gps/1e3, Z_gps, Uz_gps, color = 'blue', label='GPS data',units='inches',
	#                     scale_units='width', width=2e-2, scale=10)
	# quiver2 = ax1.quiver(X_gps/1e3, Y_gps/1e3, Z_gps, Uz_gps_model, color = 'red', label='model',
	#                      units='inches',scale_units='width', width=2e-2, scale=10)
	# #
	# ax1.plot(Uz_gps, 'b*', color = 'blue', label='GPS data')
	# ax1.plot(Uz_gps_model, 'ro', color = 'red', label='model')
	#ax1.plot(Uz_gps_model/Uz_gps, 'ro', color = 'red', label = 'model')

	ax1.legend(fontsize='small', loc=1)
	ax1.set_ylim([50, 250])
	ax1.set_xlim([30 , 200])

	#fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_gps_surface_up.png', dpi = 400)
	#fig.savefig('figures/3d_gps_up_scatter.png', dpi = 400)

def plot_gps_surface_xyz():
	fig = plt.figure()
	fig, plot = plt.subplots(3, sharex = True)

	#
	plot[0].plot(Ux_gps, 'b*', color = 'blue', label='GPS data')
	plot[0].plot(Ux_gps_model, 'ro', color = 'red', label='model')
	plot[0].plot(Z_gps, '--k', label = 'model')

	plot[1].plot(Uy_gps, 'b*', color = 'blue', label='GPS data')
	plot[1].plot(Uy_gps_model, 'ro', color = 'red', label='model')
	plot[1].plot(Z_gps, '--k', label = 'model')

	plot[2].plot(Uz_gps, 'b*', color = 'blue', label='GPS data')
	plot[2].plot(Uz_gps_model, 'ro', color = 'red', label='model')
	plot[2].plot(Z_gps, '--k', label = 'model')
	#ax1.plot(Uz_gps_model/Uz_gps, 'ro', color = 'red', label = 'model')

	plot[0].legend(fontsize = 'small', loc = 4)
	# ax1.set_ylim([0, 250])
	# ax1.set_xlim([0 , 250])

	# fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_gps_surface_compare.png', dpi = 400)

	fig = plt.figure()
	fig, plot = plt.subplots(3, sharex = True)

	plot[0].plot((Ux_gps_model-Ux_gps)/Ux_gps, '.b', color = 'blue', label='GPS data')
	plot[0].plot(Z_gps, '--k', label = 'model')

	plot[1].plot((Uy_gps_model-Uy_gps)/Uy_gps, '.b', color = 'blue', label='GPS data')
	plot[1].plot(Z_gps, '--k', label = 'model')

	plot[2].plot((Uz_gps_model-Uz_gps)/Uz_gps, '.b', color = 'blue', label='GPS data')
	plot[2].plot(Z_gps, '--k', label = 'model')
	#ax1.plot(Uz_gps_model/Uz_gps, 'ro', color = 'red', label = 'model')

	#plot[0].legend(fontsize = 'small', loc = 1)
	# ax1.set_ylim([0, 250])
	# ax1.set_xlim([0 , 250])

	# fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_gps_surface_compare_ratio.png', dpi = 400)

def plot_slab_depth():
	fig = plt.figure()
	ax1 = fig.add_subplot(111)


	ax1.plot(X_slab / 1e3, Z_slab / 1e3, '.k', label = 'GPS data')
	ax1.plot(X_slab_model / 1e3, Z_slab_model / 1e3, '.r', label = 'model',)

	ax1.legend(fontsize = 'small', loc = 1)
	# ax1.set_ylim([0, 250])
	# ax1.set_xlim([0 , 250])

	# fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_slab.png', dpi = 400)

def plot_slab_contour():
	fig = plt.figure()
	ax1 = fig.add_subplot(111)

	#triang = tri.Triangulation(X_slab, Y_slab, slab_mesh.cells())
	triang = tri.Triangulation(1e-3*X_slab, 1e-3*Y_slab)


	Z = U_mag_slab[:]
	Z_slab_plot = 1e-3*Z_slab
	Zmin, Zmax = Z.min(), Z.max()

	contour_steps = 10
	num_lines = 10
	#orig_cmap = cm.coolwarm
	#midpoint = 1 - Zmax / (Zmax + abs(Zmin))
	#shifted_cmap = shiftedColorMap(orig_cmap, midpoint = midpoint)

	zero = np.array([0])

	# levels_10 = np.array([Z_10.min(), Z_10.min()/2, Z_10.min()/4, Z_10.min()/8, Z_10.min()/16,
	#                      0, Z_10.max()/16, Z_10.max()/8, Z_10.max()/4, Z_10.max()/2, Z_10.max()])
	levels = np.linspace(Zmin, Zmax, num = contour_steps, endpoint = True)
	levels = np.linspace(0.0, 3., num = 13, endpoint = True)

	lines = np.linspace(Zmin, Zmax, num = num_lines, endpoint = True)
	#lines_slab = np.linspace(Z_slab.min, Z_slab.max(), num = 5, endpoint = True)
	lines_slab = -1.0*np.array([50,40,30,20,10])


	skip = (slice(None, None, 7))


	cpf = ax1.tricontourf(triang, Z, levels, cmap = cm.jet, alpha=.5)
	#cp = ax1.tricontour(triang, Z, lines, linewidth = 10 ,colors = 'black', linestyles = 'solid')
	quiver10 = ax1.quiver(1e-3*X_slab[skip], 1e-3*Y_slab[skip], Ux_slab[skip], Uy_slab[skip],  units='inches',
	                     scale_units='width', width=1.5e-2, scale=30e0)

	cp = ax1.tricontour(triang, Z_slab_plot, lines_slab, linewidth = 10 ,colors = 'black', linestyles = 'dashed')


	#ax1.legend(fontsize = 'small', loc = 1)
	plt.axis('scaled')
	ax1.set_ylim([0, 250])
	ax1.set_xlim([0, 130])

	#set tickmarks
	majorLocator = MultipleLocator(25)
	majorFormatter = FormatStrFormatter('%d')
	minorLocator = MultipleLocator(5)
	ax1.xaxis.set_major_locator(majorLocator)
	ax1.xaxis.set_major_formatter(majorFormatter)
	# for the minor ticks, use no labels; default NullFormatter
	ax1.xaxis.set_minor_locator(minorLocator)
	ax1.tick_params(which = 'both', direction = 'out')

	fig.colorbar(cpf)


	# fig.savefig('figures/3d_gps_surface_XY.png', dpi = 400)
	fig.savefig('figures/3d_slab.png', dpi = 400)



#points = 1e3*np.array([[40, 140], [50, 140], [60.5, 140], [70, 140], [80, 140], [90, 140], [100, 140], [110, 140], [120, 140], [130, 140]])
#points = 1e3*np.array([[78, 140], [96, 129], [61, 160], [114,124]])
points = 1e3*np.array([[83.5, 141]])


splice = 1
endtime = 300

#plot_surf_k()
#plot_gps_elastic_compare(20)
#plot_surf_k(wellpoints, splice)
#plot_pressure_contour()
#plot_sea_flux()
#plot_gps_surface()
#plot_gps_surface_xyz()
#plot_gps_u(endtime, 1)
#plot_slab_depth()
plot_slab_contour()

