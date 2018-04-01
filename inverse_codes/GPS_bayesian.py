import petsc4py as p4p
from dolfin import *
import sys
sys.path.append('/fenics/shared/local_lib')
from hippylib import *
import scipy.sparse as sp
import time
import logging
import numpy as np
import matplotlib.pyplot as plt
import openpyxl as pyxl
import time


parameters.reorder_dofs_serial = False
dir_path = '/fenics/shared/'

def u_boundary(x, on_boundary):
    return on_boundary

def GPSlatlon2XY(data_sheet, origin, theta):
    """
    latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
    into numerical models.

    INPUTS:
    [data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
    [origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
    [theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
                counter-clockwise from latlon to XY

    Assumes first column is longitude and second column is latitude

    OUTPUTS:
    Returns transformed coordinates as numpy vectors X, Y in kilometers
    """
    
    lon = np.array([[data_sheet.cell(row=i, column=1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    lat = np.array([[data_sheet.cell(row=i, column=2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    
    lon_u = np.array([[data_sheet.cell(row=i, column=5).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    lat_u = np.array([[data_sheet.cell(row=i, column=6).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    Uz = np.array([[data_sheet.cell(row=i, column=4).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    
    lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
    lat_in_km = (lat - origin[1]) * 111
    
    rho_u = np.sqrt(np.power(lon_u, 2) + np.power(lat_u, 2))
    theta_new_u = np.arctan2(lat_u, lon_u) - theta
    
    rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
    theta_new = np.arctan2(lat_in_km, lon_in_km) - theta
    
    X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)
    Ux, Uy = rho_u * np.cos(theta_new_u), rho_u * np.sin(theta_new_u)
    
    return 1e3 * X, 1e3 * Y, 1e-3*Ux, 1e-3*Uy, 1e-3*Uz

class Elasticity:
    def __init__(self, mesh, boundaries, Vh, u0true, targets, u_gps, prior, invGnoise, point_var, synthetic):
        """
        Construct a model by providing
        - the mesh
        - the finite element spaces for the STATE/ADJOINT variable and the PARAMETER variable
        - the Prior information
        """
        self.mesh = mesh
        self.boundaries = boundaries
        self.Vh = Vh
        self.ds = Measure('ds', domain=self.mesh, subdomain_data=self.boundaries)
        
        ################ PARAMETERS ###############
        self.synthetic = synthetic
        self.point_variance = point_var
        self.invGnoise = invGnoise

        self.st_dev_prior = 2; # standard deviation of prior
        self.correlation_length = 2e4 # length over which points share 10% correlation
        self.corr = -np.log(.1)/self.correlation_length #regularization parameter
        
        self.area = assemble(Constant(1.)*self.ds(2))
  
        self.h = CellSize(self.mesh)
        self.delta_b = Constant(1e-6)#*self.h #Dirichlet control regularization on bottom boundary
        #self.gamma = 1e-6 # Tikinov regularization parameter
    
        self.E = Expression('(5.7972e-10*(-(pow(x[2],3)))) - (7.25283e-5*(-(pow(x[2],2)))) + (3.8486e-6*(-x[2])) + 6.94e4', degree = 1) # Young's modulus
        self.nu = 0.25  # Poisson's ratio
        self.mu   = self.E / (2.0*(1.0 + self.nu)) # shear modulus
        self.lmbda = self.E*self.nu / ((1.0 + self.nu)*(1.0 - 2.0*self.nu))# Lame's parameter
        
        
        # Initialize Expressions
        self.dim = 3
        self.u0true = u0true
        self.f = Constant(('0.0','0.0','0.0'))
        self.I = Identity(self.dim)
        self.n = FacetNormal(mesh)
        self.T = self.I - outer(self.n,self.n)
        
        self.b0 = Constant(0.0)
        self.b30 = Constant((0.0,0.0,0.0))

        self.bc1 = DirichletBC(self.Vh[STATE], self.b30, self.boundaries, 1) # ocean
        self.bc2 = DirichletBC(self.Vh[STATE], self.b30, self.boundaries, 2) # land
        
        self.bc3 = DirichletBC(self.Vh[STATE].sub(0), self.b0, self.boundaries, 6) #back
        self.bc4 = DirichletBC(self.Vh[STATE], self.b30, self.boundaries, 7) # mantle
        self.bc5 = DirichletBC(self.Vh[STATE], self.b30, u_boundary) # adj var = 0 on all boundaries
        
        self.bcstate = [self.bc3] # u = 0 on sides and bottom for state eq. and everywhere for adjoint
        self.bc_adj = [self.bc3]
        
        # Assemble the observation operator, B,
        self.prior = prior
     
        self.B = assemblePointwiseObservation(Vh[STATE],targets)
        #coords = self.Vh[STATE].tabulate_dof_coordinates().reshape((self.Vh[STATE].dim(), -1))
        #self.B = self.build_B(coords, targets[:,0], targets[:,1])
 
        self.invG = self.build_invG(self.invGnoise)

       
        if self.synthetic:
            self.u_obsall = Vector()
            self.computeObservation(self.u_obsall)
            self.u_obs = self.B * self.u_obsall
            
        if not self.synthetic:
            self.u_obs = Vector()
            self.B.init_vector(self.u_obs, 0)
            self.u_obs.set_local(u_gps)
            
            
        print "obs max, min: ", self.u_obs.max(), self.u_obs.min()


        if not self.point_variance:
            self.noise_variance = 1e-2
        
        #print "noise: ", self.noise_variance, MAX
        #quit()
        
        self.A = []
        self.At = []
        self.C = []
        self.Raa = []
        self.Wau = []

    def extract_coords(self):
        if not os.path.exists(self.var_path):
            os.makedirs(self.var_path)

        self.coordinates = self.Vh1.tabulate_dof_coordinates().reshape((self.Vdim, -1))
        self.x_all, self.y_all, self.z_all = self.coordinates[:, 0], self.coordinates[:, 1], self.coordinates[:, 2]

        bctest0 = DirichletBC(self.Vh1, (1, 1, 1), self.boundaries, 1)  # ocean surface
        bctest1 = DirichletBC(self.Vh1, (2, 2, 2), self.boundaries, 2)  # top surface
        bctest2 = DirichletBC(self.Vh1, (3, 3, 3), self.boundaries, 3)  # slab surface

        ptest = Function(self.Vh1)
        bctest0.apply(ptest.vector())
        bctest1.apply(ptest.vector())
        bctest2.apply(ptest.vector())

        self.ocean_dofs = np.where(ptest.vector() == 1)[0]
        self.surface_dofs = np.where(ptest.vector() == 2)[0]
        self.slab_dofs = np.where(ptest.vector() == 3)[0]



        self.size_gps = self.targets.shape[0]
        size_w_gps = 3 * self.size_gps
        indices_gps = np.empty((self.size_gps, 3))
        for j in range(0, self.size_gps):
            indice_gps = np.where(np.sqrt(
                pow(self.x_all[:] - self.targets[j,0], 2) + pow(self.y_all[:] - self.targets[j,1], 2) + pow(
                    self.z_all[:] - 0.0, 2)) < 0.1)[0]
            indices_gps[j, :] = indice_gps
            self.GPS_dofs = indices_gps.reshape((size_w_gps), order = 'F')
        # print indice_gps

        XYZ_slab = np.vstack((self.x_all[self.slab_dofs], self.y_all[self.slab_dofs], self.z_all[self.slab_dofs])).T

        np.save(self.var_path + "XYZ_slab", XYZ_slab)
        np.save(self.var_path + "x_all_3D", self.x_all)
        np.save(self.var_path + "y_all_3D", self.y_all)
        np.save(self.var_path + "z_all_3D", self.z_all)
        np.save(self.var_path + "GPS_dofs", self.GPS_dofs)
        np.save(self.var_path + "surface_dofs_u_3D", self.surface_dofs)
        np.save(self.var_path + "ocean_dofs_u_3D", self.ocean_dofs)
        np.save(self.var_path + "slab_dofs_u_3D", self.slab_dofs)

    def build_B(self, coords, X, Y):
    
        #coords = np.around(coords)
        sizeU = coords.shape[0]
        nobs = X.shape[0]

        Bp4p = p4p.PETSc.Mat()
        Bp4p.create(p4p.PETSc.COMM_WORLD)
        Bp4p.setSizes([3 * nobs, sizeU])
        Bp4p.setType('aij')  # sparse
        Bp4p.setPreallocationNNZ(5)
        Bp4p.setUp()
        Istart, Iend = Bp4p.getOwnershipRange()

        for j in range(0, nobs):
            ind = np.where(np.sqrt(pow((coords[:, 0] - X[j]), 2) + pow((coords[:, 1] - Y[j]), 2) + pow((coords[:, 2]), 2)) < 2.0)[0]
            Bp4p[j, ind[0]], Bp4p[j + nobs, ind[1]], Bp4p[j + 2 * nobs, ind[2]] = 1.0, 1.0, 1.0

        Bp4p.assemblyBegin()
        Bp4p.assemblyEnd()

        B = PETScMatrix(Bp4p)
        
        return B

    def build_invG(self, invGnoise):
    

        invGp4p = p4p.PETSc.Mat()
        invGp4p.create(p4p.PETSc.COMM_WORLD)
        invGp4p.setSizes([invGnoise.shape[0], invGnoise.shape[1]])
        invGp4p.setType('aij')  # sparse
        #invGp4p.setPreallocationNNZ(5)
        invGp4p.setUp()
    
        for j in range(0, invGnoise.shape[0]):
            
            invGp4p[j, j] = invGnoise[j, j]
    
        invGp4p.assemblyBegin()
        invGp4p.assemblyEnd()
    
        invG = PETScMatrix(invGp4p)
    
        return invG

    def generate_vector(self, component="ALL"):
        """
        Return the list x=[u,a,p] where:
        - u is any object that describes the state variable
        - a is a Vector object that describes the parameter.
          (Need to support linear algebra operations)
        - p is any object that describes the adjoint variable
        
        If component is STATE, PARAMETER, or ADJOINT return x[component]
        """
        if component == "ALL":
            x = [Vector(), Vector(), Vector()]
            self.B.init_vector(x[STATE],1)
            self.prior.init_vector(x[PARAMETER],0) # I think this initializing a vetor to be compatible with the R prior matrix in a given dimension
            self.B.init_vector(x[ADJOINT], 1)
        elif component == STATE:
            x = Vector()
            self.B.init_vector(x,1)
        elif component == PARAMETER: # This is compatible with the prior matrix bc the the gradient needs the
            x = Vector()
            self.prior.init_vector(x,0)
        elif component == ADJOINT:
            x = Vector()
            self.B.init_vector(x,1)
            
        return x
    
    def init_parameter(self, a):
        """
        Reshape a so that it is compatible with the parameter variable
        """
        self.prior.init_vector(a,0)

    ##### WEAK FORM DEFINITIONS #####
    def strain(self, w): # strain = 1/2 (grad u + grad u^T)
        return sym(nabla_grad(w))
    
    def elasticity(self, u, v): # elasticity weak form with slip boundary conditions
        return inner(2.0*self.mu*self.strain(u), self.strain(v))*dx + inner(self.lmbda*nabla_div(u), nabla_div(v))*dx\
               + (1.0/self.delta_b)*inner(self.T*u ,v)*self.ds(3) \
               + dot(u,self.n)*dot(v,self.n)*self.ds(3)

    def constraintB (self, u, v): # slip boundary condition
        return (1.0/self.delta_b)*inner(self.T*u ,v)*self.ds(3)
    #def reg(self, u, v): # regularization
        #return self.gamma*inner(nabla_grad(u), nabla_grad(v))*ds(3) # mass matrix or Ident_zero

    def assembleA(self,x, assemble_adjoint = False, assemble_rhs = False):
        """
        Assemble the matrices and rhs for the forward/adjoint problems
        """
              
        trial = TrialFunction(self.Vh[STATE])
        test = TestFunction(self.Vh[STATE])
        u0 = vector2Function(x[PARAMETER], self.Vh[PARAMETER])
        
        Avarf = self.elasticity(trial, test)  #Weak form of the forward problem
        
        if not assemble_adjoint:
            bform = self.constraintB(u0, test)
            Matrix, rhs = assemble_system(Avarf, bform, self.bcstate)
        
        else:

            # Assemble the adjoint of A (i.e. the transpose of A)
            bform = inner(self.f, test)*dx
            Matrix, _ = assemble_system(adjoint(Avarf), bform, self.bc_adj)
            Bu = -(self.B*x[STATE])
            Bu += self.u_obs  # These are the measured or synthetic observations
            rhs = Vector()
            self.B.init_vector(rhs, 1)

            if self.point_variance:
                GBu = self.invG * Bu
                self.B.transpmult(GBu,rhs)

            else:
                self.B.transpmult(Bu,rhs)
                rhs *= 1.0/self.noise_variance
            


        if assemble_rhs:
            return Matrix, rhs
        else:
            return Matrix
    
    def assembleC(self, x):
        """
        Assemble the derivative of the forward problem with respect to the parameter
        """
        trial = TrialFunction(self.Vh[PARAMETER]) #u0
        test = TestFunction(self.Vh[STATE]) # v_tilda
        Cvarf = -self.constraintB(trial, test)
        C = assemble(Cvarf)
        #print "||c||", x[PARAMETER].norm("l2"), "||s||", x[STATE].norm("l2"), "||C||", C.norm("linf")

        [bc.zero(C) for bc in self.bcstate]
        return C

    def assembleWau(self, x): # ZERO
        """
        Assemble the derivative of the parameter equation with respect to the state
        """
        trial = TrialFunction(self.Vh[STATE])
        test  = TestFunction(self.Vh[PARAMETER])
        varf = inner(trial,test)*dx
        Wau = assemble(varf)
        Wau.zero()
        return Wau
    
    def assembleRaa(self, x): # ZERO
        """
        Assemble the derivative of the parameter equation with respect to the parameter (Newton method)
        """
        trial = TrialFunction(self.Vh[PARAMETER])
        test  = TestFunction(self.Vh[PARAMETER])
        varf = inner(trial,test)*dx
        Raa = assemble(varf)
        Raa.zero()
        return Raa

    def computeObservation(self, u_o):
        """
        Compute the syntetic observation
        """
        ut = interpolate(self.u0true, Vh[PARAMETER])
        x = [self.generate_vector(STATE), ut.vector(), None]
        A, b = self.assembleA(x, assemble_rhs = True)

        A.init_vector(u_o, 1)
        solve(A, u_o, b)
        self.u0_fwdsolve = u_o

        # Create noisy data, ud
        MAX = u_o.norm("linf")
        noise = .01 * MAX * np.random.normal(0, 1, len(u_o.array()))
        u_o.set_local(u_o.array() + noise)
                
        #plot(vector2Function(u_o, Vh[STATE]), title = "Observation")
        #self.noise_variance = (.01*MAX)**2
        #print self.noise_variance

        self.paraviewpath = 'results_cont/'
        self.u0file = File(self.paraviewpath + 'u_obs.pvd')
        self.u0file << vector2Function(u_o, Vh[STATE], name = "noisy displacement")
        #quit()

    def cost(self, x):
        """
        Given the list x = [u,a,p] which describes the state, parameter, and
        adjoint variable compute the cost functional as the sum of 
        the misfit functional and the regularization functional.
        
        Return the list [cost functional, regularization functional, misfit functional]
        
        Note: p is not needed to compute the cost functional
        """        
        assert x[STATE] != None
                       
        diff = self.B*x[STATE]
        diff -= self.u_obs

        if self.point_variance:
            Gdiff = self.invG*diff
            misfit = 0.5 * diff.inner(Gdiff)

        else:
            misfit = (.5 / self.noise_variance) * diff.inner(diff)

        Rdiff_x = Vector()
        self.prior.init_vector(Rdiff_x,0)
        diff_x = x[PARAMETER] - self.prior.mean
        self.prior.R.mult(diff_x, Rdiff_x)
        reg = .5 * diff_x.inner(Rdiff_x)
        
        c = misfit + reg #### COST FUNCTIONAL ###################
        
        return c, reg, misfit

    def solveFwd(self, out, x, tol=1e-9):
        """
        Solve the forward problem.
        """
        A, b = self.assembleA(x, assemble_rhs = True)
        A.init_vector(out, 1)
        #solver = PETScKrylovSolver("cg", amg_method()) # try linear solver
        solver  = PETScLUSolver("umfpack")
        #solver.parameters["relative_tolerance"] = tol
        solver.set_operator(A)
        nit = solver.solve(out,b)
        #print "FWD", (self.A*out - b).norm("l2")/b.norm("l2"), nit

    def solveAdj(self, out, x, tol=1e-9):
        """
        Solve the adjoint problem.
        """
        At, badj = self.assembleA(x, assemble_adjoint = True,assemble_rhs = True)
        At.init_vector(out, 1)
        solver  = PETScLUSolver("umfpack")

        #solver = PETScKrylovSolver("cg", amg_method())
        #solver.parameters["relative_tolerance"] = tol
        solver.set_operator(At)
        nit = solver.solve(out,badj)
        
        #print "ADJ", (self.At*out - badj).norm("l2")/badj.norm("l2"), nit
    
    def evalGradientParameter(self,x, mg):
        """
        Evaluate the gradient for the variation parameter equation at the point x=[u,a,p].
        Parameters:
        - x = [u,a,p] the point at which to evaluate the gradient.
        - mg the variational gradient (g, atest) being atest a test function in the parameter space
          (Output parameter)
        
        Returns the norm of the gradient in the correct inner product g_norm = sqrt(g,g)
        """ 
        C = self.assembleC(x)

        self.prior.init_vector(mg,0)
        C.transpmult(x[ADJOINT], mg)
        Rdx = Vector()
        self.prior.init_vector(Rdx,0)
        dx = x[PARAMETER] - self.prior.mean
        self.prior.R.mult(dx, Rdx)   
        mg.axpy(1., Rdx)
        
        g = Vector()
        self.prior.init_vector(g,1)
        
        self.prior.Msolver.solve(g, mg)
        g_norm = sqrt( g.inner(mg) )
        
        return g_norm
           
    def setPointForHessianEvaluations(self, x):  
        """
        Specify the point x = [u,a,p] at which the Hessian operator (or the Gauss-Newton approximation)
        need to be evaluated.
        """      
        self.A  = self.assembleA(x)
        self.At = self.assembleA(x, assemble_adjoint=True )
        self.C  = self.assembleC(x)
        self.Wau = self.assembleWau(x)
        self.Raa = self.assembleRaa(x)

    def solveFwdIncremental(self, sol, rhs, tol):
        """
        Solve the incremental forward problem for a given rhs
        """
        solver  = PETScLUSolver("umfpack")

        #solver = PETScKrylovSolver("cg", amg_method())
        #solver.parameters["relative_tolerance"] = tol

        solver.set_operator(self.A)
        self.A.init_vector(sol,1)
        nit = solver.solve(sol,rhs)
        #print "FwdInc", (self.A*sol-rhs).norm("l2")/rhs.norm("l2"), nit
        
    def solveAdjIncremental(self, sol, rhs, tol):
        """
        Solve the incremental adjoint problem for a given rhs
        """
        
        solver  = PETScLUSolver("umfpack")
        
        #solver = PETScKrylovSolver("cg", amg_method())
        #solver.parameters["relative_tolerance"] = tol
        
        solver.set_operator(self.At)
        self.At.init_vector(sol,1)
        nit = solver.solve(sol, rhs)
        #print "AdjInc", (self.At*sol-rhs).norm("l2")/rhs.norm("l2"), nit
    
    def applyC(self, da, out):
        self.C.mult(da,out)
    
    def applyCt(self, dp, out):
        self.C.transpmult(dp,out)

    def applyWuu(self, du, out, gn_approx=False):
        
        # out = model.generate_vector(ADJOINT)
        help = Vector()
        self.B.init_vector(help, 0)
        self.B.mult(du, help) #du = uhat = model.generate_vector(STATE) in reducedHessian.py

        if self.point_variance:
            Ghelp = self.invG*help
            self.B.transpmult(Ghelp, out)
        else:
            self.B.transpmult(help, out)
            out *= 1.0/self.noise_variance

    def applyWua(self, da, out):
        self.Wau.transpmult(da,out)

    def applyWau(self, du, out):
        self.Wau.mult(du, out)
    
    def applyR(self, da, out):
        self.prior.R.mult(da, out)
        
    def Rsolver(self):        
        return self.prior.Rsolver
    
    def applyRaa(self, da, out):
        self.Raa.mult(da, out)


################### SET UP MESH, MODEL PARAMETERS AND INPUT DATA ##########################################


if __name__ == "__main__":
    set_log_active(False)

    start = time.time()

    sep = "\n"+"#"*80+"\n"
    print sep, "Set up the mesh and finite element spaces: ",(time.time() - start) / 60.0, "minutes", sep
    
    ############ IMPORT MESH (made with gmsh)

    path = "/fenics/shared/meshes"
    mesh_size = 'coarse'  # fine, med, coarse

    mesh = Mesh(path + '/3D_' + mesh_size + '.xml')
    boundaries = MeshFunction("size_t", mesh, path + '/3D_' + mesh_size + '_facet_region.xml')

    Vh2 = VectorFunctionSpace(mesh, 'Lagrange', 1)
    Vh1 = VectorFunctionSpace(mesh, 'Lagrange', 1)
    Vh = [Vh1, Vh1, Vh1]
    print "Number of dofs: STATE={0}, PARAMETER={1}, ADJOINT={2}".format(Vh[STATE].dim(), Vh[PARAMETER].dim(), Vh[ADJOINT].dim())

    dim = 3
    
    ################ USER INPUTS ###############

    rel_noise = .02
    correlation_length = 5e3  # length over which points share 10% correlation
    gamma = 10e2
    point_var = True
    synthetic = True
    
    sigmab = 2e4
    xcenter = 80e3
    ycenter = 130e3
    u0d, u0s = -4.0 , 0.0

    ################ DATA ###############


    print sep, "Set up the location of observation, Prior Information, and model: ",(time.time() - start) / 60.0, "minutes", sep
    nobs = 37
    np.random.seed(seed=1)

    origin = np.array([-85.21, 8.64])  # origin of top surface of mesh - used as point of rotation
    theta = 0.733  # angle of rotation between latitude line and fault trace
    data_file = dir_path + 'data/Data.xlsx'
    wb = pyxl.load_workbook(data_file)
    gps_data = wb['GPS_data']

    X_gps, Y_gps, Ux_gps, Uy_gps, Uz_gps = GPSlatlon2XY(gps_data, origin, theta)
    Z_gps = np.zeros(len(X_gps))
    
    
    targets = np.concatenate((X_gps.reshape(nobs,1), Y_gps.reshape(nobs,1), Z_gps.reshape(nobs,1)),axis=1)
    #might need to arange as [pt1x,pt1y,pt1x..... pt37x, pt37y, pt37z]
    
    ##### Use my B operator
    #u_gps = np.concatenate((Ux_gps.reshape(nobs,1), Uy_gps.reshape(nobs,1),
    #                        Uz_gps.reshape(nobs,1)),axis=0).reshape(3*nobs,)# order="F")
    
    ####Use hippylib B operator
    u_gps = np.concatenate((Ux_gps.reshape(nobs,1), Uy_gps.reshape(nobs,1),
                            Uz_gps.reshape(nobs,1)),axis=1).reshape(3*nobs,)# order="F")


    print "Number of observation points: {0}".format(ntargets)

    ##### NOISE MATRIX #####
    
    
    #might need to arange as [pt1x,pt1y,pt1x..... pt37x, pt37y, pt37z]
    stddev = .001*np.array([4.4,1.7,0.8,1.2,0.8,1.7,1.3,0.9,1,0.8,1.3,0.8,1.9,0.8,0.8,0.8,0.8,1,6.5,4.6,6.5,7.8,5.8,4.5,6.2,4.8,7.5,5.1,
                            5.3,6.3,6.6,6.3,4.6,5.1,6.4,5.2,5.2,5.1,1.6,0.7,0.9,0.7,1.2,0.9,0.6,0.9,0.7,1,0.6,1.7,0.6,0.7,0.7,0.7,0.8,
                            2.5,2.6,2.5,4,2.2,2,1.9,2.3,2.3,2.2,2.1,3,2.7,2.5,2.1,2.1,2.6,2.3,2.1,15.7,6.6,2.8,4.3,2.8,5.3,4.1,2.7,4,3.8,
                            4.4,2.6,8.2,2.6,3.2,2.8,3,3.5,11.4,8.9,10.3,15.1,10,8.4,9.9,8.6,10.6,9.4,8.7,15,10.7,10,8,8.3,10.5,9.1,8.4])

    stddev = stddev.reshape((nobs, 3),order='F').reshape(3*nobs,)


    Gnoise = np.diag(stddev**2)
    invGnoise = np.linalg.inv(Gnoise)

    
    #gives good inversion
    #correlation_length = 5e3  # length over which points share 10% correlation
    #gamma = 2.5e7 # double gamma cuts max slip by factor of 4
    #delta = gamma/(correlation_length**2)

    #Gives good prior range
    #correlation_length = 5e3  # length over which points share 10% correlation
    #gamma = 1e0

    delta = gamma/(correlation_length**2)

    prior = BiLaplacianPrior(Vh[PARAMETER], gamma, delta)

    noise = Vector()
    prior.init_vector(noise, "noise")
    noise_size = noise.array().shape[0]
    noise.set_local(rel_noise*np.random.randn(noise_size))
    

    s_prior = Function(Vh[PARAMETER])
    prior.sample(noise, s_prior.vector())

    File("results_point/prior_sample.pvd") << s_prior

    bcprior = DirichletBC(Vh[PARAMETER], (1,1,1), boundaries, 2)  # land surface
    priortest = Function(Vh[PARAMETER])
    bcprior.apply(priortest.vector())
    dof_surf_prior = np.where(priortest.vector() == 1)[0]

    #ds = Measure('ds', domain=mesh, subdomain_data=boundaries)
    #area = assemble(Constant(1.) * ds(2))

    print "Max surface deformation of Prior: ", s_prior.vector()[dof_surf_prior].max()
    print "Min surface deformation of Prior: ", s_prior.vector()[dof_surf_prior].min()

        
    print "Prior regularization: (delta - gamma*Laplacian)^order: delta={0}, gamma={1}, order={2}".format(delta, gamma,2)
    

    if synthetic:
        u0true = Expression(('u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigma,2)) + (pow((x[1] - \
    ycenter),2)/(2*pow(sigma,2))))) ', 'u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigma,2)) + \
    (pow((x[1] - ycenter),2)/(2*pow(sigma,2)))))', '0.0'), sigma=sigmab,xcenter=xcenter, ycenter=ycenter, u0d=u0d, u0s=u0s, degree=1)
    else:
        u0true = Expression(('0.0 ', '0.0', '0.0'), degree=1)


    prior_mean_expression = Expression(('0.0', '0.0','0.0'), degree=1)
    prior.mean = interpolate(prior_mean_expression, Vh[PARAMETER]).vector()
    


    ############### SOLVE #####################

    model = Elasticity(mesh, boundaries, Vh, u0true, targets, u_gps, prior, invGnoise, point_var, synthetic)
    
    print sep, "Test the gradient and the Hessian of the model: ",(time.time() - start) / 60.0, "minutes", sep

    #a0 = Function(Vh[PARAMETER])
    #a0.vector()[:] = model.u0_fwdsolve
    a0 = interpolate(Expression(('sin(x[0]/1e4)','sin(x[1]/1e4)','sin(x[2]/1e4)'), degree = 1), Vh[PARAMETER])

    modelVerify(model, a0.vector(), 1e-12)
    
    print sep, "Find the MAP point: ",(time.time() - start) / 60.0, "minutes", sep

    x = model.generate_vector()
    [u, a, p] = x
    model.solveFwd(u, [u, a, p], 1e-12)
    model.solveAdj(p, [u, a, p], 1e-12)
    mg = model.generate_vector(PARAMETER)
    grad_norm = model.evalGradientParameter([u, a, p], mg)

    print "(g,g) = ", grad_norm

    H = ReducedHessian(model, 1e-12, gauss_newton_approx=True, misfit_only=True)

    k = 80
    p = 10
    print "Double Pass Algorithm. Requested eigenvectors: {0}; Oversampling {1}.".format(k, p)
    print  (time.time() - start) / 60.0, "minutes"
    Omega = np.random.randn(a.array().shape[0], k + p)
    d, U = doublePassG(H, prior.R, prior.Rsolver, Omega, k)
    
    print "eigenvalues", d

    posterior = GaussianLRPosterior(prior, d, U)

    H.misfit_only = False

    solver = CGSolverSteihaug()
    solver.set_operator(H)
    solver.set_preconditioner(posterior.Hlr)
    solver.parameters["print_level"] = 1
    solver.parameters["rel_tolerance"] = 1e-12
    solver.solve(a, -mg)
    model.solveFwd(u, [u, a, p], 1e-12)

    total_cost, reg_cost, misfit_cost = model.cost([u, a, p])
    #total_cost, reg_cost, misfit_cost = model.cost_point([u, a, p])
    print "Total cost {0:5g}; Reg Cost {1:5g}; Misfit {2:5g}".format(total_cost, reg_cost, misfit_cost)

    # a0 = prior.mean.copy()
    # solver = ReducedSpaceNewtonCG(model)
    # solver.parameters["rel_tolerance"] = 1e-9
    # solver.parameters["abs_tolerance"] = 1e-12
    # solver.parameters["max_iter"]      = 25
    # solver.parameters["inner_rel_tolerance"] = 1e-15
    # solver.parameters["c_armijo"] = 1e-4
    # solver.parameters["GN_iter"] = 5
    # solver.parameters["cg_coarse_tolerance"] = 1e-3
    #
    # x = solver.solve(a0)
    #
    # if solver.converged:
    #     print "\nConverged in ", solver.it, " iterations."
    # else:
    #     print "\nNot Converged"
    #
    # print "Termination reason: ", solver.termination_reasons[solver.reason]
    # print "Final gradient norm: ", solver.final_grad_norm
    # print "Final cost: ", solver.final_cost
    #
    # print sep, "Compute the low rank Gaussian Approximation of the posterior", sep
    # model.setPointForHessianEvaluations(x)
    # Hmisfit = ReducedHessian(model, solver.parameters["inner_rel_tolerance"], gauss_newton_approx=False, misfit_only=True)
    # k = 50
    # p = 20
    # print "Double Pass Algorithm. Requested eigenvectors: {0}; Oversampling {1}.".format(k,p)
    # Omega = np.random.randn(x[PARAMETER].array().shape[0], k+p)
    # #d, U = singlePassG(Hmisfit, model.R, model.Rsolver, Omega, k, check_Bortho=True, check_Aortho=True, check_residual=True)
    # d, U = doublePassG(Hmisfit, prior.R, prior.Rsolver, Omega, k, check_Bortho=False, check_Aortho=False, check_residual=False)
    # posterior = GaussianLRPosterior(prior, d, U)

    print "Max deformation of Solution: ", a.array()[:].max()
    print "Min deformation of Solution: ", a.array()[:].min()
    print "prior max/post max: ", s_prior.vector()[:].min() / a.array()[:].min()
    
    if synthetic:
        print "true_max/post_max: ", model.u0_fwdsolve.array()[:].min()/a.array()[:].min()
    #print np.linalg.norm(model.u0_fwdsolve.array()[dof_surf_prior], np.inf)/ np.linalg.norm(a.array()[dof_surf_prior], np.inf)


    posterior.mean = x[PARAMETER]
    
    # post_tr, prior_tr, corr_tr = posterior.trace(method="Estimator", tol=1e-1, min_iter=20, max_iter=100)
    # print "Posterior trace {0:5e}; Prior trace {1:5e}; Correction trace {2:5e}".format(post_tr, prior_tr, corr_tr)
    # post_pw_variance, pr_pw_variance, corr_pw_variance = posterior.pointwise_variance("Exact")

    print sep, "Save State, Parameter, Adjoint, and observation in paraview: ",(time.time() - start) / 60.0, "minutes", sep
    xxname = ["State", "boundary_slip", "Adjoint"]
    xx = [Function(Vh[i], x[i], name=xxname[i]) for i in range(len(Vh))]
    File("results_point/elasticity_state.pvd") << xx[STATE]
    File("results_point/slip_inversion.pvd") << xx[PARAMETER]
    sliptrue = interpolate(model.u0true, Vh[PARAMETER])
    sliptrue.rename("slip true", "ignore_this")
    File("results_point/slip_true.pvd") << sliptrue
    File("results_point/elasticity_adjoint.pvd") << xx[ADJOINT]
    
    # exportPointwiseObservation(targets, model.u_obs, "results/poisson_observation.vtp")
    #
    # fid = File("results/pointwise_variance.pvd")
    # fid << Function(Vh[PARAMETER], post_pw_variance, name="Posterior")
    # fid << Function(Vh[PARAMETER], pr_pw_variance, name="Prior")
    # fid << Function(Vh[PARAMETER], corr_pw_variance, name="Correction")
    #
    
    print sep, "Generate samples from Prior and Posterior\n","Export generalized Eigenpairs", sep
    fid_prior = File("samples/sample_prior_point.pvd")
    fid_post  = File("samples/sample_post_point.pvd")
    nsamples = 5
    noise = Vector()
    posterior.init_vector(noise,"noise")
    noise_size = noise.array().shape[0]
    s_prior = Function(Vh[PARAMETER], name="sample_prior")
    s_post = Function(Vh[PARAMETER], name="sample_post")
    for i in range(nsamples):
        noise.set_local(rel_noise * np.random.randn(noise_size))
        posterior.sample(noise, s_prior.vector(), s_post.vector())
        fid_prior << s_prior
        fid_post << s_post
        print "sample: ", i, "--" ,(time.time() - start) / 60.0, "minutes"
        
    #Save eigenvalues for printing:
    posterior.exportU(Vh[PARAMETER], "hmisfit/evect_point.pvd")
    np.savetxt("hmisfit/eigevalues_point.dat", d)
    
    print sep, "THE END: ",(time.time() - start) / 60.0, "minutes" , sep
