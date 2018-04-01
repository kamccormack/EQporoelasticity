import sys
sys.path.append('/home/fenics/shared/local_lib')
import petsc4py as p4p
from petsc4py.PETSc import Mat
from dolfin import *
from hippylib import *
import scipy.sparse as sp
from scipy.spatial.distance import pdist, squareform
import logging
import numpy as np
import matplotlib.pyplot as plt
import openpyxl as pyxl
import time
from block import *
from block.dolfin_util import *
from block.iterative import *
from block.algebraic.petsc import *
import time

parameters.reorder_dofs_serial = False

"""
TO DO:

print out misfit, cost, reg
build matern type reg matrix.... maybe not?

"""
dir_path = '/home/fenics/shared/'

def u_boundary(x, on_boundary):
    return on_boundary

def GPSlatlon2XY(data_sheet, origin, theta):
    """
    latlon2XY - converts from [longitude, latitude] to user defined XY coordinate system for importing data
    into numerical models.

    INPUTS:
    [data_sheet] = excel sheet containing data. Default is to ignore the first row (containing headers).
    [origin] = origin of XY coordinate system in [lon,lat] format - used as point of rotation
    [theta] = rotation angle in radians between line of latitude and fault trace. Positive theta rotates
                counter-clockwise from latlon to XY

    Assumes first column is longitude and second column is latitude

    OUTPUTS:
    Returns transformed coordinates as numpy vectors X, Y in kilometers
    """
    
    lon = np.array([[data_sheet.cell(row=i, column=1).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    lat = np.array([[data_sheet.cell(row=i, column=2).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    
    lon_u = np.array([[data_sheet.cell(row=i, column=5).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    lat_u = np.array([[data_sheet.cell(row=i, column=6).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    Uz = np.array([[data_sheet.cell(row=i, column=4).value] for i in range(2, data_sheet.max_row + 1)]).reshape(
        data_sheet.max_row - 1, )
    
    lon_in_km = (lon - origin[0]) * 111 * np.cos(lat * np.pi / 180)
    lat_in_km = (lat - origin[1]) * 111
    
    rho_u = np.sqrt(np.power(lon_u, 2) + np.power(lat_u, 2))
    theta_new_u = np.arctan2(lat_u, lon_u) - theta
    
    rho = np.sqrt(np.power(lon_in_km, 2) + np.power(lat_in_km, 2))
    theta_new = np.arctan2(lat_in_km, lon_in_km) - theta
    
    X, Y = rho * np.cos(theta_new), rho * np.sin(theta_new)
    Ux, Uy = rho_u * np.cos(theta_new_u), rho_u * np.sin(theta_new_u)
    
    return 1e3 * X, 1e3 * Y, 1e-3*Ux, 1e-3*Uy, 1e-3*Uz

def checkerboard(coords_b, size):
	X, Y = coords_b[:,0], coords_b[:,1]
	check_slip = np.zeros(X.shape[0],)
	xblocks = (np.round(X.max()/ size)/2 + 1).astype(int)
	yblocks = (np.round(Y.max() / size)/2 + 1).astype(int)
	for i in range (0, xblocks):
		for j in range(0, yblocks):
			checkdof = np.where((X > 2*size*i) & (X < 2*size*i + size)
			                    & (Y > 2*size*j + size) & (Y < 2*size*j + 2*size))[0]

			check_slip[checkdof] = -1.0
	for i in range (0, xblocks):
		for j in range(0, yblocks):
			checkdof = np.where((X > 2*size*i +size) & (X < 2*size*i + 2*size)
			                    & (Y > 2*size*j) & (Y < 2*size*j + size))[0]

			check_slip[checkdof] = -1.0

	return check_slip

# Class HessianOperator to perform Hessian apply to a vector
class HessianOperator():
    cgiter = 0
    def __init__(self, model, use_gaussnewton=True):
        self.R = R
        self.Raa = Raa
        self.C = C
        self.A = A
        self.adj_A = adj_A
        self.W = W
        self.Wua = Wua
        self.use_gaussnewton = use_gaussnewton

        # incremental state
        self.du = Vector()
        self.A.init_vector(self.du,0)

        #incremental adjoint
        self.dp = Vector()
        self.adj_A.init_vector(self.dp,0)

        # auxiliary vectors
        self.CT_dp = Vector()
        self.C.init_vector(self.CT_dp, 1)
        self.Wua_du = Vector()
        self.Wua.init_vector(self.Wua_du, 1)

    def init_vector(self, v, dim):
        self.R.init_vector(v,dim)

    # Hessian performed on x, output as generic vector y - Where does this get used???
    def mult(self, v, y):
        self.cgiter += 1
        y.zero()
        if self.use_gaussnewton:
            self.mult_GaussNewton(v,y)
        else:
            self.mult_Newton(v,y)

    # define (Gauss-Newton) Hessian apply H * v
    def mult_GaussNewton(self, v, y):

        #incremental forward
        rhs = -(self.C * v)
        #bc_adj.apply(rhs)
        [bc.apply(rhs) for bc in bc_adj]
        solve (self.A, self.du, rhs)

        #incremental adjoint
        rhs = - (self.W * self.du)
        #bc_adj.apply(rhs)
        [bc.apply(rhs) for bc in bc_adj]
        solve (self.adj_A, self.dp, rhs)

        # Reg/Prior term
        self.R.mult(v,y)

        # Misfit term
        self.C.transpmult(self.dp, self.CT_dp)
        y.axpy(1, self.CT_dp)

    # define (Newton) Hessian apply H * v
    def mult_Newton(self, v, y):

        #incremental forward
        rhs = -(self.C * v)
        #bc_adj.apply(rhs)
        [bc.apply(rhs) for bc in bc_adj]
        solve (self.A, self.du, rhs)

        #incremental adjoint
        rhs = -(self.W * self.du) -  self.Wua * v
        #bc_adj.apply(rhs)
        [bc.apply(rhs) for bc in bc_adj]
        solve (self.adj_A, self.dp, rhs)

        #Reg/Prior term
        self.R.mult(v,y)
        y.axpy(1., Raa*v)

        #Misfit term
        self.C.transpmult(self.dp, self.CT_dp)
        y.axpy(1., self.CT_dp)
        self.Wua.transpmult(self.du, self.Wua_du)
        y.axpy(1., self.Wua_du)

class Elasticity:
    def __init__(self, mesh, boundaries, Vh, Vh1, u0true, targets, u_gps, invGnoise, point_var, synthetic, matern_reg):
        """
        Construct a model by providing
        - the mesh
        - the finite element spaces for the STATE/ADJOINT variable and the PARAMETER variable
        - the Prior information
        """
        self.results_path = 'results_deterministic/numpy_results/'
        self.var_path = 'results_deterministic/numpy_variables/'
        self.paraviewpath = 'results_deterministic/paraview/'

        if not os.path.exists(self.results_path):
            os.makedirs(self.results_path)
        if not os.path.exists(self.var_path):
            os.makedirs(self.var_path)
        if not os.path.exists(self.paraviewpath):
            os.makedirs(self.paraviewpath)


        self.mesh = mesh
        self.boundaries = boundaries
        self.Vh = Vh
        self.Vh1 = Vh1
        self.targets = targets
        self.Vdim = self.Vh1.dim()
        self.u, self.u0, self.p = TrialFunction(self.Vh[STATE]), \
                                  TrialFunction(self.Vh[PARAMETER]), TrialFunction(self.Vh[ADJOINT])
        self.u_test, self.u0_test, self.p_test = TestFunction(self.Vh[STATE]), \
                                                 TestFunction(self.Vh[PARAMETER]), TestFunction(self.Vh[ADJOINT])
        self.ds = Measure('ds', domain=self.mesh, subdomain_data=self.boundaries)
        
        ################ PARAMETERS ###############
        self.synthetic = synthetic
        self.point_variance = point_var
        self.matern_reg = matern_reg
        self.invGnoise = invGnoise
        self.correlation_length = 2e4 # length over which points share 10% correlation
        self.corr = -np.log(.1)/self.correlation_length #regularization parameter
        self.area = assemble(Constant(1.)*self.ds(2))
        self.h = CellSize(self.mesh)


        self.delta_b = 2e-6 * self.h  # Dirichlet control regularization on bottom boundary
        self.gamma = 5e2 # Laplace regularization constant
        self.beta = 1e-9 # mass matrix multiplier in reg.

        #self.delta_b = 15e-5 * self.h  # Dirichlet control regularization on bottom boundary
        #self.gamma = 2e2 # Laplace regularization constant
        #self.beta = 1e-9 # mass matrix multiplier in reg.

        print "Gamma {0:5g}; Beta {1:5g}".format(self.gamma, self.beta)

        # self.delta_b = 5e-6 * self.h  # Dirichlet control regularization on bottom boundary
        # self.gamma = 4e3 # Laplace regularization constant
        # self.beta = 5e-9 # mass matrix multiplier in reg.

        self.extract_coords()
        #self.E = Expression('347 * pow(-x[2], 0.516) + 5.324e4', degree = 1)
        self.E = Expression('500 * pow(-x[2], 0.516) + 2e4', degree = 1)

        self.nu = 0.4  # Poisson's ratio
        self.mu   = self.E / (2.0*(1.0 + self.nu)) # shear modulus
        self.lmbda = self.E*self.nu / ((1.0 + self.nu)*(1.0 - 2.0*self.nu))# Lame's parameter
        
        
        # Initialize Expressions
        self.dim = 3
        self.f = Constant(('0.0','0.0','0.0'))
        self.I = Identity(self.dim)
        self.n = FacetNormal(mesh)
        self.T = self.I - outer(self.n,self.n)

        # Set boundary conditions
        self.b0 = Constant(0.0)
        self.ud1 = Constant((0.0, 0.0, 0.0))
        self.bc1 = DirichletBC(self.Vh[STATE].sub(0), self.b0, self.boundaries, 6) # back
        self.bc2 = DirichletBC(self.Vh[STATE].sub(2), self.b0, self.boundaries, 7) # mantle
        self.bcall = DirichletBC(self.Vh[STATE], self.ud1, u_boundary) # adj var = 0 on all boundaries
        self.bcstate = [self.bc1, self.bc2] # u = 0 on sides and bottom for state eq. and everywhere for adjoint
        self.bcadj = [self.bc1, self.bc2]

        #coords = Vh[STATE].tabulate_dof_coordinates().reshape((Vh[STATE].dim(), -1))
        #self.B = self.build_B(coords, targets[:,0], targets[:,1])
        self.B = assemblePointwiseObservation(Vh[STATE],targets)
        self.invG = self.build_invG(self.invGnoise)

        if not self.point_variance:
            self.noise_variance = 1e-4

        bcprior = DirichletBC(Vh[PARAMETER], (1., 0., 0.), boundaries, 3)  # fault surface
        priortest = Function(Vh[PARAMETER])
        bcprior.apply(priortest.vector())
        self.dof_slab = np.where(priortest.vector() == 1.)[0]#.tolist()
        self.coords_all = Vh[STATE].tabulate_dof_coordinates().reshape((Vh[STATE].dim(), -1))
        self.coords_b = self.coords_all[self.dof_slab, :]

        check_slip = checkerboard(self.coords_b, 40e3)
        check_slip_func = Function(FunctionSpace(mesh, 'Lagrange', 1))
        check_slip_func.vector()[:] = 0.0
        check_slip_func.vector()[self.dof_slab] = check_slip
        self.check_slip = Expression(('check', '0.0', '-1.0*check'), check= check_slip_func, degree = 1)

        #self.u0true = u0true
        self.u0true = self.check_slip
 
        self.assemble_system()
        print (time.time() - start) / 60.0, "minutes"

        self.AA = block_mat([[self.E_state, self.C_state,       0      ],   # u
                             [    0       , self.RR_grad,   self.C_grad],   # v
                             [self.Wuu    ,      0      ,    self.E_adj]])  # u0

        if self.synthetic:
            self.u_obs = self.computeObservation()
            #quit()
        if not self.synthetic:
            self.u_obs = Vector()
            self.B.init_vector(self.u_obs, 0)
            self.u_obs.set_local(u_gps) #BUG

        self.assemble_rhs()
        self.bb = block_vec([ 0, 0, self.rhs_adjoint])
        print "obs max, min: ", self.u_obs.max(), self.u_obs.min()

        Eps = MumpsSolver(self.E_state)
        Epa = MumpsSolver(self.E_adj)
        #Rp = MumpsSolver(self.RR_grad)
        #Rp = ILU(self.RR_grad) #LU solver works too (sometimes)
        Rp = Jacobi(self.RR_grad)
        #Eps = Jacobi(self.E_state)
        #Epa = Jacobi(self.E_adj)

        #Rp = AMG(self.RR_grad) #LU solver works too (sometimes)
        #Eps = AMG(self.E_state)
        #Epa = AMG(self.E_adj)

        
        # I = assemble(inner(self.u0, self.u0_test)*dx)
        # Ip = LumpedInvDiag(I)

        #Ip = LumpedInvDiag(self.RR_grad)

        # # Create preconditioners:
        AApre = block_mat([[Eps, 0, 0 ],
                           [0, Rp, 0 ],
                           [0, 0, Epa]])


        # # Create the block inverse, using the preconditioned Minimum Residual method
        # # (suitable for symmetric indefinite problems).
        #self.AAinv = ConjGrad(self.AA, precond = AApre, tolerance = 1e-10, maxiter = 500, show = 2)
        #self.AAinv = TFQMR(self.AA, precond = AApre, tolerance = 1e-10, maxiter = 500, show = 2)
        self.AAinv = BiCGStab(self.AA, precond = AApre)
        #self.AAinv = LGMRES(self.AA, precond = AApre)

    def extract_coords(self):
        if not os.path.exists(self.var_path):
            os.makedirs(self.var_path)

        self.coordinates = self.Vh1.tabulate_dof_coordinates().reshape((self.Vdim, -1))
        self.x_all, self.y_all, self.z_all = self.coordinates[:, 0], self.coordinates[:, 1], self.coordinates[:, 2]

        bctest0 = DirichletBC(self.Vh1, (1, 1, 1), self.boundaries, 1)  # ocean surface
        bctest1 = DirichletBC(self.Vh1, (2, 2, 2), self.boundaries, 2)  # top surface
        bctest2 = DirichletBC(self.Vh1, (3, 3, 3), self.boundaries, 3)  # slab surface

        ptest = Function(self.Vh1)
        bctest0.apply(ptest.vector())
        bctest1.apply(ptest.vector())
        bctest2.apply(ptest.vector())

        self.ocean_dofs = np.where(ptest.vector() == 1)[0]
        self.surface_dofs = np.where(ptest.vector() == 2)[0]
        self.slab_dofs = np.where(ptest.vector() == 3)[0]



        self.size_gps = self.targets.shape[0]
        size_w_gps = 3 * self.size_gps
        indices_gps = np.empty((self.size_gps, 3))
        for j in range(0, self.size_gps):
            indice_gps = np.where(np.sqrt(
                pow(self.x_all[:] - self.targets[j,0], 2) + pow(self.y_all[:] - self.targets[j,1], 2) + pow(
                    self.z_all[:] - 0.0, 2)) < 0.1)[0]
            indices_gps[j, :] = indice_gps
            self.GPS_dofs = indices_gps.reshape((size_w_gps), order = 'F')
        # print indice_gps

        XYZ_slab = np.vstack((self.x_all[self.slab_dofs], self.y_all[self.slab_dofs], self.z_all[self.slab_dofs])).T

        np.save(self.var_path + "XYZ_slab", XYZ_slab)
        np.save(self.var_path + "x_all_3D", self.x_all)
        np.save(self.var_path + "y_all_3D", self.y_all)
        np.save(self.var_path + "z_all_3D", self.z_all)
        np.save(self.var_path + "GPS_dofs", self.GPS_dofs)
        np.save(self.var_path + "surface_dofs_u_3D", self.surface_dofs)
        np.save(self.var_path + "ocean_dofs_u_3D", self.ocean_dofs)
        np.save(self.var_path + "slab_dofs_u_3D", self.slab_dofs)

    def build_B(self, coords, X, Y):
    
        #coords = np.around(coords)
        sizeU = coords.shape[0]
        nobs = X.shape[0]

        Bp4p = p4p.PETSc.Mat()
        Bp4p.create(p4p.PETSc.COMM_WORLD)
        Bp4p.setSizes([3 * nobs, sizeU])
        Bp4p.setType('aij')  # sparse
        Bp4p.setPreallocationNNZ(5)
        Bp4p.setUp()
        Istart, Iend = Bp4p.getOwnershipRange()


        for j in range(0, nobs):
            ind = np.where(np.sqrt(pow((coords[:, 0] - X[j]), 2) + pow((coords[:, 1] - Y[j]), 2) + pow((coords[:, 2]), 2)) < 2.0)[0]
            Bp4p[j, ind[0]], Bp4p[j + nobs, ind[1]], Bp4p[j + 2 * nobs, ind[2]] = 1.0, 1.0, 1.0

        Bp4p.assemblyBegin()
        Bp4p.assemblyEnd()

        B = PETScMatrix(Bp4p)
        #B = Matrix(PETScMatrix(Bp4p))

        
        return B

    def build_invG(self, invGnoise):
    

        invGp4p = p4p.PETSc.Mat()
        invGp4p.create(p4p.PETSc.COMM_WORLD)
        invGp4p.setSizes([invGnoise.shape[0], invGnoise.shape[1]])
        invGp4p.setType('aij')  # sparse
        #invGp4p.setPreallocationNNZ(5)
        invGp4p.setUp()
    
        for j in range(0, invGnoise.shape[0]):
            
            invGp4p[j, j] = invGnoise[j, j]
    
        invGp4p.assemblyBegin()
        invGp4p.assemblyEnd()
    
        invG = PETScMatrix(invGp4p)
    
        return invG

    def build_Cov_R(self):
        
        bcprior = DirichletBC(self.Vh[PARAMETER], (1., 1., 1.), self.boundaries, 3)  # land surface
        priortest = Function(self.Vh[PARAMETER])
        bcprior.apply(priortest.vector())
        self.dof_slab_all = np.where(priortest.vector() == 1.)[0]#.tolist()

        rho = squareform(pdist(self.coords_b[:, :3], 'euclidean'))
        R = np.exp(-self.corr * rho)
        
        print rho[0:10, 1]
        print R[0:10,1]
        quit()
        Rinv = np.linalg.inv(R)
        RR = np.empty((self.dof_slab_all.shape[0], self.dof_slab_all.shape[0]))
        RI = np.identity(3)
   
        for i in range(0,R.shape[0]):
            RRi = np.kron(RI, Rinv[i,:])
            RR[3*(i+1)-3:3*(i+1),:] = RRi
            
        
        RRall = np.empty((self.coords_all.shape[0], self.coords_all.shape[0]))
        RRall[self.dof_slab_all,:][:, self.dof_slab_all] = RR
        RRall += assemble(self.beta * inner(self.u0, self.u0_test)*dx).array() #add mass matrix
        
        
        RRp4p = p4p.PETSc.Mat()
        RRp4p.create(p4p.PETSc.COMM_WORLD)
        RRp4p.setSizes([RRall.shape[0], RRall.shape[1]])
        RRp4p.setType('dense')
        RRp4p.setUp()
        
        RRp4p[:,:] = RRall[:,:] #bug here?

        RRp4p.assemblyBegin()
        RRp4p.assemblyEnd()
        RR_block = Matrix(PETScMatrix(RRp4p))

        return RR_block
 
    def build_Wuu(self):

        Bm = as_backend_type(self.B).mat()
        BT = Bm.transpose(Mat())
   
        if self.point_variance:
            BtG_mat = BT.matMult(self.invG.mat())
            BtGB_mat = BtG_mat.matMult(Bm)
            BtGB = Matrix(PETScMatrix(BtGB_mat))
            Wuu = BtGB #* du
            self.BtG = Matrix(PETScMatrix(BtG_mat))

        else:
            BtB_mat = BT.matMult(Bm)
            BtB = Matrix(PETScMatrix((1.0 / self.noise_variance) * BtB_mat))
            Wuu = BtB
            self.BtG = Matrix(PETScMatrix((1.0 / self.noise_variance) * BT))
            
        return Wuu

    def strain(self, w):  # strain = 1/2 (grad u + grad u^T)
        return sym(nabla_grad(w))
    
    def elasticity(self, u, v): # elasticity weak form with slip boundary conditions
        return inner(2.0*self.mu*self.strain(u), self.strain(v))*dx + inner(self.lmbda*nabla_div(u), nabla_div(v))*dx\
               + (1.0/self.delta_b)*inner(self.T*u ,v)*self.ds(3) \
               + dot(u,self.n)*dot(v,self.n)*self.ds(3)

    def constraintB(self, u, v): # slip boundary condition
        return (1.0/self.delta_b)*inner(self.T*u, v)*self.ds(3)

    def regularization(self, u, v): # regularization
        varf = self.gamma * inner(nabla_grad(u), nabla_grad(v)) * self.ds(3) \
               + self.beta * inner(u,v)*dx #tiny mass matrix to make block invertable

        return varf

    def regularizationTV(self, u, v): # regularization
        eps = 1e-9
        varf = self.gamma * (inner(nabla_grad(u), nabla_grad(v)) + eps) * self.ds(3) \
                + self.beta * inner(u,v)*dx #tiny mass matrix to make block invertable

        return varf

    def assemble_system(self):

        self.Wuu = self.build_Wuu()

        self.E_state = assemble(self.elasticity(self.u, self.p_test))
        [bc.apply(self.E_state) for bc in self.bcstate]

        self.E_adj = assemble(self.elasticity(self.p, self.u_test))
        [bc.apply(self.E_adj) for bc in self.bcadj]

        self.C_state = assemble(-self.constraintB(self.u0, self.p_test))
        [bc.zero(self.C_state) for bc in self.bcstate]

        # Take transpose of C_state for C_grad
        self.C_grad = Transpose(self.C_state)
        
        if matern_reg:
            self.RR_grad = self.build_Cov_R() #+ assemble(self.beta * inner(self.u0, self.u0_test)*dx)
            
        else:
            u0 = Function(self.Vh[PARAMETER])
            u0_test = Function(self.Vh[PARAMETER])
            RR = assemble(self.regularization(self.u0, self.u0_test))
            #RR_TV = assemble(self.regularizationTV(self.u0, self.u0_test))
            self.RR_grad = RR

    def assemble_rhs(self):
        self.rhs_adjoint = self.BtG * self.u_obs

    def generate_vector(self, component="ALL"):
        """
        Return the list x=[u,a,p] where:
        - u is any object that describes the state variable
        - a is a Vector object that describes the parameter.
          (Need to support linear algebra operations)
        - p is any object that describes the adjoint variable
        
        If component is STATE, PARAMETER, or ADJOINT return x[component]
        """
        if component == "ALL":
            x = [Vector(), Vector(), Vector()]
            self.B.init_vector(x[STATE],1)
            self.R.init_vector(x[PARAMETER],0) # I think this initializing a vetor to be compatible with the R prior matrix in a given dimension
            self.B.init_vector(x[ADJOINT], 1)
        elif component == STATE:
            x = Vector()
            self.B.init_vector(x,1)
        elif component == PARAMETER: # This is compatible with the prior matrix bc the the gradient needs the
            x = Vector()
            self.R.init_vector(x,0)
        elif component == ADJOINT:
            x = Vector()
            self.B.init_vector(x,1)
            
        return x

    def computeObservation(self):
        """
        Compute the syntetic observation
        """
        
        u_o = Vector()
        A = self.E_state
        b = assemble(self.constraintB(self.u0true, self.p_test))
        [bc.apply(A) for bc in self.bcstate]
        [bc.apply(b) for bc in self.bcstate]

        A.init_vector(u_o, 1)
        solver = LinearSolver('mumps')
        solver.solve(A, u_o, b)
        self.u0_fwdsolve = u_o

        # Create noisy data, ud
        MAX = u_o.norm("linf")
        noise = .01 * MAX * np.random.normal(0, 1, len(u_o.array()))
        u_o.set_local(u_o.array())# + noise)

        u_obs = self.B * u_o  # check that this is no ZERO

        self.paraviewpath = 'results_deterministic/'
        self.u0file = File(self.paraviewpath + 'u_true.pvd')
        self.u0file << vector2Function(self.u0_fwdsolve, Vh[STATE], name = "true displacement")
        
        return u_obs

    def cost(self, x):
        """
        Given the list x = [u,a,p] which describes the state, parameter, and
        adjoint variable compute the cost functional as the sum of 
        the misfit functional and the regularization functional.
        
        Return the list [cost functional, regularization functional, misfit functional]
        
        Note: p is not needed to compute the cost functional
        """        
        assert x[STATE] != None
                       
        diff = self.B*x[STATE]
        diff -= self.u_obs

        if self.point_variance:
            Gdiff = self.invG*diff
            misfit = 0.5 * diff.inner(Gdiff)

        else:
            misfit = (.5 / self.noise_variance) * diff.inner(diff)

        rx = x[PARAMETER]
        reg = 0.5 * rx.inner(self.RR_grad * rx)

        c = misfit + reg
        
        return c, reg, misfit

    def solveFwd(self, out, x, tol=1e-9):
        """
        Solve the forward problem.
        """
        A, b = self.assembleA(x, assemble_rhs = True)
        A.init_vector(out, 1)
        #solver = PETScKrylovSolver("cg", amg_method()) # try linear solver
        solver  = PETScLUSolver("umfpack")
        #solver.parameters["relative_tolerance"] = tol
        solver.set_operator(A)
        nit = solver.solve(out,b)
        #print "FWD", (self.A*out - b).norm("l2")/b.norm("l2"), nit

    def solveAdj(self, out, x, tol=1e-9):
        """
        Solve the adjoint problem.
        """
        At, badj = self.assembleA(x, assemble_adjoint = True,assemble_rhs = True)
        At.init_vector(out, 1)
        
        solver  = PETScLUSolver("umfpack")

        #solver = PETScKrylovSolver("cg", amg_method())
        #solver.parameters["relative_tolerance"] = tol
        solver.set_operator(At)
        nit = solver.solve(out,badj)
        
        #print "ADJ", (self.At*out - badj).norm("l2")/badj.norm("l2"), nit

################### SET UP MESH, MODEL PARAMETERS AND INPUT DATA ##########################################

if __name__ == "__main__":
    set_log_active(False)

    start = time.time()

    sep = "\n"+"#"*80+"\n"
    print sep, "Set up the mesh and finite element spaces: ",(time.time() - start) / 60.0, "minutes", sep
    
    ############ IMPORT MESH (made with gmsh) #######################

    path = "/home/fenics/shared/meshes"
    mesh_size = 'med_inv'  # fine, med, coarse

    mesh = Mesh(path + '/3D_' + mesh_size + '.xml')
    boundaries = MeshFunction("size_t", mesh, path + '/3D_' + mesh_size + '_facet_region.xml')

    Vh1 = VectorFunctionSpace(mesh, 'Lagrange', 1)
    Vh = [Vh1, Vh1, Vh1]
    print "Number of dofs: STATE={0}, PARAMETER={1}, ADJOINT={2}".\
        format(Vh[STATE].dim(), Vh[PARAMETER].dim(), Vh[ADJOINT].dim())

    dim = 3
    udim = Vh[STATE].dim()

    ################ USER INPUTS ###############

    point_var = True
    synthetic = False
    matern_reg = False
    
    sigmab = 2e4
    xcenter = 80e3
    ycenter = 130e3
    u0d, u0s = -4.0, -0.0

    ################ DATA ###############


    print sep, "Set up the location of observation, Prior Information, and model: ",(time.time() - start) / 60.0, "minutes", sep
    nobs = 37

    #origin = np.array([-85.21, 8.64])  # origin of top surface of mesh - used as point of rotation
    #theta = 0.733  # angle of rotation between latitude line and fault trace
    origin = np.array([-85.1688, 8.6925])  # origin of top surface of mesh - used as point of rotation
    theta = 0.816  # angle of rotation between latitude line and fault trace (42 deg)

    data_file = dir_path + 'data/Data.xlsx'
    wb = pyxl.load_workbook(data_file)
    gps_data = wb['GPS_data']

    X_gps, Y_gps, Ux_gps, Uy_gps, Uz_gps = GPSlatlon2XY(gps_data, origin, theta)
    Z_gps = np.zeros(len(X_gps))

    targets = np.concatenate((X_gps.reshape(nobs,1), Y_gps.reshape(nobs,1), Z_gps.reshape(nobs,1)),axis=1)

    ##### Use my B operator
    #u_gps = np.concatenate((Ux_gps.reshape(nobs,1), Uy_gps.reshape(nobs,1),
    #                        Uz_gps.reshape(nobs,1)),axis=0).reshape(3*nobs,)# order="F")

    ####Use hippylib B operator
    u_gps = np.concatenate((Ux_gps.reshape(nobs, 1), Uy_gps.reshape(nobs, 1),
                            Uz_gps.reshape(nobs, 1)), axis = 1).reshape(3 * nobs, )  # order="F")

    print "Number of observation points: {0}".format(targets.shape[0])

    ##### NOISE MATRIX #####
    stddev = .001 * np.array([4.4, 1.7, 0.8, 1.2, 0.8, 1.7, 1.3, 0.9, 1., 0.8, 1.3, 0.8, 1.9, 0.8, 0.8, 0.8,
                              0.8, 1., 6.5, 4.6, 6.5, 7.8, 5.8,4.5, 6.2, 4.8, 7.5, 5.1,5.3, 6.3, 6.6, 6.3, 4.6,
                              5.1, 6.4, 5.2, 5.2, 5.1, 1.6, 0.7, 0.9, 0.7, 1.2, 0.9, 0.6, 0.9, 0.7, 1, 0.6, 1.7,
                              0.6, 0.7, 0.7, 0.7, 0.8,2.5, 2.6, 2.5, 4., 2.2, 2, 1.9, 2.3, 2.3, 2.2, 2.1, 3, 2.7,
                              2.5, 2.1, 2.1, 2.6, 2.3, 2.1, 15.7, 6.6, 2.8, 4.3,2.8, 5.3, 4.1, 2.7, 4., 3.8,4.4,
                              2.6, 8.2, 2.6, 3.2, 2.8, 3., 3.5, 11.4, 8.9, 10.3, 15.1, 10., 8.4, 9.9, 8.6, 10.6,
                              9.4, 8.7, 15., 10.7, 10,8, 8.3, 10.5, 9.1, 8.4])
    stddev = stddev.reshape((nobs, 3), order = 'F').reshape(3 * nobs, ) #Use hippylib B operator
    Gnoise = np.diag(stddev ** 2)
    invGnoise = np.linalg.inv(Gnoise)

    if synthetic:
        u0true = Expression(('u0d * exp(-(pow((x[0] - xcenter),2)/(pow(sigma,2)) + (pow((x[1] - \
    ycenter),2)/(2*pow(sigma,2))))) ', 'u0s * exp(-(pow((x[0] - xcenter),2)/(pow(sigma,2)) + \
    (pow((x[1] - ycenter),2)/(2*pow(sigma,2)))))', '0.0'), sigma=sigmab,xcenter=xcenter, ycenter=ycenter, u0d=u0d, u0s=u0s, degree=1)
    else:
        u0true = Expression(('0.0 ', '0.0', '0.0'), degree=1)

    ############### SOLVE #####################

    model = Elasticity(mesh, boundaries, Vh, Vh1, u0true, targets, u_gps, invGnoise, point_var, synthetic, matern_reg)

    sol_u, sol_u0, sol_v = model.AAinv * model.bb

    np.save(model.results_path + 'u0_slab_array', sol_u0[model.slab_dofs])
    np.save(model.results_path + 'u0_array', sol_u0)
    np.save(model.results_path + 'u_array', sol_u)

    total_cost, reg_cost, misfit_cost = model.cost([sol_u, sol_u0, sol_v])
    print "Total cost {0:5g}; Reg Cost {1:5g}; Misfit {2:5g}".format(total_cost, reg_cost, misfit_cost)

    print "state max/min: ", sol_u.array().max(), sol_u.array().min()
    print "u0 max/min: ", sol_u0.array().max(), sol_u0.array().min()

    u0 = vector2Function(sol_u0, Vh[PARAMETER])
    u_adj = vector2Function(sol_v, Vh[ADJOINT])
    ufwd = vector2Function(sol_u, Vh[STATE])

    File(model.paraviewpath + "sol_forward.pvd") << ufwd
    File(model.paraviewpath+"sol_adjoint.pvd") << u_adj
    File(model.paraviewpath+"u0_inversion.pvd") << u0

    output_file = HDF5File(mesh.mpi_comm(), model.results_path + 'u0slab.h5', 'w')
    output_file.write(u0, 'slip')
    output_file.write(ufwd, 'ufwd')
    output_file.write(mesh, 'mesh')
    output_file.close()
    print sep, "Solution saved as function ",(time.time() - start) / 60.0, "minutes", sep


    ############### HESSIAN APPLY #####################

    #Hess_Apply = HessianOperator(model, use_gaussnewton = (iter < 6))



    print sep, "ALL DONE!: ",(time.time() - start) / 60.0, "minutes", sep
    quit()
